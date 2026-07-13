"""
export_uat_report.py v1.0
UAT 測試報告匯出工具
將 Ragic 週期採購各表單資料匯出為 Excel UAT 報告
"""
import argparse
import logging
import sys
import os
from datetime import datetime

sys.path.insert(0, os.path.dirname(__file__))
from ragic_client import RagicClient

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    raise ImportError("請先安裝 openpyxl：pip install openpyxl")

VERSION = "v1.0"
ICON = "📄"

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

UAT_CASES = [
    {"編號": "TC-001", "功能": "週期設定", "測試項目": "每月週期產生批次", "預期結果": "成功產生1筆批次", "狀態": ""},
    {"編號": "TC-002", "功能": "週期設定", "測試項目": "雙週週期產生批次", "預期結果": "成功產生1筆批次", "狀態": ""},
    {"編號": "TC-003", "功能": "批次管理", "測試項目": "防止重複產生批次", "預期結果": "重複操作顯示錯誤", "狀態": ""},
    {"編號": "TC-004", "功能": "請購作業", "測試項目": "請購數量預設0", "預期結果": "所有明細數量=0", "狀態": ""},
    {"編號": "TC-005", "功能": "請購作業", "測試項目": "空白數量無法送出", "預期結果": "阻止送出並顯示錯誤", "狀態": ""},
    {"編號": "TC-006", "功能": "請購作業", "測試項目": "逾期補填申請", "預期結果": "寫入異常紀錄", "狀態": ""},
    {"編號": "TC-007", "功能": "採購彙整", "測試項目": "彙整數量=各部門合計", "預期結果": "彙整量正確加總", "狀態": ""},
    {"編號": "TC-008", "功能": "採購彙整", "測試項目": "調整量≠需求量需填原因", "預期結果": "未填原因阻止轉採購單", "狀態": ""},
    {"編號": "TC-009", "功能": "採購作業", "測試項目": "主管退回請購單", "預期結果": "狀態改為退回並通知", "狀態": ""},
    {"編號": "TC-010", "功能": "驗收作業", "測試項目": "採購調整後驗收", "預期結果": "驗收數量可部分驗收", "狀態": ""},
    {"編號": "TC-011", "功能": "驗收作業", "測試項目": "驗收差異寫入異常紀錄", "預期結果": "差異原因必填且記錄", "狀態": ""},
    {"編號": "TC-012", "功能": "請款作業", "測試項目": "驗收異常授權放行請款", "預期結果": "異常紀錄存在才可請款", "狀態": ""},
    {"編號": "TC-013", "功能": "請款作業", "測試項目": "分攤金額合計=發票金額", "預期結果": "不符時差異原因必填", "狀態": ""},
    {"編號": "TC-014", "功能": "報表查詢", "測試項目": "月度進貨數量報表", "預期結果": "正確彙總當月驗收量", "狀態": ""},
    {"編號": "TC-015", "功能": "權限測試", "測試項目": "請購人員無法進入採購彙整", "預期結果": "無法看到彙整選單", "狀態": ""},
]


def export_uat_excel(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UAT測試案例"

    header = ["編號", "功能", "測試項目", "預期結果", "實際結果", "狀態", "測試人員", "測試日期", "備註"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E6DA4")

    for col, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    col_widths = [10, 15, 35, 35, 35, 10, 12, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for row_idx, case in enumerate(UAT_CASES, 2):
        ws.cell(row=row_idx, column=1, value=case["編號"])
        ws.cell(row=row_idx, column=2, value=case["功能"])
        ws.cell(row=row_idx, column=3, value=case["測試項目"])
        ws.cell(row=row_idx, column=4, value=case["預期結果"])
        ws.cell(row=row_idx, column=5, value="")
        ws.cell(row=row_idx, column=6, value="待測")
        ws.cell(row=row_idx, column=7, value="")
        ws.cell(row=row_idx, column=8, value="")
        ws.cell(row=row_idx, column=9, value="")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{len(UAT_CASES)+1}"
    wb.save(output_path)
    logger.info(f"{ICON} [{VERSION}] UAT 報表已匯出：{output_path}")


def main():
    parser = argparse.ArgumentParser(description=f"Ragic 週期採購 UAT 報表匯出 {VERSION}")
    parser.add_argument("--output", default="UAT_週期採購_測試報表.xlsx", help="輸出檔名")
    args = parser.parse_args()

    export_uat_excel(args.output)


if __name__ == "__main__":
    main()
