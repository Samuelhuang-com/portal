"""
create_import_excels.py v1.1
讀取真實料號明細表，生成 10 張 Ragic 表單匯入 Excel
- 01 料號主檔：合併兩份明細，去重後取聯集
- 02 料號對照表：日曜天地 + 春大直 各廠商 × 單價
- 03-10：以真實料號作為範例資料
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DOCS = os.path.join(os.path.dirname(__file__), '..', 'docs')
OUT  = os.path.join(os.path.dirname(__file__), '..', 'templates', 'import_forms')
os.makedirs(OUT, exist_ok=True)

NICHINYO = os.path.join(DOCS, '日曜天地設料號明細表.xlsx')
CHUNDARR = os.path.join(DOCS, '春大直設料號明細表.xlsx')

# ── 樣式工具 ─────────────────────────────────────────────
def _border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def make_header(ws, headers, col_widths, color='1F5C99'):
    hf   = Font(bold=True, color='FFFFFF', size=10)
    hfill= PatternFill('solid', fgColor=color)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hf; c.fill = hfill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _border()
        ws.column_dimensions[get_column_letter(i)].width = col_widths[i-1] if i<=len(col_widths) else 14
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = 'A2'

def add_rows(ws, rows, start=2):
    alt = PatternFill('solid', fgColor='EEF3FA')
    for ri, row in enumerate(rows, start):
        bg = alt if ri%2==0 else None
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _border()
            if bg: c.fill = bg
            c.alignment = Alignment(vertical='center')
    if rows:
        ws.auto_filter.ref = ws.dimensions

def note_sheet(wb, lines):
    ws = wb.create_sheet('★匯入說明')
    ws.column_dimensions['A'].width = 85
    for i, line in enumerate(lines, 1):
        c = ws.cell(row=i, column=1, value=line)
        ws.row_dimensions[i].height = 17
        if line.startswith('【'):
            c.font = Font(bold=True, color='1F5C99', size=11)
        elif line.startswith('▶') or line.startswith('注意'):
            c.font = Font(bold=True, color='C00000')
        elif line.startswith('  '):
            c.font = Font(color='555555')

# ── 讀取料號明細 ──────────────────────────────────────────
def read_items(path, company):
    """
    回傳 list of dict:
    {料號, 類別, 位置, 品名, 單位, 廠商, 單價, 最大庫存量, 最小庫存量, mini_order, 公司別}
    跳過 header 行和空料號行
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    items = []
    for sname in wb.sheetnames:
        if '編碼' in sname:
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        # 找 header 行（含「料號」欄）
        header_row = None
        for r in rows:
            if '料號' in r:
                header_row = r
                break
        if not header_row:
            continue
        # 建立欄位索引（跳過首個 None 欄）
        idx = {v: i for i, v in enumerate(header_row) if v is not None}
        get = lambda r, k: r[idx[k]] if k in idx and idx[k] < len(r) else None

        for r in rows[rows.index(header_row)+1:]:
            no = get(r, '料號')
            if not no or str(no).strip() == '':
                continue
            no = str(no).strip()
            name = get(r, '品名')
            if not name:
                continue  # 跳過無品名的行
            items.append({
                '料號':     no,
                '類別':     get(r, '類別') or '',
                '位置':     get(r, '位置') or '',
                '品名':     str(name).strip(),
                '單位':     get(r, '單位') or '',
                '廠商':     get(r, '廠商') or '',
                '單價':     get(r, '單價'),
                '最大庫存': get(r, '最大庫存量'),
                '最小庫存': get(r, '最小庫存量'),
                'mini_order': get(r, 'mini order ') or get(r, 'mini order'),
                '公司別':   company,
                'Sheet':    sname,
            })
    return items

print("讀取料號資料...")
items_n = read_items(NICHINYO, '日曜天地')
items_c = read_items(CHUNDARR, '春大直')
print(f"  日曜天地：{len(items_n)} 筆")
print(f"  春大直  ：{len(items_c)} 筆")

all_items = items_n + items_c

# ── 建主檔字典（料號→代表記錄，優先春大直）────────────────
master = {}
for it in items_c:
    master[it['料號']] = it
for it in items_n:
    if it['料號'] not in master:
        master[it['料號']] = it

master_list = sorted(master.values(), key=lambda x: x['料號'])
print(f"  合併料號主檔：{len(master_list)} 筆")


# ════════════════════════════════════════════════════════
# 01 週期採購料號主檔
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '主表'
make_header(ws,
    ['料號','品名','類別','位置','單位','預設數量','MOQ','是否啟用','是否週期品'],
    [12, 38, 20, 16, 8, 12, 10, 12, 12])
rows01 = []
for it in master_list:
    moq = it['mini_order'] if it['mini_order'] else ''
    rows01.append([
        it['料號'], it['品名'], it['類別'], it['位置'],
        it['單位'], 0, moq, 'Y', 'Y'
    ])
add_rows(ws, rows01)
note_sheet(wb, [
    '【01 週期採購料號主檔 匯入說明】',
    f'來源：日曜天地({len(items_n)}筆) + 春大直({len(items_c)}筆)，合併去重後共 {len(master_list)} 筆',
    '',
    '匯入步驟：新增表單 → 勾選「用既有Excel來建立表單」→ 上傳此檔(主表Sheet)',
    '表單名稱請填：週期採購料號主檔',
    '',
    '欄位型態設定（匯入後在 Design Mode 修改）：',
    '  料號       → Auto Number（或保留現有編碼，改為 Text）',
    '  品名       → Text，必填',
    '  類別       → Select（可從資料中自動建選項）',
    '  位置       → Text',
    '  單位       → Select',
    '  預設數量   → Number，預設值=0',
    '  MOQ        → Number，預設值=0',
    '  是否啟用   → Checkbox（Y=勾選）',
    '  是否週期品 → Checkbox（Y=勾選）',
    '',
    '▶ 注意：料號已沿用兩份明細表的原始編碼（E/C/G/S開頭）',
])
path01 = os.path.join(OUT, '01_週期採購料號主檔.xlsx')
wb.save(path01)
print(f"[OK] 01_週期採購料號主檔.xlsx  ({len(rows01)} 筆)")


# ════════════════════════════════════════════════════════
# 02 週期採購料號對照表
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '主表'
make_header(ws,
    ['對照ID','料號','公司別','廠商','供應商品號','供應商品名','單價'],
    [10, 12, 12, 16, 16, 38, 10])
rows02 = []
for it in sorted(all_items, key=lambda x: (x['料號'], x['公司別'])):
    rows02.append([
        '',          # 對照ID：留空
        it['料號'],
        it['公司別'],
        it['廠商'],
        '',          # 供應商品號：留空（可後續補填）
        it['品名'],
        it['單價'],
    ])
add_rows(ws, rows02)
note_sheet(wb, [
    '【02 週期採購料號對照表 匯入說明】',
    f'來源：日曜天地({len(items_n)}筆) + 春大直({len(items_c)}筆)，共 {len(rows02)} 筆對照',
    '',
    '此表記錄：同一料號在不同公司使用不同廠商/單價',
    '',
    '欄位型態設定：',
    '  對照ID     → Auto Number（留空自動產生）',
    '  料號       → Link → 週期採購料號主檔',
    '  公司別     → Select（日曜天地 / 春大直）',
    '  廠商       → Text（或 Link 廠商主檔）',
    '  供應商品號 → Text（可後續補填）',
    '  供應商品名 → Text（帶入廠商品名，Lookup或手填）',
    '  單價       → Number',
    '',
    '▶ 此表為 Workflow 自動帶入採購單價的來源',
    '▶ 同一料號可有多筆（不同公司或不同廠商）',
])
path02 = os.path.join(OUT, '02_週期採購料號對照表.xlsx')
wb.save(path02)
print(f"[OK] 02_週期採購料號對照表.xlsx  ({len(rows02)} 筆)")


# ── 取幾筆真實料號作後續表單範例 ─────────────────────────
sample_items_n = [x for x in master_list if x['公司別']=='春大直'][:6]
sample_items_all = master_list[:6]
def sno(i): return sample_items_all[i]['料號'] if i < len(sample_items_all) else 'E0101001'
def sname(i): return sample_items_all[i]['品名'][:12] if i < len(sample_items_all) else '品名'
def sunit(i): return sample_items_all[i]['單位'] if i < len(sample_items_all) else '個'
def sprice(i):
    p = sample_items_all[i]['單價'] if i < len(sample_items_all) else 100
    try: return float(p) if p else 0
    except: return 0


# ════════════════════════════════════════════════════════
# 03 週期採購週期設定
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '主表'
make_header(ws,
    ['週期ID','週期名稱','週期頻率','開放起始日','截止日','適用單位','狀態'],
    [10, 28, 14, 14, 14, 28, 10])
add_rows(ws, [
    ['','2026年06月每月週期','每月','2026-06-01','2026-06-10','日曜天地,春大直','啟用'],
    ['','2026年07月每月週期','每月','2026-07-01','2026-07-10','日曜天地,春大直','啟用'],
    ['','2026年Q3雙月週期', '每兩個月','2026-07-01','2026-07-15','春大直','啟用'],
])
note_sheet(wb, [
    '【03 週期採購週期設定 匯入說明】',
    '表單名稱：週期採購週期設定',
    '',
    '欄位型態設定：',
    '  週期ID     → Auto Number',
    '  週期頻率   → Select：每月 / 雙週 / 每兩個月 / 自訂',
    '  開放起始日/截止日 → Date',
    '  適用單位   → Multi-Select 或 Text（逗號分隔）',
    '  狀態       → Select：啟用 / 停用，預設=啟用',
    '',
    '▶ Action Button：「產生批次」→ 安裝 cycle_generate_batch.js',
])
wb.save(os.path.join(OUT, '03_週期採購週期設定.xlsx'))
print("[OK] 03_週期採購週期設定.xlsx")


# ════════════════════════════════════════════════════════
# 04 週期採購批次
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '主表'
make_header(ws,
    ['批次號','週期ID','批次名稱','開放日期','截止日期','是否已產生請購','狀態'],
    [12, 12, 28, 14, 14, 18, 10])
add_rows(ws, [
    ['','CYC001','2026年06月批次','2026-06-01','2026-06-10','N','開放'],
    ['','CYC002','2026年07月批次','2026-07-01','2026-07-10','N','開放'],
])
note_sheet(wb, [
    '【04 週期採購批次 匯入說明】',
    '表單名稱：週期採購批次',
    '',
    '欄位型態設定：',
    '  批次號           → Auto Number',
    '  週期ID           → Link → 週期採購週期設定',
    '  開放日期/截止日期 → Date',
    '  是否已產生請購   → Checkbox，預設=未勾（N）',
    '  狀態             → Select：開放 / 關閉 / 完成',
    '',
    '▶ Action Button：「產生各部門請購單」→ batch_generate_requests.js',
    '▶ Action Button：「產生彙整單」→ batch_generate_summary.js',
])
wb.save(os.path.join(OUT, '04_週期採購批次.xlsx'))
print("[OK] 04_週期採購批次.xlsx")


# ════════════════════════════════════════════════════════
# 05 週期採購請購單（主表 + 子表）
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active; ws1.title = '主表'
make_header(ws1,
    ['請購單號','批次號','公司別','請購部門','成本中心','會計科目',
     '請購總金額','狀態','送出人','送出時間','簽核人','簽核時間'],
    [12,12,12,14,14,14,14,10,12,18,12,18])
add_rows(ws1, [
    ['','BAT001','日曜天地','工務部','CC-E-NIC','E01',0,'待填','','','',''],
    ['','BAT001','日曜天地','清潔部','CC-C-NIC','C01',0,'待填','','','',''],
    ['','BAT001','春大直',  '工務部','CC-E-CHU','E01',0,'待填','','','',''],
    ['','BAT001','春大直',  '清潔部','CC-C-CHU','C01',0,'待填','','','',''],
])
ws2 = wb.create_sheet('子表_請購明細')
make_header(ws2,
    ['請購單號(關聯)','統購料號','品名','單位','會計科目','單價','請購數量','小計'],
    [16,14,38,8,14,10,12,12], color='2E6DA4')
detail_rows = []
for i in range(min(5, len(sample_items_all))):
    detail_rows.append(['REQ001', sno(i), sname(i), sunit(i), 'E01' if sno(i).startswith('E') else 'C01', sprice(i), 0, 0])
add_rows(ws2, detail_rows)
note_sheet(wb, [
    '【05 週期採購請購單 匯入說明】',
    '',
    '▶ 步驟1：用「主表」Sheet 匯入，表單名稱：週期採購請購單',
    '▶ 步驟2：Design Mode 新增子表「請購明細」，依「子表_請購明細」Sheet 建立欄位',
    '',
    '主表欄位設定：',
    '  請購單號     → Auto Number',
    '  批次號       → Link → 週期採購批次',
    '  公司別       → Select（日曜天地 / 春大直）',
    '  請購部門     → Select（或 Link ORG_UNIT）',
    '  成本中心     → Lookup（從部門帶入）',
    '  會計科目     → Lookup（從部門帶入）',
    '  請購總金額   → Formula: SUM(請購明細.小計)',
    '  狀態         → Select：待填/草稿/已送出/簽核中/已核准/退回/已彙整',
    '  送出時間/簽核時間 → DateTime',
    '',
    '子表「請購明細」欄位設定：',
    '  統購料號     → Link → 週期採購料號主檔',
    '  品名/單位/會計科目 → Lookup（從料號主檔自動帶入）',
    '  單價         → Lookup（從料號對照表，依公司別）',
    '  請購數量     → Number，預設值=0，必填',
    '  小計         → Formula: 單價 * 請購數量',
    '',
    '注意：Before Submit Workflow → request_validate_submit.js',
])
wb.save(os.path.join(OUT, '05_週期採購請購單.xlsx'))
print("[OK] 05_週期採購請購單.xlsx")


# ════════════════════════════════════════════════════════
# 06 週期採購彙整單
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '主表'
make_header(ws,
    ['彙整單號','批次號','料號','品名','廠商','需求總量','調整量','調整原因','狀態'],
    [12,12,14,36,16,12,12,28,12])
sum_rows = []
for i in range(min(4, len(sample_items_all))):
    it = sample_items_all[i]
    sum_rows.append(['', 'BAT001', it['料號'], it['品名'][:20], it['廠商'], 0, 0, '', '草稿'])
add_rows(ws, sum_rows)
note_sheet(wb, [
    '【06 週期採購彙整單 匯入說明】',
    '表單名稱：週期採購彙整單',
    '',
    '欄位型態設定：',
    '  彙整單號     → Auto Number',
    '  批次號       → Link → 週期採購批次',
    '  料號         → Link → 週期採購料號主檔',
    '  品名         → Lookup（從料號主檔帶入）',
    '  廠商         → Lookup（從料號對照表帶入）',
    '  需求總量/調整量 → Number',
    '  調整原因     → Text（調整量≠需求量時必填）',
    '  狀態         → Select：草稿 / 已轉採購單',
    '',
    '▶ 此表單通常由 batch_generate_summary.js 自動產生，匯入為測試用',
    '▶ Action Button：「轉採購單」→ summary_to_po.js',
])
wb.save(os.path.join(OUT, '06_週期採購彙整單.xlsx'))
print("[OK] 06_週期採購彙整單.xlsx")


# ════════════════════════════════════════════════════════
# 07 週期採購採購單（主表 + 子表）
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active; ws1.title = '主表'
make_header(ws1,
    ['採購單號','彙整單號','廠商','採購人員','預計到貨日','採購總金額','狀態'],
    [12,12,16,14,14,14,12])
# 取前3個不同廠商做範例
vendors = []
seen_v = set()
for it in sample_items_all:
    v = it['廠商']
    if v and v not in seen_v:
        vendors.append(it)
        seen_v.add(v)
    if len(vendors) >= 2: break
po_rows = []
for i, it in enumerate(vendors):
    po_rows.append(['', f'SUM00{i+1}', it['廠商'], '採購人員', '2026-06-20', 0, '草稿'])
add_rows(ws1, po_rows)

ws2 = wb.create_sheet('子表_採購明細')
make_header(ws2,
    ['採購單號(關聯)','料號','品名','單位','訂購數量','單價','小計'],
    [16,14,36,8,12,12,12], color='2E6DA4')
po_detail = []
for i in range(min(4, len(sample_items_all))):
    it = sample_items_all[i]
    p = sprice(i)
    po_detail.append(['PO001', it['料號'], it['品名'][:20], it['單位'], 0, p, 0])
add_rows(ws2, po_detail)
note_sheet(wb, [
    '【07 週期採購採購單 匯入說明】',
    '',
    '▶ 步驟1：用「主表」Sheet 匯入，表單名稱：週期採購採購單',
    '▶ 步驟2：Design Mode 新增子表「採購明細」',
    '',
    '主表欄位設定：',
    '  採購單號   → Auto Number',
    '  彙整單號   → Link → 週期採購彙整單',
    '  廠商       → Link → 廠商主檔（或 Text）',
    '  預計到貨日 → Date',
    '  採購總金額 → Formula: SUM(採購明細.小計)',
    '  狀態       → Select：草稿/已發出/部分到貨/完全到貨/取消',
    '',
    '子表「採購明細」欄位設定：',
    '  料號     → Link → 週期採購料號主檔',
    '  品名/單位 → Lookup',
    '  訂購數量/單價 → Number',
    '  小計     → Formula: 訂購數量 * 單價',
    '',
    '▶ Action Button：「產生驗收單」→ po_to_receiving.js',
])
wb.save(os.path.join(OUT, '07_週期採購採購單.xlsx'))
print("[OK] 07_週期採購採購單.xlsx")


# ════════════════════════════════════════════════════════
# 08 週期採購驗收單（主表 + 子表）
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active; ws1.title = '主表'
make_header(ws1,
    ['驗收單號','採購單號','驗收人員','驗收日期','狀態','備註'],
    [12,12,14,14,14,30])
add_rows(ws1, [
    ['','PO001','驗收人員','2026-06-22','待驗收',''],
])
ws2 = wb.create_sheet('子表_驗收明細')
make_header(ws2,
    ['驗收單號(關聯)','料號','品名','單位','驗收數量','發票數量','差異數量','差異原因'],
    [16,14,36,8,12,12,12,28], color='2E6DA4')
recv_detail = []
for i in range(min(4, len(sample_items_all))):
    it = sample_items_all[i]
    recv_detail.append(['RECV001', it['料號'], it['品名'][:20], it['單位'], 0, 0, 0, ''])
add_rows(ws2, recv_detail)
note_sheet(wb, [
    '【08 週期採購驗收單 匯入說明】',
    '',
    '▶ 步驟1：用「主表」Sheet 匯入，表單名稱：週期採購驗收單',
    '▶ 步驟2：Design Mode 新增子表「驗收明細」',
    '',
    '主表欄位設定：',
    '  驗收單號   → Auto Number',
    '  採購單號   → Link → 週期採購採購單',
    '  驗收日期   → Date',
    '  狀態       → Select：待驗收 / 驗收完成 / 驗收異常',
    '',
    '子表「驗收明細」欄位設定：',
    '  料號       → Link → 週期採購料號主檔',
    '  品名/單位  → Lookup',
    '  驗收數量/發票數量 → Number',
    '  差異數量   → Formula: 驗收數量 - 發票數量',
    '  差異原因   → Text（差異≠0時 Workflow 強制必填）',
    '',
    '▶ Action Button：「產生請款單」→ receiving_to_payment.js',
])
wb.save(os.path.join(OUT, '08_週期採購驗收單.xlsx'))
print("[OK] 08_週期採購驗收單.xlsx")


# ════════════════════════════════════════════════════════
# 09 週期採購請款單（主表 + 子表）
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws1 = wb.active; ws1.title = '主表'
make_header(ws1,
    ['請款單號','驗收單號','發票號碼','發票日期','發票金額','狀態','財務處理人員'],
    [12,12,16,14,14,12,14])
add_rows(ws1, [
    ['','RECV001','INV-2026-001','2026-06-25',0,'草稿','財務人員'],
])
ws2 = wb.create_sheet('子表_費用分攤明細')
make_header(ws2,
    ['請款單號(關聯)','公司別','部門','成本中心','會計科目','分攤金額','差異原因'],
    [16,12,14,14,14,14,28], color='2E6DA4')
add_rows(ws2, [
    ['PAY001','日曜天地','工務部','CC-E-NIC','E01',0,''],
    ['PAY001','日曜天地','清潔部','CC-C-NIC','C01',0,''],
    ['PAY001','春大直',  '工務部','CC-E-CHU','E01',0,''],
])
note_sheet(wb, [
    '【09 週期採購請款單 匯入說明】',
    '',
    '▶ 步驟1：用「主表」Sheet 匯入，表單名稱：週期採購請款單',
    '▶ 步驟2：Design Mode 新增子表「費用分攤明細」',
    '',
    '主表欄位設定：',
    '  請款單號   → Auto Number',
    '  驗收單號   → Link → 週期採購驗收單',
    '  發票日期   → Date',
    '  發票金額   → Number',
    '  狀態       → Select：草稿/已提交/付款中/已付款',
    '',
    '子表「費用分攤明細」欄位設定：',
    '  公司別/部門 → Select',
    '  成本中心/會計科目 → Text 或 Lookup',
    '  分攤金額   → Number',
    '  差異原因   → Text（SUM(分攤)≠發票金額時必填）',
    '',
    '注意：分攤合計必須等於發票金額，receiving_to_payment.js 會自動建立初稿',
])
wb.save(os.path.join(OUT, '09_週期採購請款單.xlsx'))
print("[OK] 09_週期採購請款單.xlsx")


# ════════════════════════════════════════════════════════
# 10 週期採購異常紀錄
# ════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active; ws.title = '主表'
make_header(ws,
    ['紀錄ID','關聯類型','關聯單號','事件類型','說明','操作人員','建立時間','原始值','變更後值'],
    [10,14,14,14,44,12,18,20,20])
audit_items = sample_items_all[:2]
add_rows(ws, [
    ['','驗收單','RECV001','驗收差異',
     f'料號{audit_items[0]["料號"]} {audit_items[0]["品名"][:12]} 驗收差異',
     '驗收人員','2026-06-22 14:30','10','8'],
    ['','請購單','REQ001','補填','截止後補填申請，主管授權放行','主管','2026-06-11 09:00','',''],
    ['','採購單','PO001','缺貨',
     f'料號{audit_items[1]["料號"] if len(audit_items)>1 else "E0102001"} 廠商缺貨',
     '採購人員','2026-06-18 11:00','',''],
])
note_sheet(wb, [
    '【10 週期採購異常紀錄 匯入說明】',
    '表單名稱：週期採購異常紀錄',
    '',
    '欄位型態設定：',
    '  紀錄ID     → Auto Number',
    '  關聯類型   → Select：請購單 / 採購單 / 驗收單 / 請款單',
    '  事件類型   → Select：補填 / 逾期 / 缺貨 / 替代品 / 驗收差異 / 請款差異',
    '  說明       → Text，必填',
    '  操作人員   → Text（或設為當前登入使用者）',
    '  建立時間   → DateTime',
    '  原始值/變更後值 → Text',
    '',
    '▶ 此表單使用者不可手動新增，所有記錄由 Workflow JS 自動寫入',
    '▶ 建議在 Portal 設定：僅 Admin 角色可見此表單',
    '注意：匯入後請在 Design Mode 關閉「新增」按鈕（只允許 API 寫入）',
])
wb.save(os.path.join(OUT, '10_週期採購異常紀錄.xlsx'))
print("[OK] 10_週期採購異常紀錄.xlsx")


print()
print(f"全部完成！路徑：{os.path.abspath(OUT)}")
print(f"料號主檔：{len(master_list)} 筆，對照表：{len(rows02)} 筆")
