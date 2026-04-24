"""
大直工務部 API Router
Prefix: /api/v1/dazhi-repair

端點：
  GET  /raw-fields        — 回傳 Ragic 第一筆資料欄位名稱（debug 用）
  GET  /ping              — 快速連線診斷
  GET  /sync              — 同步診斷：直接抓 Ragic 回傳摘要（debug 用）
  POST /sync              — 觸發背景同步：Ragic → SQLite
  GET  /years             — 回傳資料中的年份清單
  GET  /filter-options    — 回傳過濾條件選項（類型/樓層/狀態）
  GET  /dashboard         — Dashboard KPI + 圖表資料
  GET  /stats/repair      — 4.1 報修統計（月份×7指標）
  GET  /stats/fee         — 金額統計（費用類型 × 月份交叉表）
  GET  /stats/closing     — 4.2 結案時間統計
  GET  /stats/type        — 4.3 報修類型統計
  GET  /stats/room        — 4.4 本月客房報修表
  GET  /detail            — 明細清單（分頁+排序+搜尋）
  GET  /export            — 匯出 Excel（串流回傳）

設計：
  - 所有資料端點（/years 以下）從本地 SQLite 讀取（DazhiRepairCase 表）
  - GET /sync 仍直連 Ragic，僅供診斷用
  - POST /sync 觸發背景同步（Ragic → SQLite），立即回傳
"""
from __future__ import annotations

import io
import logging
import time
from collections import Counter
from datetime import datetime
from typing import Optional

logger = logging.getLogger(__name__)

import httpx
from fastapi import APIRouter, BackgroundTasks, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.models.dazhi_repair import DazhiRepairCase
from app.services import dazhi_repair_service as svc

router = APIRouter()


# ── 共用 DB 讀取 helper ────────────────────────────────────────────────────────

def load_cases_from_db(db: Session) -> list[DazhiRepairCase]:
    """從本地 SQLite 載入所有大直報修案件（ORM 物件與 RepairCase 介面相容）。"""
    return db.query(DazhiRepairCase).all()


# ── /raw-fields ───────────────────────────────────────────────────────────────

@router.get("/raw-fields", summary="回傳 Ragic 第一筆欄位名稱（debug 用）")
async def get_raw_fields():
    """
    回傳 Ragic 第一筆資料的 key 列表。
    用途：確認實際欄位名稱，方便調整 dazhi_repair_service.py 中的 RK_* 常數。
    """
    return await svc.fetch_raw_fields()


# ── /ping ─────────────────────────────────────────────────────────────────────

@router.get("/ping", summary="快速連線診斷（5 秒 timeout，直接回傳裸 HTTP 結果）")
async def ping_ragic():
    """
    直接對 Ragic URL 發一次 GET（limit=1），5 秒內沒回應就 timeout。
    回傳 status_code / elapsed_ms / body_preview / error，
    讓你立即判斷是連線問題還是欄位/資料問題。
    不走 RagicAdapter 的 pagination 迴圈。
    """
    base_url = (
        f"https://{settings.RAGIC_DAZHI_REPAIR_SERVER_URL}"
        f"/{settings.RAGIC_DAZHI_REPAIR_ACCOUNT}"
        f"/{settings.RAGIC_DAZHI_REPAIR_PATH}"
    )
    api_key  = settings.RAGIC_API_KEY
    pageid   = settings.RAGIC_DAZHI_REPAIR_PAGEID

    results = []

    for label, params in [
        ("with_pageid",    {"api": "", "limit": 1, "PAGEID": pageid}),
        ("without_pageid", {"api": "", "limit": 1}),
    ]:
        t0 = time.monotonic()
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(5.0), verify=False) as client:
                resp = await client.get(
                    base_url,
                    headers={"Authorization": f"Basic {api_key}", "Accept": "application/json"},
                    params=params,
                )
            elapsed = int((time.monotonic() - t0) * 1000)
            try:
                body = resp.json()
                body_type = type(body).__name__
                if isinstance(body, dict):
                    record_ids = [k for k in body if k.lstrip("-").isdigit()]
                    record_count = len(record_ids)
                    first_record_id = record_ids[0] if record_ids else None
                    first_record_raw = body[first_record_id] if first_record_id else {}
                    preview_keys = list(body.keys())[:5]
                    body_preview = {k: type(body[k]).__name__ for k in preview_keys}
                else:
                    body_preview = str(body)[:200]
                    record_count = 0
                    first_record_id = None
                    first_record_raw = {}
            except Exception:
                body_type = "not_json"
                body_preview = resp.text[:300]
                record_count = 0
                first_record_id = None
                first_record_raw = {}
            results.append({
                "test":             label,
                "status_code":      resp.status_code,
                "elapsed_ms":       elapsed,
                "body_type":        body_type,
                "record_count":     record_count,
                "body_preview":     body_preview,
                "first_record_id":  first_record_id,
                "first_record_raw": first_record_raw,
                "error":            None,
            })
        except httpx.TimeoutException:
            elapsed = int((time.monotonic() - t0) * 1000)
            results.append({
                "test":       label,
                "elapsed_ms": elapsed,
                "error":      f"TIMEOUT after {elapsed}ms — Ragic 沒有在 5 秒內回應",
                "tip":        "可能原因：(1) 伺服器端封鎖對外 HTTPS (2) Ragic 帳號/路徑錯誤 (3) API Key 無此 Sheet 存取權",
            })
        except Exception as exc:
            elapsed = int((time.monotonic() - t0) * 1000)
            results.append({
                "test":       label,
                "elapsed_ms": elapsed,
                "error":      str(exc),
            })

    return {
        "ragic_base_url": base_url,
        "pageid":         pageid,
        "api_key_prefix": api_key[:8] + "..." if len(api_key) > 8 else api_key,
        "results":        results,
    }


# ── /sync ─────────────────────────────────────────────────────────────────────

@router.get("/sync", summary="同步診斷：直接抓 Ragic 回傳統計摘要（debug 用）")
async def sync_diagnostic():
    """
    完整從 Ragic 抓取所有資料並回傳診斷摘要，用於：
    1. 確認 Ragic 連線是否正常
    2. 確認資料筆數與年份分布
    3. 確認欄位名稱對應（field_names = 第一筆原始 key）
    4. 顯示 parse 失敗數（title 有值但 occurred_at 為 None）

    注意：此端點直連 Ragic，僅供診斷；正常頁面資料讀本地 SQLite。
    觸發實際同步請用 POST /sync。
    """
    field_info = await svc.fetch_raw_fields()
    all_cases  = await svc.fetch_all_cases()

    year_dist = dict(sorted(Counter(
        c.year for c in all_cases if c.year is not None
    ).items()))

    no_date_count = sum(
        1 for c in all_cases if c.occurred_at is None and c.title
    )

    recent = sorted(
        [c for c in all_cases if c.occurred_at],
        key=lambda c: c.occurred_at,  # type: ignore
        reverse=True,
    )[:3]
    recent_samples = [
        {
            "ragic_id":  c.ragic_id,
            "case_no":   c.case_no,
            "title":     c.title[:30] if c.title else "",
            "occurred_at": c.occurred_at.strftime("%Y/%m/%d") if c.occurred_at else "",
            "status":    c.status,
        }
        for c in recent
    ]

    return {
        "ok":               True,
        "total_parsed":     len(all_cases),
        "no_date_count":    no_date_count,
        "year_distribution": year_dist,
        "field_names":      field_info.get("fields", []),
        "sample_raw":       field_info.get("sample", {}),
        "recent_samples":   recent_samples,
        "ragic_url": (
            f"https://{svc.settings.RAGIC_DAZHI_REPAIR_SERVER_URL}"
            f"/{svc.settings.RAGIC_DAZHI_REPAIR_ACCOUNT}"
            f"/{svc.settings.RAGIC_DAZHI_REPAIR_PATH}"
            f"?PAGEID={svc.settings.RAGIC_DAZHI_REPAIR_PAGEID}"
        ),
    }


@router.post("/sync", summary="觸發背景同步：Ragic → SQLite（非阻塞）")
async def sync_from_ragic(background_tasks: BackgroundTasks):
    """
    將大直工務報修資料從 Ragic 同步到本地 SQLite（背景執行）。
    立即回傳，不等待同步完成。
    """
    from app.services.dazhi_repair_sync import sync_from_ragic as do_sync
    background_tasks.add_task(do_sync)
    return {"success": True, "message": "同步已在背景啟動"}


# ── /years ─────────────────────────────────────────────────────────────────────

@router.get("/years", summary="資料中的年份清單")
def get_years(db: Session = Depends(get_db)):
    all_cases = load_cases_from_db(db)
    return {"years": svc.get_years(all_cases)}


# ── /filter-options ───────────────────────────────────────────────────────────

@router.get("/filter-options", summary="過濾條件選項")
def get_filter_options(db: Session = Depends(get_db)):
    all_cases = load_cases_from_db(db)
    return svc.get_filter_options(all_cases)


# ── /dashboard ────────────────────────────────────────────────────────────────

@router.get("/dashboard", summary="Dashboard KPI + 圖表資料")
def get_dashboard(
    year:  int = Query(..., description="年度"),
    month: int = Query(0,   description="月份（0=全年）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    eff_month = month if month > 0 else datetime.now().month
    return svc.compute_dashboard(all_cases, year, eff_month)


# ── /stats/repair ─────────────────────────────────────────────────────────────

@router.get("/stats/repair", summary="4.1 報修統計（全年月份×6指標）")
def get_repair_stats(
    year: int = Query(..., description="年度"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_repair_stats(all_cases, year)


# ── /stats/fee ────────────────────────────────────────────────────────────────

@router.get("/stats/fee", summary="金額統計（費用類型 × 月份交叉表）")
def get_fee_stats(
    year: int = Query(..., description="年度"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_fee_stats(all_cases, year)


# ── /stats/closing ────────────────────────────────────────────────────────────

@router.get("/stats/closing", summary="4.2 結案時間統計（小型/中大型）")
def get_closing_stats(
    year:  int           = Query(..., description="年度"),
    month: Optional[int] = Query(None, description="月份（不傳=全年）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_closing_time(all_cases, year, month)


# ── /stats/type ───────────────────────────────────────────────────────────────

@router.get("/stats/type", summary="4.3 報修類型統計（類型×月份）")
def get_type_stats(
    year:  int           = Query(..., description="年度"),
    month: Optional[int] = Query(None, description="聚焦月份（選填）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_type_stats(all_cases, year, month)


# ── /stats/room ───────────────────────────────────────────────────────────────

@router.get("/stats/room", summary="4.4 本月客房報修表（房號×分類）")
def get_room_stats(
    year:  int = Query(..., description="年度"),
    month: int = Query(..., description="月份（必選）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_room_repair_table(all_cases, year, month)


# ── /detail ───────────────────────────────────────────────────────────────────

@router.get("/detail", summary="明細清單（分頁+排序+搜尋）")
def get_detail(
    year:        Optional[int] = Query(None),
    month:       Optional[int] = Query(None),
    repair_type: Optional[str] = Query(None),
    floor:       Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    keyword:     Optional[str] = Query(None),
    page:        int            = Query(1, ge=1),
    page_size:   int            = Query(50, ge=1, le=200),
    sort_by:     str            = Query("occurred_at"),
    sort_desc:   bool           = Query(True),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.query_detail(
        all_cases,
        year=year, month=month,
        repair_type=repair_type, floor=floor, status=status,
        keyword=keyword,
        page=page, page_size=page_size,
        sort_by=sort_by, sort_desc=sort_desc,
    )


# ── /images/{ragic_id} ────────────────────────────────────────────────────────

@router.get("/images/{ragic_id}", summary="取得單筆案件的圖片 URL 清單")
async def get_case_images(ragic_id: str):
    """
    取得指定案件的圖片 URL。
    若傳入的是顯示用案件編號（如 In202604-046），從整表回傳中找出對應記錄。
    """
    from app.services.ragic_data_service import parse_images
    adapter = svc._get_adapter()
    IMAGE_FIELDS = ["上傳圖片", "上傳圖片.1", "維修照上傳", "維修照"]
    server = settings.RAGIC_DAZHI_REPAIR_SERVER_URL

    try:
        resp = await adapter.fetch_one(ragic_id)
        if not isinstance(resp, dict):
            return {"ragic_id": ragic_id, "images": []}

        # ── Step 1: 從回傳中找出正確的數字 _ragicId ─────────────────────────
        numeric_id: str | None = None

        # 情形 A：resp = {ragic_id: {...}} （數字 ID 直接命中）
        direct = resp.get(ragic_id) or resp.get(str(ragic_id))
        if isinstance(direct, dict):
            numeric_id = str(ragic_id)
        else:
            # 情形 B：resp = {數字key: {record...}, ...} （整表資料，顯示用 ID 無效）
            for _key, record in resp.items():
                if isinstance(record, dict):
                    if (record.get("報修單編號") == ragic_id or
                            str(record.get("_ragicId", "")) == ragic_id):
                        numeric_id = str(record.get("_ragicId", _key))
                        logger.info(f"[DazhiRepair] images: {ragic_id} → _ragicId={numeric_id}")
                        break

        if not numeric_id:
            logger.warning(f"[DazhiRepair] images: no record found for {ragic_id}")
            return {"ragic_id": ragic_id, "images": []}

        # ── Step 2: 用數字 ID 重新發 detail 請求（才有附件欄位）────────────
        if numeric_id != str(ragic_id):
            resp2 = await adapter.fetch_one(numeric_id)
            if isinstance(resp2, dict):
                data = resp2.get(numeric_id) or resp2.get(str(numeric_id))
                if not isinstance(data, dict):
                    # detail 可能直接是 flat dict
                    data = {k: v for k, v in resp2.items() if not str(k).lstrip("-").isdigit()}
            else:
                data = {}
        else:
            data = direct or {}

        if not data:
            logger.warning(f"[DazhiRepair] images: empty detail for {ragic_id} (numeric={numeric_id})")
            return {"ragic_id": ragic_id, "images": []}

        # ── Step 3: 解析圖片欄位 ─────────────────────────────────────────
        images: list[dict] = []
        for fk in IMAGE_FIELDS:
            val = data.get(fk)
            if val:
                logger.info(f"[DazhiRepair] images {ragic_id} field {fk!r}={repr(val)[:150]}")
                parsed = parse_images(val, server=server, account=settings.RAGIC_DAZHI_REPAIR_ACCOUNT)
                for img in parsed:
                    if not any(r["url"] == img["url"] for r in images):
                        images.append(img)

        logger.info(f"[DazhiRepair] images {ragic_id} (id={numeric_id}): found {len(images)}")
        return {"ragic_id": ragic_id, "images": images}

    except Exception as exc:
        logger.warning(f"[DazhiRepair] images {ragic_id} failed: {exc}")
    return {"ragic_id": ragic_id, "images": []}


# ── /export ───────────────────────────────────────────────────────────────────

@router.get("/export", summary="匯出 Excel")
def export_excel(
    year:        Optional[int] = Query(None),
    month:       Optional[int] = Query(None),
    repair_type: Optional[str] = Query(None),
    floor:       Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    keyword:     Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """匯出過濾後的明細資料為 Excel 檔案（.xlsx）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="請安裝 openpyxl：pip install openpyxl")

    all_cases = load_cases_from_db(db)
    result = svc.query_detail(
        all_cases,
        year=year, month=month,
        repair_type=repair_type, floor=floor, status=status,
        keyword=keyword,
        page=1, page_size=9999,
    )
    items = result["items"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "報修明細"

    # ── 標題行 ────────────────────────────────────────────────────────────────
    headers = [
        "報修編號", "標題", "報修人姓名", "報修類型", "發生樓層",
        "發生時間", "負責單位", "花費工時", "處理狀況",
        "委外費用", "維修費用", "總費用",
        "驗收者", "驗收", "結案人",
        "扣款事項", "扣款費用", "財務備註",
        "結案時間", "結案天數", "是否完成",
    ]

    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border

    # ── 資料行 ────────────────────────────────────────────────────────────────
    STATUS_FILLS = {
        True:  PatternFill(start_color="F6FFED", end_color="F6FFED", fill_type="solid"),
        False: PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid"),
    }

    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.get("case_no", ""),
            item.get("title", ""),
            item.get("reporter_name", ""),
            item.get("repair_type", ""),
            item.get("floor", ""),
            item.get("occurred_at", ""),
            item.get("responsible_unit", ""),
            item.get("work_hours", 0),
            item.get("status", ""),
            item.get("outsource_fee", 0),
            item.get("maintenance_fee", 0),
            item.get("total_fee", 0),
            item.get("acceptor", ""),
            item.get("accept_status", ""),
            item.get("closer", ""),
            item.get("deduction_item", ""),
            item.get("deduction_fee", 0),
            item.get("finance_note", ""),
            item.get("completed_at", ""),
            item.get("close_days", ""),
            "是" if item.get("is_completed") else "否",
        ]
        is_done = item.get("is_completed", False)
        fill = STATUS_FILLS[is_done]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    col_widths = [12, 30, 12, 10, 10, 18, 12, 8, 10, 10, 10, 10, 10, 8, 10, 14, 10, 18, 14, 8, 8]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"dazhi_repair_{year or 'all'}_{month or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
