"""
核准請購單月報表 API Router

端點：
  GET  /api/v1/purchase-report/approved/monthly      — 月報表明細（品項級別，分頁）
  GET  /api/v1/purchase-report/approved/summary      — KPI 統計（9 項，主單級別）
  GET  /api/v1/purchase-report/approved/departments  — 部門彙總表
  GET  /api/v1/purchase-report/approved/export       — Excel 匯出（品項級別）
  POST /api/v1/purchase-report/sync                  — 手動觸發同步（Master + Detail）
  GET  /api/v1/purchase-report/sync/status           — 同步狀態查詢

權限：
  開發期間：system_admin_only（只有 system_admin 可存取）
  上線後：purchase_report_view（透過角色管理介面設定）
"""
import io
import asyncio
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, Query, BackgroundTasks, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.core.database import get_db, SessionLocal
from app.core.time import twnow
from app.dependencies import require_permission
from app.models.module_sync_log import ModuleSyncLog
from app.models.purchase_request import (
    ApprovedPurchaseRequest,
    ApprovedPurchaseRequestItem,
    DEPT_SHEETS,
)

router = APIRouter()

# 開發期間使用 system_admin_only；上線後改為 purchase_report_view
_PERM = "system_admin_only"


# ─────────────────────────────────────────────────────────────────────────────
# GET /debug/sample — 診斷用（查看實際 status / approved_date 欄位值）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/debug/sample")
def debug_sample(
    db:    Session = Depends(get_db),
    limit: int = Query(20),
):  # ← 診斷用，暫不需認證
    """診斷：列出最新 N 筆主單的 status / approved_date / last_updated_at 實際值。"""
    rows = (
        db.query(
            ApprovedPurchaseRequest.id,
            ApprovedPurchaseRequest.department_display,
            ApprovedPurchaseRequest.purchase_no,
            ApprovedPurchaseRequest.status,
            ApprovedPurchaseRequest.approved_date,
            ApprovedPurchaseRequest.last_updated_at,
            ApprovedPurchaseRequest.request_date,
        )
        .order_by(ApprovedPurchaseRequest.id.desc())
        .limit(limit)
        .all()
    )
    # 統計 status 分布
    from sqlalchemy import func as _func
    status_dist = (
        db.query(
            ApprovedPurchaseRequest.status,
            _func.count(ApprovedPurchaseRequest.id).label("cnt"),
        )
        .group_by(ApprovedPurchaseRequest.status)
        .all()
    )
    # approved_date 非 NULL 的筆數
    approved_cnt = (
        db.query(_func.count(ApprovedPurchaseRequest.id))
        .filter(ApprovedPurchaseRequest.approved_date != None)
        .scalar()
        or 0
    )
    return {
        "status_distribution": [{"status": r.status, "count": r.cnt} for r in status_dist],
        "approved_date_not_null": approved_cnt,
        "sample": [
            {
                "id":                 r.id,
                "department":         r.department_display,
                "purchase_no":        r.purchase_no,
                "status":             r.status,
                "approved_date":      r.approved_date.isoformat() if r.approved_date else None,
                "last_updated_at":    r.last_updated_at.isoformat() if r.last_updated_at else None,
                "request_date":       r.request_date.isoformat() if r.request_date else None,
            }
            for r in rows
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# 工具函式
# ─────────────────────────────────────────────────────────────────────────────

def _build_date_filter(
    q,
    year_month: Optional[str] = None,
    year_month_from: Optional[str] = None,
    year_month_to: Optional[str] = None,
):
    """
    日期篩選：status=F + approved_date 符合單月 / 全年 / 區間
      - year_month:      YYYY-MM（單月，精確比對）
      - year_month_from + year_month_to：YYYY-MM（閉區間）
    """
    q = q.filter(ApprovedPurchaseRequest.status == "F")
    ym_col = func.strftime("%Y-%m", ApprovedPurchaseRequest.approved_date)
    if year_month_from and year_month_to:
        q = q.filter(ym_col >= year_month_from, ym_col <= year_month_to)
    elif year_month_from:
        q = q.filter(ym_col >= year_month_from)
    elif year_month:
        q = q.filter(ym_col == year_month)
    return q


def _date_label(
    year_month: Optional[str] = None,
    year_month_from: Optional[str] = None,
    year_month_to: Optional[str] = None,
) -> str:
    """產生適合放入回應 year_month 欄位的標籤字串。"""
    if year_month_from and year_month_to:
        if (year_month_from.endswith("-01")
                and year_month_to.endswith("-12")
                and year_month_from[:4] == year_month_to[:4]):
            return f"{year_month_from[:4]}年度"
        return f"{year_month_from}~{year_month_to}"
    return year_month or ""


def _apply_dept_filter(q, department: Optional[str]):
    if department:
        return q.filter(ApprovedPurchaseRequest.department_display == department)
    return q


def _apply_account_filter(q, account_category: Optional[str]):
    if account_category:
        return q.filter(ApprovedPurchaseRequest.account_category == account_category)
    return q


def _apply_search(q, keyword: Optional[str]):
    if keyword:
        like = f"%{keyword}%"
        return q.filter(
            ApprovedPurchaseRequest.description.ilike(like)
            | ApprovedPurchaseRequest.purchase_no.ilike(like)
            | ApprovedPurchaseRequest.applicant.ilike(like)
        )
    return q


def _ragic_url(order: ApprovedPurchaseRequest) -> str:
    """建構 Ragic 內頁連結。"""
    detail_path = ""
    for d in DEPT_SHEETS:
        if d["list_path"] == order.ragic_sheet_path:
            detail_path = d["detail_path"]
            break
    if not detail_path:
        return ""
    return f"https://ap12.ragic.com/soutlet001/{detail_path}/{order.ragic_record_id}"


def _format_item_row(order: ApprovedPurchaseRequest, item: ApprovedPurchaseRequestItem) -> dict:
    """將 order + item 合併為月報表一列（key 名稱與前端 PurchaseReportItem type 對齊）。"""
    return {
        # 主單欄位（同一主單的每個品項重複顯示）
        "order_id":           order.id,
        "company":            order.company,
        "department_display": order.department_display,   # ← 前端 type: department_display
        "purchase_no":        order.purchase_no,
        "request_date":       order.request_date.isoformat()  if order.request_date  else None,
        "approved_date":      order.approved_date.isoformat() if order.approved_date else None,
        "applicant":          order.applicant,
        "description":        order.description,
        "account_category":   order.account_category,
        "amount":             order.amount,
        "amount_tax":         order.amount_tax,
        "amount_total":       order.amount_total,
        "vendor1":            order.vendor1,
        "vendor2":            order.vendor2,
        "vendor3":            order.vendor3,
        "status":             order.status,
        "remark":             order.remark,
        "ragic_url":          _ragic_url(order),
        # 品項欄位（每列不同）
        "item_id":            item.id,
        "seq":                item.seq,
        "product_name":       item.product_name,
        "qty":                item.qty,
        "unit":               item.unit,
        "selected_vendor":    item.selected_vendor,
        "selected_unit_price": item.selected_unit_price,
        "selected_amount":    item.selected_amount,
        "is_confirmed":       item.is_confirmed,
        "item_remark":        item.item_remark,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /monthly — 月報表明細（品項級別）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/monthly")
def get_monthly_report(
    year_month:       Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from:  Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:    Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    company:          str           = Query("樂群"),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    applicant:        Optional[str] = Query(None),
    q:                Optional[str] = Query(None, description="關鍵字（說明/單號/申請人）"),
    page:             int           = Query(1, ge=1),
    per_page:         int           = Query(50, ge=1, le=200),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """
    月報表明細—以品項為單位（每個品項一列）。
    同一張請購單的主單欄位在各品項列重複顯示。
    """
    order_q = (
        db.query(ApprovedPurchaseRequest)
        .filter(ApprovedPurchaseRequest.company == company)
    )
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_category)
    if applicant:
        order_q = order_q.filter(
            ApprovedPurchaseRequest.applicant.ilike(f"%{applicant}%")
        )
    order_q = _apply_search(order_q, q)

    order_ids = [r.id for r in order_q.with_entities(ApprovedPurchaseRequest.id).all()]

    if not order_ids:
        return {"total": 0, "page": page, "per_page": per_page, "items": []}

    item_q = (
        db.query(ApprovedPurchaseRequestItem, ApprovedPurchaseRequest)
        .join(
            ApprovedPurchaseRequest,
            ApprovedPurchaseRequestItem.order_id == ApprovedPurchaseRequest.id,
        )
        .filter(ApprovedPurchaseRequestItem.order_id.in_(order_ids))
        .order_by(
            ApprovedPurchaseRequest.approved_date.desc(),
            ApprovedPurchaseRequest.purchase_no,
            ApprovedPurchaseRequestItem.seq,
        )
    )

    total = item_q.count()
    rows  = item_q.offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_format_item_row(order, item) for item, order in rows],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /summary — KPI 統計（主單級別）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/summary")
def get_summary(
    year_month:       Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from:  Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:    Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    company:          str           = Query("樂群"),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """KPI 統計（9 項，以主單為計算單位）。"""
    label = _date_label(year_month, year_month_from, year_month_to)
    order_q = (
        db.query(ApprovedPurchaseRequest)
        .filter(ApprovedPurchaseRequest.company == company)
    )
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_category)

    orders = order_q.all()
    if not orders:
        return _empty_summary(label, company, department)

    order_ids = [o.id for o in orders]
    item_count = (
        db.query(func.count(ApprovedPurchaseRequestItem.id))
        .filter(ApprovedPurchaseRequestItem.order_id.in_(order_ids))
        .scalar()
        or 0
    )

    rej_count = (
        db.query(func.count(ApprovedPurchaseRequest.id))
        .filter(
            ApprovedPurchaseRequest.company == company,
            ApprovedPurchaseRequest.status == "REJ",
            func.strftime("%Y-%m", ApprovedPurchaseRequest.request_date) == year_month,
        )
        .scalar()
        or 0
    )

    total_amount = sum(o.amount or 0 for o in orders)
    total_tax    = sum(o.amount_tax or 0 for o in orders)
    order_count  = len(orders)

    top_order = max(orders, key=lambda o: o.amount or 0)

    dept_count: dict = {}
    dept_amount: dict = {}
    for o in orders:
        dept_count[o.department_display]  = dept_count.get(o.department_display, 0)  + 1
        dept_amount[o.department_display] = dept_amount.get(o.department_display, 0) + (o.amount or 0)

    top_dept_by_count  = max(dept_count,  key=lambda d: dept_count[d],  default="")
    top_dept_by_amount = max(dept_amount, key=lambda d: dept_amount[d], default="")

    dept_ratio = [
        {
            "department_display": dept,
            "order_count":        cnt,
            "total_amount":       dept_amount[dept],
            "total_tax":          0,
            "amount_ratio":       round(dept_amount[dept] / total_amount * 100, 1) if total_amount else 0,
        }
        for dept, cnt in sorted(dept_count.items(), key=lambda x: -x[1])
    ]

    return {
        "year_month":           label,
        "company":              company,
        "department":           department,
        "order_count":          order_count,
        "total_amount":         total_amount,
        "total_tax":            total_tax,
        "item_count":           item_count,
        "dept_count":           len(dept_count),
        "avg_amount":           round(total_amount / order_count) if order_count else 0,
        "top_order": {
            "purchase_no": top_order.purchase_no,
            "department":  top_order.department_display,
            "amount":      top_order.amount,
            "description": (top_order.description or "")[:50],
        },
        "top_dept_by_count":  top_dept_by_count,
        "top_dept_by_amount": top_dept_by_amount,
        "dept_summary":       dept_ratio,
        "rej_count":          rej_count,
    }


def _empty_summary(year_month, company, department):
    return {
        "year_month": year_month, "company": company, "department": department,
        "order_count": 0, "total_amount": 0, "total_tax": 0,
        "item_count": 0, "dept_count": 0, "avg_amount": 0,
        "top_order": None, "top_dept_by_count": "", "top_dept_by_amount": "",
        "dept_summary": [], "rej_count": 0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /departments — 部門彙總表
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/departments")
def get_departments(
    year_month:       Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from:  Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:    Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    company:          str           = Query("樂群"),
    account_category: Optional[str] = Query(None),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """各部門請購筆數 + 金額彙總（已核准）。"""
    base_q = (
        db.query(ApprovedPurchaseRequest)
        .filter(ApprovedPurchaseRequest.company == company)
    )
    base_q = _build_date_filter(base_q, year_month, year_month_from, year_month_to)
    base_q = _apply_account_filter(base_q, account_category)

    rows = (
        base_q.with_entities(
            ApprovedPurchaseRequest.department_display,
            func.count(ApprovedPurchaseRequest.id).label("order_count"),
            func.sum(ApprovedPurchaseRequest.amount).label("total_amount"),
            func.sum(ApprovedPurchaseRequest.amount_tax).label("total_tax"),
        )
        .group_by(ApprovedPurchaseRequest.department_display)
        .order_by(func.sum(ApprovedPurchaseRequest.amount).desc())
        .all()
    )
    return [
        {
            "department_display": r.department_display,
            "order_count":        r.order_count,
            "total_amount":       r.total_amount or 0,
            "total_tax":          r.total_tax    or 0,
        }
        for r in rows
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /export — Excel 匯出（品項級別）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/export")
def export_excel(
    year_month:       Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from:  Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:    Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    company:          str           = Query("樂群"),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    applicant:        Optional[str] = Query(None),
    q:                Optional[str] = Query(None),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """Excel 匯出（品項級別，同月報表欄位）。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 套件未安裝，無法匯出 Excel")

    label = _date_label(year_month, year_month_from, year_month_to)
    order_q = (
        db.query(ApprovedPurchaseRequest)
        .filter(ApprovedPurchaseRequest.company == company)
    )
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_category)
    if applicant:
        order_q = order_q.filter(
            ApprovedPurchaseRequest.applicant.ilike(f"%{applicant}%")
        )
    order_q = _apply_search(order_q, q)

    order_ids = [r.id for r in order_q.with_entities(ApprovedPurchaseRequest.id).all()]
    if not order_ids:
        raise HTTPException(status_code=404, detail="所選條件無資料")

    rows = (
        db.query(ApprovedPurchaseRequestItem, ApprovedPurchaseRequest)
        .join(ApprovedPurchaseRequest,
              ApprovedPurchaseRequestItem.order_id == ApprovedPurchaseRequest.id)
        .filter(ApprovedPurchaseRequestItem.order_id.in_(order_ids))
        .order_by(
            ApprovedPurchaseRequest.approved_date.desc(),
            ApprovedPurchaseRequest.purchase_no,
            ApprovedPurchaseRequestItem.seq,
        )
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"請購單_{label}"

    HEADER_FILL  = PatternFill("solid", fgColor="1B3A5C")
    MASTER_FILL  = PatternFill("solid", fgColor="D5E8F0")
    ITEM_FILL    = PatternFill("solid", fgColor="FFF3E0")
    THIN_BORDER  = Border(
        left  =Side(style="thin"), right =Side(style="thin"),
        top   =Side(style="thin"), bottom=Side(style="thin"),
    )

    MASTER_COLS = ["公司", "部門", "請購單號", "申請日期", "核准日期（簽核完成日）",
                   "申請人", "說明（請購事由）", "會科", "全案小計（未稅）", "營業稅", "備註"]
    ITEM_COLS   = ["項次", "品名", "數量", "單位", "擬定廠商", "擬定單價", "擬定金額"]
    headers     = MASTER_COLS + ITEM_COLS + ["Ragic連結"]

    for col_idx, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = THIN_BORDER

    ws.row_dimensions[1].height = 30

    for row_idx, (item, order) in enumerate(rows, 2):
        data = [
            order.company,
            order.department_display,
            order.purchase_no,
            order.request_date.isoformat() if order.request_date else "",
            order.approved_date.isoformat() if order.approved_date else "",
            order.applicant or "",
            (order.description or "")[:200],
            order.account_category or "",
            order.amount or 0,
            order.amount_tax or 0,
            order.remark or "",
            item.seq,
            item.product_name or "",
            item.qty or "",
            item.unit or "",
            item.selected_vendor or "",
            item.selected_unit_price or "",
            item.selected_amount or "",
            _ragic_url(order),
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx <= len(MASTER_COLS):
                cell.fill = PatternFill("solid", fgColor="EBF5FF" if row_idx % 2 == 0 else "FFFFFF")
            else:
                cell.fill = PatternFill("solid", fgColor="FFF8F0" if row_idx % 2 == 0 else "FFFBF5")
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if isinstance(val, int) and col_idx in (9, 10, 17, 18):
                cell.number_format = '#,##0'

    col_widths = [8, 10, 18, 12, 18, 10, 35, 18, 14, 10, 20, 6, 25, 8, 8, 18, 12, 12, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename_cn = f"核准請購單月報表_{label}.xlsx"
    filename_safe = f"purchase_report_{label.replace('~', '_')}.xlsx"
    encoded = quote(filename_cn, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_safe}"; '
                f"filename*=UTF-8''{encoded}"
            )
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# POST /sync — 手動觸發同步
# ─────────────────────────────────────────────────────────────────────────────

_sync_lock = asyncio.Lock()


@router.post("/sync")
async def trigger_sync(
    background_tasks: BackgroundTasks,
    full_resync: bool = False,
    _: object = Depends(require_permission(_PERM)),
):
    """
    手動觸發核准請購單雙層同步（背景執行，立即回傳 202）。
    full_resync=True 時強制重新同步所有 Detail 資料。
    """
    from app.services.purchase_request_sync import sync_from_ragic

    if _sync_lock.locked():
        return {"status": "already_running", "message": "同步已在執行中，請稍後再試"}

    async def _run():
        async with _sync_lock:
            await sync_from_ragic(full_resync=full_resync)

    background_tasks.add_task(_run)
    return {
        "status":  "started",
        "message": f"同步已啟動（full_resync={full_resync}）",
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /sync/status — 同步狀態查詢
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/sync/status")
def get_sync_status(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    """查詢最近同步記錄 + 各部門最後同步時間。"""
    recent_logs = (
        db.query(ModuleSyncLog)
        .filter(ModuleSyncLog.module_name == "核准請購單")
        .order_by(ModuleSyncLog.started_at.desc())
        .limit(5)
        .all()
    )

    dept_rows = (
        db.query(
            ApprovedPurchaseRequest.department_display,
            func.count(ApprovedPurchaseRequest.id).label("total"),
        )
        .group_by(ApprovedPurchaseRequest.department_display)
        .all()
    )
    synced_rows = (
        db.query(
            ApprovedPurchaseRequest.department_display,
            func.count(ApprovedPurchaseRequest.id).label("synced"),
        )
        .filter(ApprovedPurchaseRequest.detail_synced == True)
        .group_by(ApprovedPurchaseRequest.department_display)
        .all()
    )
    synced_map = {r.department_display: r.synced for r in synced_rows}

    pending_detail_count = (
        db.query(func.count(ApprovedPurchaseRequest.id))
        .filter(ApprovedPurchaseRequest.detail_synced == False)
        .scalar()
        or 0
    )

    return {
        "is_running":           _sync_lock.locked(),
        "pending_detail_count": pending_detail_count,
        "recent_logs": [
            {
                "id":         log.id,
                "module":     log.module_name,
                "trigger":    log.triggered_by,
                "status":     log.status,
                "message":    log.error_msg or f"fetched={log.fetched} upserted={log.upserted}",
                "created_at": log.started_at.isoformat() if log.started_at else None,
            }
            for log in recent_logs
        ],
        "dept_stats": [
            {
                "department_display": r.department_display,
                "total":              r.total,
                "detail_synced":      synced_map.get(r.department_display, 0),
                "pending":            r.total - synced_map.get(r.department_display, 0),
            }
            for r in dept_rows
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /config/departments — 可用部門清單
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/config/departments")
def get_available_departments(
    company: str = Query("樂群"),
    db:      Session = Depends(get_db),
    _:       object  = Depends(require_permission(_PERM)),
):
    """回傳資料庫中實際有資料的部門清單（供前端篩選器使用）。"""
    rows = (
        db.query(ApprovedPurchaseRequest.department_display)
        .filter(ApprovedPurchaseRequest.company == company)
        .distinct()
        .order_by(ApprovedPurchaseRequest.department_display)
        .all()
    )
    return [r[0] for r in rows if r[0]]


@router.get("/config/account-categories")
def get_account_categories(
    company: str = Query("樂群"),
    db:      Session = Depends(get_db),
    _:       object  = Depends(require_permission(_PERM)),
):
    """回傳資料庫中實際有資料的會科清單（供前端篩選器使用），依名稱排序。"""
    rows = (
        db.query(ApprovedPurchaseRequest.account_category)
        .filter(
            ApprovedPurchaseRequest.company == company,
            ApprovedPurchaseRequest.account_category.isnot(None),
            ApprovedPurchaseRequest.account_category != "",
        )
        .distinct()
        .order_by(ApprovedPurchaseRequest.account_category)
        .all()
    )
    return [r[0] for r in rows if r[0]]


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/available-months — 有資料的年月清單（供前端 DatePicker 使用）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/available-months")
def get_available_months(
    company: str = Query("樂群"),
    db:      Session = Depends(get_db),
    _:       object  = Depends(require_permission(_PERM)),
):
    """
    回傳資料庫中實際有核准記錄（status=F + approved_date 非 NULL）的年月清單，降冪排序。
    前端用此清單作為 DatePicker 的可選範圍，並預設選最新一筆。
    """
    rows = (
        db.query(
            func.strftime("%Y-%m", ApprovedPurchaseRequest.approved_date).label("ym")
        )
        .filter(
            ApprovedPurchaseRequest.company == company,
            ApprovedPurchaseRequest.status  == "F",
            ApprovedPurchaseRequest.approved_date.isnot(None),
        )
        .distinct()
        .order_by(
            func.strftime("%Y-%m", ApprovedPurchaseRequest.approved_date).desc()
        )
        .all()
    )
    return [r.ym for r in rows if r.ym]


# ─────────────────────────────────────────────────────────────────────────────
# 共用序列化 helper（order 主單 dict）
# ─────────────────────────────────────────────────────────────────────────────

def _order_to_dict(o: ApprovedPurchaseRequest) -> dict:
    return {
        "id":                 o.id,
        "purchase_no":        o.purchase_no,
        "department_display": o.department_display,
        "account_category":   o.account_category,
        "applicant":          o.applicant,
        "description":        o.description,
        "amount":             o.amount,
        "amount_tax":         o.amount_tax,
        "amount_total":       o.amount_total,
        "status":             o.status,
        "vendor1":            o.vendor1,
        "vendor2":            o.vendor2,
        "vendor3":            o.vendor3,
        "remark":             o.remark,
        "request_date":       o.request_date.isoformat()       if o.request_date    else None,
        "approved_date":      o.approved_date.isoformat()      if o.approved_date   else None,
        "last_updated_at":    o.last_updated_at.strftime("%Y/%m/%d %H:%M") if o.last_updated_at else None,
        "detail_synced":      o.detail_synced,
        "ragic_sheet_path":   o.ragic_sheet_path,
        "ragic_record_id":    o.ragic_record_id,
        "ragic_url":          _ragic_url(o),
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/orders — 請購單清單（訂單級，分頁）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/orders")
def get_approved_orders(
    year_month:       Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from:  Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:    Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    status:           Optional[str] = Query(None, description="F=已核准 / N=待審 / REJ=退回"),
    keyword:          Optional[str] = Query(None, description="全文搜尋（說明/單號/申請人/廠商）"),
    page:             int            = Query(1,  ge=1),
    per_page:         int            = Query(20, le=100),
    company:          str            = Query("樂群"),
    db:               Session        = Depends(get_db),
    _:                object         = Depends(require_permission(_PERM)),
):
    """請購單主單清單（不展開品項），對應 Ragic list_path 清單視圖。"""
    ym_col = func.strftime("%Y-%m", ApprovedPurchaseRequest.approved_date)
    q = db.query(ApprovedPurchaseRequest).filter(
        ApprovedPurchaseRequest.company == company
    )
    # 日期篩選（不強制 status=F，保留原行為可篩全狀態）
    if year_month_from and year_month_to:
        q = q.filter(ym_col >= year_month_from, ym_col <= year_month_to)
    elif year_month_from:
        q = q.filter(ym_col >= year_month_from)
    elif year_month:
        q = q.filter(ym_col == year_month)
    if department:
        q = q.filter(ApprovedPurchaseRequest.department_display == department)
    if account_category:
        q = q.filter(ApprovedPurchaseRequest.account_category == account_category)
    if status:
        q = q.filter(ApprovedPurchaseRequest.status == status)
    if keyword:
        like = f"%{keyword}%"
        q = q.filter(
            ApprovedPurchaseRequest.description.ilike(like)
            | ApprovedPurchaseRequest.purchase_no.ilike(like)
            | ApprovedPurchaseRequest.applicant.ilike(like)
            | ApprovedPurchaseRequest.vendor1.ilike(like)
            | ApprovedPurchaseRequest.vendor2.ilike(like)
        )

    total = q.count()
    orders = (
        q.order_by(ApprovedPurchaseRequest.last_updated_at.desc().nullsfirst())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )
    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_order_to_dict(o) for o in orders],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/orders/{order_id} — 單筆請購單完整資料（含品項）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/orders/{order_id}")
def get_approved_order_detail(
    order_id: int,
    db:       Session = Depends(get_db),
    _:        object  = Depends(require_permission(_PERM)),
):
    """單筆請購單完整資料（含品項清單），用於前端 Detail Drawer。"""
    order = db.query(ApprovedPurchaseRequest).filter(
        ApprovedPurchaseRequest.id == order_id
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="請購單不存在")

    items = (
        db.query(ApprovedPurchaseRequestItem)
        .filter(ApprovedPurchaseRequestItem.order_id == order_id)
        .order_by(ApprovedPurchaseRequestItem.seq)
        .all()
    )

    def item_dict(i: ApprovedPurchaseRequestItem) -> dict:
        return {
            "id":                  i.id,
            "seq":                 i.seq,
            "product_name":        i.product_name,
            "qty":                 i.qty,
            "unit":                i.unit,
            "vendor1_price":       i.vendor1_price,
            "vendor2_price":       i.vendor2_price,
            "vendor3_price":       i.vendor3_price,
            "selected_vendor":     i.selected_vendor,
            "selected_unit_price": i.selected_unit_price,
            "selected_amount":     i.selected_amount,
            "is_confirmed":        i.is_confirmed,
            "item_remark":         i.item_remark,
        }

    return {
        "order": _order_to_dict(order),
        "items": [item_dict(i) for i in items],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /debug/raw-keys — 列出第一筆的 raw_data_json 所有欄位 key（不需認證）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/debug/raw-keys")
def debug_raw_keys(
    db:         Session = Depends(get_db),
    department: str = Query("管理部"),
    limit:      int = Query(3),
):
    """診斷：列出指定部門前 N 筆的 raw_data_json 所有 key，找出日期欄位的真實名稱。"""
    import json as _json
    rows = (
        db.query(
            ApprovedPurchaseRequest.id,
            ApprovedPurchaseRequest.purchase_no,
            ApprovedPurchaseRequest.department_display,
            ApprovedPurchaseRequest.raw_data_json,
        )
        .filter(ApprovedPurchaseRequest.department_display == department)
        .order_by(ApprovedPurchaseRequest.id.desc())
        .limit(limit)
        .all()
    )
    result = []
    for r in rows:
        raw = {}
        if r.raw_data_json:
            try:
                raw = _json.loads(r.raw_data_json)
            except Exception:
                pass
        # 只回傳 key + 值（不含子 dict）方便找日期欄位
        flat = {k: v for k, v in raw.items() if not isinstance(v, dict)}
        result.append({
            "id":           r.id,
            "purchase_no":  r.purchase_no,
            "department":   r.department_display,
            "raw_keys":     flat,
        })
    return result


# ─────────────────────────────────────────────────────────────────────────────
# GET /audit/anomalies — 請購資料異常列表（分頁）
# GET /audit/summary   — 請購資料異常 KPI 摘要
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/audit/anomalies")
def get_purchase_audit_anomalies(
    year_month:      Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from: Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:   Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    department:      Optional[str] = Query(None),
    company:         str           = Query("樂群"),
    rule_code:       Optional[str] = Query(None, description="規則代碼篩選，如 R01"),
    page:            int           = Query(1,  ge=1),
    per_page:        int           = Query(20, ge=1, le=200),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    """請購資料異常明細（分頁），來源固定為 purchase。"""
    from app.services.audit_service import get_anomalies
    return get_anomalies(
        db,
        source="purchase",
        year_month=year_month,
        year_month_from=year_month_from,
        year_month_to=year_month_to,
        department=department,
        company=company,
        rule_code=rule_code,
        page=page,
        per_page=per_page,
    )


@router.get("/audit/summary")
def get_purchase_audit_summary(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None),
    year_month_to:   Optional[str] = Query(None),
    department:      Optional[str] = Query(None),
    company:         str           = Query("樂群"),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    """請購資料異常各規則計數 KPI。"""
    from app.services.audit_service import get_audit_summary
    return get_audit_summary(
        db,
        source="purchase",
        year_month=year_month,
        year_month_from=year_month_from,
        year_month_to=year_month_to,
        department=department,
        company=company,
    )
