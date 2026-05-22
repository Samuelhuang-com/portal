"""
工作日誌 API
Prefix: /api/v1/work-journal

彙整 10 個模組當日工作記錄，依人員分組回傳。
日誌時間基準：報修類用 occurred_at，保養/巡檢類用各自的 scheduled_date / inspection_date / maint_date。

回傳格式：
{
  "date": "2026/05/15",
  "persons": [
    { "person": "王大明", "rows": [ JournalRow, ... ] },
    { "person": "未指定", "rows": [...] }
  ],
  "total_rows": 38
}
"""

import io
import re
from datetime import date as _date, timedelta, datetime
from collections import defaultdict
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Query, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.dependencies import get_current_user

# ── Models ────────────────────────────────────────────────────────────────────
from app.models.schedule                import ScheduleDetail, ShiftType
from app.models.dazhi_repair            import DazhiRepairCase
from app.models.luqun_repair            import LuqunRepairCase
from app.models.periodic_maintenance    import PeriodicMaintenanceBatch, PeriodicMaintenanceItem
from app.models.ihg_room_maintenance    import IHGRoomMaintenanceMaster
from app.models.hotel_daily_inspection  import HotelDIBatch
from app.models.mall_periodic_maintenance import MallPeriodicMaintenanceBatch, MallPeriodicMaintenanceItem
from app.models.full_building_maintenance import FullBldgPMBatch, FullBldgPMItem
from app.models.mall_facility_inspection  import MallFIBatch
from app.models.b1f_inspection  import B1FInspectionBatch
from app.models.b2f_inspection  import B2FInspectionBatch
from app.models.b4f_inspection  import B4FInspectionBatch
from app.models.rf_inspection   import RFInspectionBatch
from app.models.hotel_meter_readings import HotelMRBatch
from app.models.other_tasks          import OtherTask

router = APIRouter(dependencies=[Depends(get_current_user)])

# ── 常數 ─────────────────────────────────────────────────────────────────────

CATEGORIES = ["現場報修", "上級交辦", "緊急事件", "例行維護", "每日巡檢"]

SOURCE_LABEL = {
    "dazhi":        "飯店工務",
    "luqun":        "商場工務",
    "hotel_pm":     "飯店週期保養",
    "ihg":          "IHG客房保養",
    "hotel_di":     "飯店每日巡檢",
    "mall_pm":      "商場週期保養",
    "full_bldg_pm": "整棟保養",
    "mall_fi":      "商場設施巡檢",
    "full_bi":      "整棟巡檢",
    "hotel_mr":     "飯店水電錶抄表",
    "other_tasks":  "主管交辦/緊急事件",
}

SORT_ORDER = list(SOURCE_LABEL.keys())

# ── Ragic 直連 URL ─────────────────────────────────────────────────────────────
# 全部模組均位於 ap12.ragic.com / soutlet001
_RAGIC_BASE    = "https://ap12.ragic.com/soutlet001"

_SOURCE_PATH: dict[str, str] = {
    "dazhi":        "lequn-public-works/8",
    "luqun":        "luqun-public-works-repair-reporting-system/6",
    "hotel_pm":     "periodic-maintenance/8",
    "ihg":          "periodic-maintenance/4",
    "mall_pm":      "periodic-maintenance/18",
    "full_bldg_pm": "periodic-maintenance/21",
    "other_tasks":  "other-tasks/1",
}

# 多 Sheet 來源：sheet_key → Ragic path
_HOTEL_DI_PATHS: dict[str, str] = {
    "rf":     "main-project-inspection/17",
    "4f-10f": "main-project-inspection/18",
    "4f":     "main-project-inspection/19",
    "2f":     "main-project-inspection/20",
    "1f":     "main-project-inspection/21",
}
_MALL_FI_PATHS: dict[str, str] = {
    "4f":      "mall-facility-inspection/2",
    "3f":      "mall-facility-inspection/3",
    "1f-3f":   "mall-facility-inspection/4",
    "1f":      "mall-facility-inspection/5",
    "b1f-b4f": "mall-facility-inspection/7",
}
_FULL_BI_PATHS: dict[str, str] = {
    "b1f": "full-building-inspection/4",
    "b2f": "full-building-inspection/3",
    "b4f": "full-building-inspection/2",
    "rf":  "full-building-inspection/1",
}
_HOTEL_MR_PATHS: dict[str, str] = {
    "building-electric": "hotel-routine-inspection/11",
    "mall-ac-electric":  "hotel-routine-inspection/12",
    "tenant-electric":   "hotel-routine-inspection/14",
    "tenant-water":      "hotel-routine-inspection/15",
}


def _ragic_url(path: str, ragic_id: str) -> str:
    """組出 Ragic 記錄直連 URL。
    多 Sheet 來源的 ragic_id 格式為 '{sheet_key}_{numeric_id}'，取最後一段數字作 record ID。
    """
    if not path or not ragic_id:
        return ""
    # 取尾段：'rf_2431' → '2431'；'2431' → '2431'
    numeric_id = ragic_id.rsplit("_", 1)[-1] if "_" in ragic_id else ragic_id
    return f"{_RAGIC_BASE}/{path}/{numeric_id}"


# ── 工具函數 ───────────────────────────────────────────────────────────────────

def _parse_wm(val: str) -> Optional[int]:
    """'2.50 小時' / '30 分鐘' / '8'（無單位→分鐘）→ int 分鐘，解析不到 → None
    注意：巡檢模組 work_hours 欄位儲存的是純數字分鐘，無單位標記，因此無單位時直接視為分鐘。
    """
    if not val:
        return None
    s = str(val).strip()
    m = re.search(r"[\d.]+", s)
    if not m:
        return None
    num = float(m.group())
    if "分" in s:
        return round(num) if num > 0 else None
    if "時" in s:                          # 含「小時」字樣 → ×60
        mins = round(num * 60)
        return mins if mins > 0 else None
    # 無單位 → 視為分鐘（保全/巡檢 DB 欄位直接儲存分鐘數）
    return round(num) if num > 0 else None


def _clean_time(t: str) -> str:
    """'2026/05/15 09:00' → '09:00'，已是 HH:MM 則直接回傳"""
    if not t:
        return ""
    t = t.strip()
    if " " in t:
        t = t.split(" ")[-1]
    if re.match(r"^\d{1,2}:\d{2}$", t):
        return t
    return t[:5] if len(t) >= 5 else t


def _calc_wm_from_times(start: str, end: str) -> Optional[int]:
    """由 HH:MM 起/迄計算工時分鐘；跨午夜視為 +1 天；失敗回 None"""
    s = _clean_time(start)
    e = _clean_time(end)
    if not s or not e:
        return None
    try:
        sh, sm = map(int, s.split(":"))
        eh, em = map(int, e.split(":"))
        total = (eh * 60 + em) - (sh * 60 + sm)
        if total < 0:
            total += 24 * 60   # 跨午夜
        return total if total > 0 else None
    except Exception:
        return None


def _persons(name: str) -> list[str]:
    """executor_name 可能含多人（空格分隔），拆成 list；空白 → ['未指定']"""
    if not name or not name.strip():
        return ["未指定"]
    parts = [p.strip() for p in name.strip().split() if p.strip()]
    return parts if parts else ["未指定"]


def _date_str(year: int, month: int, day: int) -> str:
    return f"{year}/{month:02d}/{day:02d}"


def _make_row(
    source: str,
    category: str,
    task: str,
    person: str,
    est_min: Optional[int] = None,
    start_time: str = "",
    end_time: str = "",
    work_min: Optional[int] = None,
    remark: str = "",
    report: str = "",
    ragic_id: str = "",
    ragic_url: str = "",
    venue: str = "",
    detail: Optional[dict] = None,
) -> dict:
    st = _clean_time(start_time)
    et = _clean_time(end_time)
    # 無工時時，由起迄時間自動計算
    if work_min is None:
        work_min = _calc_wm_from_times(st, et)
    return {
        "source":       source,
        "source_label": SOURCE_LABEL.get(source, source),
        "category":     category,
        "task":         task.strip(),
        "person":       person or "未指定",
        "est_min":      est_min,
        "start_time":   st,
        "end_time":     et,
        "work_min":     work_min,
        "remark":       (remark or "").strip(),
        "report":       (report or "").strip(),
        "ragic_id":     ragic_id,
        "ragic_url":    ragic_url,
        "venue":        venue,
        "detail":       detail or {},
    }


# ── 10 個模組 helper ──────────────────────────────────────────────────────────

def _fetch_dazhi(db: Session, year: int, month: int, day: int) -> list[dict]:
    """大直工務報修：occurred_at 日期口徑"""
    rows = []
    target = _date(year, month, day)
    for c in db.query(DazhiRepairCase).all():
        if not c.occurred_at:
            continue
        if (c.status or "").strip() == "取消":
            continue
        if c.occurred_at.date() != target:
            continue
        person     = (c.responsible_unit or "").strip() or "未指定"
        task       = " ".join(filter(None, [c.repair_type, c.floor, c.title]))
        wm         = round(c.work_hours * 60) if c.work_hours and c.work_hours > 0 else None
        occ        = c.occurred_at.strftime("%Y/%m/%d %H:%M") if c.occurred_at else ""
        start_t    = c.occurred_at.strftime("%H:%M")  if c.occurred_at  else ""
        end_t      = c.completed_at.strftime("%H:%M") if c.completed_at else ""
        rows.append(_make_row(
            source="dazhi",
            category="現場報修",
            task=task or "(無說明)",
            person=person,
            start_time=start_t,
            end_time=end_t,
            work_min=wm,
            remark=c.finance_note or "",
            ragic_id=c.ragic_id,
            ragic_url=_ragic_url(_SOURCE_PATH["dazhi"], c.ragic_id),
            detail={
                "報修編號":  c.case_no or "",
                "標題":      c.title or "",
                "報修人姓名": c.reporter_name or "",
                "報修類型":  c.repair_type or "",
                "發生樓層":  c.floor or "",
                "發生時間":  occ,
                "負責單位":  c.responsible_unit or "",
                "花費工時":  f"{c.work_hours:.2f} hr" if c.work_hours else "",
                "處理狀況":  c.status or "",
                "委外費用":  str(int(c.outsource_fee or 0)),
                "維修費用":  str(int(c.maintenance_fee or 0)),
                "總費用":    str(int(c.total_fee or 0)),
                "驗收者":    c.acceptor or "",
                "驗收":      c.accept_status or "",
                "結案人":    c.closer or "",
                "結案時間":  c.completed_at.strftime("%Y/%m/%d") if c.completed_at else "",
                "結案天數":  f"{c.close_days:.1f} 天" if c.close_days else "",
                "扣款事項":  c.deduction_item or "",
                "扣款費用":  str(int(c.deduction_fee or 0)),
                "財務備註":  c.finance_note or "",
            },
        ))
    return rows


def _fetch_luqun(db: Session, year: int, month: int, day: int) -> list[dict]:
    """商場工務報修：occurred_at 日期口徑"""
    rows = []
    target = _date(year, month, day)
    for c in db.query(LuqunRepairCase).all():
        if not c.occurred_at:
            continue
        if (c.status or "").strip() == "取消":
            continue
        if c.occurred_at.date() != target:
            continue
        person     = (c.responsible_unit or "").strip() or "未指定"
        task       = " ".join(filter(None, [c.repair_type, c.floor, c.title]))
        wm         = round(c.work_hours * 60) if c.work_hours and c.work_hours > 0 else None
        occ        = c.occurred_at.strftime("%Y/%m/%d %H:%M") if c.occurred_at else ""
        start_t    = c.occurred_at.strftime("%H:%M")  if c.occurred_at  else ""
        end_t      = c.completed_at.strftime("%H:%M") if c.completed_at else ""
        rows.append(_make_row(
            source="luqun",
            category="現場報修",
            task=task or "(無說明)",
            person=person,
            start_time=start_t,
            end_time=end_t,
            work_min=wm,
            remark=c.mgmt_response or c.finance_note or "",
            ragic_id=c.ragic_id,
            ragic_url=_ragic_url(_SOURCE_PATH["luqun"], c.ragic_id),
            detail={
                "報修編號":  c.case_no or "",
                "標題":      c.title or "",
                "報修人姓名": c.reporter_name or "",
                "報修類型":  c.repair_type or "",
                "發生樓層":  c.floor or "",
                "發生時間":  occ,
                "負責單位":  c.responsible_unit or "",
                "花費工時":  f"{c.work_hours:.2f} hr" if c.work_hours else "",
                "處理狀況":  c.status or "",
                "委外費用":  str(int(c.outsource_fee or 0)),
                "維修費用":  str(int(c.maintenance_fee or 0)),
                "總費用":    str(int(c.total_fee or 0)),
                "驗收者":    c.acceptor or "",
                "驗收":      c.accept_status or "",
                "結案人":    c.closer or "",
                "結案時間":  c.completed_at.strftime("%Y/%m/%d") if c.completed_at else "",
                "結案天數":  f"{c.close_days:.1f} 天" if c.close_days else "",
                "扣款事項":  c.deduction_item or "",
                "扣款費用":  str(int(c.deduction_fee or 0)),
                "管理回應":  c.mgmt_response or "",
                "財務備註":  c.finance_note or "",
            },
        ))
    return rows


def _fetch_hotel_pm(db: Session, year: int, month: int, day: int) -> list[dict]:
    """飯店週期保養：batch.period_month + item.scheduled_date"""
    rows = []
    period_month = f"{year}/{month:02d}"
    sched_day    = f"{month:02d}/{day:02d}"

    batches = (
        db.query(PeriodicMaintenanceBatch)
        .filter(PeriodicMaintenanceBatch.period_month == period_month)
        .all()
    )
    if not batches:
        return rows
    batch_map = {b.ragic_id: b for b in batches}
    batch_ids = set(batch_map.keys())

    items = (
        db.query(PeriodicMaintenanceItem)
        .filter(
            PeriodicMaintenanceItem.batch_ragic_id.in_(batch_ids),
            PeriodicMaintenanceItem.scheduled_date == sched_day,
        )
        .all()
    )
    for item in items:
        batch = batch_map.get(item.batch_ragic_id)
        task = " ".join(filter(None, [item.task_name, item.location]))
        wm = None
        if item.ragic_work_minutes and item.ragic_work_minutes > 0:
            wm = int(item.ragic_work_minutes)
        elif item.estimated_minutes and item.estimated_minutes > 0:
            wm = int(item.estimated_minutes)
        est = item.estimated_minutes if item.estimated_minutes else None
        for person in _persons(item.executor_name):
            rows.append(_make_row(
                source="hotel_pm",
                category="例行維護",
                task=task or "(無說明)",
                person=person,
                est_min=est,
                start_time=item.start_time or "",
                end_time=item.end_time or "",
                work_min=wm,
                remark=item.result_note or "",
                ragic_id=item.ragic_id,
                ragic_url=_ragic_url(_SOURCE_PATH["hotel_pm"], item.ragic_id),
                detail={
                    "日誌編號":  (batch.journal_no if batch else "") or "",
                    "保養月份":  (batch.period_month if batch else "") or "",
                    "類別":      item.category or "",
                    "頻率":      item.frequency or "",
                    "區域":      item.location or "",
                    "排定日期":  item.scheduled_date or "",
                    "排定人員":  item.scheduler_name or "",
                    "執行人員":  item.executor_name or "",
                    "完成狀況":  "已完成" if item.is_completed else "未完成",
                    "執行結果":  item.result_note or "",
                    "異常說明":  item.abnormal_note if getattr(item, "abnormal_flag", False) else "",
                },
            ))
    return rows


def _fetch_ihg(db: Session, year: int, month: int, day: int) -> list[dict]:
    """IHG客房保養：maint_date YYYY/MM/DD"""
    rows = []
    date_prefix = _date_str(year, month, day)
    for rec in (
        db.query(IHGRoomMaintenanceMaster)
        .filter(IHGRoomMaintenanceMaster.maint_date == date_prefix)
        .all()
    ):
        person = (rec.assignee_name or "").strip() or "未指定"
        task   = f"IHG客房保養 {rec.room_no}".strip()
        # 工時：優先用 DB 欄位 work_minutes；fallback raw_json 解析
        raw = rec.get_raw() if hasattr(rec, "get_raw") else {}
        wm_float: float | None = None
        if rec.work_minutes is not None and rec.work_minutes > 0:
            wm_float = rec.work_minutes
        else:
            try:
                mins_val = raw.get("工時計算", "")
                if mins_val and re.search(r"[\d.]+", str(mins_val)):
                    wm_float = float(re.search(r"[\d.]+", str(mins_val)).group())
            except Exception:
                pass
        wm = round(wm_float) if wm_float and wm_float > 0 else None
        # 起迄時間：DB 欄位優先，fallback raw_json（未重新同步的舊記錄）
        def _str_or_raw(db_val: str, raw_key: str) -> str:
            return db_val.strip() if db_val and db_val.strip() else str(raw.get(raw_key) or "").strip()
        ihg_start = _clean_time(_str_or_raw(rec.start_time or "", "保養時間起"))
        ihg_end   = _clean_time(_str_or_raw(rec.end_time   or "", "保養時間迄"))
        rows.append(_make_row(
            source="ihg",
            category="例行維護",
            task=task,
            person=person,
            est_min=wm,
            start_time=ihg_start,
            end_time=ihg_end,
            work_min=wm,
            remark=rec.notes or "",
            ragic_id=rec.ragic_id,
            ragic_url=_ragic_url(_SOURCE_PATH["ihg"], rec.ragic_id),
            detail={
                "房號":       rec.room_no or "",
                "樓層":       rec.floor or "",
                "保養類型":   rec.maint_type or "",
                "保養人員":   rec.assignee_name or "",
                "複核人員":   rec.checker_name or "",
                "保養日期":   rec.maint_date or "",
                "完成日期":   rec.completion_date or "",
                "狀態":       rec.status or "",
                "保養時間起": ihg_start,
                "保養時間迄": ihg_end,
                "工時（分鐘）": str(wm) if wm else "",
                "備註":       rec.notes or "",
            },
        ))
    return rows


def _fetch_hotel_di(db: Session, year: int, month: int, day: int) -> list[dict]:
    """飯店每日巡檢：hotel_di_inspection_batch"""
    rows = []
    date_str = _date_str(year, month, day)
    for b in (
        db.query(HotelDIBatch)
        .filter(HotelDIBatch.inspection_date == date_str)
        .all()
    ):
        person = (b.inspector_name or "").strip() or "未指定"
        task   = f"{b.sheet_name} 巡檢" if b.sheet_name else "飯店每日巡檢"
        wm     = _parse_wm(b.work_hours)
        rows.append(_make_row(
            source="hotel_di",
            category="每日巡檢",
            task=task,
            person=person,
            start_time=b.start_time or "",
            end_time=b.end_time or "",
            work_min=wm,
            ragic_id=b.ragic_id,
            ragic_url=_ragic_url(_HOTEL_DI_PATHS.get(b.sheet_key, ""), b.ragic_id),
            detail={
                "巡檢表名稱": b.sheet_name or "",
                "巡檢人員":   b.inspector_name or "",
                "巡檢日期":   b.inspection_date or "",
                "開始時間":   b.start_time or "",
                "結束時間":   b.end_time or "",
                "工時":       f"{wm} min" if wm else "",
            },
        ))
    return rows


def _fetch_mall_pm(db: Session, year: int, month: int, day: int) -> list[dict]:
    """商場週期保養：mall_pm_batch + mall_pm_batch_item"""
    rows = []
    period_month = f"{year}/{month:02d}"
    sched_day    = f"{month:02d}/{day:02d}"

    batches = (
        db.query(MallPeriodicMaintenanceBatch)
        .filter(MallPeriodicMaintenanceBatch.period_month == period_month)
        .all()
    )
    if not batches:
        return rows
    batch_map = {b.ragic_id: b for b in batches}
    batch_ids = set(batch_map.keys())

    items = (
        db.query(MallPeriodicMaintenanceItem)
        .filter(
            MallPeriodicMaintenanceItem.batch_ragic_id.in_(batch_ids),
            MallPeriodicMaintenanceItem.scheduled_date == sched_day,
        )
        .all()
    )
    for item in items:
        batch = batch_map.get(item.batch_ragic_id)
        task = " ".join(filter(None, [item.task_name, item.location]))
        est  = item.estimated_minutes if item.estimated_minutes else None
        wm   = int(item.estimated_minutes) if item.estimated_minutes and item.estimated_minutes > 0 else None
        for person in _persons(item.executor_name):
            rows.append(_make_row(
                source="mall_pm",
                category="例行維護",
                task=task or "(無說明)",
                person=person,
                est_min=est,
                start_time=item.start_time or "",
                end_time=item.end_time or "",
                work_min=wm,
                remark=item.result_note or "",
                ragic_id=item.ragic_id,
                ragic_url=_ragic_url(_SOURCE_PATH["mall_pm"], item.ragic_id),
                detail={
                    "日誌編號":  (batch.journal_no if batch else "") or "",
                    "保養月份":  (batch.period_month if batch else "") or "",
                    "類別":      item.category or "",
                    "頻率":      item.frequency or "",
                    "區域":      item.location or "",
                    "排定日期":  item.scheduled_date or "",
                    "排定人員":  item.scheduler_name or "",
                    "執行人員":  item.executor_name or "",
                    "完成狀況":  "已完成" if item.is_completed else "未完成",
                    "執行結果":  item.result_note or "",
                    "異常說明":  item.abnormal_note if getattr(item, "abnormal_flag", False) else "",
                },
            ))
    return rows


def _fetch_full_bldg_pm(db: Session, year: int, month: int, day: int) -> list[dict]:
    """整棟保養：full_bldg_pm_batch + full_bldg_pm_batch_item"""
    rows = []
    period_month = f"{year}/{month:02d}"
    sched_day    = f"{month:02d}/{day:02d}"

    batches = (
        db.query(FullBldgPMBatch)
        .filter(FullBldgPMBatch.period_month == period_month)
        .all()
    )
    if not batches:
        return rows
    batch_map = {b.ragic_id: b for b in batches}
    batch_ids = set(batch_map.keys())

    items = (
        db.query(FullBldgPMItem)
        .filter(
            FullBldgPMItem.batch_ragic_id.in_(batch_ids),
            FullBldgPMItem.scheduled_date == sched_day,
        )
        .all()
    )
    for item in items:
        batch = batch_map.get(item.batch_ragic_id)
        task = " ".join(filter(None, [item.task_name, item.location]))
        est  = item.estimated_minutes if item.estimated_minutes else None
        wm   = int(item.estimated_minutes) if item.estimated_minutes and item.estimated_minutes > 0 else None
        for person in _persons(item.executor_name):
            rows.append(_make_row(
                source="full_bldg_pm",
                category="例行維護",
                task=task or "(無說明)",
                person=person,
                est_min=est,
                start_time=item.start_time or "",
                end_time=item.end_time or "",
                work_min=wm,
                remark=item.result_note or "",
                ragic_id=item.ragic_id,
                ragic_url=_ragic_url(_SOURCE_PATH["full_bldg_pm"], item.ragic_id),
                detail={
                    "日誌編號":  (batch.journal_no if batch else "") or "",
                    "保養月份":  (batch.period_month if batch else "") or "",
                    "類別":      item.category or "",
                    "頻率":      item.frequency or "",
                    "區域":      item.location or "",
                    "排定日期":  item.scheduled_date or "",
                    "排定人員":  item.scheduler_name or "",
                    "執行人員":  item.executor_name or "",
                    "完成狀況":  "已完成" if item.is_completed else "未完成",
                    "執行結果":  item.result_note or "",
                    "異常說明":  item.abnormal_note if getattr(item, "abnormal_flag", False) else "",
                },
            ))
    return rows


def _fetch_mall_fi(db: Session, year: int, month: int, day: int) -> list[dict]:
    """商場設施巡檢：mall_fi_inspection_batch"""
    rows = []
    date_str = _date_str(year, month, day)
    for b in (
        db.query(MallFIBatch)
        .filter(MallFIBatch.inspection_date == date_str)
        .all()
    ):
        person = (b.inspector_name or "").strip() or "未指定"
        task   = f"{b.sheet_name} 設施巡檢" if b.sheet_name else "商場設施巡檢"
        wm     = _parse_wm(b.work_hours)
        rows.append(_make_row(
            source="mall_fi",
            category="每日巡檢",
            task=task,
            person=person,
            start_time=b.start_time or "",
            end_time=b.end_time or "",
            work_min=wm,
            ragic_id=b.ragic_id,
            ragic_url=_ragic_url(_MALL_FI_PATHS.get(b.sheet_key, ""), b.ragic_id),
            detail={
                "巡檢表名稱": b.sheet_name or "",
                "巡檢人員":   b.inspector_name or "",
                "巡檢日期":   b.inspection_date or "",
                "開始時間":   b.start_time or "",
                "結束時間":   b.end_time or "",
                "工時":       f"{wm} min" if wm else "",
            },
        ))
    return rows


def _fetch_full_bi(db: Session, year: int, month: int, day: int) -> list[dict]:
    """整棟巡檢：B1F / B2F / B4F / RF inspection_batch"""
    rows = []
    date_str = _date_str(year, month, day)
    SHEETS = [
        (B1FInspectionBatch, "B1F 整棟巡檢", "b1f"),
        (B2FInspectionBatch, "B2F 整棟巡檢", "b2f"),
        (B4FInspectionBatch, "B4F 整棟巡檢", "b4f"),
        (RFInspectionBatch,  "RF 整棟巡檢",  "rf"),
    ]
    for Model, label, bi_key in SHEETS:
        for b in (
            db.query(Model)
            .filter(Model.inspection_date == date_str)
            .all()
        ):
            person = (b.inspector_name or "").strip() or "未指定"
            wm     = _parse_wm(b.work_hours)
            rows.append(_make_row(
                source="full_bi",
                category="每日巡檢",
                task=label,
                person=person,
                start_time=b.start_time or "",
                end_time=b.end_time or "",
                work_min=wm,
                ragic_id=b.ragic_id,
                ragic_url=_ragic_url(_FULL_BI_PATHS[bi_key], b.ragic_id),
                detail={
                    "巡檢表名稱": label,
                    "巡檢人員":   b.inspector_name or "",
                    "巡檢日期":   b.inspection_date or "",
                    "開始時間":   b.start_time or "",
                    "結束時間":   b.end_time or "",
                    "工時":       f"{wm} min" if wm else "",
                },
            ))
    return rows


def _fetch_hotel_mr(db: Session, year: int, month: int, day: int) -> list[dict]:
    """飯店水電錶抄表：hotel_mr_batch（4 張 Sheet，每張獨立一行）"""
    rows = []
    date_str = _date_str(year, month, day)
    for b in (
        db.query(HotelMRBatch)
        .filter(HotelMRBatch.record_date == date_str)
        .all()
    ):
        person = (b.recorder_name or "").strip() or "未指定"
        task   = f"{b.sheet_name} 抄表" if b.sheet_name else "水電錶抄表"
        wm     = _parse_wm(b.work_hours)
        rows.append(_make_row(
            source="hotel_mr",
            category="每日巡檢",
            task=task,
            person=person,
            start_time=b.start_time or "",
            end_time=b.end_time or "",
            work_min=wm,
            ragic_id=b.ragic_id,
            ragic_url=_ragic_url(_HOTEL_MR_PATHS.get(b.sheet_key, ""), b.ragic_id),
            detail={
                "抄表表名稱": b.sheet_name or "",
                "抄表人員":   b.recorder_name or "",
                "抄表日期":   b.record_date or "",
                "開始時間":   b.start_time or "",
                "結束時間":   b.end_time or "",
                "工時":       f"{wm} min" if wm else "",
            },
        ))
    return rows


def _fetch_other_tasks(db: Session, year: int, month: int, day: int) -> list[dict]:
    """主管交辦／緊急事件：created_at 日期口徑"""
    rows = []
    target = _date(year, month, day)
    for rec in db.query(OtherTask).all():
        if not rec.created_at:
            continue
        if rec.created_at.date() != target:
            continue
        wm = round(rec.work_hours * 60) if rec.work_hours and rec.work_hours > 0 else None
        created_str = rec.created_at.strftime("%Y/%m/%d %H:%M")
        updated_str = rec.updated_at.strftime("%Y/%m/%d %H:%M") if rec.updated_at else ""
        ragic_url_val = _ragic_url(_SOURCE_PATH["other_tasks"], rec.ragic_id) if rec.ragic_id else ""
        for person in _persons(rec.engineer):
            rows.append(_make_row(
                source="other_tasks",
                category=rec.task_type or "上級交辦",
                task=rec.description or "(無說明)",
                person=person,
                work_min=wm,
                remark=rec.notes or "",
                ragic_id=rec.ragic_id or "",
                ragic_url=ragic_url_val,
                venue=rec.venue or "",
                detail={
                    "歸屬":         rec.venue or "",
                    "屬性":         rec.task_type or "",
                    "交辦主管":     rec.supervisor or "",
                    "工程人員":     rec.engineer or "",
                    "建立日期":     created_str,
                    "最後更新日期": updated_str,
                    "狀態":         rec.status or "",
                    "維修工時":     f"{rec.work_hours:.2f} hr" if rec.work_hours else "",
                    "問題說明":     rec.description or "",
                    "備註":         rec.notes or "",
                },
            ))
    return rows


# ── 端點 ─────────────────────────────────────────────────────────────────────

def _build_daily(db: Session, year: int, month: int, day: int) -> dict:
    """內部共用：聚合指定日期的全模組工作記錄，回傳 WorkJournalDaily dict。"""
    all_rows: list[dict] = []
    all_rows += _fetch_dazhi(db, year, month, day)
    all_rows += _fetch_luqun(db, year, month, day)
    all_rows += _fetch_hotel_pm(db, year, month, day)
    all_rows += _fetch_ihg(db, year, month, day)
    all_rows += _fetch_hotel_di(db, year, month, day)
    all_rows += _fetch_mall_pm(db, year, month, day)
    all_rows += _fetch_full_bldg_pm(db, year, month, day)
    all_rows += _fetch_mall_fi(db, year, month, day)
    all_rows += _fetch_full_bi(db, year, month, day)
    all_rows += _fetch_hotel_mr(db, year, month, day)
    all_rows += _fetch_other_tasks(db, year, month, day)

    person_map: dict[str, list[dict]] = defaultdict(list)
    for r in all_rows:
        person_map[r["person"]].append(r)

    named = sorted(
        [p for p in person_map if p != "未指定"],
        key=lambda p: sum(r["work_min"] or 0 for r in person_map[p]),
        reverse=True,
    )
    persons_order = named + (["未指定"] if "未指定" in person_map else [])

    def _row_sort_key(r: dict):
        st = r["start_time"]
        has_time = bool(st and re.match(r"^\d{1,2}:\d{2}$", st))
        src_order = SORT_ORDER.index(r["source"]) if r["source"] in SORT_ORDER else 99
        return (0 if has_time else 1, st or "99:99", src_order)

    persons_data = []
    for p in persons_order:
        sorted_rows = sorted(person_map[p], key=_row_sort_key)
        for i, row in enumerate(sorted_rows, 1):
            row["seq"] = i
        persons_data.append({"person": p, "rows": sorted_rows})

    return {
        "date":       _date_str(year, month, day),
        "persons":    persons_data,
        "total_rows": len(all_rows),
    }


@router.get("/daily", summary="工作日誌 — 指定日期所有人員工作記錄")
def get_work_journal_daily(
    year:  int = Query(..., ge=2020, le=2030),
    month: int = Query(..., ge=1,    le=12),
    day:   int = Query(..., ge=1,    le=31),
    db:    Session = Depends(get_db),
):
    """聚合 10 個模組當日工作記錄，依人員分組回傳。"""
    return _build_daily(db, year, month, day)


@router.get("/range", summary="工作日誌 — 日期區間（最多 31 天）")
def get_work_journal_range(
    date_from: str = Query(..., description="起始日期 YYYY-MM-DD"),
    date_to:   str = Query(..., description="結束日期 YYYY-MM-DD"),
    db: Session = Depends(get_db),
):
    """
    聚合 date_from～date_to 區間的工作記錄（最多 31 天）。
    僅回傳有記錄的日期（total_rows > 0）。
    整月查詢：date_from=YYYY-MM-01, date_to=YYYY-MM-{last_day}。
    """
    try:
        start = _date.fromisoformat(date_from)
        end   = _date.fromisoformat(date_to)
    except ValueError:
        return {"date_from": date_from, "date_to": date_to, "days": [], "total_rows": 0}

    if end < start:
        start, end = end, start
    if (end - start).days > 30:
        end = start + timedelta(days=30)

    days = []
    cur  = start
    while cur <= end:
        daily = _build_daily(db, cur.year, cur.month, cur.day)
        if daily["total_rows"] > 0:
            days.append(daily)
        cur += timedelta(days=1)

    return {
        "date_from": start.isoformat(),
        "date_to":   end.isoformat(),
        "days":      days,
        "total_rows": sum(d["total_rows"] for d in days),
    }


@router.get("/export-excel", summary="工作日誌匯出 Excel（每人一 Sheet）")
def export_work_journal_excel(
    date_from: str = Query(..., description="起始日期 YYYY-MM-DD"),
    date_to:   str = Query(..., description="結束日期 YYYY-MM-DD"),
    person:    Optional[str] = Query(None, description="指定人員；不傳則匯出全員"),
    db: Session = Depends(get_db),
):
    """匯出指定日期區間工作日誌為 Excel；全員模式每人一 Sheet，最多 93 天。"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl 未安裝")

    try:
        start = _date.fromisoformat(date_from)
        end   = _date.fromisoformat(date_to)
    except ValueError:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="日期格式錯誤，需 YYYY-MM-DD")

    if end < start:
        start, end = end, start
    if (end - start).days > 92:
        end = start + timedelta(days=92)

    # ── 收集資料 ───────────────────────────────────────────────────────────────
    days_data = []
    cur = start
    while cur <= end:
        daily = _build_daily(db, cur.year, cur.month, cur.day)
        if daily["total_rows"] > 0:
            days_data.append(daily)
        cur += timedelta(days=1)

    # ── 人員清單 ───────────────────────────────────────────────────────────────
    if person:
        persons_order = [person]
    else:
        seen: set = set()
        persons_order = []
        for daily in days_data:
            for pd in daily["persons"]:
                if pd["person"] not in seen:
                    persons_order.append(pd["person"])
                    seen.add(pd["person"])

    # ── 班別對照表：{ staff_name: { "YYYY-MM-DD": { name, start, end } } } ────────
    # 查詢區間內所有相關人員的班表明細，join ShiftType 取名稱與時間
    _shift_type_map: dict[str, dict] = {
        s.code: {"name": s.name, "start": s.start_time, "end": s.end_time}
        for s in db.query(ShiftType).filter(ShiftType.is_deleted == False).all()
    }
    _details = (
        db.query(ScheduleDetail)
        .filter(
            ScheduleDetail.work_date >= start,
            ScheduleDetail.work_date <= end,
            ScheduleDetail.is_deleted == False,
        )
        .all()
    )
    # shift_lookup[staff_name][date_iso] = "班別名稱　HH:MM-HH:MM"
    shift_lookup: dict[str, dict[str, str]] = {}
    for _d in _details:
        _sn = _d.staff_name
        _ds = _d.work_date.isoformat()          # "2026-05-20"
        _st = _shift_type_map.get(_d.shift_code, {})
        _name  = _st.get("name", "") or _d.shift_code
        _start = _d.start_time or _st.get("start", "")
        _end   = _d.end_time   or _st.get("end",   "")
        _time_range = f"{_start}-{_end}" if _start and _end else (_start or "")
        _label = f"{_name}　{_time_range}".rstrip() if _time_range else _name
        shift_lookup.setdefault(_sn, {})[_ds] = _label

    def _shift_label(pname: str, date_iso: str) -> str:
        """取某人某天的班別顯示字串，查不到回空字串。"""
        return shift_lookup.get(pname, {}).get(date_iso, "")

    # ── openpyxl helpers ──────────────────────────────────────────────────────────
    from openpyxl.utils import get_column_letter

    THIN   = Side(style="thin")
    MEDIUM = Side(style="medium")

    def _border(l="thin", r="thin", t="thin", b="thin"):
        def _s(x):
            return Side(style=x) if x else Side(style=None)
        return Border(left=_s(l), right=_s(r), top=_s(t), bottom=_s(b))

    def _apply_row(ws, ri, n_cols, fill=None, font=None, alignment=None, border_fn=None, height=None):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=ri, column=c)
            if fill:      cell.fill      = fill
            if font:      cell.font      = font
            if alignment: cell.alignment = alignment
            if border_fn: cell.border    = border_fn(c)
        if height is not None:
            ws.row_dimensions[ri].height = height

    # ── 飯店 / 商場 來源集合（供工作事項前綴用）────────────────────────────────────
    _HOTEL_SRC = {"dazhi", "hotel_pm", "ihg", "hotel_di"}

    def _venue_prefix(source: str, venue: str = "") -> str:
        """回傳工作事項前綴（飯：/ 商：）。
        other_tasks 以「歸屬」欄位為準；其餘來源依 source 判斷。"""
        if source == "other_tasks":
            return "飯：" if venue == "飯店" else "商："
        return "飯：" if source in _HOTEL_SRC else "商："

    def _row_height(task: str, remark: str, report: str) -> float:
        import math
        def _lines(text: str, col_w: float) -> int:
            """估算折行數：依欄寬（Excel 字元單位），中文字算 2、ASCII 算 1；
            文字中若含換行符則逐段計算後加總。"""
            if not text:
                return 1
            total = 0
            for seg in text.replace('\r\n', '\n').replace('\r', '\n').split('\n'):
                w = sum(2 if ord(c) > 127 else 1 for c in seg)
                total += max(1, math.ceil(w / col_w)) if w else 1
            return total
        max_lines = max(_lines(task, 46), _lines(remark, 22), _lines(report, 34))
        return max(20.0, max_lines * 15.0) + 2

    # ── 樣式常數 ──────────────────────────────────────────────────────────────────
    NCOLS       = 13
    CAT_COLS    = {"現場報修": 2, "上級交辦": 3, "緊急事件": 4, "例行維護": 5, "每日巡檢": 6}
    COL_WIDTHS  = {1:5, 2:7, 3:7, 4:7, 5:7, 6:7, 7:46, 8:9, 9:10, 10:10, 11:9, 12:22, 13:34}

    TITLE_FONT  = Font(bold=True, size=18)
    META_FONT   = Font(size=10)
    HDR_FILL    = PatternFill(start_color="D9EAF7",  end_color="D9EAF7",  fill_type="solid")
    HDR_FONT    = Font(bold=True, size=10)
    DATE_FILL   = PatternFill(start_color="1B3A5C",  end_color="1B3A5C",  fill_type="solid")
    DATE_FONT   = Font(bold=True, size=11, color="FFFFFF")
    TOTAL_FILL  = PatternFill(start_color="EBF3FB",  end_color="EBF3FB",  fill_type="solid")
    FOOTER_FILL = PatternFill(start_color="D9EAF7",  end_color="D9EAF7",  fill_type="solid")

    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)

    # ── 標題文字（支援 YYYY-MM-DD 與 YYYY/MM/DD 兩種格式） ────────────────────────
    def _parse_dt(s: str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                pass
        raise ValueError(f"Cannot parse date: {s}")

    try:
        df_dt = _parse_dt(date_from)
        dt_dt = _parse_dt(date_to)
        if df_dt.year == dt_dt.year and df_dt.month == dt_dt.month:
            yyyymm_txt = f"{df_dt.month}月"          # e.g. "5月" — matches paper form
        else:
            yyyymm_txt = f"{df_dt.year}/{df_dt.month:02d}~{dt_dt.year}/{dt_dt.month:02d}"
        date_display = date_from if date_from == date_to else f"{date_from} ~ {date_to}"
    except Exception:
        yyyymm_txt   = date_from
        date_display = f"{date_from} ~ {date_to}"

    is_multi_day = len(days_data) > 1

    # ── 建立 Workbook ─────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for pname in persons_order:
        ws = wb.create_sheet(title=(pname or "未指定")[:31])

        # 欄寬
        for ci, w in COL_WIDTHS.items():
            ws.column_dimensions[get_column_letter(ci)].width = w

        # ── Row 1: 大標題 ─────────────────────────────────────────────────────────
        # 單日模式：人員名稱後附上班別名稱 + 上下班時間
        if not is_multi_day:
            _single_date_iso = start.isoformat()   # 單日 start == end
            _sl = _shift_label(pname, _single_date_iso)
            title_val = f"{yyyymm_txt} 飯店每日工作日誌 - {pname}　{_sl}".rstrip() if _sl \
                        else f"{yyyymm_txt} 飯店每日工作日誌 - {pname}"
        else:
            title_val = f"{yyyymm_txt} 飯店每日工作日誌 - {pname}"
        c1 = ws.cell(row=1, column=1, value=title_val)
        c1.font = TITLE_FONT
        c1.alignment = CENTER
        ws.merge_cells("A1:M1")
        ws.row_dimensions[1].height = 28

        # ── Row 2: 日期行 ─────────────────────────────────────────────────────────
        # G 欄寬 48，把班別+時間合在同一格避免 H 欄（寬9）換行
        ws.cell(row=2, column=1, value="日期").font = META_FONT
        ws.cell(row=2, column=1).alignment = CENTER
        ws.cell(row=2, column=2, value=f"　{date_display}").font = META_FONT
        ws.cell(row=2, column=2).alignment = LEFT
        ws.merge_cells("B2:F2")
        ws.cell(row=2, column=7, value="早班值班　08:30-17:30").font = META_FONT
        ws.cell(row=2, column=7).alignment = LEFT
        ws.merge_cells("G2:H2")
        ws.cell(row=2, column=9, value="簽名：").font = META_FONT
        ws.cell(row=2, column=9).alignment = LEFT
        ws.merge_cells("I2:M2")
        ws.row_dimensions[2].height = 20

        # ── Row 3: 天氣行 ─────────────────────────────────────────────────────────
        ws.cell(row=3, column=1, value="天氣").font = META_FONT
        ws.cell(row=3, column=1).alignment = CENTER
        ws.cell(row=3, column=2, value="□ 晴").font = META_FONT
        ws.cell(row=3, column=2).alignment = CENTER
        ws.cell(row=3, column=3, value="□ 陰").font = META_FONT
        ws.cell(row=3, column=3).alignment = CENTER
        ws.cell(row=3, column=4, value="□ 雨").font = META_FONT
        ws.cell(row=3, column=4).alignment = LEFT
        ws.merge_cells("D3:F3")
        ws.cell(row=3, column=7, value="晚班值班　14:00-23:00").font = META_FONT
        ws.cell(row=3, column=7).alignment = LEFT
        ws.merge_cells("G3:H3")
        ws.cell(row=3, column=9, value="簽名：").font = META_FONT
        ws.cell(row=3, column=9).alignment = LEFT
        ws.merge_cells("I3:M3")
        ws.row_dimensions[3].height = 20

        # ── Row 4: 溫度行 ─────────────────────────────────────────────────────────
        ws.cell(row=4, column=1, value="溫度").font = META_FONT
        ws.cell(row=4, column=1).alignment = CENTER
        ws.cell(row=4, column=2, value="　　　度").font = META_FONT
        ws.cell(row=4, column=2).alignment = LEFT
        ws.merge_cells("B4:F4")
        ws.merge_cells("G4:M4")
        ws.row_dimensions[4].height = 20

        # ── Row 5: 空白分隔 ───────────────────────────────────────────────────────
        ws.merge_cells("A5:M5")
        ws.row_dimensions[5].height = 8

        # ── Row 6: 欄位表頭 ───────────────────────────────────────────────────────
        HDR_LABELS = [
            "項次", "現場\n報修", "上級\n交辦", "緊急\n事件", "例行\n維護", "每日\n巡檢",
            "工作事項", "預估\n耗時\n(min)", "執行時間\n起", "執行時間\n迄", "工時\n(min)", "備註", "回報事項",
        ]
        for ci, label in enumerate(HDR_LABELS, 1):
            cell = ws.cell(row=6, column=ci, value=label)
            cell.fill      = HDR_FILL
            cell.font      = HDR_FONT
            cell.alignment = CENTER
            l_style = "medium" if ci == 1    else "thin"
            r_style = "medium" if ci == NCOLS else "thin"
            cell.border = _border(l=l_style, r=r_style, t="medium", b="medium")
        ws.row_dimensions[6].height = 42

        # ── 資料行 ────────────────────────────────────────────────────────────────
        row_idx   = 7
        total_min = 0
        global_seq = 0  # sequential item number (resets per day for multi-day)

        for daily in days_data:
            person_rows_for_day: list = []
            for pd in daily["persons"]:
                if pd["person"] == pname:
                    person_rows_for_day = pd["rows"]
                    break
            if not person_rows_for_day:
                continue

            # 多日模式：日期分隔行（日期 + 當日班別名稱 + 上班時間）
            if is_multi_day:
                # daily["date"] 格式 "2026/05/20" → 轉 ISO 查班別
                _day_iso = daily["date"].replace("/", "-")
                _sl = _shift_label(pname, _day_iso)
                _sep_val = f"  {daily['date']}　{_sl}".rstrip() if _sl \
                           else f"  {daily['date']}"
                date_cell = ws.cell(row=row_idx, column=1, value=_sep_val)
                date_cell.fill      = DATE_FILL
                date_cell.font      = DATE_FONT
                date_cell.alignment = LEFT
                ws.merge_cells(f"A{row_idx}:L{row_idx}")
                ws.row_dimensions[row_idx].height = 22
                row_idx += 1
                day_seq = 0  # reset seq per day

            # 依起始時間排序
            sorted_rows = sorted(
                person_rows_for_day,
                key=lambda r: r.get("start_time") or "99:99",
            )

            for r in sorted_rows:
                if is_multi_day:
                    day_seq += 1
                    seq_val = day_seq
                else:
                    global_seq += 1
                    seq_val = global_seq

                cat = r.get("category", "")
                wm  = r.get("work_min")
                if isinstance(wm, int):
                    total_min += wm
                cat_col = CAT_COLS.get(cat)

                # 項次
                a = ws.cell(row=row_idx, column=1, value=seq_val)
                a.alignment = CENTER
                a.font      = Font(size=10)
                a.border    = _border(l="medium")

                # 類別勾選欄 B-F
                for c in range(2, 7):
                    mark = "✓" if c == cat_col else ""
                    cell = ws.cell(row=row_idx, column=c, value=mark)
                    cell.alignment = CENTER
                    cell.font      = Font(size=10, bold=(mark == "✓"))
                    cell.border    = _border()

                # 工作事項（前綴飯：/ 商：）
                _task_val = _venue_prefix(r.get("source", ""), r.get("venue", "")) + r.get("task", "")
                g = ws.cell(row=row_idx, column=7, value=_task_val)
                g.alignment = LEFT
                g.font      = Font(size=10)
                g.border    = _border()

                # 預估耗時
                h = ws.cell(row=row_idx, column=8, value=r.get("est_min") or "")
                h.alignment = CENTER
                h.font      = Font(size=10)
                h.border    = _border()

                # 起
                i = ws.cell(row=row_idx, column=9, value=r.get("start_time") or "")
                i.alignment = CENTER
                i.font      = Font(size=10)
                i.border    = _border()

                # 迄
                j = ws.cell(row=row_idx, column=10, value=r.get("end_time") or "")
                j.alignment = CENTER
                j.font      = Font(size=10)
                j.border    = _border()

                # 工時(min)
                k = ws.cell(row=row_idx, column=11, value=wm if wm is not None else "")
                k.alignment = CENTER
                k.font      = Font(size=10, bold=True, color="1B3A5C") if wm is not None else Font(size=10)
                k.border    = _border()

                # 備註
                l_cell = ws.cell(row=row_idx, column=12, value=r.get("remark", ""))
                l_cell.alignment = LEFT
                l_cell.font      = Font(size=10)
                l_cell.border    = _border()

                # 回報事項
                m_cell = ws.cell(row=row_idx, column=13, value=r.get("report", ""))
                m_cell.alignment = LEFT
                m_cell.font      = Font(size=10, color="D46B08") if r.get("report") else Font(size=10)
                m_cell.border    = _border(r="medium")

                ws.row_dimensions[row_idx].height = _row_height(
                    _task_val,
                    r.get("remark", "") or "",
                    r.get("report", "") or "",
                )
                row_idx += 1

        # ── 合計行 ────────────────────────────────────────────────────────────────
        total_row = row_idx
        for c in range(1, NCOLS + 1):
            cell = ws.cell(row=total_row, column=c)
            cell.fill   = TOTAL_FILL
            cell.border = _border(
                l="medium" if c == 1     else "thin",
                r="medium" if c == NCOLS else "thin",
                t="medium", b="medium",
            )
            cell.alignment = CENTER
            cell.font      = Font(size=10)

        # 合計工時欄（L:M 合併）
        tc = ws.cell(row=total_row, column=12, value=f"合計工時：{total_min} min")
        tc.font      = Font(bold=True, size=10, color="1B3A5C")
        tc.alignment = RIGHT
        ws.merge_cells(f"L{total_row}:M{total_row}")
        ws.row_dimensions[total_row].height = 22
        row_idx += 1

        # ── Footer: 主管覆核 / 值班人員簽名 / 備註 ───────────────────────────────
        for fr in range(row_idx, row_idx + 3):
            ws.row_dimensions[fr].height = 22

        # Footer — 先設樣式再 merge，避免對 MergedCell 設 fill/border 時 AttributeError
        # 每格先設 fill + border，再做 merge_cells（merge 只影響顯示，不影響已設的樣式）
        for fr in range(row_idx, row_idx + 3):
            for c in range(1, NCOLS + 1):
                cell = ws.cell(row=fr, column=c)
                cell.fill   = FOOTER_FILL
                cell.border = _border(
                    l="medium" if c == 1     else "thin",
                    r="medium" if c == NCOLS else "thin",
                    t="thin",
                    b="medium",
                )

        # 主管覆核
        ws.cell(row=row_idx, column=1, value="主管覆核").font      = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=row_idx, column=1).alignment = CENTER
        ws.cell(row=row_idx, column=1).fill      = PatternFill("solid", fgColor="1B3A5C")
        ws.merge_cells(f"A{row_idx}:D{row_idx + 2}")

        # 值班人員簽名
        ws.cell(row=row_idx, column=5, value="值班人員簽名").font      = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=row_idx, column=5).alignment = CENTER
        ws.cell(row=row_idx, column=5).fill      = PatternFill("solid", fgColor="1B3A5C")
        ws.merge_cells(f"E{row_idx}:H{row_idx + 2}")

        # 備註
        ws.cell(row=row_idx, column=9, value="備註").font      = Font(bold=True, size=10, color="FFFFFF")
        ws.cell(row=row_idx, column=9).alignment = CENTER
        ws.cell(row=row_idx, column=9).fill      = PatternFill("solid", fgColor="1B3A5C")
        ws.merge_cells(f"I{row_idx}:M{row_idx + 2}")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label         = f"{date_from}_{date_to}" + (f"_{person}" if person else "")
    filename_safe = f"work_journal_{date_from}_{date_to}.xlsx"   # ASCII only — no Chinese chars
    filename_cn   = f"工作日誌_{label}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_safe}"; '
                f"filename*=UTF-8''{quote(filename_cn)}"
            ),
        },
    )
