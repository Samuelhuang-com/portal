"""
合約管理系統 - API 路由層

router 掛載於 /api/v1/contract（見 main.py）

端點一覽：
  GET    /                       合約列表（分頁、篩選、排序）
  POST   /                       新增合約
  GET    /stats                  合約統計
  POST   /sync                   立即同步（v2 Portal 版回傳空結果）
  GET    /dashboard/kpi          Dashboard KPI 指標
  GET    /dashboard/by-dept      Dashboard 部門金額分組
  GET    /budget-analysis        預算執行率分析（依預算科目聚合）
  GET    /{contract_id}          單筆合約詳情
  PUT    /{contract_id}          編輯合約
  DELETE /{contract_id}          邏輯刪除合約

  GET    /vendors/options         廠商下拉（用於表單）
  GET    /vendors                 廠商列表
  POST   /vendors                 新增廠商
  GET    /vendors/{vendor_id}             單筆廠商
  PUT    /vendors/{vendor_id}             編輯廠商
  DELETE /vendors/{vendor_id}             刪除廠商
  GET    /vendors/{vendor_id}/performance 廠商績效（準時率、爭議率、評分）

  GET    /budget-categories/options  預算科目下拉（用於表單）
  GET    /budget-categories          預算科目列表
  POST   /budget-categories          新增預算科目
  PUT    /budget-categories/{id}     編輯預算科目
  DELETE /budget-categories/{id}     刪除預算科目

注意：FastAPI 路由按宣告順序匹配，靜態路徑必須在動態路徑（{id}）之前宣告。
"""

import io
import uuid
from pathlib import Path
from datetime import date
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, status, Query, File, UploadFile
from fastapi.concurrency import run_in_threadpool
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.exceptions import ContractManagementException
from app.dependencies import get_current_user, get_user_permissions
from app.models.user import User
from app.models.contract import ContractAttachment
from app.services.contract_service import (
    ContractService,
    VendorService,
    BudgetCategoryService,
    ClaimService,
    ContractItemService,
    RenewalService,
    generate_contract_excel,
    generate_claims_excel,
    get_budget_analysis,
)
from app.schemas.contract import (
    ContractCreate,
    ContractUpdate,
    ContractDetailResponse,
    ContractListResponse,
    ContractStatsResponse,
    DashboardKPIResponse,
    DashboardByDeptResponse,
    VendorCreate,
    VendorUpdate,
    VendorResponse,
    VendorListResponse,
    VendorPagedResponse,
    BudgetCategoryCreate,
    BudgetCategoryUpdate,
    BudgetCategoryResponse,
    BudgetCategoryListResponse,
    BudgetCategoryPagedResponse,
    ContractClaimCreate,
    ContractClaimUpdate,
    ContractClaimResponse,
    ContractClaimReviewRequest,
    ContractClaimBatchReviewRequest,
    ContractItemCreate,
    ContractItemUpdate,
    ContractItemResponse,
    RenewalCreate,
    RenewalReview,
    RenewalResponse,
    VendorImportResult,
    ContractApprovalRequest,
    CostAllocationItem,
    CostAllocationResponse,
    # H1~H4
    ContractTemplateCreate,
    ContractTemplateUpdate,
    ContractTemplateResponse,
    ContractChangeLogResponse,
    PaymentScheduleCreate,
    PaymentScheduleUpdate,
    PaymentScheduleResponse,
    ContractAuditLogResponse,
    # K2
    SlaMetricCreate,
    SlaMetricUpdate,
    SlaMetricResponse,
    SlaRecordCreate,
    SlaRecordResponse,
    SlaSummaryResponse,
    # I1~I4
    ApprovalStageResponse,
    ApprovalConfigCreate,
    ApprovalConfigUpdate,
    ApprovalConfigResponse,
    StageReviewRequest,
    AcceptanceCreate,
    AcceptanceUpdate,
    AcceptanceResponse,
    DepositCreate,
    DepositUpdate,
    DepositResponse,
    CostSummaryResponse,
    # 原合約複製續約 + 上下層級查詢（2026-07-21）
    ContractChainNode,
)

router = APIRouter(tags=["合約管理"])


# ─────────────────────────────────────────────────────────────────────────────
# 合約列表 / 新增
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "",
    response_model=ContractListResponse,
    summary="查詢合約列表",
)
def list_contracts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    page: int = Query(1, ge=1, description="頁碼"),
    size: int = Query(20, ge=1, le=200, description="每頁筆數"),
    search: Optional[str] = Query(None, description="搜尋編號或名稱"),
    status: Optional[str] = Query(None, description="合約狀態"),
    vendor_id: Optional[str] = Query(None, description="廠商ID"),
    risk_level: Optional[str] = Query(None, description="風險等級"),
    budget_year: Optional[int] = Query(None, description="預算年度"),
    responsible_dept: Optional[str] = Query(None, description="負責部門"),
    manager: Optional[str] = Query(None, description="管理人帳號（J5 個人化篩選）"),
    renewal_filter: Optional[str] = Query(
        None,
        description="續約鏈篩選：is_copy=只看複製續約產生的合約；has_copies=只看已被複製續約過的合約",
    ),
    sort_by: str = Query("updated_at", description="排序欄位"),
    sort_order: str = Query("desc", description="排序順序"),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_view")

        skip = (page - 1) * size
        contracts, total = ContractService.list_contracts(
            db,
            skip=skip,
            limit=size,
            search=search,
            status=status,
            vendor_id=vendor_id,
            risk_level=risk_level,
            budget_year=budget_year,
            responsible_dept=responsible_dept,
            manager=manager,
            renewal_filter=renewal_filter,
            sort_by=sort_by,
            sort_order=sort_order,
        )
        return ContractListResponse(total=total, page=page, size=size, items=contracts)

    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "",
    response_model=ContractDetailResponse,
    status_code=status.HTTP_201_CREATED,
    summary="新增合約",
)
def create_contract(
    contract_data: ContractCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")

        return ContractService.create_contract(db, contract_data)

    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 統計 / 同步（靜態路徑，必須在 /{contract_id} 之前）
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/stats",
    response_model=ContractStatsResponse,
    summary="合約統計",
)
def get_contract_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ContractService.get_stats(db)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/sync",
    summary="立即同步（Portal v2 版）",
)
def sync_contracts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Portal v2 無 Ragic 依賴，此端點保留以維持前端相容性。
    回傳 synced=0 表示無需同步。
    """
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_admin" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足，需要 contract_admin")
    return {"success": True, "synced": 0, "errors": []}


# ─────────────────────────────────────────────────────────────────────────────
# Dashboard（靜態路徑，必須在 /{contract_id} 之前）
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/dashboard/kpi",
    response_model=DashboardKPIResponse,
    summary="Dashboard KPI 指標",
)
def get_dashboard_kpi(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: Optional[int] = Query(None, description="預算年度（預設當年）"),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ContractService.get_dashboard_kpi(db, budget_year)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get(
    "/dashboard/by-dept",
    response_model=DashboardByDeptResponse,
    summary="Dashboard 部門金額分組",
)
def get_dashboard_by_dept(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: Optional[int] = Query(None, description="預算年度（預設當年）"),
    company: Optional[str] = Query(None, description="簽約公司名稱過濾（空=全部）"),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ContractService.get_by_dept(db, budget_year, company or None)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# G2 — 行事曆事件端點（靜態路徑）
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/calendar-events",
    summary="G2 合約行事曆事件（到期日 / 請款日 / 續約截止）",
)
def get_calendar_events(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    year:  int = Query(..., description="年份"),
    month: int = Query(..., ge=1, le=12, description="月份"),
):
    """
    回傳指定月份的所有行事曆事件：
      - type='expiry'  : 合約到期日（end_date）
      - type='claim'   : 請款記錄日期（claim_date）
      - type='renewal' : 續約申請截止（renewal_end_date）
    """
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")

        import calendar as _cal
        from datetime import date as _date
        from app.models.contract import Contract, ContractClaim, ContractRenewal

        _, last_day = _cal.monthrange(year, month)
        d_start = _date(year, month, 1)
        d_end   = _date(year, month, last_day)
        s_start = d_start.isoformat()   # "YYYY-MM-DD"（VARCHAR 比較用）
        s_end   = d_end.isoformat()

        events = []

        # 合約到期日（end_date = Date 欄位，用 date 物件比較）
        for c in db.query(Contract).filter(
            Contract.end_date >= d_start,
            Contract.end_date <= d_end,
            Contract.contract_status.notin_(["已終止"]),
        ).all():
            events.append({
                "date": str(c.end_date)[:10],
                "type": "expiry",
                "color": "#FF4D4F",
                "label": f"到期：{c.contract_name}",
                "contract_id": c.contract_id,
                "contract_name": c.contract_name,
            })

        # 請款日（claim_date = VARCHAR YYYY-MM-DD，字串比較）
        for cl in db.query(ContractClaim).filter(
            ContractClaim.claim_date >= s_start,
            ContractClaim.claim_date <= s_end,
        ).all():
            events.append({
                "date": cl.claim_date[:10],
                "type": "claim",
                "color": "#4BA8E8",
                "label": f"請款：{cl.contract_id} ${float(cl.amount):,.0f}",
                "contract_id": cl.contract_id,
                "claim_id": cl.id,
                "amount": float(cl.amount),
                "status": cl.status,
            })

        # 續約申請截止（renewal_end_date = VARCHAR YYYY-MM-DD）
        for r in db.query(ContractRenewal).filter(
            ContractRenewal.renewal_end_date >= s_start,
            ContractRenewal.renewal_end_date <= s_end,
            ContractRenewal.status == "已核准",
        ).all():
            events.append({
                "date": r.renewal_end_date[:10],
                "type": "renewal",
                "color": "#52C41A",
                "label": f"續約到期：{r.contract_id}",
                "contract_id": r.contract_id,
                "renewal_id": r.id,
            })

        return {"year": year, "month": month, "events": events, "total": len(events)}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 到期預警清單（靜態路徑，必須在 /{contract_id} 之前）
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/expiring",
    summary="即將到期合約清單",
)
def list_expiring_contracts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    days: int = Query(90, ge=1, le=365, description="未來幾天內到期（預設 90）"),
):
    """
    傳回到期日在今天 ~ 今天+days 之間且狀態不是「已終止」的合約。
    依剩餘天數升冪排列。
    """
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ContractService.get_expiring(db, days)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 合約列表匯出 Excel
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/export",
    summary="匯出合約列表 Excel（串流）",
)
def export_contracts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    search: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    vendor_id: Optional[str] = Query(None),
    risk_level: Optional[str] = Query(None),
    budget_year: Optional[int] = Query(None),
    responsible_dept: Optional[str] = Query(None),
):
    """帶入與列表頁相同的篩選條件，匯出全部符合記錄（不分頁）。"""
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        contracts, _ = ContractService.list_contracts(
            db, skip=0, limit=10000,
            search=search, status=status, vendor_id=vendor_id,
            risk_level=risk_level, budget_year=budget_year,
            responsible_dept=responsible_dept,
        )
        excel_bytes = generate_contract_excel(contracts, db=db)  # F8: 傳入 db 以查詢費用分攤
        filename = f"合約列表_{date.today().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"匯出失敗：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 預算執行率分析
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/budget-analysis",
    summary="預算執行率分析（依預算科目聚合）",
)
def get_budget_analysis_endpoint(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: int = Query(2026, ge=2020, le=2100, description="預算年度"),
):
    """
    依預算科目（L1 / L2）聚合：
    - contract_count：合約數量
    - total_claimed：請款總額
    - paid_amount：已付款金額
    - approved_amount：已核准待撥款金額
    - pending_amount：待審核金額
    """
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return get_budget_analysis(db, budget_year)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 單筆合約（動態路徑，必須在所有靜態路徑之後）
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/vendors/options",
    response_model=List[VendorListResponse],
    summary="廠商下拉選項（用於合約表單）",
)
def list_vendors_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    search: Optional[str] = Query(None),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        vendors, _ = VendorService.list_vendors(db, skip=0, limit=2000, search=search)
        return [VendorListResponse(vendor_id=v.vendor_id, vendor_name=v.vendor_name) for v in vendors]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/vendors/import",
    response_model=VendorImportResult,
    summary="廠商 Excel 批次匯入（upsert）",
)
def import_vendors(
    file: UploadFile = File(..., description="廠商 Excel 檔（.xlsx）"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_vendor_manage" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")

        if not file.filename or not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="請上傳 .xlsx 格式的 Excel 檔案")

        file_bytes = file.file.read()
        result = VendorService.import_vendors_from_excel(db, file_bytes)
        return VendorImportResult(**result)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"匯入失敗：{str(e)}")


@router.get(
    "/vendors",
    response_model=VendorPagedResponse,
    summary="查詢廠商列表",
)
def list_vendors(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    search: Optional[str] = Query(None),
    vendor_type: Optional[str] = Query(None),
    risk_level: Optional[str] = Query(None),
    is_critical: Optional[bool] = Query(None),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_vendor_manage" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")

        skip = (page - 1) * size
        vendors, total = VendorService.list_vendors(
            db,
            skip=skip,
            limit=size,
            search=search,
            vendor_type=vendor_type,
            risk_level=risk_level,
            is_critical=is_critical,
        )
        return VendorPagedResponse(total=total, items=vendors)

    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/vendors",
    response_model=VendorResponse,
    status_code=status.HTTP_201_CREATED,
    summary="新增廠商",
)
def create_vendor(
    vendor_data: VendorCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_vendor_manage" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return VendorService.create_vendor(db, vendor_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get("/vendors/concentration", summary="J1 廠商集中度分析（靜態路徑，必須在 /{vendor_id} 之前）")
def vendor_concentration_inline(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: Optional[int] = Query(None, description="預算年度篩選（None=全部）"),
    threshold: float = Query(30.0, ge=1, le=100, description="集中度警示閾值（%，預設 30%）"),
):
    from app.models.contract import Contract
    from sqlalchemy import func as _func

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    q = db.query(
        Contract.vendor_id,
        Contract.vendor_name,
        _func.count(Contract.contract_id).label("contract_count"),
        _func.sum(Contract.total_amount_tax_included).label("total_amount"),
    ).filter(Contract.contract_status.notin_(["已終止"]))
    if budget_year:
        q = q.filter(Contract.budget_year == budget_year)
    q = q.group_by(Contract.vendor_id, Contract.vendor_name)

    rows = q.all()
    grand_total = sum(float(r.total_amount or 0) for r in rows)

    result = []
    for r in sorted(rows, key=lambda x: float(x.total_amount or 0), reverse=True):
        amount = float(r.total_amount or 0)
        pct = round(amount / grand_total * 100, 2) if grand_total > 0 else 0.0
        result.append({
            "vendor_id": r.vendor_id,
            "vendor_name": r.vendor_name,
            "contract_count": r.contract_count,
            "total_amount": amount,
            "percentage": pct,
            "is_high_concentration": pct >= threshold,
        })

    return {
        "budget_year": budget_year,
        "threshold": threshold,
        "grand_total": grand_total,
        "vendor_count": len(result),
        "high_concentration_count": sum(1 for r in result if r["is_high_concentration"]),
        "items": result,
    }


@router.get(
    "/vendors/{vendor_id}",
    response_model=VendorResponse,
    summary="查詢單筆廠商",
)
def get_vendor(
    vendor_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_vendor_manage" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return VendorService.get_vendor(db, vendor_id)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.put(
    "/vendors/{vendor_id}",
    response_model=VendorResponse,
    summary="編輯廠商",
)
def update_vendor(
    vendor_id: str,
    vendor_data: VendorUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_vendor_manage" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return VendorService.update_vendor(db, vendor_id, vendor_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.delete(
    "/vendors/{vendor_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="刪除廠商",
)
def delete_vendor(
    vendor_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_vendor_manage" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        VendorService.delete_vendor(db, vendor_id)
        return None
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.get(
    "/vendors/{vendor_id}/performance",
    summary="廠商績效評分（準時率、爭議率、平均處理天數）",
)
def get_vendor_performance(
    vendor_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        return VendorService.get_vendor_performance(db, vendor_id)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 預算科目 — options 必須在 /{category_id} 之前
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/budget-categories/options",
    response_model=List[BudgetCategoryListResponse],
    summary="預算科目下拉選項（用於合約表單）",
)
def list_budget_categories_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: Optional[int] = Query(None),
    dept: Optional[str] = Query(None),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        categories, _ = BudgetCategoryService.list_budget_categories(
            db, skip=0, limit=2000, budget_year=budget_year, dept=dept, is_enabled=True
        )
        return [
            BudgetCategoryListResponse(
                id=c.id,
                budget_year=c.budget_year,
                category_l1=c.category_l1,
                category_l2=c.category_l2,
                accounting_code=c.accounting_code,
            )
            for c in categories
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get(
    "/budget-categories",
    response_model=BudgetCategoryPagedResponse,
    summary="查詢預算科目列表",
)
def list_budget_categories(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    budget_year: Optional[int] = Query(None),
    dept: Optional[str] = Query(None),
    category_l1: Optional[str] = Query(None),
    is_enabled: Optional[bool] = Query(None),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")

        skip = (page - 1) * size
        categories, total = BudgetCategoryService.list_budget_categories(
            db,
            skip=skip,
            limit=size,
            budget_year=budget_year,
            dept=dept,
            category_l1=category_l1,
            is_enabled=is_enabled,
        )
        return BudgetCategoryPagedResponse(total=total, items=categories)

    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/budget-categories",
    response_model=BudgetCategoryResponse,
    status_code=status.HTTP_201_CREATED,
    summary="新增預算科目",
)
def create_budget_category(
    category_data: BudgetCategoryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_admin" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_admin")
        return BudgetCategoryService.create_budget_category(db, category_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.put(
    "/budget-categories/{category_id}",
    response_model=BudgetCategoryResponse,
    summary="編輯預算科目",
)
def update_budget_category(
    category_id: int,
    category_data: BudgetCategoryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_admin" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_admin")
        return BudgetCategoryService.update_budget_category(db, category_id, category_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.delete(
    "/budget-categories/{category_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="刪除預算科目",
)
def delete_budget_category(
    category_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_admin" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_admin")
        BudgetCategoryService.delete_budget_category(db, category_id)
        return None
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


# ─────────────────────────────────────────────────────────────────────────────
# 請款 / 核銷記錄（靜態路徑，必須在 /{contract_id} 之前）
# ─────────────────────────────────────────────────────────────────────────────

@router.post(
    "/claims/{claim_id}/review",
    response_model=ContractClaimResponse,
    summary="請款審核（核准 / 拒絕 / 付款 / 重送）",
)
def review_claim(
    claim_id: int,
    review_data: ContractClaimReviewRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        actor = getattr(current_user, "username", "") or getattr(current_user, "email", "")
        return ClaimService.review_claim(db, claim_id, review_data, actor)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.post(
    "/claims/batch-review",
    response_model=dict,
    summary="批次審核請款（approve / reject）",
)
def batch_review_claims(
    batch_data: ContractClaimBatchReviewRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = getattr(current_user, "permissions", [])
        if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        actor = getattr(current_user, "username", "") or getattr(current_user, "email", "")
        return ClaimService.batch_review_claims(db, batch_data, actor)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get(
    "/claims/export",
    summary="匯出請款清單 Excel（串流）",
)
def export_claims(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    contract_id: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
):
    """帶入與請款頁相同的篩選條件，匯出全部符合記錄（不分頁）。"""
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        claims, _ = ClaimService.list_claims(
            db, contract_id=contract_id, status=status,
            skip=0, limit=10000,
        )
        excel_bytes = generate_claims_excel(claims)
        filename = f"請款清單_{date.today().strftime('%Y%m%d')}.xlsx"
        return StreamingResponse(
            io.BytesIO(excel_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"匯出失敗：{str(e)}")


@router.get(
    "/claims/stats",
    response_model=dict,
    summary="請款統計（各狀態筆數/金額、當月請款）",
)
def get_claims_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ClaimService.get_stats(db)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get(
    "/claims",
    response_model=dict,
    summary="請款記錄清單（可按合約篩選）",
)
def list_claims(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    contract_id: Optional[str] = Query(None),
    status_filter: Optional[str] = Query(None, alias="status"),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        skip = (page - 1) * size
        items, total = ClaimService.list_claims(db, contract_id, status_filter, skip, size)
        return {"total": total, "page": page, "size": size, "items": [i.dict() for i in items]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/claims",
    response_model=ContractClaimResponse,
    status_code=status.HTTP_201_CREATED,
    summary="新增請款記錄",
)
def create_claim(
    claim_data: ContractClaimCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ClaimService.create_claim(db, claim_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get(
    "/claims/{claim_id}",
    response_model=ContractClaimResponse,
    summary="單筆請款記錄",
)
def get_claim(
    claim_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ClaimService.get_claim(db, claim_id)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.put(
    "/claims/{claim_id}",
    response_model=ContractClaimResponse,
    summary="更新請款記錄",
)
def update_claim(
    claim_id: int,
    claim_data: ContractClaimUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ClaimService.update_claim(db, claim_id, claim_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.delete(
    "/claims/{claim_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="刪除請款記錄",
)
def delete_claim(
    claim_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        ClaimService.delete_claim(db, claim_id)
        return None
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


# ─────────────────────────────────────────────────────────────────────────────
# 請款附件  /claims/{claim_id}/attachments  （靜態路徑優先於動態 {claim_id}）
# ─────────────────────────────────────────────────────────────────────────────

CLAIM_UPLOAD_DIR = Path("uploads/claims")
ATTACHMENT_ALLOWED_TYPES = {
    "application/pdf",
    "image/jpeg", "image/jpg", "image/png", "image/webp",
}
ATTACHMENT_MAX_SIZE = 20 * 1024 * 1024   # 20 MB


def _get_attachment_or_404(db, attachment_id: int):
    from app.models.contract import ContractClaimAttachment
    att = db.query(ContractClaimAttachment).filter(ContractClaimAttachment.id == attachment_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="附件不存在")
    return att


@router.post(
    "/claims/{claim_id}/attachments",
    summary="上傳請款附件（PDF / 圖片）",
    status_code=status.HTTP_201_CREATED,
)
async def upload_claim_attachment(
    claim_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractClaim, ContractClaimAttachment

    # 權限
    user_permissions = await run_in_threadpool(get_user_permissions, current_user.id, db)
    if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    # 請款存在
    claim = await run_in_threadpool(db.query(ContractClaim).filter(ContractClaim.id == claim_id).first)
    if not claim:
        raise HTTPException(status_code=404, detail="請款記錄不存在")

    # 類型檢查
    content_type = file.content_type or ""
    if content_type not in ATTACHMENT_ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="只接受 PDF / JPG / PNG / WEBP 格式")

    content = await file.read()
    if len(content) > ATTACHMENT_MAX_SIZE:
        raise HTTPException(status_code=400, detail="檔案大小不得超過 20 MB")

    # 儲存
    CLAIM_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "file").suffix or ".bin"
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    (CLAIM_UPLOAD_DIR / stored_name).write_bytes(content)

    att = ContractClaimAttachment(
        claim_id=claim_id,
        stored_filename=stored_name,
        original_filename=file.filename or stored_name,
        content_type=content_type,
        file_size=len(content),
        uploader=current_user.username,
    )

    def _save():
        db.add(att)
        db.commit()
        db.refresh(att)

    await run_in_threadpool(_save)

    return {
        "id": att.id,
        "claim_id": att.claim_id,
        "original_filename": att.original_filename,
        "content_type": att.content_type,
        "file_size": att.file_size,
        "uploader": att.uploader,
        "created_at": att.created_at.isoformat(),
        "download_url": f"/api/v1/contract/claims/attachments/{att.id}/download",
    }


@router.get(
    "/claims/{claim_id}/attachments",
    summary="列出請款附件",
)
def list_claim_attachments(
    claim_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractClaimAttachment

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    attachments = (
        db.query(ContractClaimAttachment)
        .filter(ContractClaimAttachment.claim_id == claim_id)
        .order_by(ContractClaimAttachment.created_at)
        .all()
    )
    return [
        {
            "id": a.id,
            "claim_id": a.claim_id,
            "original_filename": a.original_filename,
            "content_type": a.content_type,
            "file_size": a.file_size,
            "uploader": a.uploader,
            "created_at": a.created_at.isoformat(),
            "download_url": f"/api/v1/contract/claims/attachments/{a.id}/download",
        }
        for a in attachments
    ]


@router.get(
    "/claims/attachments/{attachment_id}/download",
    summary="下載 / 預覽請款附件",
)
def download_claim_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    att = _get_attachment_or_404(db, attachment_id)
    file_path = CLAIM_UPLOAD_DIR / att.stored_filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="附件檔案遺失")

    # 瀏覽器直接預覽（PDF / 圖片），使用 inline；下載用 attachment
    disposition = "inline" if att.content_type.startswith("image/") or att.content_type == "application/pdf" else "attachment"
    encoded_name = att.original_filename.encode("utf-8").decode("latin-1", errors="replace")
    return FileResponse(
        str(file_path),
        media_type=att.content_type,
        headers={"Content-Disposition": f"{disposition}; filename*=UTF-8''{att.original_filename}"},
    )


@router.delete(
    "/claims/attachments/{attachment_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="刪除請款附件",
)
def delete_claim_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_claims_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    att = _get_attachment_or_404(db, attachment_id)
    file_path = CLAIM_UPLOAD_DIR / att.stored_filename
    if file_path.exists():
        file_path.unlink(missing_ok=True)
    db.delete(att)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 單筆合約 CRUD  /{contract_id}
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/{contract_id}",
    response_model=ContractDetailResponse,
    summary="單筆合約詳情",
)
def get_contract(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        return ContractService.get_contract(db, contract_id)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.put(
    "/{contract_id}",
    response_model=ContractDetailResponse,
    summary="編輯合約",
)
def update_contract(
    contract_id: str,
    contract_data: ContractUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")

        # H2：在更新前抓舊值快照
        from app.models.contract import Contract, ContractChangeLog, ContractAuditLog
        old_contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        old_snapshot: dict = {}
        if old_contract:
            for col in Contract.__table__.columns:
                old_snapshot[col.name] = getattr(old_contract, col.name, None)

        result = ContractService.update_contract(db, contract_id, contract_data)

        # H2：寫入有差異的欄位變更歷程
        _FIELD_LABELS: dict[str, str] = {
            "contract_name": "合約名稱", "contract_type": "合約類型",
            "contract_status": "合約狀態", "responsible_dept": "權責部門",
            "using_depts": "使用部門", "vendor_id": "廠商編號",
            "vendor_name": "廠商名稱", "start_date": "合約起日",
            "end_date": "合約迄日", "notification_days": "通知天數",
            "auto_renewal": "自動續約", "currency": "幣別",
            "total_amount_tax_included": "合約總額（含稅）",
            "monthly_fixed_amount": "每月固定金額", "pricing_method": "計價方式",
            "needs_purchase_order": "需請購單", "can_claim_without_po": "可無購請款",
            "needs_allocation": "需分攤", "allocation_method": "分攤方式",
            "budget_year": "預算年度", "budget_category_l1": "預算大項",
            "budget_category_l2": "預算細項", "accounting_code": "會計科目",
            "budget_source": "預算來源", "budget_control_method": "預算控管",
            "require_acceptance": "需驗收", "risk_level": "風險等級",
            "manager": "管理人", "reviewer": "覆核人",
            "signing_company": "簽約公司", "signing_dept": "簽約部門",
            "budget_company": "預算公司", "budget_dept": "預算部門",
            "pricing_spec": "計價規格", "remarks": "備註",
        }
        if old_snapshot:
            new_contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
            if new_contract:
                for field, label in _FIELD_LABELS.items():
                    old_val = old_snapshot.get(field)
                    new_val = getattr(new_contract, field, None)
                    if str(old_val) != str(new_val):
                        db.add(ContractChangeLog(
                            contract_id=contract_id,
                            field_name=field,
                            field_label=label,
                            old_value=str(old_val) if old_val is not None else None,
                            new_value=str(new_val) if new_val is not None else None,
                            operator=current_user.username if hasattr(current_user, "username") else str(current_user.id),
                        ))

        # H4：寫入稽核日誌
        update_fields = [k for k, v in contract_data.dict(exclude_unset=True).items()]
        db.add(ContractAuditLog(
            contract_id=contract_id,
            action="update",
            resource="contract",
            resource_id=contract_id,
            operator=current_user.username if hasattr(current_user, "username") else str(current_user.id),
            payload_summary=f"更新欄位：{', '.join(update_fields[:10])}{'...' if len(update_fields) > 10 else ''}",
            result="success",
        ))
        db.commit()

        return result
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.delete(
    "/{contract_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    summary="刪除合約",
)
def delete_contract(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")
        ContractService.delete_contract(db, contract_id)
        return None
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 合約項目 (Line Items)  /{contract_id}/items
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/{contract_id}/items",
    summary="查詢合約項目清單",
)
def list_contract_items(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        items = ContractItemService.list_items(db, contract_id)
        return {"total": len(items), "items": [i.dict() for i in items]}
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.post(
    "/{contract_id}/items",
    response_model=ContractItemResponse,
    summary="新增合約項目",
)
def create_contract_item(
    contract_id: str,
    item_data: ContractItemCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")
        return ContractItemService.create_item(db, contract_id, item_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.put(
    "/{contract_id}/items/{item_id}",
    response_model=ContractItemResponse,
    summary="更新合約項目",
)
def update_contract_item(
    contract_id: str,
    item_id: int,
    item_data: ContractItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")
        return ContractItemService.update_item(db, contract_id, item_id, item_data)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


@router.delete(
    "/{contract_id}/items/{item_id}",
    summary="刪除合約項目",
)
def delete_contract_item(
    contract_id: str,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")
        ContractItemService.delete_item(db, contract_id, item_id)
        return {"success": True}
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise


# ─────────────────────────────────────────────────────────────────────────────
# F3 費用分攤  /{contract_id}/cost-allocations
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/{contract_id}/cost-allocations",
    response_model=List[CostAllocationResponse],
    summary="查詢合約費用分攤明細",
)
def get_cost_allocations(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        from app.models.contract import ContractCostAllocation
        rows = db.query(ContractCostAllocation).filter(
            ContractCostAllocation.contract_id == contract_id
        ).order_by(ContractCostAllocation.id).all()
        return [CostAllocationResponse.from_orm(r) for r in rows]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.put(
    "/{contract_id}/cost-allocations",
    response_model=List[CostAllocationResponse],
    summary="整批覆寫合約費用分攤明細",
)
def upsert_cost_allocations(
    contract_id: str,
    items: List[CostAllocationItem],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """整批覆寫：刪除舊有分攤記錄，再整批插入新記錄。"""
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")
        # 驗證：若有比例行，所有比例行加總須 ≈ 100
        pct_rows = [i for i in items if i.allocation_type == "percentage"]
        if pct_rows:
            total = sum(r.value for r in pct_rows)
            if abs(total - 100) > 0.01:
                raise HTTPException(
                    status_code=400,
                    detail=f"比例行加總為 {total:.2f}%，必須等於 100%"
                )
        from app.models.contract import ContractCostAllocation
        # 整批覆寫
        db.query(ContractCostAllocation).filter(
            ContractCostAllocation.contract_id == contract_id
        ).delete()
        new_rows = []
        for item in items:
            row = ContractCostAllocation(
                contract_id=contract_id,
                company_name=item.company_name,
                allocation_type=item.allocation_type,
                value=item.value,
            )
            db.add(row)
            new_rows.append(row)
        db.commit()
        for r in new_rows:
            db.refresh(r)
        return [CostAllocationResponse.from_orm(r) for r in new_rows]
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"儲存失敗：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 合約續約  renewals
# ─────────────────────────────────────────────────────────────────────────────

@router.get(
    "/renewals/all",
    summary="查詢所有續約申請（跨合約）",
)
def list_all_renewals(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=200),
    status: Optional[str] = Query(None),
    contract_id: Optional[str] = Query(None),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        from app.models.contract import ContractRenewal
        from app.schemas.contract import RenewalResponse
        q = db.query(ContractRenewal)
        if status:
            q = q.filter(ContractRenewal.status == status)
        if contract_id:
            q = q.filter(ContractRenewal.contract_id == contract_id)
        total = q.count()
        items = q.order_by(ContractRenewal.id.desc()).offset((page - 1) * size).limit(size).all()
        return {"total": total, "page": page, "size": size, "items": [
            RenewalResponse.from_orm(r) for r in items
        ]}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get(
    "/{contract_id}/renewals",
    summary="查詢單一合約的續約申請清單",
)
def list_renewals_by_contract(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        from app.models.contract import ContractRenewal
        from app.schemas.contract import RenewalResponse
        items = db.query(ContractRenewal).filter(
            ContractRenewal.contract_id == contract_id
        ).order_by(ContractRenewal.id.desc()).all()
        return [RenewalResponse.from_orm(r) for r in items]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/{contract_id}/renewals",
    summary="申請合約續約",
)
def apply_for_renewal(
    contract_id: str,
    renewal_data: RenewalCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise HTTPException(status_code=404, detail=f"合約 {contract_id} 不存在")
        from app.models.contract import ContractRenewal
        from app.schemas.contract import RenewalResponse
        renewal = ContractRenewal(
            contract_id=contract_id,
            renewal_start_date=renewal_data.renewal_start_date,
            renewal_end_date=renewal_data.renewal_end_date,
            new_amount=renewal_data.new_amount,
            renewal_reason=renewal_data.renewal_reason,
            remarks=renewal_data.remarks,
            applicant=current_user.full_name or current_user.email,
            applicant_dept=renewal_data.applicant_dept or "",
            status="待審核",
        )
        db.add(renewal)
        db.commit()
        db.refresh(renewal)
        return RenewalResponse.from_orm(renewal)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"申請失敗：{str(e)}")


@router.post(
    "/renewals/{renewal_id}/review",
    summary="審核 / 拒絕 / 撤回續約申請",
)
def review_renewal(
    renewal_id: int,
    review_data: RenewalReview,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_approve" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_approve")
        from app.models.contract import ContractRenewal
        from app.schemas.contract import RenewalResponse
        renewal = db.query(ContractRenewal).filter(ContractRenewal.id == renewal_id).first()
        if not renewal:
            raise HTTPException(status_code=404, detail=f"續約申請 #{renewal_id} 不存在")
        action = review_data.action
        if action == "approve":
            renewal.status = "已核准"
            renewal.reviewer = current_user.full_name or current_user.email
            renewal.review_comment = review_data.review_comment
        elif action == "reject":
            renewal.status = "已拒絕"
            renewal.reviewer = current_user.full_name or current_user.email
            renewal.review_comment = review_data.review_comment
        elif action == "withdraw":
            renewal.status = "已撤回"
        else:
            raise HTTPException(status_code=400, detail=f"不支援的操作：{action}，請使用 approve / reject / withdraw")
        db.commit()
        db.refresh(renewal)
        return RenewalResponse.from_orm(renewal)
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"審核失敗：{str(e)}")


# ════════════════════════════════════════════════════════════════════════════
# 原合約複製續約 + 上下層級查詢（2026-07-21）
# ════════════════════════════════════════════════════════════════════════════

@router.post(
    "/{contract_id}/copy-renew",
    response_model=ContractDetailResponse,
    status_code=status.HTTP_201_CREATED,
    summary="複製原合約建立續約新合約",
)
def copy_renew_contract(
    contract_id: str,
    contract_data: ContractCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    以原合約（path 中的 contract_id）為來源，複製其欄位與子資料
    （合約項目／費用分攤／付款計劃）建立一份新合約；
    新合約標記 renewed_from_contract_id 指向原合約，供「上下層級」TAB 查詢。
    body 為完整的新合約資料（前端表單已預先帶入原合約值，使用者可自行編輯，
    含新的合約編號），驗證規則與「新增合約」端點相同。
    """
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_create_edit")

        return ContractService.copy_renew_contract(db, contract_id, contract_data)

    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.get(
    "/{contract_id}/renewal-chain",
    response_model=List[ContractChainNode],
    summary="查詢合約上下層級續約鏈",
)
def get_renewal_chain(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    查詢合約完整續約鏈（往上溯源到最源頭合約 + 往下所有代，依起日排序）。
    """
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足，需要 contract_view")

        chain = ContractService.get_renewal_chain(db, contract_id)
        return [
            ContractChainNode(
                contract_id=c.contract_id,
                contract_name=c.contract_name,
                contract_status=c.contract_status,
                start_date=c.start_date,
                end_date=c.end_date,
                total_amount_tax_included=float(c.total_amount_tax_included),
                renewed_from_contract_id=c.renewed_from_contract_id,
                is_current=(c.contract_id == contract_id),
            )
            for c in chain
        ]

    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ════════════════════════════════════════════════════════════════════════════
# 合約審核流程端點
# ════════════════════════════════════════════════════════════════════════════

@router.post(
    "/{contract_id}/submit",
    response_model=ContractDetailResponse,
    summary="送審（草稿 → 審核中）",
)
def submit_contract_for_review(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_view" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足")
        submitter = getattr(current_user, "username", "") or ""
        result = ContractService.submit_for_review(db, contract_id, submitter)

        # I1：送審後自動建立多層審核關卡
        from app.models.contract import Contract, ContractApprovalStage, ContractApprovalConfig
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if contract:
            # 查本合約類型 + 通用(*) 的設定，取啟用的關卡
            configs = (
                db.query(ContractApprovalConfig)
                .filter(
                    ContractApprovalConfig.contract_type.in_([contract.contract_type, "*"]),
                    ContractApprovalConfig.is_enabled == True,
                )
                .order_by(ContractApprovalConfig.stage_order)
                .all()
            )
            if configs:
                # 計算本次送審是第幾輪
                from sqlalchemy import func as _func
                max_round = db.query(_func.max(ContractApprovalStage.submission_round)).filter(
                    ContractApprovalStage.contract_id == contract_id
                ).scalar() or 0
                new_round = max_round + 1
                for cfg in configs:
                    db.add(ContractApprovalStage(
                        contract_id=contract_id,
                        submission_round=new_round,
                        stage_order=cfg.stage_order,
                        stage_name=cfg.stage_name,
                        assigned_to=cfg.assigned_to,
                        status="待審核",
                    ))
                db.commit()

        return result
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/{contract_id}/approve",
    response_model=ContractDetailResponse,
    summary="核准合約（審核中 → 生效中）",
)
def approve_contract(
    contract_id: str,
    body: ContractApprovalRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_approve" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足（需 contract_approve 權限）")
        approver = getattr(current_user, "username", "") or ""
        return ContractService.approve_contract(db, contract_id, approver, body.comment)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


@router.post(
    "/{contract_id}/reject",
    response_model=ContractDetailResponse,
    summary="拒絕合約（審核中 → 草稿）",
)
def reject_contract(
    contract_id: str,
    body: ContractApprovalRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    try:
        user_permissions = get_user_permissions(current_user.id, db)
        if "*" not in user_permissions and "contract_approve" not in user_permissions:
            raise HTTPException(status_code=403, detail="權限不足（需 contract_approve 權限）")
        rejector = getattr(current_user, "username", "") or ""
        return ContractService.reject_contract(db, contract_id, rejector, body.comment)
    except ContractManagementException as e:
        raise HTTPException(status_code=e.status_code, detail={"message": e.message, "error_code": e.error_code})
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"伺服器錯誤：{str(e)}")


# ─────────────────────────────────────────────────────────────────────────────
# 合約本體附件  /{contract_id}/attachments
# 靜態前綴 /doc-attachments/{id}/... 需在動態路由之前（已在頂端聲明 router order）
# ─────────────────────────────────────────────────────────────────────────────

CONTRACT_UPLOAD_DIR = Path("uploads/contracts")
_CONTRACT_ATTACH_ALLOWED = {
    "application/pdf",
    "image/jpeg", "image/png", "image/webp",
}
_CONTRACT_ATTACH_MAX_BYTES = 20 * 1024 * 1024  # 20 MB


def _get_contract_attachment_or_404(db: Session, attachment_id: int) -> ContractAttachment:
    att = db.query(ContractAttachment).filter(ContractAttachment.id == attachment_id).first()
    if not att:
        raise HTTPException(status_code=404, detail="附件不存在")
    return att


@router.post(
    "/{contract_id}/attachments",
    summary="上傳合約附件",
)
async def upload_contract_attachment(
    contract_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_permissions = await run_in_threadpool(get_user_permissions, current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足（需 contract_create_edit）")

    from app.models.contract import Contract
    contract_exists = await run_in_threadpool(db.query(Contract).filter(Contract.contract_id == contract_id).first)
    if not contract_exists:
        raise HTTPException(status_code=404, detail="合約不存在")

    content = await file.read()
    if len(content) > _CONTRACT_ATTACH_MAX_BYTES:
        raise HTTPException(status_code=413, detail="檔案超過 20MB 限制")
    if file.content_type not in _CONTRACT_ATTACH_ALLOWED:
        raise HTTPException(status_code=415, detail="僅支援 PDF / JPG / PNG / WEBP")

    CONTRACT_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
    suffix = Path(file.filename or "file").suffix or ".bin"
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    (CONTRACT_UPLOAD_DIR / stored_name).write_bytes(content)

    att = ContractAttachment(
        contract_id=contract_id,
        stored_filename=stored_name,
        original_filename=file.filename or stored_name,
        content_type=file.content_type or "application/octet-stream",
        file_size=len(content),
        uploader=getattr(current_user, "username", "") or "",
    )

    def _save():
        db.add(att)
        db.commit()
        db.refresh(att)

    await run_in_threadpool(_save)

    return {
        "id": att.id,
        "original_filename": att.original_filename,
        "content_type": att.content_type,
        "file_size": att.file_size,
        "uploader": att.uploader,
        "created_at": att.created_at.isoformat() if att.created_at else None,
        "download_url": f"/api/v1/contract/doc-attachments/{att.id}/download",
    }


@router.get(
    "/{contract_id}/attachments",
    summary="列出合約附件",
)
def list_contract_attachments(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    attachments = (
        db.query(ContractAttachment)
        .filter(ContractAttachment.contract_id == contract_id)
        .order_by(ContractAttachment.created_at)
        .all()
    )
    return [
        {
            "id": a.id,
            "original_filename": a.original_filename,
            "content_type": a.content_type,
            "file_size": a.file_size,
            "uploader": a.uploader,
            "created_at": a.created_at.isoformat() if a.created_at else None,
            "download_url": f"/api/v1/contract/doc-attachments/{a.id}/download",
        }
        for a in attachments
    ]


@router.get(
    "/doc-attachments/{attachment_id}/download",
    summary="下載合約附件",
)
def download_contract_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    att = _get_contract_attachment_or_404(db, attachment_id)
    file_path = CONTRACT_UPLOAD_DIR / att.stored_filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="附件檔案遺失")
    return FileResponse(
        path=str(file_path),
        media_type=att.content_type,
        filename=att.original_filename,
    )


@router.delete(
    "/doc-attachments/{attachment_id}",
    summary="刪除合約附件",
)
def delete_contract_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足（需 contract_create_edit）")

    att = _get_contract_attachment_or_404(db, attachment_id)
    file_path = CONTRACT_UPLOAD_DIR / att.stored_filename
    if file_path.exists():
        file_path.unlink()
    db.delete(att)
    db.commit()
    return {"success": True}


# ─────────────────────────────────────────────────────────────────────────────
# H1 — 合約範本管理  /templates
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/templates", summary="H1 合約範本清單")
def list_templates(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    enabled_only: bool = Query(False, description="只回傳啟用中的範本"),
):
    from app.models.contract import ContractTemplate
    q = db.query(ContractTemplate)
    if enabled_only:
        q = q.filter(ContractTemplate.is_enabled == True)
    items = q.order_by(ContractTemplate.contract_type, ContractTemplate.name).all()
    return {"total": len(items), "templates": [ContractTemplateResponse.from_orm(t).dict() for t in items]}


@router.post("/templates", summary="H1 新增合約範本")
def create_template(
    body: ContractTemplateCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractTemplate, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    existing = db.query(ContractTemplate).filter(ContractTemplate.name == body.name).first()
    if existing:
        raise HTTPException(status_code=409, detail=f"範本名稱「{body.name}」已存在")
    t = ContractTemplate(**body.dict())
    db.add(t)
    operator = current_user.username if hasattr(current_user, "username") else str(current_user.id)
    db.add(ContractAuditLog(
        action="create", resource="template", resource_id=body.name,
        operator=operator, payload_summary=f"新增範本：{body.name}（{body.contract_type}）", result="success",
    ))
    db.commit()
    db.refresh(t)
    return ContractTemplateResponse.from_orm(t)


@router.put("/templates/{template_id}", summary="H1 修改合約範本")
def update_template(
    template_id: int,
    body: ContractTemplateUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractTemplate, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    t = db.query(ContractTemplate).filter(ContractTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="範本不存在")
    for field, val in body.dict(exclude_unset=True).items():
        setattr(t, field, val)
    operator = current_user.username if hasattr(current_user, "username") else str(current_user.id)
    db.add(ContractAuditLog(
        action="update", resource="template", resource_id=str(template_id),
        operator=operator, payload_summary=f"修改範本 id={template_id}", result="success",
    ))
    db.commit()
    db.refresh(t)
    return ContractTemplateResponse.from_orm(t)


@router.delete("/templates/{template_id}", status_code=status.HTTP_204_NO_CONTENT, summary="H1 刪除合約範本")
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractTemplate, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    t = db.query(ContractTemplate).filter(ContractTemplate.id == template_id).first()
    if not t:
        raise HTTPException(status_code=404, detail="範本不存在")
    operator = current_user.username if hasattr(current_user, "username") else str(current_user.id)
    db.add(ContractAuditLog(
        action="delete", resource="template", resource_id=str(template_id),
        operator=operator, payload_summary=f"刪除範本：{t.name}", result="success",
    ))
    db.delete(t)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# H2 — 合約變更歷程  /{contract_id}/change-logs
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/change-logs", summary="H2 合約欄位變更歷程")
def list_change_logs(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(100, ge=1, le=500),
):
    from app.models.contract import ContractChangeLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    logs = (
        db.query(ContractChangeLog)
        .filter(ContractChangeLog.contract_id == contract_id)
        .order_by(ContractChangeLog.operated_at.desc())
        .limit(limit)
        .all()
    )
    return {"total": len(logs), "logs": [ContractChangeLogResponse.from_orm(l).dict() for l in logs]}


# ─────────────────────────────────────────────────────────────────────────────
# H3 — 分期付款計劃  /{contract_id}/payment-schedules
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/payment-schedules", summary="H3 分期付款計劃清單")
def list_payment_schedules(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractPaymentSchedule
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    items = (
        db.query(ContractPaymentSchedule)
        .filter(ContractPaymentSchedule.contract_id == contract_id)
        .order_by(ContractPaymentSchedule.due_date)
        .all()
    )
    return {"total": len(items), "schedules": [PaymentScheduleResponse.from_orm(i).dict() for i in items]}


@router.post("/{contract_id}/payment-schedules", summary="H3 新增付款里程碑")
def create_payment_schedule(
    contract_id: str,
    body: PaymentScheduleCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractPaymentSchedule, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    if not db.query(Contract).filter(Contract.contract_id == contract_id).first():
        raise HTTPException(status_code=404, detail="合約不存在")
    ps = ContractPaymentSchedule(contract_id=contract_id, **body.dict())
    db.add(ps)
    operator = current_user.username if hasattr(current_user, "username") else str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="create", resource="payment_schedule",
        operator=operator, payload_summary=f"新增里程碑：{body.milestone_name} {body.due_date} ${body.amount:,.0f}", result="success",
    ))
    db.commit()
    db.refresh(ps)
    return PaymentScheduleResponse.from_orm(ps)


@router.put("/{contract_id}/payment-schedules/{schedule_id}", summary="H3 更新付款里程碑")
def update_payment_schedule(
    contract_id: str,
    schedule_id: int,
    body: PaymentScheduleUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractPaymentSchedule, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    ps = db.query(ContractPaymentSchedule).filter(
        ContractPaymentSchedule.id == schedule_id,
        ContractPaymentSchedule.contract_id == contract_id,
    ).first()
    if not ps:
        raise HTTPException(status_code=404, detail="付款里程碑不存在")
    for field, val in body.dict(exclude_unset=True).items():
        setattr(ps, field, val)
    operator = current_user.username if hasattr(current_user, "username") else str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="update", resource="payment_schedule",
        resource_id=str(schedule_id), operator=operator,
        payload_summary=f"更新里程碑 id={schedule_id} 狀態={body.status or '—'}", result="success",
    ))
    db.commit()
    db.refresh(ps)
    return PaymentScheduleResponse.from_orm(ps)


@router.delete("/{contract_id}/payment-schedules/{schedule_id}", status_code=status.HTTP_204_NO_CONTENT, summary="H3 刪除付款里程碑")
def delete_payment_schedule(
    contract_id: str,
    schedule_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractPaymentSchedule, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    ps = db.query(ContractPaymentSchedule).filter(
        ContractPaymentSchedule.id == schedule_id,
        ContractPaymentSchedule.contract_id == contract_id,
    ).first()
    if not ps:
        raise HTTPException(status_code=404, detail="付款里程碑不存在")
    operator = current_user.username if hasattr(current_user, "username") else str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="delete", resource="payment_schedule",
        resource_id=str(schedule_id), operator=operator,
        payload_summary=f"刪除里程碑：{ps.milestone_name}", result="success",
    ))
    db.delete(ps)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# H4 — 操作稽核日誌  /{contract_id}/audit-logs
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/audit-logs", summary="H4 合約操作稽核日誌")
def list_audit_logs(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(100, ge=1, le=500),
):
    from app.models.contract import ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    logs = (
        db.query(ContractAuditLog)
        .filter(ContractAuditLog.contract_id == contract_id)
        .order_by(ContractAuditLog.operated_at.desc())
        .limit(limit)
        .all()
    )
    return {"total": len(logs), "logs": [ContractAuditLogResponse.from_orm(l).dict() for l in logs]}


# ─────────────────────────────────────────────────────────────────────────────
# I1 — 多層審核關卡設定  /approval-configs
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approval-configs", summary="I1 審核關卡設定清單")
def list_approval_configs(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractApprovalConfig
    cfgs = db.query(ContractApprovalConfig).order_by(
        ContractApprovalConfig.contract_type,
        ContractApprovalConfig.stage_order,
    ).all()
    return {"total": len(cfgs), "configs": [ApprovalConfigResponse.from_orm(c).dict() for c in cfgs]}


@router.post("/approval-configs", summary="I1 新增審核關卡設定")
def create_approval_config(
    body: ApprovalConfigCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractApprovalConfig
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_approve" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足（需 contract_approve）")
    cfg = ContractApprovalConfig(**body.dict())
    db.add(cfg)
    db.commit()
    db.refresh(cfg)
    return ApprovalConfigResponse.from_orm(cfg)


@router.put("/approval-configs/{config_id}", summary="I1 修改審核關卡設定")
def update_approval_config(
    config_id: int,
    body: ApprovalConfigUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractApprovalConfig
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_approve" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    cfg = db.query(ContractApprovalConfig).filter(ContractApprovalConfig.id == config_id).first()
    if not cfg:
        raise HTTPException(status_code=404, detail="設定不存在")
    for k, v in body.dict(exclude_unset=True).items():
        setattr(cfg, k, v)
    db.commit()
    db.refresh(cfg)
    return ApprovalConfigResponse.from_orm(cfg)


@router.delete("/approval-configs/{config_id}", status_code=status.HTTP_204_NO_CONTENT, summary="I1 刪除審核關卡設定")
def delete_approval_config(
    config_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractApprovalConfig
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_approve" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    cfg = db.query(ContractApprovalConfig).filter(ContractApprovalConfig.id == config_id).first()
    if not cfg:
        raise HTTPException(status_code=404, detail="設定不存在")
    db.delete(cfg)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# I1 — 合約審核關卡操作  /{contract_id}/approval-stages
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/approval-stages", summary="I1 查詢合約審核關卡")
def list_approval_stages(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    latest_only: bool = Query(True, description="只回傳最新一輪送審的關卡（預設 True）"),
):
    from app.models.contract import ContractApprovalStage
    from sqlalchemy import func as _func
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    q = db.query(ContractApprovalStage).filter(ContractApprovalStage.contract_id == contract_id)
    if latest_only:
        max_round = db.query(_func.max(ContractApprovalStage.submission_round)).filter(
            ContractApprovalStage.contract_id == contract_id
        ).scalar() or 0
        q = q.filter(ContractApprovalStage.submission_round == max_round)

    stages = q.order_by(ContractApprovalStage.submission_round, ContractApprovalStage.stage_order).all()
    return {"total": len(stages), "stages": [ApprovalStageResponse.from_orm(s).dict() for s in stages]}


@router.post("/{contract_id}/approval-stages/{stage_id}/approve", summary="I1 核准審核關卡")
def approve_approval_stage(
    contract_id: str,
    stage_id: int,
    body: StageReviewRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractApprovalStage
    from sqlalchemy import func as _func
    from app.core.time import twnow

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_approve" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足（需 contract_approve）")

    stage = db.query(ContractApprovalStage).filter(
        ContractApprovalStage.id == stage_id,
        ContractApprovalStage.contract_id == contract_id,
    ).first()
    if not stage:
        raise HTTPException(status_code=404, detail="審核關卡不存在")
    if stage.status != "待審核":
        raise HTTPException(status_code=400, detail=f"此關卡目前狀態為「{stage.status}」，無法再次操作")

    reviewer = getattr(current_user, "username", "") or str(current_user.id)
    now = twnow()
    stage.status = "已核准"
    stage.reviewer = reviewer
    stage.comment = body.comment
    stage.reviewed_at = now

    # 檢查同輪是否全部關卡已核准 → 若是則合約升為「生效中」
    all_stages = db.query(ContractApprovalStage).filter(
        ContractApprovalStage.contract_id == contract_id,
        ContractApprovalStage.submission_round == stage.submission_round,
    ).all()
    all_approved = all(s.status == "已核准" or s.id == stage_id for s in all_stages)

    contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
    if all_approved and contract and contract.contract_status == "審核中":
        contract.contract_status = "生效中"
        contract.approved_by = reviewer
        contract.approved_at = now
        contract.approval_comment = body.comment
        contract.updated_at = now

    db.commit()
    return {
        "success": True,
        "stage_id": stage_id,
        "status": "已核准",
        "contract_promoted": all_approved,
    }


@router.post("/{contract_id}/approval-stages/{stage_id}/reject", summary="I1 拒絕審核關卡")
def reject_approval_stage(
    contract_id: str,
    stage_id: int,
    body: StageReviewRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractApprovalStage
    from app.core.time import twnow

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_approve" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足（需 contract_approve）")

    stage = db.query(ContractApprovalStage).filter(
        ContractApprovalStage.id == stage_id,
        ContractApprovalStage.contract_id == contract_id,
    ).first()
    if not stage:
        raise HTTPException(status_code=404, detail="審核關卡不存在")
    if stage.status != "待審核":
        raise HTTPException(status_code=400, detail=f"此關卡目前狀態為「{stage.status}」，無法操作")

    reviewer = getattr(current_user, "username", "") or str(current_user.id)
    now = twnow()
    stage.status = "已拒絕"
    stage.reviewer = reviewer
    stage.comment = body.comment
    stage.reviewed_at = now

    # 後續關卡全部取消
    same_round_stages = db.query(ContractApprovalStage).filter(
        ContractApprovalStage.contract_id == contract_id,
        ContractApprovalStage.submission_round == stage.submission_round,
        ContractApprovalStage.stage_order > stage.stage_order,
        ContractApprovalStage.status == "待審核",
    ).all()
    for s in same_round_stages:
        s.status = "已取消"

    # 合約退回草稿
    contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
    if contract and contract.contract_status == "審核中":
        contract.contract_status = "草稿"
        contract.approved_by = reviewer
        contract.approved_at = None
        contract.approval_comment = body.comment
        contract.updated_at = now

    db.commit()
    return {"success": True, "stage_id": stage_id, "status": "已拒絕", "contract_reverted": True}


# ─────────────────────────────────────────────────────────────────────────────
# I2 — 驗收管理  /{contract_id}/acceptances
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/acceptances", summary="I2 驗收記錄清單")
def list_acceptances(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractAcceptance
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    items = (
        db.query(ContractAcceptance)
        .filter(ContractAcceptance.contract_id == contract_id)
        .order_by(ContractAcceptance.acceptance_date.desc())
        .all()
    )
    return {"total": len(items), "acceptances": [AcceptanceResponse.from_orm(i).dict() for i in items]}


@router.post("/{contract_id}/acceptances", summary="I2 新增驗收記錄")
def create_acceptance(
    contract_id: str,
    body: AcceptanceCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractAcceptance, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    if not db.query(Contract).filter(Contract.contract_id == contract_id).first():
        raise HTTPException(status_code=404, detail="合約不存在")
    acc = ContractAcceptance(contract_id=contract_id, **body.dict())
    db.add(acc)
    operator = getattr(current_user, "username", "") or str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="create", resource="acceptance",
        operator=operator, payload_summary=f"新增驗收：{body.acceptance_name} {body.acceptance_date}", result="success",
    ))
    db.commit()
    db.refresh(acc)
    return AcceptanceResponse.from_orm(acc)


@router.put("/{contract_id}/acceptances/{acceptance_id}", summary="I2 更新驗收記錄")
def update_acceptance(
    contract_id: str,
    acceptance_id: int,
    body: AcceptanceUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractAcceptance, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    acc = db.query(ContractAcceptance).filter(
        ContractAcceptance.id == acceptance_id,
        ContractAcceptance.contract_id == contract_id,
    ).first()
    if not acc:
        raise HTTPException(status_code=404, detail="驗收記錄不存在")
    for k, v in body.dict(exclude_unset=True).items():
        setattr(acc, k, v)
    operator = getattr(current_user, "username", "") or str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="update", resource="acceptance",
        resource_id=str(acceptance_id), operator=operator,
        payload_summary=f"更新驗收 id={acceptance_id} 狀態={body.status or '—'}", result="success",
    ))
    db.commit()
    db.refresh(acc)
    return AcceptanceResponse.from_orm(acc)


@router.delete("/{contract_id}/acceptances/{acceptance_id}", status_code=status.HTTP_204_NO_CONTENT, summary="I2 刪除驗收記錄")
def delete_acceptance(
    contract_id: str,
    acceptance_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractAcceptance
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    acc = db.query(ContractAcceptance).filter(
        ContractAcceptance.id == acceptance_id,
        ContractAcceptance.contract_id == contract_id,
    ).first()
    if not acc:
        raise HTTPException(status_code=404, detail="驗收記錄不存在")
    db.delete(acc)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# I3 — 保證金追蹤  /{contract_id}/deposits
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/deposits", summary="I3 保證金清單")
def list_deposits(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractDeposit
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    items = (
        db.query(ContractDeposit)
        .filter(ContractDeposit.contract_id == contract_id)
        .order_by(ContractDeposit.expected_return_date)
        .all()
    )
    return {"total": len(items), "deposits": [DepositResponse.from_orm(i).dict() for i in items]}


@router.post("/{contract_id}/deposits", summary="I3 新增保證金記錄")
def create_deposit(
    contract_id: str,
    body: DepositCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractDeposit, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    if not db.query(Contract).filter(Contract.contract_id == contract_id).first():
        raise HTTPException(status_code=404, detail="合約不存在")
    dep = ContractDeposit(contract_id=contract_id, **body.dict())
    db.add(dep)
    operator = getattr(current_user, "username", "") or str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="create", resource="deposit",
        operator=operator, payload_summary=f"新增保證金：{body.deposit_type} ${body.deposit_amount:,.0f}", result="success",
    ))
    db.commit()
    db.refresh(dep)
    return DepositResponse.from_orm(dep)


@router.put("/{contract_id}/deposits/{deposit_id}", summary="I3 更新保證金記錄")
def update_deposit(
    contract_id: str,
    deposit_id: int,
    body: DepositUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractDeposit, ContractAuditLog
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    dep = db.query(ContractDeposit).filter(
        ContractDeposit.id == deposit_id,
        ContractDeposit.contract_id == contract_id,
    ).first()
    if not dep:
        raise HTTPException(status_code=404, detail="保證金記錄不存在")
    for k, v in body.dict(exclude_unset=True).items():
        setattr(dep, k, v)
    operator = getattr(current_user, "username", "") or str(current_user.id)
    db.add(ContractAuditLog(
        contract_id=contract_id, action="update", resource="deposit",
        resource_id=str(deposit_id), operator=operator,
        payload_summary=f"更新保證金 id={deposit_id} 狀態={body.status or '—'}", result="success",
    ))
    db.commit()
    db.refresh(dep)
    return DepositResponse.from_orm(dep)


@router.delete("/{contract_id}/deposits/{deposit_id}", status_code=status.HTTP_204_NO_CONTENT, summary="I3 刪除保證金記錄")
def delete_deposit(
    contract_id: str,
    deposit_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractDeposit
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    dep = db.query(ContractDeposit).filter(
        ContractDeposit.id == deposit_id,
        ContractDeposit.contract_id == contract_id,
    ).first()
    if not dep:
        raise HTTPException(status_code=404, detail="保證金記錄不存在")
    db.delete(dep)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# I4 — 年化費用計算  /{contract_id}/cost-summary
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/cost-summary", summary="I4 年化費用計算摘要")
def get_cost_summary(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractClaim
    from sqlalchemy import func as _func
    from datetime import date as _date

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合約不存在")

    # 合約期間（天數 / 月數）
    try:
        start = contract.start_date
        end   = contract.end_date
        duration_days   = (end - start).days if hasattr(end - start, "days") else 0
        duration_months = round(duration_days / 30.44, 2)
    except Exception:
        duration_days   = 0
        duration_months = 0.0

    total_amount = float(contract.total_amount_tax_included or 0)
    monthly_fixed = float(contract.monthly_fixed_amount) if contract.monthly_fixed_amount else None

    # 年化金額（月費類型）
    annual_amount = round(monthly_fixed * 12, 2) if monthly_fixed else None

    # 月攤提（總額 / 合約月數）
    monthly_amortization = round(total_amount / duration_months, 2) if duration_months > 0 else None

    # 請款金額統計
    claimed_total = float(
        db.query(_func.sum(ContractClaim.amount)).filter(
            ContractClaim.contract_id == contract_id,
        ).scalar() or 0
    )
    approved_total = float(
        db.query(_func.sum(ContractClaim.amount)).filter(
            ContractClaim.contract_id == contract_id,
            ContractClaim.status.in_(["已核准", "已付款"]),
        ).scalar() or 0
    )
    claimed_percentage = round(approved_total / total_amount * 100, 2) if total_amount > 0 else 0.0
    remaining_amount   = round(total_amount - approved_total, 2)

    return CostSummaryResponse(
        contract_id=contract_id,
        contract_name=contract.contract_name,
        total_amount=total_amount,
        monthly_fixed_amount=monthly_fixed,
        annual_amount=annual_amount,
        monthly_amortization=monthly_amortization,
        duration_days=duration_days,
        duration_months=duration_months,
        claimed_total=claimed_total,
        approved_total=approved_total,
        claimed_percentage=claimed_percentage,
        remaining_amount=remaining_amount,
        is_monthly_contract=monthly_fixed is not None,
    )


# ─────────────────────────────────────────────────────────────────────────────
# J2 — 合約成本趨勢  /analytics/cost-trend
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/analytics/cost-trend", summary="J2 合約成本趨勢圖資料")
def cost_trend(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: int = Query(2026, description="預算年度"),
    granularity: str = Query("month", description="粒度：month / quarter"),
    company: Optional[str] = Query(None, description="簽約公司篩選"),
    dept: Optional[str] = Query(None, description="負責部門篩選"),
):
    from app.models.contract import Contract, ContractClaim
    from sqlalchemy import func as _func, extract, case

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    # 合約金額：依 start_date 歸月/季
    c_q = db.query(Contract).filter(
        Contract.budget_year == budget_year,
        Contract.contract_status.notin_(["已終止"]),
    )
    if company:
        c_q = c_q.filter(Contract.signing_company == company)
    if dept:
        c_q = c_q.filter(Contract.responsible_dept == dept)
    all_contracts = c_q.all()

    # 請款金額：依 claim_date 歸月/季
    cl_q = db.query(ContractClaim).join(
        Contract, ContractClaim.contract_id == Contract.contract_id
    ).filter(
        Contract.budget_year == budget_year,
        ContractClaim.status.in_(["已核准", "已付款"]),
    )
    if company:
        cl_q = cl_q.filter(Contract.signing_company == company)
    if dept:
        cl_q = cl_q.filter(Contract.responsible_dept == dept)
    all_claims = cl_q.all()

    # 分組統計
    def period_key(date_str: str) -> str:
        if not date_str:
            return "unknown"
        try:
            import datetime
            d = date_str if isinstance(date_str, str) else str(date_str)
            year, month = int(d[:4]), int(d[5:7])
            if granularity == "quarter":
                q = (month - 1) // 3 + 1
                return f"{year}-Q{q}"
            return f"{year}-{month:02d}"
        except Exception:
            return "unknown"

    def period_label(key: str) -> str:
        if granularity == "quarter":
            return key  # "2026-Q1"
        try:
            _, m = key.split("-")
            return f"{int(m)}月"
        except Exception:
            return key

    # 依合約 start_date 統計合約總額
    contract_by_period: dict[str, float] = {}
    for c in all_contracts:
        k = period_key(str(c.start_date))
        contract_by_period[k] = contract_by_period.get(k, 0) + float(c.total_amount_tax_included or 0)

    # 依請款 claim_date 統計請款金額
    claim_by_period: dict[str, float] = {}
    for cl in all_claims:
        k = period_key(cl.claim_date)
        claim_by_period[k] = claim_by_period.get(k, 0) + float(cl.amount or 0)

    # 合併所有期間
    all_keys = sorted(set(list(contract_by_period.keys()) + list(claim_by_period.keys())))
    result = [
        {
            "period": k,
            "label": period_label(k),
            "contract_amount": round(contract_by_period.get(k, 0), 2),
            "claimed_amount": round(claim_by_period.get(k, 0), 2),
        }
        for k in all_keys if k != "unknown"
    ]

    return {
        "budget_year": budget_year,
        "granularity": granularity,
        "company": company,
        "dept": dept,
        "data": result,
    }


# ─────────────────────────────────────────────────────────────────────────────
# J3 — 月度/季度報表  /reports/summary  +  /reports/summary/export
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/reports/summary", summary="J3 月度/季度合約報表")
def reports_summary(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: int = Query(2026, description="預算年度"),
    period_type: str = Query("monthly", description="monthly / quarterly"),
):
    from app.models.contract import Contract, ContractClaim

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    def _period(date_str: str) -> str:
        try:
            d = str(date_str)
            year, month = int(d[:4]), int(d[5:7])
            if period_type == "quarterly":
                return f"{year}-Q{(month-1)//3+1}"
            return f"{year}-{month:02d}"
        except Exception:
            return "unknown"

    def _label(key: str) -> str:
        if period_type == "quarterly":
            return key
        try:
            _, m = key.split("-")
            return f"{int(m)}月"
        except Exception:
            return key

    # 新簽合約（依 start_date）
    contracts = db.query(Contract).filter(Contract.budget_year == budget_year).all()
    # 請款（依 claim_date）
    claims = db.query(ContractClaim).join(
        Contract, ContractClaim.contract_id == Contract.contract_id
    ).filter(Contract.budget_year == budget_year).all()

    periods: dict[str, dict] = {}
    for c in contracts:
        k = _period(str(c.start_date))
        if k == "unknown":
            continue
        if k not in periods:
            periods[k] = {
                "period": k, "label": _label(k),
                "new_contracts": 0, "new_amount": 0.0,
                "claim_count": 0, "claim_amount": 0.0,
                "approved_amount": 0.0,
            }
        periods[k]["new_contracts"] += 1
        periods[k]["new_amount"] += float(c.total_amount_tax_included or 0)

    for cl in claims:
        k = _period(cl.claim_date)
        if k == "unknown":
            continue
        if k not in periods:
            periods[k] = {
                "period": k, "label": _label(k),
                "new_contracts": 0, "new_amount": 0.0,
                "claim_count": 0, "claim_amount": 0.0,
                "approved_amount": 0.0,
            }
        periods[k]["claim_count"] += 1
        periods[k]["claim_amount"] += float(cl.amount or 0)
        if cl.status in ("已核准", "已付款"):
            periods[k]["approved_amount"] += float(cl.amount or 0)

    rows = sorted(periods.values(), key=lambda x: x["period"])
    # 四捨五入
    for r in rows:
        r["new_amount"]      = round(r["new_amount"], 2)
        r["claim_amount"]    = round(r["claim_amount"], 2)
        r["approved_amount"] = round(r["approved_amount"], 2)

    return {
        "budget_year": budget_year,
        "period_type": period_type,
        "rows": rows,
        "totals": {
            "new_contracts":   sum(r["new_contracts"] for r in rows),
            "new_amount":      round(sum(r["new_amount"] for r in rows), 2),
            "claim_count":     sum(r["claim_count"] for r in rows),
            "claim_amount":    round(sum(r["claim_amount"] for r in rows), 2),
            "approved_amount": round(sum(r["approved_amount"] for r in rows), 2),
        },
    }


@router.get("/reports/summary/export", summary="J3 月度/季度報表 Excel 匯出")
def export_summary_report(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    budget_year: int = Query(2026),
    period_type: str = Query("monthly"),
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse as _StreamingResponse

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    # 重用 summary 邏輯
    from app.models.contract import Contract, ContractClaim

    def _period(date_str: str) -> str:
        try:
            d = str(date_str)
            year, month = int(d[:4]), int(d[5:7])
            if period_type == "quarterly":
                return f"{year}-Q{(month-1)//3+1}"
            return f"{year}-{month:02d}"
        except Exception:
            return "unknown"

    contracts = db.query(Contract).filter(Contract.budget_year == budget_year).all()
    claims = db.query(ContractClaim).join(
        Contract, ContractClaim.contract_id == Contract.contract_id
    ).filter(Contract.budget_year == budget_year).all()

    periods: dict[str, dict] = {}
    for c in contracts:
        k = _period(str(c.start_date))
        if k == "unknown":
            continue
        if k not in periods:
            periods[k] = {"period": k, "new_contracts": 0, "new_amount": 0.0, "claim_count": 0, "claim_amount": 0.0, "approved_amount": 0.0}
        periods[k]["new_contracts"] += 1
        periods[k]["new_amount"] += float(c.total_amount_tax_included or 0)

    for cl in claims:
        k = _period(cl.claim_date)
        if k == "unknown":
            continue
        if k not in periods:
            periods[k] = {"period": k, "new_contracts": 0, "new_amount": 0.0, "claim_count": 0, "claim_amount": 0.0, "approved_amount": 0.0}
        periods[k]["claim_count"] += 1
        periods[k]["claim_amount"] += float(cl.amount or 0)
        if cl.status in ("已核准", "已付款"):
            periods[k]["approved_amount"] += float(cl.amount or 0)

    rows = sorted(periods.values(), key=lambda x: x["period"])

    # 建立 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{budget_year}年{'月度' if period_type=='monthly' else '季度'}報表"

    header_fill = PatternFill("solid", fgColor="1B3A5C")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ["期間", "新簽合約數", "新簽金額（含稅）", "請款筆數", "請款金額（含稅）", "已核准金額"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 20

    for ri, r in enumerate(rows, 2):
        ws.cell(row=ri, column=1, value=r["period"])
        ws.cell(row=ri, column=2, value=r["new_contracts"])
        ws.cell(row=ri, column=3, value=r["new_amount"])
        ws.cell(row=ri, column=4, value=r["claim_count"])
        ws.cell(row=ri, column=5, value=r["claim_amount"])
        ws.cell(row=ri, column=6, value=r["approved_amount"])

    # 合計列
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True)
    for ci in range(2, 7):
        ws.cell(row=total_row, column=ci, value=sum(r[list(r.keys())[ci-1]] for r in rows)).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"contract_report_{budget_year}_{period_type}.xlsx"
    return _StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# J4 — 批次操作  /batch/*
# ─────────────────────────────────────────────────────────────────────────────

@router.patch("/batch/manager", summary="J4 批次更新管理人")
def batch_update_manager(
    body: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractAuditLog
    from app.core.time import twnow

    contract_ids: list = body.get("contract_ids", [])
    new_manager: str = body.get("manager", "")
    if not contract_ids:
        raise HTTPException(status_code=400, detail="contract_ids 不可為空")
    if not new_manager:
        raise HTTPException(status_code=400, detail="manager 不可為空")
    if len(contract_ids) > 100:
        raise HTTPException(status_code=400, detail="單次批次上限 100 筆")

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    updated = db.query(Contract).filter(Contract.contract_id.in_(contract_ids)).all()
    operator = getattr(current_user, "username", "") or str(current_user.id)
    now = twnow()
    for c in updated:
        c.manager = new_manager
        c.updated_at = now
    db.add(ContractAuditLog(
        action="batch_update", resource="contract",
        operator=operator,
        payload_summary=f"批次更新管理人→{new_manager}，共 {len(updated)} 筆：{', '.join(contract_ids[:5])}{'...' if len(contract_ids)>5 else ''}",
        result="success",
    ))
    db.commit()
    return {"success": True, "updated": len(updated)}


@router.post("/batch/submit", summary="J4 批次送審")
def batch_submit(
    body: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractAuditLog, ContractApprovalStage, ContractApprovalConfig
    from sqlalchemy import func as _func
    from app.core.time import twnow

    contract_ids: list = body.get("contract_ids", [])
    if not contract_ids:
        raise HTTPException(status_code=400, detail="contract_ids 不可為空")
    if len(contract_ids) > 50:
        raise HTTPException(status_code=400, detail="批次送審上限 50 筆")

    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    operator = getattr(current_user, "username", "") or str(current_user.id)
    now = twnow()
    succeeded, failed = [], []

    for cid in contract_ids:
        c = db.query(Contract).filter(Contract.contract_id == cid).first()
        if not c:
            failed.append({"contract_id": cid, "reason": "不存在"})
            continue
        if c.contract_status != "草稿":
            failed.append({"contract_id": cid, "reason": f"狀態為「{c.contract_status}」，僅草稿可送審"})
            continue
        c.contract_status = "審核中"
        c.updated_at = now
        # 建立多關卡記錄
        configs = db.query(ContractApprovalConfig).filter(
            ContractApprovalConfig.contract_type.in_([c.contract_type, "*"]),
            ContractApprovalConfig.is_enabled == True,
        ).order_by(ContractApprovalConfig.stage_order).all()
        if configs:
            max_round = db.query(_func.max(ContractApprovalStage.submission_round)).filter(
                ContractApprovalStage.contract_id == cid
            ).scalar() or 0
            for cfg in configs:
                db.add(ContractApprovalStage(
                    contract_id=cid, submission_round=max_round + 1,
                    stage_order=cfg.stage_order, stage_name=cfg.stage_name,
                    assigned_to=cfg.assigned_to, status="待審核",
                ))
        succeeded.append(cid)

    db.add(ContractAuditLog(
        action="batch_submit", resource="contract", operator=operator,
        payload_summary=f"批次送審 {len(succeeded)} 筆成功，{len(failed)} 筆失敗",
        result="success" if not failed else "partial",
    ))
    db.commit()
    return {"success": True, "submitted": len(succeeded), "failed": failed}


# ─────────────────────────────────────────────────────────────────────────────
# K2 — SLA 指標定義  /{contract_id}/sla-metrics
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/sla-metrics", summary="K2 SLA 指標清單")
def list_sla_metrics(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractSlaMetric
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    items = (
        db.query(ContractSlaMetric)
        .filter(ContractSlaMetric.contract_id == contract_id)
        .order_by(ContractSlaMetric.id)
        .all()
    )
    return {"total": len(items), "metrics": [SlaMetricResponse.from_orm(m).dict() for m in items]}


@router.post("/{contract_id}/sla-metrics", summary="K2 新增 SLA 指標")
def create_sla_metric(
    contract_id: str,
    body: SlaMetricCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import Contract, ContractSlaMetric
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    if not db.query(Contract).filter(Contract.contract_id == contract_id).first():
        raise HTTPException(status_code=404, detail="合約不存在")
    m = ContractSlaMetric(contract_id=contract_id, **body.dict())
    db.add(m)
    db.commit()
    db.refresh(m)
    return SlaMetricResponse.from_orm(m)


@router.put("/{contract_id}/sla-metrics/{metric_id}", summary="K2 修改 SLA 指標")
def update_sla_metric(
    contract_id: str,
    metric_id: int,
    body: SlaMetricUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractSlaMetric
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    m = db.query(ContractSlaMetric).filter(
        ContractSlaMetric.id == metric_id,
        ContractSlaMetric.contract_id == contract_id,
    ).first()
    if not m:
        raise HTTPException(status_code=404, detail="SLA 指標不存在")
    for k, v in body.dict(exclude_unset=True).items():
        setattr(m, k, v)
    db.commit()
    db.refresh(m)
    return SlaMetricResponse.from_orm(m)


@router.delete("/{contract_id}/sla-metrics/{metric_id}",
               status_code=status.HTTP_204_NO_CONTENT, summary="K2 刪除 SLA 指標")
def delete_sla_metric(
    contract_id: str,
    metric_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractSlaMetric
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    m = db.query(ContractSlaMetric).filter(
        ContractSlaMetric.id == metric_id,
        ContractSlaMetric.contract_id == contract_id,
    ).first()
    if not m:
        raise HTTPException(status_code=404, detail="SLA 指標不存在")
    db.delete(m)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# K2 — SLA 達成記錄  /{contract_id}/sla-records
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/sla-records", summary="K2 SLA 達成記錄清單")
def list_sla_records(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    metric_id: Optional[int] = Query(None, description="篩選特定指標"),
):
    from app.models.contract import ContractSlaRecord
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")
    q = db.query(ContractSlaRecord).filter(ContractSlaRecord.contract_id == contract_id)
    if metric_id:
        q = q.filter(ContractSlaRecord.metric_id == metric_id)
    records = q.order_by(ContractSlaRecord.period_label).all()
    return {"total": len(records), "records": [SlaRecordResponse.from_orm(r).dict() for r in records]}


@router.post("/{contract_id}/sla-records", summary="K2 登錄 SLA 達成記錄")
def create_sla_record(
    contract_id: str,
    body: SlaRecordCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractSlaMetric, ContractSlaRecord
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    metric = db.query(ContractSlaMetric).filter(
        ContractSlaMetric.id == body.metric_id,
        ContractSlaMetric.contract_id == contract_id,
    ).first()
    if not metric:
        raise HTTPException(status_code=404, detail="SLA 指標不存在")

    # 自動判斷是否達標（數值類型：實際值 >= 目標值 視為達標）
    actual = float(body.actual_value)
    target = float(metric.target_value)
    achieved = actual >= target

    operator = getattr(current_user, "username", "") or str(current_user.id)
    rec = ContractSlaRecord(
        metric_id=body.metric_id,
        contract_id=contract_id,
        period_label=body.period_label,
        period_start=body.period_start,
        period_end=body.period_end,
        actual_value=body.actual_value,
        target_value=metric.target_value,
        achieved=achieved,
        notes=body.notes,
        recorded_by=operator,
    )
    db.add(rec)
    db.commit()
    db.refresh(rec)
    return SlaRecordResponse.from_orm(rec)


@router.delete("/{contract_id}/sla-records/{record_id}",
               status_code=status.HTTP_204_NO_CONTENT, summary="K2 刪除 SLA 記錄")
def delete_sla_record(
    contract_id: str,
    record_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractSlaRecord
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_create_edit" not in user_permissions:
        raise HTTPException(status_code=403, detail="権限不足")
    rec = db.query(ContractSlaRecord).filter(
        ContractSlaRecord.id == record_id,
        ContractSlaRecord.contract_id == contract_id,
    ).first()
    if not rec:
        raise HTTPException(status_code=404, detail="SLA 記錄不存在")
    db.delete(rec)
    db.commit()
    return None


# ─────────────────────────────────────────────────────────────────────────────
# K2 — SLA 達成率摘要  /{contract_id}/sla-summary
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/{contract_id}/sla-summary", summary="K2 SLA 達成率摘要")
def get_sla_summary(
    contract_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.contract import ContractSlaMetric, ContractSlaRecord
    user_permissions = get_user_permissions(current_user.id, db)
    if "*" not in user_permissions and "contract_view" not in user_permissions:
        raise HTTPException(status_code=403, detail="權限不足")

    metrics = db.query(ContractSlaMetric).filter(
        ContractSlaMetric.contract_id == contract_id,
        ContractSlaMetric.is_enabled == True,
    ).all()

    metrics_summary = []
    all_rates = []

    for m in metrics:
        records = db.query(ContractSlaRecord).filter(
            ContractSlaRecord.metric_id == m.id
        ).order_by(ContractSlaRecord.period_label).all()

        total = len(records)
        achieved_count = sum(1 for r in records if r.achieved)
        rate = round(achieved_count / total * 100, 1) if total > 0 else None

        if rate is not None:
            all_rates.append(rate)

        metrics_summary.append({
            "metric_id": m.id,
            "metric_name": m.metric_name,
            "metric_type": m.metric_type,
            "target_value": float(m.target_value),
            "target_unit": m.target_unit,
            "measurement_period": m.measurement_period,
            "record_count": total,
            "achieved_count": achieved_count,
            "achievement_rate": rate,
            "trend": [
                {
                    "period": r.period_label,
                    "actual": float(r.actual_value),
                    "target": float(r.target_value),
                    "achieved": r.achieved,
                }
                for r in records[-12:]   # 最近 12 期
            ],
        })

    overall = round(sum(all_rates) / len(all_rates), 1) if all_rates else None
    return {
        "contract_id": contract_id,
        "metrics": metrics_summary,
        "overall_achievement_rate": overall,
    }
