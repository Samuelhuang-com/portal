"""
IHG 客房保養 API Router
Prefix: /api/v1/ihg-room-maintenance

端點說明：
  GET  /matrix          — 年度矩陣表（房號 × 月份）
  GET  /section-matrix  — 月份區段矩陣（房號 × 類別，V/▲/X）
  GET  /stats           — KPI 統計卡
  GET  /records         — 原始記錄清單（含篩選）
  GET  /debug-raw       — Ragic 原始欄位結構（除錯用）
  GET  /{ragic_id}      — 單筆明細（含子表格）
  POST /sync            — 觸發同步

矩陣表邏輯：
  - 以房號為行、月份（1-12）為欄
  - 每格顯示狀態、日期、人員
  - 狀態判斷：
      completed  → 已完成（green）
      overdue    → 逾期（red）    = 本月應保養但超過今日且未完成
      scheduled  → 本月應保養（yellow/blue）
      pending    → 未完成（grey）

統計卡：
  全年應保養數、已完成數、未完成數、逾期數、完成率
"""
import calendar as _calendar
import json
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.time import twnow
from app.dependencies import get_current_user, require_roles
from app.models.ihg_room_maintenance import (
    IHGRoomMaintenanceMaster,
    IHGRoomMaintenanceDetail,
    IHGRoomMaintenanceSection,
)
from app.services.ihg_room_maintenance_sync import (
    sync_from_ragic, _derive_floor, _parse_minutes,
    IHG_SERVER_URL, IHG_ACCOUNT, IHG_SHEET_PATH,
)
from app.services.ragic_verify_utils import (
    read_portal_count_and_last_sync, read_portal_ragic_ids,
    fetch_ragic_count, fetch_ragic_url_map_single,
    build_verify_count_response, build_verify_diff_response,
)

router = APIRouter(dependencies=[Depends(get_current_user)])

# ── IHG 規範房號清單（單一來源，新增/移除房號只改這裡）────────────────────────
CANONICAL_ROOMS: list[str] = [
    # 5F
    "501","502","503","505","506","507","508","509","510","511","512","513",
    "515","516","517","518","519","520","521","522","523","525","526","527",
    "528","529","530","531",
    # 6F
    "601","602","603","605","606","607","608","609","610","611","612","613",
    "615","616","617","618","619","620","621","622","623","625","626","627",
    "628","629","630","631",
    # 7F
    "701","702","703","705","706","707","708","709","710","711","712","713",
    "715","716","717","718","719","720","721","722","723","725","726","727",
    "728","729","730","731",
    # 8F
    "801","803","805","806","807","808","809","810","811","812","813",
    "815","816","817","818","819","820","821","822","823","825","826","827",
    "828","829","830","831",
    # 9F
    "909","910","911","912","913","915","916","917","918","919","920","921",
    "922","923","925","926","927","928","929","930","931",
    # 10F
    "1013","1015","1016","1017","1018","1019","1020","1021","1022","1023",
    "1025","1026","1027","1028","1029","1030","1031",
]
CANONICAL_ROOM_SET: frozenset[str] = frozenset(CANONICAL_ROOMS)

# ── 保養檢查欄位：忽略清單 ────────────────────────────────────────────────────
# 明確不屬於保養檢查項目的欄位名稱（精確匹配）
_IGNORE_FIELD_NAMES: frozenset[str] = frozenset({
    # 使用者指定忽略
    "項目", "更換日期", "費用", "設備保養上傳照片", "設備",
    # 已知 metadata 欄位
    "保養月份", "保養人員", "保養時間起", "保養時間迄", "保養日期",
    "工時計算", "房號", "複核人員", "是否有陽台",
})

def _is_check_field(key: str, value: object) -> bool:
    """
    判斷是否為有效的保養檢查項目欄位。
    符合條件的欄位值：正常 / 當時維護完成 / 等待維護(待料中) / ""（未檢查）
    """
    if key in _IGNORE_FIELD_NAMES:
        return False
    if "上傳照片" in key:
        return False
    if isinstance(value, str) and value in ("正常", "當時維護完成", "等待維護(待料中)", ""):
        return True
    # None 視為未檢查
    if value is None:
        return True
    return False


# ── 輔助 ─────────────────────────────────────────────────────────────────────

_RAGIC_BASE = "https://ap12.ragic.com/soutlet001"
_IHG_RAGIC_PATH = "periodic-maintenance/4"


def _ragic_url(ragic_id: str) -> str:
    """組出 IHG 客房保養 Ragic 記錄直連 URL。"""
    if not ragic_id:
        return ""
    numeric_id = ragic_id.rsplit("_", 1)[-1] if "_" in ragic_id else ragic_id
    return f"{_RAGIC_BASE}/{_IHG_RAGIC_PATH}/{numeric_id}"


def _cell_status(rec: IHGRoomMaintenanceMaster) -> str:
    """
    計算矩陣格基礎狀態（check 欄位結果會在 /matrix 再覆蓋）
    Ragic 無逾期概念，只區分：本月應保養 / 待排程
    最終有效狀態：completed / abnormal / scheduled / pending
    """
    today = twnow()
    try:
        if rec.maint_year and rec.maint_month:
            rec_year  = int(rec.maint_year)
            rec_month = int(rec.maint_month)
            if (rec_year, rec_month) == (today.year, today.month):
                return "scheduled"
    except (ValueError, TypeError):
        pass
    return "pending"


def _room_sort_key(room_no: str) -> tuple:
    """房號排序：先依樓層數字、再依房間序號"""
    digits = "".join(c for c in room_no if c.isdigit())
    try:
        n = int(digits)
        floor = n // 100
        seq   = n % 100
        return (floor, seq)
    except ValueError:
        return (999, 0)


def _calc_record_status(raw_json_str: Optional[str]) -> str:
    """
    從 raw_json 字串計算保養記錄狀態（與 /matrix 邏輯一致）。
    回傳值：'completed' | 'abnormal' | 'pending'
    """
    normal_c = done_c = maint_c = 0
    try:
        raw_data = json.loads(raw_json_str or "{}")
        for k, v in raw_data.items():
            if not _is_check_field(k, v):
                continue
            val = v if isinstance(v, str) else ""
            if val == "正常":
                normal_c += 1
            elif val == "當時維護完成":
                done_c += 1
            elif val == "等待維護(待料中)":
                maint_c += 1
    except Exception:
        pass
    if maint_c > 0:
        return "abnormal"
    if normal_c + done_c > 0:
        return "completed"
    return "pending"


# ── GET /debug-raw ────────────────────────────────────────────────────────────

@router.get("/debug-raw", summary="[除錯] 查看 Ragic IHG Sheet 4 原始欄位結構", dependencies=[Depends(require_roles("system_admin", "module_manager"))])
async def debug_raw():
    """
    直接回傳 Ragic Sheet 4 第一筆記錄的原始 key/value，
    用於確認欄位 ID 與中文 label。
    ⚠️ 建議同步後刪除或加 auth 保護。
    """
    from app.services.ragic_adapter import RagicAdapter
    from app.services.ihg_room_maintenance_sync import IHG_SERVER_URL, IHG_ACCOUNT, IHG_SHEET_PATH

    adapter = RagicAdapter(
        sheet_path=IHG_SHEET_PATH,
        server_url=IHG_SERVER_URL,
        account=IHG_ACCOUNT,
    )
    try:
        raw_data = await adapter.fetch_all()
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Ragic 拉取失敗：{exc}")

    total = len(raw_data)
    first_id, first_rec = next(iter(raw_data.items()), (None, {}))

    # 取第一筆完整記錄（含子表格）
    detail_structure = {}
    if first_id:
        try:
            full = await adapter.fetch_one(str(first_id))
            if str(first_id) in full and len(full) == 1:
                full = full[str(first_id)]
            detail_structure = {
                k: repr(v)[:120] for k, v in full.items()
            }
        except Exception as exc:
            detail_structure = {"error": str(exc)}

    return {
        "sheet_path": IHG_SHEET_PATH,
        "server_url": IHG_SERVER_URL,
        "account": IHG_ACCOUNT,
        "total_records": total,
        "first_ragic_id": first_id,
        "master_keys": {k: repr(v)[:80] for k, v in first_rec.items()},
        "full_record_keys": detail_structure,
    }


# ── POST /sync ────────────────────────────────────────────────────────────────

@router.post("/sync", summary="從 Ragic 同步 IHG 客房保養資料（背景執行）", dependencies=[Depends(require_roles("system_admin", "module_manager"))])
async def sync_records(background_tasks: BackgroundTasks):
    """觸發背景同步：Ragic Sheet 4 → ihg_rm_master + ihg_rm_detail"""
    background_tasks.add_task(sync_from_ragic)
    return {"success": True, "message": "IHG 客房保養同步已在背景啟動"}


# ── /verify-count／/verify-diff ─────────────────────────────────────────────────

@router.get("/verify-count", summary="與 Ragic 數量比對（管理員）",
            dependencies=[Depends(require_roles("system_admin"))])
async def verify_count(db: Session = Depends(get_db)):
    portal_count, last_synced_at = await read_portal_count_and_last_sync(
        db, IHGRoomMaintenanceMaster, "IHG客房保養"
    )
    try:
        ragic_count = await fetch_ragic_count(IHG_SHEET_PATH, IHG_SERVER_URL, IHG_ACCOUNT)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Ragic 連線失敗：{exc}")
    return build_verify_count_response("IHG客房保養", portal_count, ragic_count, last_synced_at)


@router.get("/verify-diff", summary="與 Ragic 明細差集比對（管理員）",
            dependencies=[Depends(require_roles("system_admin"))])
async def verify_diff(db: Session = Depends(get_db)):
    try:
        ragic_url_map = await fetch_ragic_url_map_single(IHG_SHEET_PATH, IHG_SERVER_URL, IHG_ACCOUNT)
    except Exception as exc:
        raise HTTPException(status_code=502, detail=f"Ragic 連線失敗：{exc}")
    portal_ids = await read_portal_ragic_ids(db, IHGRoomMaintenanceMaster)
    return build_verify_diff_response(ragic_url_map, portal_ids)


# ── GET /stats ────────────────────────────────────────────────────────────────

@router.get("/stats", summary="IHG 客房保養 KPI 統計")
def get_stats(
    year:  Optional[str] = Query(None, description="篩選年度，如 2026；空白=當年"),
    month: Optional[str] = Query(None, description="篩選月份（不補零，如 4）；空白=全年"),
    db: Session = Depends(get_db),
):
    """
    回傳統計卡：
      total_scheduled  — 有執行的房間數（distinct room_no；指定月份時為當月房間數，否則全年）
      completed        — 已完成房間數
      pending          — 未完成數（不含異常）
      abnormal         — 異常數
      completion_rate  — 完成率（%）
      work_hours       — 工時合計（小時，來自 raw_json「工時計算」欄位加總 / 60）
    """
    if not year:
        year = str(twnow().year)

    q = db.query(IHGRoomMaintenanceMaster).filter(
        IHGRoomMaintenanceMaster.maint_year == year,
        IHGRoomMaintenanceMaster.room_no.in_(CANONICAL_ROOM_SET),
    )
    if month:
        q = q.filter(IHGRoomMaintenanceMaster.maint_month == month.zfill(2))
    all_recs = q.all()

    # 去重：若同一房號同一月份有多筆 Ragic 記錄，只取最後一筆（與矩陣表行為一致）
    dedup: dict[tuple[str, str], "IHGRoomMaintenanceMaster"] = {}
    for r in all_recs:
        key = (r.room_no, r.maint_month or "")
        dedup[key] = r  # 後者覆蓋前者，與 /matrix 寫入 room_map 的行為一致

    # 有執行的房間數 = distinct room_no（不受月份去重影響）
    total = len({r.room_no for r in dedup.values()})

    # 工時加總：優先用 DB 欄位 work_minutes，fallback 解析 raw_json（舊資料相容）
    total_work_minutes = 0.0
    for r in all_recs:
        if r.work_minutes is not None:
            total_work_minutes += r.work_minutes
        else:
            try:
                raw_data = json.loads(r.raw_json or "{}")
                v = _parse_minutes(raw_data.get("工時計算"))
                if v is not None:
                    total_work_minutes += v
            except Exception:
                pass

    # 以 raw_json check 欄位計算各狀態（與 /matrix 邏輯完全一致，使用去重後記錄）
    completed_count = 0
    abnormal_count  = 0
    pending_count   = 0

    for r in dedup.values():
        normal_c = done_c = maint_c = unchecked_c = 0
        try:
            raw_data = json.loads(r.raw_json or "{}")
            for k, v in raw_data.items():
                if not _is_check_field(k, v):
                    continue
                val = v if isinstance(v, str) else ""
                if val == "正常":
                    normal_c += 1
                elif val == "當時維護完成":
                    done_c += 1
                elif val == "等待維護(待料中)":
                    maint_c += 1
                else:
                    unchecked_c += 1
        except Exception:
            pass

        if maint_c > 0:
            abnormal_count += 1
        elif normal_c + done_c > 0 and unchecked_c == 0:
            completed_count += 1
        else:
            pending_count += 1

    rate = round(completed_count / total * 100, 1) if total else 0.0
    work_hours = round(total_work_minutes / 60, 2)

    return {
        "year":            year,
        "month":           month,
        "total_scheduled": total,
        "completed":       completed_count,
        "abnormal":        abnormal_count,
        "pending":         pending_count,
        "completion_rate": rate,
        "work_hours":      work_hours,
        "synced_at": (
            db.query(func.max(IHGRoomMaintenanceMaster.synced_at)).scalar() or ""
        ),
    }


# ── GET /matrix ───────────────────────────────────────────────────────────────

@router.get("/matrix", summary="IHG 客房保養年度矩陣表")
def get_matrix(
    year:    Optional[str] = Query(None, description="年度，如 2026；空白=當年"),
    room_no: Optional[str] = Query(None, description="房號篩選，支援前綴匹配"),
    floor:   Optional[str] = Query(None, description="樓層篩選，如 5F"),
    cell_status: Optional[str] = Query(None, description="狀態篩選：completed/pending/overdue/scheduled"),
    db: Session = Depends(get_db),
):
    """
    回傳年度矩陣表，格式：
    {
      "year": "2026",
      "months": [1..12],
      "rooms": [
        {
          "room_no": "501",
          "floor": "5F",
          "cells": {
            "1": {"ragic_id": "...", "status": "completed", "date": "...", "assignee": "..."},
            "4": {"ragic_id": "...", "status": "pending", ...},
            ...  (只有有記錄的月份才有 key)
          }
        }
      ]
    }
    """
    if not year:
        year = str(twnow().year)

    q = db.query(IHGRoomMaintenanceMaster).filter(
        IHGRoomMaintenanceMaster.maint_year == year,
        IHGRoomMaintenanceMaster.room_no.in_(CANONICAL_ROOM_SET),
    )
    if room_no:
        q = q.filter(IHGRoomMaintenanceMaster.room_no.ilike(f"{room_no}%"))
    if floor:
        q = q.filter(IHGRoomMaintenanceMaster.floor == floor)

    all_recs = q.all()

    # ── 以規範清單為基礎初始化 room_map（保證所有規範房號都出現，即使無資料）──
    # 若有 room_no / floor 篩選條件，也要套用到初始化清單
    canonical_filtered = [
        rn for rn in CANONICAL_ROOMS
        if (not room_no or rn.startswith(room_no))
        and (not floor or _derive_floor(rn) == floor)
    ]
    room_map: dict[str, dict] = {
        rn: {"room_no": rn, "floor": _derive_floor(rn), "cells": {}}
        for rn in canonical_filtered
    }

    month_minutes: dict[str, float] = {}     # month_key → 分鐘加總
    month_orders:  dict[str, int]   = {}     # month_key → 工單（記錄）數加總
    for rec in all_recs:
        rno = rec.room_no or "??"
        # 只處理規範清單內的房號（非規範房號的 Ragic 資料跳過）
        if rno not in room_map:
            continue
        month_key = str(int(rec.maint_month)) if rec.maint_month else "?"
        cell_stat = _cell_status(rec)

        # ── 解析 raw_json 統計保養項目結果 + 工時計算 ──────────────────────
        normal_count = done_count = maint_count = unchecked_count = 0
        work_minutes: float = 0.0
        try:
            raw_data = json.loads(rec.raw_json or "{}")
            for k, v in raw_data.items():
                # 單筆工時（分鐘）— 先獨立抽取，不受 _is_check_field 影響
                if k == "工時計算":
                    parsed = _parse_minutes(v)
                    if parsed is not None:
                        work_minutes = parsed

                # 保養檢查項目計數
                if not _is_check_field(k, v):
                    continue
                val = v if isinstance(v, str) else ""
                if val == "正常":
                    normal_count += 1
                elif val == "當時維護完成":
                    done_count += 1
                elif val == "等待維護(待料中)":
                    maint_count += 1
                else:                    # "" 或 None → 未檢查
                    unchecked_count += 1
        except Exception:
            pass

        # 工時：優先 DB 欄位，fallback raw_json 解析（確保已同步記錄也正確）
        if rec.work_minutes is not None:
            work_minutes = rec.work_minutes
        if work_minutes:
            month_minutes[month_key] = month_minutes.get(month_key, 0.0) + work_minutes
        month_orders[month_key] = month_orders.get(month_key, 0) + 1

        # ── 以 check 欄位結果決定狀態（優先於日期邏輯）────────────────────
        if maint_count > 0:
            # 有「等待維護(待料中)」→ 異常
            cell_stat = "abnormal"
        elif normal_count + done_count > 0 and unchecked_count == 0:
            # 有 check 資料且全部正常/完成（無未檢查）→ 已完成
            cell_stat = "completed"
        # 否則（有未檢查 or 無 check 資料）→ 保持日期邏輯（scheduled / pending）

        # 狀態篩選（後端過濾）
        if cell_status and cell_stat != cell_status:
            continue

        # ── 同房同月多筆記錄：工時合計、records 保留各筆明細（不覆蓋）─────────
        prev_cell = room_map[rno]["cells"].get(month_key)
        rec_entry = {
            "ragic_id":     rec.ragic_id,
            "date":         rec.maint_date,
            "assignee":     rec.assignee_name,
            "work_minutes": int(work_minutes) if work_minutes else None,
        }
        records = (prev_cell["records"] if prev_cell else []) + [rec_entry]
        total_minutes = sum(r["work_minutes"] or 0 for r in records)

        room_map[rno]["cells"][month_key] = {
            "ragic_id":     rec.ragic_id,
            "status":       cell_stat,
            "date":         rec.maint_date,
            "assignee":     rec.assignee_name,
            "completion_date": rec.completion_date,
            "maint_type":   rec.maint_type,
            "notes":        rec.notes,
            "normal_count":    normal_count,
            "done_count":      done_count,
            "maint_count":     maint_count,
            "unchecked_count": unchecked_count,
            "work_minutes":    int(total_minutes) if total_minutes else None,
            "record_count":    len(records),
            "records":         records,
        }

    # 狀態篩選時，移除沒有任何符合 cell 的房號（空列無意義）
    if cell_status:
        room_map = {rn: rd for rn, rd in room_map.items() if rd["cells"]}

    # 依樓層排序
    sorted_rooms = sorted(room_map.values(), key=lambda r: _room_sort_key(r["room_no"]))

    # 取所有已出現的樓層清單（供前端篩選）
    floors = sorted(
        set(r["floor"] for r in sorted_rooms if r["floor"]),
        key=lambda f: _room_sort_key(f.replace("F", "00"))
    )

    # 分鐘 → 小時（小數點後 2 位），key 轉 int
    month_hours = {
        int(mk): round(mins / 60, 2)
        for mk, mins in month_minutes.items()
        if mk.lstrip("-").isdigit()
    }
    month_order_counts = {
        int(mk): cnt
        for mk, cnt in month_orders.items()
        if mk.lstrip("-").isdigit()
    }

    return {
        "year": year,
        "months": list(range(1, 13)),
        "floors": floors,
        "rooms": sorted_rooms,
        "month_hours": month_hours,     # {1: 12.50, 4: 10.33, ...} 只含有資料的月份
        "month_orders": month_order_counts,  # {1: 28, 4: 30, ...} 各月工單（記錄）數
    }


# ── GET /export-matrix ────────────────────────────────────────────────────────

@router.get("/export-matrix", summary="IHG 客房保養年度矩陣匯出 Excel")
def export_matrix(
    year:    Optional[str] = Query(None, description="年度，如 2026；空白=當年"),
    room_no: Optional[str] = Query(None, description="房號篩選，支援前綴匹配"),
    floor:   Optional[str] = Query(None, description="樓層篩選，如 5F"),
    cell_status: Optional[str] = Query(None, description="狀態篩選：completed/abnormal/scheduled/pending"),
    db: Session = Depends(get_db),
):
    """
    匯出年度矩陣為 Excel（兩個 Sheet）：
      Sheet 1「年度矩陣」 — 房號 × 月份（固定月份視角），格內為狀態＋保養日期，
                            底色沿用畫面色系；底部附月工時列
      Sheet 2「KPI 統計」 — 全年應保養/已完成/異常/待保養/完成率/工時合計
    篩選參數與 /matrix 相同，矩陣內容與畫面當前篩選一致；KPI 統計固定為全年。
    """
    import io
    from urllib.parse import quote

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # 複用 /matrix 與 /stats 的查詢邏輯，確保與畫面一致
    matrix = get_matrix(year=year, room_no=room_no, floor=floor, cell_status=cell_status, db=db)
    stats  = get_stats(year=matrix["year"], month=None, db=db)
    y = matrix["year"]

    # 狀態 → 顯示文字 / 色票（沿用前端年度矩陣 STATUS_CFG 色系）
    status_cfg = {
        "completed": ("已完成",     "F6FFED", "389E0D"),
        "abnormal":  ("異常",       "FFF7E6", "D46B08"),
        "scheduled": ("本月應保養", "E6F4FF", "1677FF"),
        "pending":   ("待保養",     "FAFAFA", "8C8C8C"),
    }

    wb = openpyxl.Workbook()

    # ── Sheet 1：年度矩陣 ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = f"{y}年度矩陣"[:31]

    header_fill = PatternFill("solid", start_color="1B3A5C")
    header_font = Font(bold=True, color="FFFFFF")
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["樓層", "房號"] + [f"{m}月" for m in matrix["months"]]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for ri, room in enumerate(matrix["rooms"], start=2):
        ws.cell(row=ri, column=1, value=room["floor"]).alignment   = center
        ws.cell(row=ri, column=2, value=room["room_no"]).alignment = center
        for mi, m in enumerate(matrix["months"], start=3):
            c = room["cells"].get(str(m))
            cell = ws.cell(row=ri, column=mi)
            cell.alignment = center
            if not c:
                continue
            label, bg, fg = status_cfg.get(c["status"], ("", "FFFFFF", "000000"))
            cell.value = f"{label}\n{c['date']}" if c.get("date") else label
            cell.fill  = PatternFill("solid", start_color=bg)
            cell.font  = Font(color=fg, size=10)

    # 底部：月工時列
    hours_row = len(matrix["rooms"]) + 2
    ws.cell(row=hours_row, column=1, value="月工時(小時)").font = Font(bold=True)
    for mi, m in enumerate(matrix["months"], start=3):
        v = matrix["month_hours"].get(m)
        if v is not None:
            cell = ws.cell(row=hours_row, column=mi, value=v)
            cell.font = Font(bold=True)
            cell.alignment = center

    ws.freeze_panes = "C2"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 8
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # ── Sheet 2：KPI 統計 ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("KPI 統計")
    for ci, h in enumerate(("項目", "數值"), 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = header_fill
        cell.font = header_font
    kpi_rows = [
        ("年度",                  y),
        ("全年應保養（房間數）",  stats["total_scheduled"]),
        ("已完成",                stats["completed"]),
        ("異常",                  stats["abnormal"]),
        ("待保養",                stats["pending"]),
        ("完成率（%）",           stats["completion_rate"]),
        ("工時合計（小時）",      stats["work_hours"]),
        ("統計範圍",              "全年（不受樓層/狀態篩選影響）"),
        ("匯出時間",              twnow().strftime("%Y/%m/%d %H:%M")),
    ]
    for ri, (k, v) in enumerate(kpi_rows, start=2):
        ws2.cell(row=ri, column=1, value=k)
        ws2.cell(row=ri, column=2, value=v)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 32

    # ── 回傳 ────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename_cn   = f"IHG客房保養_年度矩陣_{y}.xlsx"
    filename_safe = f"ihg_matrix_{y}.xlsx"
    encoded = quote(filename_cn, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_safe}"; '
                f"filename*=UTF-8''{encoded}"
            )
        },
    )


# ── 區段類別標準清單（顯示順序）────────────────────────────────────────────────
CANONICAL_CATEGORIES: list[str] = [
    "客房房門", "客房消防", "客房設備", "客房傢俱",
    "客房燈/電源", "客房窗", "面盆/台面", "浴厠", "浴間", "天地壁",
    "客房空調", "陽台",
]


# ── GET /section-matrix ───────────────────────────────────────────────────────

@router.get("/section-matrix", summary="客房保養區段矩陣（月份 × 類別）")
def section_matrix(
    year:  str          = Query(..., description="年度，如 2026"),
    month: str          = Query(..., description="月份，如 04 或 4"),
    floor: Optional[str] = Query(None, description="樓層篩選，如 5F"),
    db: Session = Depends(get_db),
):
    """
    回傳指定月份各房間的保養區段狀態矩陣。

    欄格值域：
      V  — 已完成
      ▲  — 當時維護完成（現場處理）
      X  — 待料中（等待維護）
      空白 — 該房間無此類別記錄

    回傳 categories 依 CANONICAL_CATEGORIES 順序排列，
    僅包含該月份至少有一筆資料的類別。
    """
    month_zf = month.zfill(2)

    # ── 取該月所有 master（只篩年月 + canonical，floor 在組合清單時套用）───
    master_rows = db.query(
        IHGRoomMaintenanceMaster.ragic_id,
        IHGRoomMaintenanceMaster.room_no,
        IHGRoomMaintenanceMaster.floor,
        IHGRoomMaintenanceMaster.maint_date,
    ).filter(
        IHGRoomMaintenanceMaster.maint_year  == year,
        IHGRoomMaintenanceMaster.maint_month == month_zf,
        IHGRoomMaintenanceMaster.room_no.in_(CANONICAL_ROOM_SET),
    ).all()

    masters_by_id   = {r.ragic_id: r for r in master_rows}
    masters_by_room = {r.room_no: r for r in master_rows}
    master_ids      = list(masters_by_id.keys())

    # ── 取所有 section 資料 ──────────────────────────────────────────────────
    sections: list[IHGRoomMaintenanceSection] = (
        db.query(IHGRoomMaintenanceSection)
        .filter(IHGRoomMaintenanceSection.master_ragic_id.in_(master_ids))
        .all()
    ) if master_ids else []

    section_map: dict[str, dict[str, str]] = {}
    for s in sections:
        section_map.setdefault(s.master_ragic_id, {})[s.category] = s.value

    # ── 組合全部 canonical 房間（含無資料者；套用 floor 篩選）──────────────
    rooms_out = []
    for rno in CANONICAL_ROOMS:
        rno_floor = _derive_floor(rno)
        if floor and rno_floor != floor:
            continue
        if rno in masters_by_room:
            mrow = masters_by_room[rno]
            room_sections = section_map.get(mrow.ragic_id, {})
            rooms_out.append({
                "room_no":    rno,
                "floor":      mrow.floor,
                "maint_date": mrow.maint_date,
                "ragic_id":   mrow.ragic_id,
                "sections":   room_sections,
                "has_data":   True,
            })
        else:
            rooms_out.append({
                "room_no":    rno,
                "floor":      rno_floor,
                "maint_date": "",
                "ragic_id":   "",
                "sections":   {},
                "has_data":   False,
            })

    total_rooms = len(rooms_out)

    # ── 各類別統計（分母含未執行房間）────────────────────────────────────────
    TRI = "▲"
    category_stats: dict[str, dict] = {}
    for cat in CANONICAL_CATEGORIES:
        v_count   = sum(1 for r in rooms_out if r["sections"].get(cat) == "V")
        tri_count = sum(1 for r in rooms_out if r["sections"].get(cat) == TRI)
        x_count   = sum(1 for r in rooms_out if r["sections"].get(cat) == "X")
        reported  = v_count + tri_count + x_count
        rate = round(v_count / total_rooms * 100, 1) if total_rooms > 0 else 0.0
        category_stats[cat] = {
            "v_count":        v_count,
            "triangle_count": tri_count,
            "x_count":        x_count,
            "reported":       reported,
            "rate":           rate,
        }

    active_cats = [c for c in CANONICAL_CATEGORIES if category_stats[c]["reported"] > 0]
    if not active_cats:
        active_cats = list(CANONICAL_CATEGORIES)

    return {
        "year":           year,
        "month":          month_zf,
        "categories":     active_cats,
        "rooms":          rooms_out,
        "category_stats": {c: category_stats[c] for c in active_cats},
        "total_rooms":    total_rooms,
    }


# ── GET /records ──────────────────────────────────────────────────────────────

@router.get("/records", summary="IHG 客房保養記錄清單（帶篩選）")
def list_records(
    year:    Optional[str] = Query(None, description="年度"),
    month:   Optional[str] = Query(None, description="月份（不補零，如 4）"),
    day:     Optional[str] = Query(None, description="日（不補零，如 5）；篩選 maint_date 的日部分"),
    room_no: Optional[str] = Query(None, description="房號（前綴匹配）"),
    floor:   Optional[str] = Query(None, description="樓層，如 5F"),
    rec_status: Optional[str] = Query(None, alias="status", description="狀態：completed / pending"),
    page:    int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    """原始清單，支援多維篩選與分頁（新增 day 篩選）"""
    q = db.query(IHGRoomMaintenanceMaster).filter(
        IHGRoomMaintenanceMaster.room_no.in_(CANONICAL_ROOM_SET)
    )
    if year:
        q = q.filter(IHGRoomMaintenanceMaster.maint_year == year)
    if month:
        q = q.filter(IHGRoomMaintenanceMaster.maint_month == month.zfill(2))
    if day:
        try:
            d = int(day)
            day_zf = str(d).zfill(2)
            # 支援格式：
            #   分隔符 / 或 -；日補零或不補零；
            #   日後可接時間（如 "2026/06/21 00:00:00"，與 /calendar 解析一致）
            patterns = []
            for sep in ("/", "-"):
                for dd in {day_zf, str(d)}:
                    patterns.append(IHGRoomMaintenanceMaster.maint_date.like(f"%{sep}{dd}"))
                    patterns.append(IHGRoomMaintenanceMaster.maint_date.like(f"%{sep}{dd} %"))
            q = q.filter(or_(*patterns))
        except (ValueError, TypeError):
            pass
    if room_no:
        q = q.filter(IHGRoomMaintenanceMaster.room_no.ilike(f"{room_no}%"))
    if floor:
        q = q.filter(IHGRoomMaintenanceMaster.floor == floor)
    if rec_status:
        if rec_status == "completed":
            q = q.filter(IHGRoomMaintenanceMaster.is_completed == True)
        elif rec_status == "pending":
            q = q.filter(IHGRoomMaintenanceMaster.is_completed == False)
        else:
            q = q.filter(IHGRoomMaintenanceMaster.status == rec_status)

    total = q.count()
    recs  = (
        q.order_by(IHGRoomMaintenanceMaster.room_no, IHGRoomMaintenanceMaster.maint_month)
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return {
        "total": total,
        "page": page,
        "per_page": per_page,
        "data": [
            {
                "ragic_id":        r.ragic_id,
                "room_no":         r.room_no,
                "floor":           r.floor,
                "maint_year":      r.maint_year,
                "maint_month":     r.maint_month,
                "maint_date":      r.maint_date,
                "status":          _cell_status(r),
                "is_completed":    r.is_completed,
                "assignee_name":   r.assignee_name,
                "checker_name":    r.checker_name,
                "completion_date": r.completion_date,
                "maint_type":      r.maint_type,
                "notes":           r.notes,
                "start_time":      r.start_time,
                "end_time":        r.end_time,
                "work_minutes":    r.work_minutes,
                "synced_at":       r.synced_at.isoformat() if r.synced_at else None,
            }
            for r in recs
        ],
    }


# ── GET /calendar ────────────────────────────────────────────────────────────

@router.get("/calendar", summary="IHG 客房保養月曆格資料（樓層 × 日）")
def get_calendar(
    year:  str = Query(..., description="年度，如 2026"),
    month: str = Query(..., description="月份，如 05 或 5"),
    db: Session = Depends(get_db),
):
    """
    回傳指定月份各樓層每日保養狀態，供月曆格 TAB 使用。

    Response 格式：
    {
      "year": "2026",
      "month": "05",
      "max_day": 31,
      "floors": ["5F","6F","7F","8F","9F","10F"],
      "kpi": {
        "total_rooms": int, "completed": int, "abnormal": int,
        "pending": int, "completion_rate": float
      },
      "calendar": {
        "5F": {
          "15": { "total": 4, "completed": 3, "abnormal": 0, "pending": 1, "ragic_ids": [...] }
        },
        ...
        "TOTAL": { ... }
      }
    }
    """
    month_zf = month.zfill(2)
    try:
        max_day = _calendar.monthrange(int(year), int(month_zf))[1]
    except Exception:
        max_day = 31

    FLOORS = ["5F", "6F", "7F", "8F", "9F", "10F"]

    # 取當月所有規範房間記錄
    all_recs = db.query(IHGRoomMaintenanceMaster).filter(
        IHGRoomMaintenanceMaster.maint_year  == year,
        IHGRoomMaintenanceMaster.maint_month == month_zf,
        IHGRoomMaintenanceMaster.room_no.in_(CANONICAL_ROOM_SET),
    ).all()

    def _empty_cell() -> dict:
        return {"total": 0, "completed": 0, "abnormal": 0, "pending": 0, "ragic_ids": []}

    # 初始化月曆格結構：floor → day_str → cell_dict
    cal: dict[str, dict[str, dict]] = {f: {} for f in FLOORS}
    cal["TOTAL"] = {}

    # 以 room_no 去重統計 KPI（後者覆蓋前者）
    room_status_map: dict[str, str] = {}

    for rec in all_recs:
        # ── 解析日 ─────────────────────────────────────────────────────────
        day_str = ""
        if rec.maint_date:
            parts = rec.maint_date.replace("-", "/").split("/")
            if len(parts) >= 3:
                try:
                    d = int(parts[2].split()[0])   # 處理 "10 00:00:00" 含時間的格式
                    if 1 <= d <= 31:
                        day_str = str(d)
                except (ValueError, IndexError):
                    pass

        floor = rec.floor or _derive_floor(rec.room_no or "")
        rec_status = _calc_record_status(rec.raw_json)

        # KPI 去重（同一房號以最後一筆覆蓋）
        room_status_map[rec.room_no] = rec_status

        if not day_str or floor not in FLOORS:
            continue

        # ── 更新樓層格 ──────────────────────────────────────────────────────
        if day_str not in cal[floor]:
            cal[floor][day_str] = _empty_cell()
        cal[floor][day_str]["total"] += 1
        cal[floor][day_str][rec_status] += 1
        cal[floor][day_str]["ragic_ids"].append(rec.ragic_id)

        # ── 更新 TOTAL 格 ───────────────────────────────────────────────────
        if day_str not in cal["TOTAL"]:
            cal["TOTAL"][day_str] = _empty_cell()
        cal["TOTAL"][day_str]["total"] += 1
        cal["TOTAL"][day_str][rec_status] += 1
        cal["TOTAL"][day_str]["ragic_ids"].append(rec.ragic_id)

    # ── KPI 統計 ─────────────────────────────────────────────────────────────
    completed_cnt = sum(1 for s in room_status_map.values() if s == "completed")
    abnormal_cnt  = sum(1 for s in room_status_map.values() if s == "abnormal")
    pending_cnt   = sum(1 for s in room_status_map.values() if s == "pending")
    total_cnt     = len(room_status_map)
    rate          = round(completed_cnt / total_cnt * 100, 1) if total_cnt else 0.0

    return {
        "year":    year,
        "month":   month_zf,
        "max_day": max_day,
        "floors":  FLOORS,
        "kpi": {
            "total_rooms":     total_cnt,
            "completed":       completed_cnt,
            "abnormal":        abnormal_cnt,
            "pending":         pending_cnt,
            "completion_rate": rate,
        },
        "calendar": cal,
    }


# ── GET /{ragic_id} ───────────────────────────────────────────────────────────

@router.get("/{ragic_id}", summary="IHG 客房保養單筆明細（含子表格）")
def get_record(ragic_id: str, db: Session = Depends(get_db)):
    """回傳主表單筆 + 所有子表格明細"""
    master = db.get(IHGRoomMaintenanceMaster, ragic_id)
    if not master:
        raise HTTPException(status_code=404, detail=f"找不到記錄 ragic_id={ragic_id}")

    details = (
        db.query(IHGRoomMaintenanceDetail)
        .filter(IHGRoomMaintenanceDetail.master_ragic_id == ragic_id)
        .order_by(IHGRoomMaintenanceDetail.seq_no)
        .all()
    )

    try:
        raw_fields = json.loads(master.raw_json or "{}")
    except Exception:
        raw_fields = {}

    # ── detail dict（中文 key，供 Drawer 第二區段顯示）─────────────────────────
    wm_str = f"{master.work_minutes} 分鐘" if master.work_minutes is not None else (
        str(raw_fields.get("工時計算", "")) or ""
    )
    start_str = (master.start_time or "").strip() or str(raw_fields.get("保養時間起", "")).strip()
    end_str   = (master.end_time   or "").strip() or str(raw_fields.get("保養時間迄", "")).strip()

    detail: dict[str, str] = {
        "房號":     master.room_no or "",
        "樓層":     master.floor or "",
        "保養年份": master.maint_year or "",
        "保養月份": master.maint_month or "",
        "保養類型": master.maint_type or "",
        "保養日期": master.maint_date or "",
        "完成日期": master.completion_date or "",
        "狀態":     "已完成" if master.is_completed else "未完成",
        "保養人員": master.assignee_name or "",
        "複核人員": master.checker_name or "",
        "保養時間起": start_str,
        "保養時間迄": end_str,
        "工時":     wm_str,
        "備註":     master.notes or "",
        "是否有陽台": str(raw_fields.get("是否有陽台", "")).strip(),
        "保養項目數": str(len(details)),
        "同步時間":  master.synced_at.strftime("%Y-%m-%d %H:%M") if master.synced_at else "",
    }
    # 去除空值，保留有意義的欄位
    detail = {k: v for k, v in detail.items() if v}

    return {
        "ragic_id":         master.ragic_id,
        "ragic_url":        _ragic_url(master.ragic_id),
        "room_no":          master.room_no,
        "floor":            master.floor,
        "maint_year":       master.maint_year,
        "maint_month":      master.maint_month,
        "maint_date":       master.maint_date,
        "status":           _cell_status(master),
        "is_completed":     master.is_completed,
        "assignee_name":    master.assignee_name,
        "checker_name":     master.checker_name,
        "completion_date":  master.completion_date,
        "maint_type":       master.maint_type,
        "notes":            master.notes,
        "start_time":       master.start_time,
        "end_time":         master.end_time,
        "work_minutes":     master.work_minutes,
        "ragic_created_at": master.ragic_created_at,
        "ragic_updated_at": master.ragic_updated_at,
        "synced_at":        master.synced_at.isoformat() if master.synced_at else None,
        "raw_fields":       raw_fields,
        "detail":           detail,
        "details": [
            {
                "ragic_id":  d.ragic_id,
                "seq_no":    d.seq_no,
                "task_name": d.task_name,
                "result":    d.result,
                "is_ok":     d.is_ok,
                "notes":     d.notes,
            }
            for d in details
        ],
    }
