"""
日曜核准請款單月報表 API Router

端點：
  GET  /api/v1/nichiyo-claim-report/approved/orders           — 請款單清單（訂單級，分頁）
  GET  /api/v1/nichiyo-claim-report/approved/orders/{id}      — 單筆詳情（含品項）
  GET  /api/v1/nichiyo-claim-report/approved/monthly          — 月報明細（品項級，分頁）
  GET  /api/v1/nichiyo-claim-report/approved/summary          — KPI 統計
  GET  /api/v1/nichiyo-claim-report/approved/departments      — 部門彙總表
  GET  /api/v1/nichiyo-claim-report/approved/export           — Excel 匯出（RFC 5987，B08）
  GET  /api/v1/nichiyo-claim-report/approved/available-months — 有資料的年月清單
  GET  /api/v1/nichiyo-claim-report/config/departments        — 部門下拉
  GET  /api/v1/nichiyo-claim-report/config/account-categories — 會科下拉
  POST /api/v1/nichiyo-claim-report/sync                      — 手動觸發同步
  GET  /api/v1/nichiyo-claim-report/sync/status               — 同步狀態
  POST /api/v1/nichiyo-claim-report/admin/reparse-claim-no    — 就地修正 null 單號

權限：nichiyo_claim.view / nichiyo_claim.export / nichiyo_claim.admin
"""
import asyncio
import io
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, Query, BackgroundTasks, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.database import get_db, SessionLocal
from app.core.time import twnow
from app.dependencies import require_permission
from app.models.module_sync_log import ModuleSyncLog
from app.models.nichiyo_claim_request import (
    NichiyoClaimRequest,
    NichiyoClaimRequestItem,
    NICHIYO_CLAIM_DEPT_SHEETS,
)

router = APIRouter()

_PERM        = "nichiyo_claim.view"
_PERM_EXPORT = "nichiyo_claim.export"
_PERM_ADMIN  = "nichiyo_claim.admin"
_COMPANY     = "日曜"


# ── 工具函式 ──────────────────────────────────────────────────────────────────

def _build_date_filter(q, year_month=None, year_month_from=None, year_month_to=None):
    """日期三模式篩選（status=F + approved_date 符合條件）"""
    q = q.filter(NichiyoClaimRequest.status == "F")
    ym_col = func.strftime("%Y-%m", NichiyoClaimRequest.approved_date)
    if year_month_from and year_month_to:
        q = q.filter(ym_col >= year_month_from, ym_col <= year_month_to)
    elif year_month_from:
        q = q.filter(ym_col >= year_month_from)
    elif year_month:
        q = q.filter(ym_col == year_month)
    return q


def _date_label(year_month=None, year_month_from=None, year_month_to=None) -> str:
    if year_month_from and year_month_to:
        if (year_month_from.endswith("-01") and year_month_to.endswith("-12")
                and year_month_from[:4] == year_month_to[:4]):
            return f"{year_month_from[:4]}年度"
        return f"{year_month_from}~{year_month_to}"
    return year_month or ""


def _apply_dept_filter(q, department: Optional[str]):
    if department:
        return q.filter(NichiyoClaimRequest.department_display == department)
    return q


def _apply_account_filter(q, account_category: Optional[str]):
    if account_category:
        return q.filter(NichiyoClaimRequest.account_category == account_category)
    return q


def _apply_search(q, keyword: Optional[str]):
    """B09：全文搜尋（事由/單號/申請人/受款者）"""
    if keyword:
        like = f"%{keyword}%"
        return q.filter(
            NichiyoClaimRequest.purpose_description.ilike(like)
            | NichiyoClaimRequest.claim_no.ilike(like)
            | NichiyoClaimRequest.applicant.ilike(like)
            | NichiyoClaimRequest.payee.ilike(like)
        )
    return q


def _ragic_url(order: NichiyoClaimRequest) -> str:
    """B11：使用 list_path（記錄所在 Sheet），不是 detail_path"""
    for d in NICHIYO_CLAIM_DEPT_SHEETS:
        if d["list_path"] == order.ragic_sheet_path:
            return f"https://ap12.ragic.com/soutlet001/{d['list_path']}/{order.ragic_record_id}"
    return ""


def _order_to_dict(order: NichiyoClaimRequest) -> dict:
    return {
        "id":                  order.id,
        "claim_no":            order.claim_no,
        "department_display":  order.department_display,
        "account_category":    order.account_category,
        "applicant":           order.applicant,
        "purpose_description": order.purpose_description,
        "payment_type":        order.payment_type,
        "subtotal":            order.subtotal,
        "tax":                 order.tax,
        "total":               order.total,
        "payable_amount":      order.payable_amount,
        "payee":               order.payee,
        "status":              order.status,
        "request_date":        order.request_date.isoformat()    if order.request_date    else None,
        "approved_date":       order.approved_date.isoformat()   if order.approved_date   else None,
        "payment_date":        order.payment_date.isoformat()    if order.payment_date    else None,
        "last_updated_at":     order.last_updated_at.strftime("%Y/%m/%d %H:%M") if order.last_updated_at else None,
        "detail_synced":       order.detail_synced,
        "ragic_sheet_path":    order.ragic_sheet_path,
        "ragic_record_id":     order.ragic_record_id,
        "ragic_url":           _ragic_url(order),
    }


def _format_item_row(order: NichiyoClaimRequest, item: NichiyoClaimRequestItem) -> dict:
    return {
        "order_id":            order.id,
        "ragic_id":            order.ragic_record_id,
        "company":             order.company,
        "department_display":  order.department_display,
        "claim_no":            order.claim_no,
        "request_date":        order.request_date.isoformat()  if order.request_date  else None,
        "approved_date":       order.approved_date.isoformat() if order.approved_date else None,
        "applicant":           order.applicant,
        "purpose_description": order.purpose_description,
        "account_category":    order.account_category,
        "subtotal":            order.subtotal,
        "tax":                 order.tax,
        "total":               order.total,
        "payable_amount":      order.payable_amount,
        "payee":               order.payee,
        "payment_type":        order.payment_type,
        "status":              order.status,
        "ragic_url":           _ragic_url(order),
        "item_id":             item.id,
        "seq":                 item.seq,
        "product_name":        item.product_name,
        "qty":                 item.qty,
        "unit":                item.unit,
        "unit_price":          item.unit_price,
        "amount":              item.amount,
        "item_remark":         item.item_remark,
    }


# ── GET /approved/orders ──────────────────────────────────────────────────────

@router.get("/approved/orders")
def get_approved_orders(
    year_month:       Optional[str] = Query(None, description="YYYY-MM（單月）"),
    year_month_from:  Optional[str] = Query(None, description="YYYY-MM（區間起）"),
    year_month_to:    Optional[str] = Query(None, description="YYYY-MM（區間迄）"),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    q:                Optional[str] = Query(None, description="關鍵字（事由/單號/申請人/受款者）"),
    page:             int           = Query(1, ge=1),
    per_page:         int           = Query(20, ge=1, le=200),  # B10：統一用 per_page
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """請款單清單（訂單級，分頁）"""
    base_q = db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.company == _COMPANY
    )
    base_q = _build_date_filter(base_q, year_month, year_month_from, year_month_to)
    base_q = _apply_dept_filter(base_q, department)
    base_q = _apply_account_filter(base_q, account_category)
    base_q = _apply_search(base_q, q)
    base_q = base_q.order_by(NichiyoClaimRequest.approved_date.desc(),
                              NichiyoClaimRequest.id.desc())

    total = base_q.count()
    rows  = base_q.offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_order_to_dict(o) for o in rows],
    }


# ── GET /approved/orders/{id} ─────────────────────────────────────────────────

@router.get("/approved/orders/{order_id}")
def get_order_detail(
    order_id: int,
    db:       Session = Depends(get_db),
    _:        object  = Depends(require_permission(_PERM)),
):
    """單筆請款單詳情（含品項列表）— 回傳 {order, items} 結構"""
    order = db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.id == order_id,
        NichiyoClaimRequest.company == _COMPANY,
    ).first()
    if not order:
        raise HTTPException(status_code=404, detail="找不到此請款單")

    items = (
        db.query(NichiyoClaimRequestItem)
        .filter(NichiyoClaimRequestItem.order_id == order_id)
        .order_by(NichiyoClaimRequestItem.seq)
        .all()
    )

    def item_dict(i: NichiyoClaimRequestItem) -> dict:
        return {
            "id":           i.id,
            "seq":          i.seq,
            "product_name": i.product_name,
            "qty":          i.qty,
            "unit":         i.unit,
            "unit_price":   i.unit_price,
            "amount":       i.amount,
            "item_remark":  i.item_remark,
        }

    # ⚠️ 必須回傳 {order, items} 結構，不可 flat dict（Drawer 會空白）
    return {
        "order": _order_to_dict(order),
        "items": [item_dict(i) for i in items],
    }


# ── GET /approved/monthly ─────────────────────────────────────────────────────

@router.get("/approved/monthly")
def get_monthly_report(
    year_month:       Optional[str] = Query(None),
    year_month_from:  Optional[str] = Query(None),
    year_month_to:    Optional[str] = Query(None),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    q:                Optional[str] = Query(None),
    page:             int           = Query(1, ge=1),
    per_page:         int           = Query(50, ge=1, le=200),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """月報明細（品項級，分頁）"""
    order_q = db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.company == _COMPANY
    )
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_category)
    order_q = _apply_search(order_q, q)

    order_ids = [r.id for r in order_q.with_entities(NichiyoClaimRequest.id).all()]
    if not order_ids:
        return {"total": 0, "page": page, "per_page": per_page, "items": []}

    item_q = (
        db.query(NichiyoClaimRequestItem, NichiyoClaimRequest)
        .join(NichiyoClaimRequest,
              NichiyoClaimRequestItem.order_id == NichiyoClaimRequest.id)
        .filter(NichiyoClaimRequestItem.order_id.in_(order_ids))
        .order_by(NichiyoClaimRequest.approved_date.desc(),
                  NichiyoClaimRequest.claim_no,
                  NichiyoClaimRequestItem.seq)
    )

    total = item_q.count()
    rows  = item_q.offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_format_item_row(order, item) for item, order in rows],
    }


# ── GET /approved/summary ─────────────────────────────────────────────────────

@router.get("/approved/summary")
def get_summary(
    year_month:       Optional[str] = Query(None),
    year_month_from:  Optional[str] = Query(None),
    year_month_to:    Optional[str] = Query(None),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    label   = _date_label(year_month, year_month_from, year_month_to)
    order_q = db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.company == _COMPANY
    )
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_category)
    orders  = order_q.all()

    if not orders:
        return {
            "year_month": label, "company": _COMPANY, "department": department,
            "order_count": 0, "total_payable": 0, "total_tax": 0,
            "item_count": 0, "dept_count": 0, "avg_payable": 0,
            "top_order": None, "top_dept_by_count": "", "top_dept_by_amount": "",
            "dept_summary": [],
        }

    order_ids  = [o.id for o in orders]
    item_count = (
        db.query(func.count(NichiyoClaimRequestItem.id))
        .filter(NichiyoClaimRequestItem.order_id.in_(order_ids))
        .scalar() or 0
    )
    total_payable = sum(o.payable_amount or 0 for o in orders)
    total_tax     = sum(o.tax or 0 for o in orders)
    order_count   = len(orders)

    top_order = max(orders, key=lambda o: o.payable_amount or 0)

    dept_count:  dict = {}
    dept_amount: dict = {}
    for o in orders:
        dept_count[o.department_display]  = dept_count.get(o.department_display, 0)  + 1
        dept_amount[o.department_display] = dept_amount.get(o.department_display, 0) + (o.payable_amount or 0)

    # B02：sorted() 加 key=lambda x: x or "" 防 None crash
    top_dept_by_count  = max(dept_count,  key=lambda d: dept_count[d],  default="")
    top_dept_by_amount = max(dept_amount, key=lambda d: dept_amount[d], default="")

    dept_ratio = [
        {
            "department_display": dept,
            "order_count":        cnt,
            "total_payable":      dept_amount[dept],
            "amount_ratio":       round(dept_amount[dept] / total_payable * 100, 1) if total_payable else 0,
        }
        for dept, cnt in sorted(
            dept_count.items(),
            key=lambda x: -(x[1]),
        )
    ]

    return {
        "year_month":         label,
        "company":            _COMPANY,
        "department":         department,
        "order_count":        order_count,
        "total_payable":      total_payable,
        "total_tax":          total_tax,
        "item_count":         item_count,
        "dept_count":         len(dept_count),
        "avg_payable":        round(total_payable / order_count) if order_count else 0,
        "top_order": {
            "claim_no":    top_order.claim_no,
            "department":  top_order.department_display,
            "payable":     top_order.payable_amount,
            "description": (top_order.purpose_description or "")[:50],
        },
        "top_dept_by_count":  top_dept_by_count,
        "top_dept_by_amount": top_dept_by_amount,
        "dept_summary":       dept_ratio,
    }


# ── GET /approved/departments ─────────────────────────────────────────────────

@router.get("/approved/departments")
def get_departments(
    year_month:       Optional[str] = Query(None),
    year_month_from:  Optional[str] = Query(None),
    year_month_to:    Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM)),
):
    """部門彙總表（已核准）"""
    base_q = db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.company == _COMPANY
    )
    base_q = _build_date_filter(base_q, year_month, year_month_from, year_month_to)
    base_q = _apply_account_filter(base_q, account_category)

    rows = (
        base_q.with_entities(
            NichiyoClaimRequest.department_display,
            func.count(NichiyoClaimRequest.id).label("order_count"),
            func.sum(NichiyoClaimRequest.payable_amount).label("total_payable"),
            func.sum(NichiyoClaimRequest.tax).label("total_tax"),
        )
        .group_by(NichiyoClaimRequest.department_display)
        .order_by(func.sum(NichiyoClaimRequest.payable_amount).desc())
        .all()
    )
    return [
        {
            "department_display": r.department_display,
            "order_count":        r.order_count,
            "total_payable":      r.total_payable or 0,
            "total_tax":          r.total_tax     or 0,
        }
        for r in rows
    ]


# ── GET /approved/export ──────────────────────────────────────────────────────

@router.get("/approved/export")
def export_excel(
    year_month:       Optional[str] = Query(None),
    year_month_from:  Optional[str] = Query(None),
    year_month_to:    Optional[str] = Query(None),
    department:       Optional[str] = Query(None),
    account_category: Optional[str] = Query(None),
    q:                Optional[str] = Query(None),
    db:               Session       = Depends(get_db),
    _:                object        = Depends(require_permission(_PERM_EXPORT)),
):
    """Excel 匯出（品項級別）B08：RFC 5987 filename* 編碼"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 套件未安裝")

    label   = _date_label(year_month, year_month_from, year_month_to)
    order_q = db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.company == _COMPANY
    )
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_category)
    order_q = _apply_search(order_q, q)

    order_ids = [r.id for r in order_q.with_entities(NichiyoClaimRequest.id).all()]
    if not order_ids:
        raise HTTPException(status_code=404, detail="所選條件無資料")

    rows = (
        db.query(NichiyoClaimRequestItem, NichiyoClaimRequest)
        .join(NichiyoClaimRequest,
              NichiyoClaimRequestItem.order_id == NichiyoClaimRequest.id)
        .filter(NichiyoClaimRequestItem.order_id.in_(order_ids))
        .order_by(NichiyoClaimRequest.approved_date.desc(),
                  NichiyoClaimRequest.claim_no,
                  NichiyoClaimRequestItem.seq)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"日曜請款單_{label}"

    HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
    THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    headers = [
        "公司", "部門", "請款單號", "申請日期", "核准日期",
        "申請人", "事由", "會科", "付款種類", "受款者",
        "小計（未稅）", "稅額", "應付款", "付款日期",
        "項次", "品名", "數量", "單位", "單價", "品項金額",
        "Ragic連結",
    ]
    for col_idx, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = THIN_BORDER
    ws.row_dimensions[1].height = 30

    for row_idx, (item, order) in enumerate(rows, 2):
        data = [
            order.company,
            order.department_display,
            order.claim_no,
            order.request_date.isoformat()  if order.request_date  else "",
            order.approved_date.isoformat() if order.approved_date else "",
            order.applicant or "",
            (order.purpose_description or "")[:200],
            order.account_category or "",
            order.payment_type or "",
            order.payee or "",
            order.subtotal or 0,
            order.tax or 0,
            order.payable_amount or 0,
            order.payment_date.isoformat()  if order.payment_date  else "",
            item.seq,
            item.product_name or "",
            item.qty or "",
            item.unit or "",
            item.unit_price or "",
            item.amount or "",
            _ragic_url(order),
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = PatternFill("solid", fgColor="EBF5FF" if row_idx % 2 == 0 else "FFFFFF")
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if isinstance(val, int) and col_idx in (11, 12, 13, 19, 20):
                cell.number_format = '#,##0'

    col_widths = [8, 10, 20, 12, 12, 10, 35, 18, 10, 18, 14, 10, 14, 12, 6, 28, 8, 8, 10, 12, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # B08：RFC 5987 中文檔名編碼
    label_safe    = label.replace("~", "_")
    filename_cn   = f"日曜核准請款單月報表_{label}.xlsx"
    filename_safe = f"nichiyo_claim_report_{label_safe}.xlsx"
    encoded       = quote(filename_cn, safe="")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_safe}"; '
                f"filename*=UTF-8''{encoded}"
            )
        },
    )


# ── GET /approved/available-months ───────────────────────────────────────────

@router.get("/approved/available-months")
def get_available_months(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(
            func.strftime("%Y-%m", NichiyoClaimRequest.approved_date).label("ym")
        )
        .filter(
            NichiyoClaimRequest.status == "F",
            NichiyoClaimRequest.company == _COMPANY,
            NichiyoClaimRequest.approved_date.isnot(None),
        )
        .distinct()
        .order_by(func.strftime("%Y-%m", NichiyoClaimRequest.approved_date).desc())
        .all()
    )
    return [r.ym for r in rows if r.ym]


# ── GET /config/* ─────────────────────────────────────────────────────────────

@router.get("/config/departments")
def get_available_departments(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(NichiyoClaimRequest.department_display)
        .filter(NichiyoClaimRequest.company == _COMPANY)
        .distinct()
        .order_by(NichiyoClaimRequest.department_display)
        .all()
    )
    # B02：sorted 防 None
    return sorted([r[0] for r in rows if r[0]], key=lambda x: x or "")


@router.get("/config/account-categories")
def get_account_categories(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(NichiyoClaimRequest.account_category)
        .filter(
            NichiyoClaimRequest.company == _COMPANY,
            NichiyoClaimRequest.account_category.isnot(None),
            NichiyoClaimRequest.account_category != "",
        )
        .distinct()
        .order_by(NichiyoClaimRequest.account_category)
        .all()
    )
    return sorted([r[0] for r in rows if r[0]], key=lambda x: x or "")


# ── POST /sync ────────────────────────────────────────────────────────────────

_sync_lock = asyncio.Lock()


@router.post("/sync")
async def trigger_sync(
    background_tasks: BackgroundTasks,
    full_resync: bool = False,
    _: object = Depends(require_permission(_PERM_ADMIN)),
):
    from app.services.nichiyo_claim_request_sync import sync_all, sync_list_only

    if _sync_lock.locked():
        return {"status": "already_running", "message": "同步已在執行中，請稍後再試"}

    async def _run():
        async with _sync_lock:
            if full_resync:
                await sync_all()
            else:
                await sync_list_only()

    background_tasks.add_task(_run)
    return {"status": "started", "message": f"日曜請款同步已啟動（full_resync={full_resync}）"}


# ── GET /sync/status ──────────────────────────────────────────────────────────

@router.get("/sync/status")
def get_sync_status(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM_ADMIN)),
):
    recent_logs = (
        db.query(ModuleSyncLog)
        .filter(ModuleSyncLog.module_name == "日曜核准請款單")
        .order_by(ModuleSyncLog.started_at.desc())
        .limit(5)
        .all()
    )
    dept_rows = (
        db.query(
            NichiyoClaimRequest.department_display,
            func.count(NichiyoClaimRequest.id).label("total"),
        )
        .filter(NichiyoClaimRequest.company == _COMPANY)
        .group_by(NichiyoClaimRequest.department_display)
        .all()
    )
    synced_rows = (
        db.query(
            NichiyoClaimRequest.department_display,
            func.count(NichiyoClaimRequest.id).label("synced"),
        )
        .filter(
            NichiyoClaimRequest.company == _COMPANY,
            NichiyoClaimRequest.detail_synced == True,
        )
        .group_by(NichiyoClaimRequest.department_display)
        .all()
    )
    synced_map    = {r.department_display: r.synced for r in synced_rows}
    pending_count = (
        db.query(func.count(NichiyoClaimRequest.id))
        .filter(
            NichiyoClaimRequest.company == _COMPANY,
            NichiyoClaimRequest.detail_synced == False,
        )
        .scalar() or 0
    )

    return {
        "is_running":           _sync_lock.locked(),
        "pending_detail_count": pending_count,
        "recent_logs": [
            {
                "id":         log.id,
                "module":     log.module_name,
                "trigger":    log.triggered_by,
                "status":     log.status,
                "message":    log.error_msg or f"fetched={log.fetched} upserted={log.upserted}",
                "created_at": log.started_at.isoformat() if log.started_at else None,
            }
            for log in recent_logs
        ],
        "dept_stats": [
            {
                "department_display": r.department_display,
                "total":              r.total,
                "detail_synced":      synced_map.get(r.department_display, 0),
                "pending":            r.total - synced_map.get(r.department_display, 0),
            }
            for r in dept_rows
        ],
    }


# ── POST /admin/reparse-claim-no ─────────────────────────────────────────────

@router.post("/admin/reparse-claim-no")
def reparse_claim_no(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM_ADMIN)),
):
    """從 raw_data_json 就地修復 claim_no 欄位。回傳 still_null 供診斷。"""
    from app.services.nichiyo_claim_request_sync import _pick_claim_no, LIST_FIELD_CANDIDATES
    total = updated = 0
    still_null: list = []

    for rec in db.query(NichiyoClaimRequest).filter(
        NichiyoClaimRequest.raw_data_json.isnot(None)
    ).all():
        total += 1
        import json as _json
        raw = _json.loads(rec.raw_data_json or "{}")
        new_val = _pick_claim_no(raw, LIST_FIELD_CANDIDATES["claim_no"]) or None
        if new_val != rec.claim_no:
            rec.claim_no = new_val
            updated += 1
        if not new_val:
            still_null.append({
                "id":              rec.id,
                "ragic_id":        rec.ragic_record_id,
                "department":      rec.department_display,
                "no_related_keys": [k for k in raw if "請" in k or "號" in k],
                "all_keys_sample": list(raw.keys())[:30],
            })

    db.commit()
    return {"total": total, "updated": updated, "still_null": still_null[:20]}


# ── GET /audit/anomalies ──────────────────────────────────────────────────────

@router.get("/audit/anomalies")
def get_audit_anomalies(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None),
    year_month_to:   Optional[str] = Query(None),
    department:      Optional[str] = Query(None),
    rule_code:       Optional[str] = Query(None),
    page:            int           = Query(1, ge=1),
    per_page:        int           = Query(20, ge=1, le=200),
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    """日曜請款稽核：分頁異常列表。"""
    from app.services.audit_service import get_nichiyo_claim_anomalies
    return get_nichiyo_claim_anomalies(
        db,
        year_month=year_month,
        year_month_from=year_month_from,
        year_month_to=year_month_to,
        department=department,
        rule_code=rule_code,
        page=page,
        per_page=per_page,
    )


# ── GET /audit/summary ────────────────────────────────────────────────────────

@router.get("/audit/summary")
def get_audit_summary(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None),
    year_month_to:   Optional[str] = Query(None),
    department:      Optional[str] = Query(None),
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    """日曜請款稽核：各規則計數 KPI。"""
    from app.services.audit_service import get_nichiyo_claim_audit_summary
    return get_nichiyo_claim_audit_summary(
        db,
        year_month=year_month,
        year_month_from=year_month_from,
        year_month_to=year_month_to,
        department=department,
    )

# ── GET /combined/departments — 日曜請購+請款 雙色部門統計 ──────────────────────

@router.get("/combined/departments")
def get_combined_departments(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None),
    year_month_to:   Optional[str] = Query(None),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    """
    日曜請購 + 日曜請款 雙色部門統計
    欄位：department_display / purchase_count / purchase_amount / purchase_tax
          / claim_count / claim_payable / claim_tax
    """
    from app.models.nichiyo_purchase_request import NichiyoPurchaseRequest

    # 請購部門彙總
    pq = db.query(
        NichiyoPurchaseRequest.department_display,
        func.count(NichiyoPurchaseRequest.id).label("p_count"),
        func.sum(NichiyoPurchaseRequest.amount).label("p_amount"),
        func.sum(NichiyoPurchaseRequest.amount_tax).label("p_tax"),
    ).filter(NichiyoPurchaseRequest.company == _COMPANY,
             NichiyoPurchaseRequest.status == "F")
    ym_p = func.strftime("%Y-%m", NichiyoPurchaseRequest.approved_date)
    if year_month_from and year_month_to:
        pq = pq.filter(ym_p >= year_month_from, ym_p <= year_month_to)
    elif year_month_from:
        pq = pq.filter(ym_p >= year_month_from)
    elif year_month:
        pq = pq.filter(ym_p == year_month)
    p_dict = {r.department_display: r
              for r in pq.group_by(NichiyoPurchaseRequest.department_display).all()}

    # 請款部門彙總
    cq = db.query(
        NichiyoClaimRequest.department_display,
        func.count(NichiyoClaimRequest.id).label("c_count"),
        func.sum(NichiyoClaimRequest.payable_amount).label("c_payable"),
        func.sum(NichiyoClaimRequest.tax).label("c_tax"),
    ).filter(NichiyoClaimRequest.company == _COMPANY,
             NichiyoClaimRequest.status == "F")
    ym_c = func.strftime("%Y-%m", NichiyoClaimRequest.approved_date)
    if year_month_from and year_month_to:
        cq = cq.filter(ym_c >= year_month_from, ym_c <= year_month_to)
    elif year_month_from:
        cq = cq.filter(ym_c >= year_month_from)
    elif year_month:
        cq = cq.filter(ym_c == year_month)
    c_dict = {r.department_display: r
              for r in cq.group_by(NichiyoClaimRequest.department_display).all()}

    all_depts = sorted(set(list(p_dict) + list(c_dict)), key=lambda x: x or "")
    return [
        {
            "department_display": dept,
            "purchase_count":  p_dict[dept].p_count             if dept in p_dict else 0,
            "purchase_amount": int(p_dict[dept].p_amount or 0)  if dept in p_dict else 0,
            "purchase_tax":    int(p_dict[dept].p_tax    or 0)  if dept in p_dict else 0,
            "claim_count":     c_dict[dept].c_count             if dept in c_dict else 0,
            "claim_payable":   int(c_dict[dept].c_payable or 0) if dept in c_dict else 0,
            "claim_tax":       int(c_dict[dept].c_tax    or 0)  if dept in c_dict else 0,
        }
        for dept in all_depts
    ]
