"""
核准請款單月報表 API Router

端點：
  GET  /api/v1/claim-report/approved/orders           — 請款單清單（訂單級，分頁）
  GET  /api/v1/claim-report/approved/orders/{id}      — 單筆請款單詳情（含品項）
  GET  /api/v1/claim-report/approved/monthly          — 月報明細（品項級，分頁）
  GET  /api/v1/claim-report/approved/summary          — KPI 統計（主單級）
  GET  /api/v1/claim-report/approved/departments      — 部門彙總表
  GET  /api/v1/claim-report/approved/export           — Excel 匯出
  GET  /api/v1/claim-report/approved/available-months — 有核准資料的年月清單
  GET  /api/v1/claim-report/config/departments        — 部門清單（下拉）
  GET  /api/v1/claim-report/config/payment-types      — 付款種類清單（下拉）
  GET  /api/v1/claim-report/config/account-subjects   — 會科清單（下拉）
  POST /api/v1/claim-report/sync                      — 手動觸發同步
  GET  /api/v1/claim-report/sync/status               — 同步狀態查詢
"""
import io
import asyncio
from typing import Optional

from fastapi import APIRouter, Depends, Query, BackgroundTasks, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.database import get_db, SessionLocal
from app.core.time import twnow
from app.dependencies import require_permission
from app.models.module_sync_log import ModuleSyncLog
from app.models.claim_request import (
    ApprovedClaimRequest,
    ApprovedClaimRequestItem,
    CLAIM_DEPT_SHEETS,
    DEPT_REQUEST_NO_LABEL,
)

router = APIRouter()

_PERM = "system_admin_only"


# ── 工具函式 ──────────────────────────────────────────────────────────────────

def _build_date_filter(
    q,
    year_month: Optional[str] = None,
    year_month_from: Optional[str] = None,
    year_month_to: Optional[str] = None,
):
    """WHERE status=F AND approved_date in date range."""
    q = q.filter(ApprovedClaimRequest.status == "F")
    ym_col = func.strftime("%Y-%m", ApprovedClaimRequest.approved_date)
    if year_month_from and year_month_to:
        q = q.filter(ym_col >= year_month_from, ym_col <= year_month_to)
    elif year_month_from:
        q = q.filter(ym_col >= year_month_from)
    elif year_month:
        q = q.filter(ym_col == year_month)
    return q


def _date_label(
    year_month: Optional[str] = None,
    year_month_from: Optional[str] = None,
    year_month_to: Optional[str] = None,
) -> str:
    """產生人類可讀的日期標籤，用於檔名與標題。"""
    if year_month_from and year_month_to:
        if (
            year_month_from.endswith("-01")
            and year_month_to.endswith("-12")
            and year_month_from[:4] == year_month_to[:4]
        ):
            return f"{year_month_from[:4]}年度"
        return f"{year_month_from}~{year_month_to}"
    return year_month or ""


def _apply_dept_filter(q, department: Optional[str]):
    if department:
        return q.filter(ApprovedClaimRequest.department_display == department)
    return q


def _apply_account_filter(q, account_subject: Optional[str]):
    if account_subject:
        return q.filter(ApprovedClaimRequest.account_subject == account_subject)
    return q


def _apply_payment_type_filter(q, payment_type: Optional[str]):
    if payment_type:
        return q.filter(ApprovedClaimRequest.payment_type == payment_type)
    return q


def _apply_search(q, keyword: Optional[str]):
    if keyword:
        like = f"%{keyword}%"
        return q.filter(
            ApprovedClaimRequest.purpose_description.ilike(like)
            | ApprovedClaimRequest.request_no.ilike(like)
            | ApprovedClaimRequest.department_request_no.ilike(like)
            | ApprovedClaimRequest.applicant.ilike(like)
            | ApprovedClaimRequest.payee.ilike(like)
        )
    return q


def _ragic_url(order: ApprovedClaimRequest) -> str:
    for d in CLAIM_DEPT_SHEETS:
        if d["list_path"] == order.ragic_sheet_path:
            return f"https://ap12.ragic.com/soutlet001/{d['detail_path']}/{order.ragic_record_id}"
    return ""


def _format_item_row(order: ApprovedClaimRequest, item: ApprovedClaimRequestItem) -> dict:
    """合併主單 + 品項為月報表一列。"""
    return {
        # 主單欄位
        "claim_id":              order.id,
        "department_display":    order.department_display,
        "request_no":            order.request_no,
        "department_request_no": order.department_request_no,
        "account_subject":       order.account_subject,
        "apply_date":            order.apply_date.isoformat() if order.apply_date else None,
        "approved_date":         order.approved_date.isoformat() if order.approved_date else None,
        "applicant":             order.applicant,
        "payment_type":          order.payment_type,
        "purpose_description":   order.purpose_description,
        "subtotal":              order.subtotal,
        "tax":                   order.tax,
        "total":                 order.total,
        "payable_amount":        order.payable_amount,
        "payee":                 order.payee,
        "status":                order.status,
        "ragic_url":             _ragic_url(order),
        # 品項欄位
        "item_id":               item.id,
        "seq":                   item.seq,
        "item_name":             item.item_name,
        "quantity":              item.quantity,
        "unit":                  item.unit,
        "proposed_vendor_amount": item.proposed_vendor_amount,
        "invoice_no":            item.invoice_no,
        "receipt_no":            item.receipt_no,
        "item_note":             item.item_note,
    }


def _empty_summary(year_month: str, company: str, department: Optional[str]) -> dict:
    return {
        "order_count":       0,
        "total_subtotal":    0,
        "total_tax":         0,
        "total_payable":     0,
        "item_count":        0,
        "dept_count":        0,
        "avg_payable":       0,
        "top_order":         None,
        "top_dept_by_count": None,
        "top_dept_by_amount":None,
        "dept_summary":      [],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/orders — 請款單清單（訂單級，分頁）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/orders")
def get_approved_orders(
    year_month:      Optional[str] = Query(None, description="YYYY-MM"),
    year_month_from: Optional[str] = Query(None, description="YYYY-MM 區間起"),
    year_month_to:   Optional[str] = Query(None, description="YYYY-MM 區間迄"),
    company:         str           = Query("樂群"),
    department:      Optional[str] = Query(None),
    account_subject: Optional[str] = Query(None),
    payment_type:    Optional[str] = Query(None, description="零用金 / 匯款"),
    status:          Optional[str] = Query(None),
    keyword:         Optional[str] = Query(None, description="全文搜尋"),
    page:            int           = Query(1, ge=1),
    per_page:        int           = Query(20, ge=1, le=200),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    q = db.query(ApprovedClaimRequest).filter(ApprovedClaimRequest.company == company)

    if year_month or year_month_from or year_month_to:
        ym_col = func.strftime("%Y-%m", ApprovedClaimRequest.approved_date)
        if year_month_from and year_month_to:
            q = q.filter(ym_col >= year_month_from, ym_col <= year_month_to)
        elif year_month_from:
            q = q.filter(ym_col >= year_month_from)
        elif year_month:
            q = q.filter(ym_col == year_month)
    if department:
        q = _apply_dept_filter(q, department)
    if account_subject:
        q = _apply_account_filter(q, account_subject)
    if payment_type:
        q = _apply_payment_type_filter(q, payment_type)
    if status:
        q = q.filter(ApprovedClaimRequest.status == status)
    if keyword:
        q = _apply_search(q, keyword)

    total = q.count()
    orders = (
        q.order_by(ApprovedClaimRequest.approved_date.desc(), ApprovedClaimRequest.id.desc())
        .offset((page - 1) * per_page)
        .limit(per_page)
        .all()
    )

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_order_to_dict(o) for o in orders],
    }


def _order_to_dict(o: ApprovedClaimRequest) -> dict:
    dept_no_label = DEPT_REQUEST_NO_LABEL.get(o.department_display, "請款編號")
    return {
        "id":                    o.id,
        "department_display":    o.department_display,
        "request_no":            o.request_no,
        "department_request_no": o.department_request_no,
        "dept_request_no_label": dept_no_label,
        "account_subject":       o.account_subject,
        "apply_date":            o.apply_date.isoformat() if o.apply_date else None,
        "approved_date":         o.approved_date.isoformat() if o.approved_date else None,
        "applicant":             o.applicant,
        "payment_type":          o.payment_type,
        "purpose_description":   o.purpose_description,
        "subtotal":              o.subtotal,
        "tax":                   o.tax,
        "total":                 o.total,
        "payable_amount":        o.payable_amount,
        "payee":                 o.payee,
        "bank_name":             o.bank_name,
        "bank_branch":           o.bank_branch,
        "bank_account":          o.bank_account,
        "payment_date":          o.payment_date.isoformat() if o.payment_date else None,
        "status":                o.status,
        "detail_synced":         o.detail_synced,
        "last_updated_at":       o.last_updated_at.isoformat() if o.last_updated_at else None,
        "ragic_sheet_path":      o.ragic_sheet_path,
        "ragic_record_id":       o.ragic_record_id,
        "ragic_url":             _ragic_url(o),
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/orders/{order_id} — 單筆請款單詳情（含品項）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/orders/{order_id}")
def get_order_detail(
    order_id: int,
    db:       Session = Depends(get_db),
    _:        object  = Depends(require_permission(_PERM)),
):
    order = db.query(ApprovedClaimRequest).filter_by(id=order_id).first()
    if not order:
        raise HTTPException(status_code=404, detail="請款單不存在")

    items = (
        db.query(ApprovedClaimRequestItem)
        .filter_by(claim_id=order_id)
        .order_by(ApprovedClaimRequestItem.seq)
        .all()
    )

    return {
        "order": _order_to_dict(order),
        "items": [
            {
                "id":                    i.id,
                "seq":                   i.seq,
                "item_name":             i.item_name,
                "quantity":              i.quantity,
                "unit":                  i.unit,
                "item_note":             i.item_note,
                "proposed_vendor_amount": i.proposed_vendor_amount,
                "invoice_no":            i.invoice_no,
                "receipt_no":            i.receipt_no,
            }
            for i in items
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/monthly — 月報明細（品項級，分頁）
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/monthly")
def get_monthly_report(
    year_month:      Optional[str] = Query(None, description="YYYY-MM"),
    year_month_from: Optional[str] = Query(None, description="YYYY-MM 區間起"),
    year_month_to:   Optional[str] = Query(None, description="YYYY-MM 區間迄"),
    company:         str           = Query("樂群"),
    department:      Optional[str] = Query(None),
    account_subject: Optional[str] = Query(None),
    payment_type:    Optional[str] = Query(None),
    q:               Optional[str] = Query(None, description="關鍵字搜尋"),
    page:            int           = Query(1, ge=1),
    per_page:        int           = Query(50, ge=1, le=200),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    order_q = db.query(ApprovedClaimRequest).filter(ApprovedClaimRequest.company == company)
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_subject)
    order_q = _apply_payment_type_filter(order_q, payment_type)
    order_q = _apply_search(order_q, q)

    order_ids = [r.id for r in order_q.with_entities(ApprovedClaimRequest.id).all()]
    if not order_ids:
        return {"total": 0, "page": page, "per_page": per_page, "items": []}

    item_q = (
        db.query(ApprovedClaimRequestItem, ApprovedClaimRequest)
        .join(ApprovedClaimRequest, ApprovedClaimRequestItem.claim_id == ApprovedClaimRequest.id)
        .filter(ApprovedClaimRequestItem.claim_id.in_(order_ids))
        .order_by(
            ApprovedClaimRequest.approved_date.desc(),
            ApprovedClaimRequest.request_no,
            ApprovedClaimRequestItem.seq,
        )
    )

    total = item_q.count()
    rows  = item_q.offset((page - 1) * per_page).limit(per_page).all()

    return {
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "items":    [_format_item_row(order, item) for item, order in rows],
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/summary — KPI 統計
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/summary")
def get_summary(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None, description="YYYY-MM 區間起"),
    year_month_to:   Optional[str] = Query(None, description="YYYY-MM 區間迄"),
    company:         str           = Query("樂群"),
    department:      Optional[str] = Query(None),
    account_subject: Optional[str] = Query(None),
    payment_type:    Optional[str] = Query(None),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    q = db.query(ApprovedClaimRequest).filter(ApprovedClaimRequest.company == company)
    q = _build_date_filter(q, year_month, year_month_from, year_month_to)
    q = _apply_dept_filter(q, department)
    q = _apply_account_filter(q, account_subject)
    q = _apply_payment_type_filter(q, payment_type)

    orders = q.all()
    if not orders:
        return _empty_summary(year_month or "", company, department)

    order_ids  = [o.id for o in orders]
    item_count = (
        db.query(func.count(ApprovedClaimRequestItem.id))
        .filter(ApprovedClaimRequestItem.claim_id.in_(order_ids))
        .scalar() or 0
    )

    total_subtotal = sum(o.subtotal or 0 for o in orders)
    total_tax      = sum(o.tax or 0 for o in orders)
    total_payable  = sum(o.payable_amount or 0 for o in orders)
    order_count    = len(orders)

    top_order     = max(orders, key=lambda o: o.payable_amount or 0)
    dept_count:  dict = {}
    dept_amount: dict = {}
    for o in orders:
        dept_count[o.department_display]  = dept_count.get(o.department_display, 0) + 1
        dept_amount[o.department_display] = dept_amount.get(o.department_display, 0) + (o.payable_amount or 0)

    top_dept_by_count  = max(dept_count,  key=lambda d: dept_count[d],  default="")
    top_dept_by_amount = max(dept_amount, key=lambda d: dept_amount[d], default="")

    dept_summary = [
        {
            "department_display": dept,
            "order_count":        dept_count[dept],
            "total_payable":      dept_amount[dept],
        }
        for dept in sorted(dept_count.keys())
    ]

    return {
        "label":              _date_label(year_month, year_month_from, year_month_to),
        "order_count":        order_count,
        "total_subtotal":     total_subtotal,
        "total_tax":          total_tax,
        "total_payable":      total_payable,
        "item_count":         item_count,
        "dept_count":         len(dept_count),
        "avg_payable":        int(total_payable / order_count) if order_count else 0,
        "top_order":          top_order.request_no if top_order else None,
        "top_dept_by_count":  top_dept_by_count,
        "top_dept_by_amount": top_dept_by_amount,
        "dept_summary":       dept_summary,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/departments — 部門彙總表
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/departments")
def get_departments(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None, description="YYYY-MM 區間起"),
    year_month_to:   Optional[str] = Query(None, description="YYYY-MM 區間迄"),
    company:         str           = Query("樂群"),
    account_subject: Optional[str] = Query(None),
    payment_type:    Optional[str] = Query(None),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    q = (
        db.query(
            ApprovedClaimRequest.department_display,
            func.count(ApprovedClaimRequest.id).label("order_count"),
            func.sum(ApprovedClaimRequest.subtotal).label("total_subtotal"),
            func.sum(ApprovedClaimRequest.tax).label("total_tax"),
            func.sum(ApprovedClaimRequest.payable_amount).label("total_payable"),
        )
        .filter(ApprovedClaimRequest.company == company)
    )
    q = _build_date_filter(q, year_month, year_month_from, year_month_to)
    q = _apply_account_filter(q, account_subject)
    q = _apply_payment_type_filter(q, payment_type)
    rows = q.group_by(ApprovedClaimRequest.department_display).all()

    return [
        {
            "department_display": r.department_display,
            "order_count":        r.order_count,
            "total_subtotal":     int(r.total_subtotal or 0),
            "total_tax":          int(r.total_tax or 0),
            "total_payable":      int(r.total_payable or 0),
        }
        for r in rows
    ]


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/available-months — 有核准資料的年月清單
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/available-months")
def get_available_months(
    company: str     = Query("樂群"),
    db:      Session = Depends(get_db),
    _:       object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(func.strftime("%Y-%m", ApprovedClaimRequest.approved_date).label("ym"))
        .filter(
            ApprovedClaimRequest.company == company,
            ApprovedClaimRequest.status  == "F",
            ApprovedClaimRequest.approved_date.isnot(None),
        )
        .group_by("ym")
        .order_by(func.strftime("%Y-%m", ApprovedClaimRequest.approved_date).desc())
        .all()
    )
    return [r.ym for r in rows if r.ym]


# ─────────────────────────────────────────────────────────────────────────────
# GET /approved/export — Excel 匯出
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/approved/export")
def export_excel(
    year_month:      Optional[str] = Query(None),
    year_month_from: Optional[str] = Query(None, description="YYYY-MM 區間起"),
    year_month_to:   Optional[str] = Query(None, description="YYYY-MM 區間迄"),
    company:         str           = Query("樂群"),
    department:      Optional[str] = Query(None),
    account_subject: Optional[str] = Query(None),
    payment_type:    Optional[str] = Query(None),
    db:              Session       = Depends(get_db),
    _:               object        = Depends(require_permission(_PERM)),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl 未安裝")

    label = _date_label(year_month, year_month_from, year_month_to)

    order_q = db.query(ApprovedClaimRequest).filter(ApprovedClaimRequest.company == company)
    order_q = _build_date_filter(order_q, year_month, year_month_from, year_month_to)
    order_q = _apply_dept_filter(order_q, department)
    order_q = _apply_account_filter(order_q, account_subject)
    order_q = _apply_payment_type_filter(order_q, payment_type)

    order_ids = [r.id for r in order_q.with_entities(ApprovedClaimRequest.id).all()]
    if not order_ids:
        raise HTTPException(status_code=404, detail="此條件無資料")

    rows = (
        db.query(ApprovedClaimRequestItem, ApprovedClaimRequest)
        .join(ApprovedClaimRequest, ApprovedClaimRequestItem.claim_id == ApprovedClaimRequest.id)
        .filter(ApprovedClaimRequestItem.claim_id.in_(order_ids))
        .order_by(
            ApprovedClaimRequest.approved_date.desc(),
            ApprovedClaimRequest.request_no,
            ApprovedClaimRequestItem.seq,
        )
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"請款單月報_{label}"[:31]

    headers = [
        "核准日期", "部門", "請款單號", "部門請款編號", "會科",
        "申請日期", "申請人", "付款種類", "事由/說明",
        "項次", "品項名稱", "數量", "單位", "擬定廠商金額",
        "發票號碼", "憑證號碼", "品項備註",
        "小計(未稅)", "營業稅", "應付款", "受款者",
    ]
    header_fill  = PatternFill("solid", start_color="E07B39")
    header_font  = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for ci, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=ci, value=h)
        cell.fill       = header_fill
        cell.font       = header_font
        cell.alignment  = header_align

    for ri, (item, order) in enumerate(rows, start=2):
        vals = [
            order.approved_date.isoformat() if order.approved_date else "",
            order.department_display,
            order.request_no,
            order.department_request_no or "",
            order.account_subject or "",
            order.apply_date.isoformat() if order.apply_date else "",
            order.applicant or "",
            order.payment_type or "",
            order.purpose_description or "",
            item.seq,
            item.item_name or "",
            item.quantity or "",
            item.unit or "",
            item.proposed_vendor_amount,
            item.invoice_no or "",
            item.receipt_no or "",
            item.item_note or "",
            order.subtotal,
            order.tax,
            order.payable_amount,
            order.payee or "",
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    col_widths = [12, 8, 20, 14, 16, 12, 10, 8, 35, 6, 30, 6, 6, 14, 14, 14, 25, 12, 10, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename_cn   = f"核准請款單月報表_{label}.xlsx"
    filename_safe = f"claim_report_{label}.xlsx"
    encoded       = quote(filename_cn, safe="")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename_safe}"; '
                f"filename*=UTF-8''{encoded}"
            )
        },
    )


# ─────────────────────────────────────────────────────────────────────────────
# Config endpoints
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/config/departments")
def get_departments_list(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(ApprovedClaimRequest.department_display)
        .filter(ApprovedClaimRequest.department_display != "")
        .distinct()
        .order_by(ApprovedClaimRequest.department_display)
        .all()
    )
    return [r.department_display for r in rows]


@router.get("/config/payment-types")
def get_payment_types(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(ApprovedClaimRequest.payment_type)
        .filter(ApprovedClaimRequest.payment_type.isnot(None))
        .filter(ApprovedClaimRequest.payment_type != "")
        .distinct()
        .order_by(ApprovedClaimRequest.payment_type)
        .all()
    )
    return [r.payment_type for r in rows]


@router.get("/config/account-subjects")
def get_account_subjects(
    company: str     = Query("樂群"),
    db:      Session = Depends(get_db),
    _:       object  = Depends(require_permission(_PERM)),
):
    rows = (
        db.query(ApprovedClaimRequest.account_subject)
        .filter(
            ApprovedClaimRequest.company == company,
            ApprovedClaimRequest.account_subject.isnot(None),
            ApprovedClaimRequest.account_subject != "",
        )
        .distinct()
        .order_by(ApprovedClaimRequest.account_subject)
        .all()
    )
    return [r.account_subject for r in rows]


# ─────────────────────────────────────────────────────────────────────────────
# POST /sync — 手動觸發同步
# ─────────────────────────────────────────────────────────────────────────────

@router.post("/sync")
async def trigger_sync(
    full_resync:    bool            = Query(False),
    background:     BackgroundTasks = BackgroundTasks(),
    _:              object          = Depends(require_permission(_PERM)),
):
    from app.services.claim_request_sync import sync_from_ragic

    async def _do():
        db = SessionLocal()
        try:
            result = await sync_from_ragic(full_resync=full_resync)
            db.add(ModuleSyncLog(
                module_name  = "核准請款單",
                started_at   = twnow(),
                finished_at  = twnow(),
                duration_sec = 0,
                status       = "success" if not result.get("errors") else "partial",
                fetched      = result.get("fetched", 0),
                upserted     = result.get("upserted", 0),
                errors_count = len(result.get("errors", [])),
                error_msg    = "; ".join(result.get("errors", [])[:3]) or None,
                triggered_by = "manual",
            ))
            db.commit()
        except Exception as exc:
            db.rollback()
        finally:
            db.close()

    background.add_task(_do)
    return {"message": "請款單同步已啟動", "task": "claim_sync"}


# ─────────────────────────────────────────────────────────────────────────────
# GET /sync/status — 同步狀態查詢
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/sync/status")
def get_sync_status(
    db: Session = Depends(get_db),
    _:  object  = Depends(require_permission(_PERM)),
):
    recent_logs = (
        db.query(ModuleSyncLog)
        .filter(ModuleSyncLog.module_name == "核准請款單")
        .order_by(ModuleSyncLog.started_at.desc())
        .limit(10)
        .all()
    )

    pending_detail_count = (
        db.query(func.count(ApprovedClaimRequest.id))
        .filter(ApprovedClaimRequest.detail_synced == False)
        .scalar() or 0
    )

    dept_stats = (
        db.query(
            ApprovedClaimRequest.department_display,
            func.count(ApprovedClaimRequest.id).label("total"),
            func.sum(
                func.cast(ApprovedClaimRequest.detail_synced, type_=func.count().type)
            ).label("detail_synced"),
        )
        .group_by(ApprovedClaimRequest.department_display)
        .all()
    )

    return {
        "recent_logs": [
            {
                "id":         l.id,
                "module":     l.module_name,
                "trigger":    l.triggered_by,
                "status":     l.status,
                "message":    l.error_msg or "",
                "created_at": l.started_at.isoformat() if l.started_at else None,
            }
            for l in recent_logs
        ],
        "pending_detail_count": pending_detail_count,
        "dept_stats": [
            {
                "department_display": r.department_display,
                "total":              r.total,
                "detail_synced":      int(r.detail_synced or 0),
                "pending":            r.total - int(r.detail_synced or 0),
            }
            for r in dept_stats
        ],
    }
