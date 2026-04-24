"""
樂群工務報修 API Router
Prefix: /api/v1/luqun-repair

端點：
  GET  /raw-fields        — 回傳 Ragic 第一筆資料欄位名稱（debug 用）
  GET  /ping              — 快速連線診斷
  GET  /sync              — 同步診斷：直接抓 Ragic 回傳統計摘要（debug 用）
  POST /sync              — 觸發背景同步：Ragic → SQLite
  GET  /years             — 回傳資料中的年份清單
  GET  /filter-options    — 回傳過濾條件選項（類型/樓層/狀態）
  GET  /dashboard         — Dashboard KPI + 圖表資料
  GET  /stats/fee         — 金額統計（費用類型 × 月份交叉表）
  GET  /stats/repair      — 4.1 報修統計（月份×6指標）
  GET  /stats/closing     — 4.2 結案時間統計
  GET  /stats/type        — 4.3 報修類型統計
  GET  /stats/room        — 4.4 本月客房報修表
  GET  /detail            — 明細清單（分頁+排序+搜尋）
  GET  /export            — 匯出 Excel（串流回傳）

設計：
  - 所有資料端點（/years 以下）從本地 SQLite 讀取（LuqunRepairCase 表）
  - GET /sync 仍直連 Ragic，僅供診斷用
  - POST /sync 觸發背景同步（Ragic → SQLite），立即回傳
"""
from __future__ import annotations

import io
import time
from collections import Counter
from datetime import datetime
from typing import Optional

import httpx
from fastapi import APIRouter, BackgroundTasks, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.database import get_db
from app.models.luqun_repair import LuqunRepairCase
from app.services import luqun_repair_service as svc

router = APIRouter()


# ── 共用 DB 讀取 helper ────────────────────────────────────────────────────────

def load_cases_from_db(db: Session) -> list[LuqunRepairCase]:
    """從本地 SQLite 載入所有樂群報修案件（ORM 物件與 RepairCase 介面相容）。"""
    return db.query(LuqunRepairCase).all()


# ── /raw-fields ───────────────────────────────────────────────────────────────

@router.get("/raw-fields", summary="回傳 Ragic 第一筆欄位名稱（debug 用）")
async def get_raw_fields():
    return await svc.fetch_raw_fields()


@router.get("/test-parse/{case_no}", summary="單筆案件解析測試：直接從 Ragic 抓取並顯示解析結果")
async def test_parse_case(case_no: str):
    """
    從 Ragic 直接抓取指定案件，用最新程式碼解析後回傳 deduction_counter_name 等欄位。
    用來確認「扣款專櫃」是否正確解析，不經過 cache。
    """
    from app.services.luqun_repair_service import RepairCase, invalidate_cache
    adapter = svc._get_adapter()
    invalidate_cache()  # 強制清快取
    try:
        raw_data = await adapter.fetch_all(limit=500)
        for rid, row in raw_data.items():
            if row.get("報修編號") == case_no or str(row.get("_ragicId", "")) == case_no:
                case = RepairCase(ragic_id=str(rid), raw=row)
                return {
                    "ragic_id": rid,
                    "case_no": case.case_no,
                    "raw_deduction_counter_field": row.get("扣款專櫃"),
                    "parsed_deduction_counter_name": case.deduction_counter_name,
                    "parsed_counter_stores": case.counter_stores,
                    "deduction_fee": case.deduction_fee,
                    "finance_note": case.finance_note[:100] if case.finance_note else "",
                    "mgmt_response": case.mgmt_response[:100] if case.mgmt_response else "",
                }
        return {"error": f"案件 {case_no} 未在主表找到（limit=500）"}
    except Exception as exc:
        return {"error": str(exc)}


@router.get("/raw-record/{case_no}", summary="直接從 Ragic 取得單筆完整 JSON（debug 用）")
async def get_raw_record(case_no: str):
    """
    從 Ragic 主表找出指定案件編號的完整原始 JSON，包含所有欄位名稱與值。
    用途：確認「扣款專櫃」等欄位在 Ragic API 中的實際 key 名稱。
    """
    adapter = svc._get_adapter()
    try:
        raw_data = await adapter.fetch_all(limit=500)
        for rid, row in raw_data.items():
            if row.get("報修編號") == case_no or row.get("報修同仁") == case_no:
                # 找到了，同時也拉 detail
                detail = await adapter.fetch_one(rid)
                detail_data = {}
                if isinstance(detail, dict):
                    detail_data = detail.get(str(rid)) or detail.get(rid) or {
                        k: v for k, v in detail.items() if not str(k).startswith("_")
                    }
                # 找所有扣款相關欄位
                deduction_keys = {k: v for k, v in row.items() if "扣" in str(k) or "款" in str(k)}
                deduction_keys_detail = {k: v for k, v in detail_data.items() if "扣" in str(k) or "款" in str(k)}
                return {
                    "ragic_id": rid,
                    "case_no": case_no,
                    "main_list_deduction_fields": deduction_keys,
                    "detail_deduction_fields": deduction_keys_detail,
                    "main_list_all_keys": list(row.keys()),
                    "detail_all_keys": list(detail_data.keys()),
                    "main_list_full": row,
                }
        return {"error": f"案件 {case_no} 未找到"}
    except Exception as exc:
        return {"error": str(exc)}


# ── /ping ─────────────────────────────────────────────────────────────────────

@router.get("/ping", summary="快速連線診斷（5 秒 timeout）")
async def ping_ragic():
    """直接對 Ragic URL 發一次 GET（limit=1），5 秒內回傳裸 HTTP 結果。"""
    base_url = (
        f"https://{settings.RAGIC_LUQUN_REPAIR_SERVER_URL}"
        f"/{settings.RAGIC_LUQUN_REPAIR_ACCOUNT}"
        f"/{settings.RAGIC_LUQUN_REPAIR_PATH}"
    )
    api_key = settings.RAGIC_API_KEY
    results = []

    for label, params in [
        ("base_url", {"api": "", "limit": 1}),
    ]:
        t0 = time.monotonic()
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(5.0), verify=False) as client:
                resp = await client.get(
                    base_url,
                    headers={"Authorization": f"Basic {api_key}", "Accept": "application/json"},
                    params=params,
                )
            elapsed = int((time.monotonic() - t0) * 1000)
            try:
                body = resp.json()
                body_type = type(body).__name__
                if isinstance(body, dict):
                    record_ids = [k for k in body if k.lstrip("-").isdigit()]
                    record_count = len(record_ids)
                    first_record_id = record_ids[0] if record_ids else None
                    first_record_raw = body[first_record_id] if first_record_id else {}
                    preview_keys = list(body.keys())[:5]
                    body_preview = {k: type(body[k]).__name__ for k in preview_keys}
                else:
                    body_preview = str(body)[:200]
                    record_count = 0
                    first_record_id = None
                    first_record_raw = {}
            except Exception:
                body_type = "not_json"
                body_preview = resp.text[:300]
                record_count = 0
                first_record_id = None
                first_record_raw = {}
            results.append({
                "test": label, "status_code": resp.status_code, "elapsed_ms": elapsed,
                "body_type": body_type, "record_count": record_count,
                "body_preview": body_preview,
                "first_record_id": first_record_id,
                "first_record_raw": first_record_raw,
                "error": None,
            })
        except httpx.TimeoutException:
            elapsed = int((time.monotonic() - t0) * 1000)
            results.append({
                "test": label, "elapsed_ms": elapsed,
                "error": f"TIMEOUT after {elapsed}ms",
                "tip": "後端無法連到 Ragic，請確認網路/防火牆設定",
            })
        except Exception as exc:
            elapsed = int((time.monotonic() - t0) * 1000)
            results.append({"test": label, "elapsed_ms": elapsed, "error": str(exc)})

    return {
        "ragic_base_url": base_url,
        "pageid": None,
        "api_key_prefix": api_key[:8] + "..." if len(api_key) > 8 else api_key,
        "results": results,
    }


# ── /sync ─────────────────────────────────────────────────────────────────────

@router.get("/sync", summary="同步診斷：直接抓 Ragic 回傳統計摘要（debug 用）")
async def sync_diagnostic():
    """
    完整從 Ragic 抓取所有資料並回傳診斷摘要（含費用合計）。

    注意：此端點直連 Ragic，僅供診斷；正常頁面資料讀本地 SQLite。
    觸發實際同步請用 POST /sync。
    """
    field_info = await svc.fetch_raw_fields()
    all_cases  = await svc.fetch_all_cases()

    year_dist = dict(sorted(Counter(
        c.year for c in all_cases if c.year is not None
    ).items()))

    no_date_count = sum(1 for c in all_cases if c.occurred_at is None and c.title)

    # ── 費用診斷 ──────────────────────────────────────────────────────────────
    total_outsource   = sum(c.outsource_fee     for c in all_cases)
    total_maintenance = sum(c.maintenance_fee   for c in all_cases)
    total_deduction   = sum(c.deduction_fee     for c in all_cases)
    total_ded_counter = sum(c.deduction_counter for c in all_cases)

    fee_samples = [
        {
            "ragic_id":          c.ragic_id,
            "case_no":           c.case_no,
            "outsource_fee":     c.outsource_fee,
            "maintenance_fee":   c.maintenance_fee,
            "deduction_fee":     c.deduction_fee,
            "deduction_counter": c.deduction_counter,
            "_raw_outsource":   c._raw.get("委外費用", ""),
            "_raw_maintenance": c._raw.get("維修費用", ""),
            "_raw_deduction":   c._raw.get("扣款費用", ""),
            "_raw_ded_counter": c._raw.get("扣款專櫃", ""),
        }
        for c in all_cases
        if c.outsource_fee > 0 or c.maintenance_fee > 0 or c.deduction_fee > 0
    ][:5]

    recent = sorted(
        [c for c in all_cases if c.occurred_at],
        key=lambda c: c.occurred_at,  # type: ignore
        reverse=True,
    )[:3]
    recent_samples = [
        {
            "ragic_id":    c.ragic_id,
            "case_no":     c.case_no,
            "title":       c.title[:30] if c.title else "",
            "occurred_at": c.occurred_at.strftime("%Y/%m/%d") if c.occurred_at else "",
            "status":      c.status,
        }
        for c in recent
    ]

    return {
        "ok":            True,
        "total_parsed":  len(all_cases),
        "no_date_count": no_date_count,
        "year_distribution": year_dist,
        "fee_totals": {
            "outsource_fee":              round(total_outsource,   2),
            "maintenance_fee":            round(total_maintenance, 2),
            "deduction_fee":              round(total_deduction,   2),
            "deduction_counter":          round(total_ded_counter, 2),
            "outsource_plus_maintenance": round(total_outsource + total_maintenance, 2),
        },
        "fee_samples":   fee_samples,
        "field_names":   field_info.get("fields", []),
        "sample_raw":    field_info.get("sample", {}),
        "recent_samples": recent_samples,
        "ragic_url": (
            f"https://{settings.RAGIC_LUQUN_REPAIR_SERVER_URL}"
            f"/{settings.RAGIC_LUQUN_REPAIR_ACCOUNT}"
            f"/{settings.RAGIC_LUQUN_REPAIR_PATH}"
        ),
    }


@router.post("/sync", summary="觸發背景同步：Ragic → SQLite（非阻塞）")
async def sync_from_ragic(background_tasks: BackgroundTasks):
    """
    將樂群工務報修資料從 Ragic 同步到本地 SQLite（背景執行）。
    立即回傳，不等待同步完成。
    """
    from app.services.luqun_repair_sync import sync_from_ragic as do_sync
    background_tasks.add_task(do_sync)
    return {"success": True, "message": "同步已在背景啟動"}


# ── /years ─────────────────────────────────────────────────────────────────────

@router.get("/years", summary="資料中的年份清單")
def get_years(db: Session = Depends(get_db)):
    all_cases = load_cases_from_db(db)
    return {"years": svc.get_years(all_cases)}


# ── /filter-options ───────────────────────────────────────────────────────────

@router.get("/filter-options", summary="過濾條件選項")
def get_filter_options(db: Session = Depends(get_db)):
    all_cases = load_cases_from_db(db)
    return svc.get_filter_options(all_cases)


# ── /dashboard ────────────────────────────────────────────────────────────────

@router.get("/dashboard", summary="Dashboard KPI + 圖表資料")
def get_dashboard(
    year:  int = Query(..., description="年度"),
    month: int = Query(0,   description="月份（0=全年）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    eff_month = month if month > 0 else datetime.now().month
    return svc.compute_dashboard(all_cases, year, eff_month)


# ── /stats/fee ────────────────────────────────────────────────────────────────

@router.get("/stats/fee", summary="金額統計（費用類型 × 月份交叉表）")
def get_fee_stats(
    year: int = Query(..., description="年度"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_fee_stats(all_cases, year)


# ── /stats/repair ─────────────────────────────────────────────────────────────

@router.get("/stats/repair", summary="4.1 報修統計（全年月份×6指標）")
def get_repair_stats(
    year: int = Query(..., description="年度"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_repair_stats(all_cases, year)


# ── /stats/closing ────────────────────────────────────────────────────────────

@router.get("/stats/closing", summary="4.2 結案時間統計（小型/中大型）")
def get_closing_stats(
    year:  int           = Query(..., description="年度"),
    month: Optional[int] = Query(None, description="月份（不傳=全年）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_closing_time(all_cases, year, month)


# ── /stats/type ───────────────────────────────────────────────────────────────

@router.get("/stats/type", summary="4.3 報修類型統計（類型×月份）")
def get_type_stats(
    year:  int           = Query(..., description="年度"),
    month: Optional[int] = Query(None, description="聚焦月份（選填）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_type_stats(all_cases, year, month)


# ── /stats/room ───────────────────────────────────────────────────────────────

@router.get("/stats/room", summary="4.4 本月客房報修表（房號×分類）")
def get_room_stats(
    year:  int = Query(..., description="年度"),
    month: int = Query(..., description="月份（必選）"),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.compute_room_repair_table(all_cases, year, month)


# ── /detail ───────────────────────────────────────────────────────────────────

@router.get("/detail", summary="明細清單（分頁+排序+搜尋）")
def get_detail(
    year:        Optional[int] = Query(None),
    month:       Optional[int] = Query(None),
    repair_type: Optional[str] = Query(None),
    floor:       Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    keyword:     Optional[str] = Query(None),
    page:        int            = Query(1, ge=1),
    page_size:   int            = Query(50, ge=1, le=200),
    sort_by:     str            = Query("occurred_at"),
    sort_desc:   bool           = Query(True),
    db: Session = Depends(get_db),
):
    all_cases = load_cases_from_db(db)
    return svc.query_detail(
        all_cases,
        year=year, month=month,
        repair_type=repair_type, floor=floor, status=status,
        keyword=keyword,
        page=page, page_size=page_size,
        sort_by=sort_by, sort_desc=sort_desc,
    )


# ── /export ───────────────────────────────────────────────────────────────────

@router.get("/export", summary="匯出 Excel")
def export_excel(
    year:        Optional[int] = Query(None),
    month:       Optional[int] = Query(None),
    repair_type: Optional[str] = Query(None),
    floor:       Optional[str] = Query(None),
    status:      Optional[str] = Query(None),
    keyword:     Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    """匯出過濾後的明細資料為 Excel 檔案（.xlsx）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="請安裝 openpyxl：pip install openpyxl")

    all_cases = load_cases_from_db(db)
    result = svc.query_detail(
        all_cases,
        year=year, month=month,
        repair_type=repair_type, floor=floor, status=status,
        keyword=keyword,
        page=1, page_size=9999,
    )
    items = result["items"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "報修明細"

    # ── 標題行 ────────────────────────────────────────────────────────────────
    headers = [
        "報修編號", "標題", "報修人姓名", "報修類型", "發生樓層",
        "發生時間", "負責單位", "花費工時", "處理狀況",
        "委外費用", "維修費用", "總費用",
        "驗收者", "驗收", "結案人",
        "扣款事項", "扣款費用", "財務備註",
        "結案時間", "結案天數", "是否完成",
    ]

    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border

    # ── 資料行 ────────────────────────────────────────────────────────────────
    STATUS_FILLS = {
        True:  PatternFill(start_color="F6FFED", end_color="F6FFED", fill_type="solid"),
        False: PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid"),
    }

    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.get("case_no", ""),
            item.get("title", ""),
            item.get("reporter_name", ""),
            item.get("repair_type", ""),
            item.get("floor", ""),
            item.get("occurred_at", ""),
            item.get("responsible_unit", ""),
            item.get("status", ""),
            item.get("outsource_fee", 0),
            item.get("maintenance_fee", 0),
            item.get("total_fee", 0),
            item.get("deduction_item", ""),
            item.get("deduction_fee", 0),
            item.get("deduction_counter_name", ""),
            item.get("closer", ""),
            item.get("acceptor", ""),
            item.get("accept_status", ""),
            item.get("finance_note", ""),
            item.get("completed_at", ""),
            item.get("close_days", ""),
            "æ¯" if item.get("is_completed") else "å¦",
        ]
        is_done = item.get("is_completed", False)
        fill = STATUS_FILLS[is_done]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    col_widths = [12, 30, 12, 10, 10, 18, 12, 8, 10, 10, 10, 12, 10, 12, 8, 10, 14, 10, 18, 14, 8]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"luqun_repair_{year or 'all'}_{month or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
