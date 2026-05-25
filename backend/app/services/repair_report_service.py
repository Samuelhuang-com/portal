"""
報修未完成報表 — Service Layer

職責：
  1. 從 DB 聚合兩個模組（飯店/商場）的未完成案件
  2. 使用與各模組 dashboard 相同的完成狀態判斷邏輯
  3. 產生 Excel 報表（openpyxl）
  4. 寄送 SMTP 郵件
  5. 排程相關輔助（讀取/更新 scheduler job）

注意：
  - 不呼叫 Ragic API，只讀取本地 DB
  - 完成狀態判斷函數直接從各模組 service 匯入，確保邏輯一致
"""
from __future__ import annotations

import calendar
import io
import logging
import smtplib
import ssl
from datetime import date, datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.dazhi_repair import DazhiRepairCase
from app.models.luqun_repair import LuqunRepairCase
from app.models.repair_report import (
    RepairReportMailLog,
    RepairReportRecipient,
    RepairReportScheduleSettings,
)

# ── 沿用各模組相同的完成/排除狀態判斷 ──────────────────────────────────────────
from app.services.dazhi_repair_service import is_completed as dazhi_is_completed
from app.services.dazhi_repair_service import is_excluded as dazhi_is_excluded
from app.services.luqun_repair_service import is_completed as luqun_is_completed
from app.services.luqun_repair_service import is_excluded as luqun_is_excluded

logger = logging.getLogger(__name__)

# 逾期閾值（等待天數 > N 天視為可能逾期）
# ⚠️ 暫定規則：等待超過 3 天標示為「可能逾期」，未來可調整
OVERDUE_THRESHOLD_DAYS = 3


# ── 欄位 Normalize ────────────────────────────────────────────────────────────

def _build_ragic_url(source: str, ragic_id: str) -> str:
    if source == "hotel":
        return (
            f"https://{settings.RAGIC_DAZHI_REPAIR_SERVER_URL}"
            f"/{settings.RAGIC_DAZHI_REPAIR_ACCOUNT}"
            f"/{settings.RAGIC_DAZHI_REPAIR_PATH}/{ragic_id}"
        )
    return (
        f"https://{settings.RAGIC_LUQUN_REPAIR_SERVER_URL}"
        f"/{settings.RAGIC_LUQUN_REPAIR_ACCOUNT}"
        f"/{settings.RAGIC_LUQUN_REPAIR_PATH}/{ragic_id}"
    )


def _normalize(c, source: str) -> dict:
    """將 ORM case 物件正規化為統一 dict 格式。"""
    now = datetime.now()
    pending_days: Optional[int] = None
    if c.occurred_at:
        delta = now - c.occurred_at
        pending_days = max(0, delta.days)

    is_overdue = (pending_days is not None and pending_days > OVERDUE_THRESHOLD_DAYS)

    return {
        "source":           source,
        "source_label":     "飯店" if source == "hotel" else "商場",
        "ragic_id":         c.ragic_id or "",
        "case_no":          c.case_no or "",
        "occurred_at":      c.occurred_at.isoformat() if c.occurred_at else None,
        "floor":            c.floor or "",
        "repair_type":      c.repair_type or "",
        "title":            c.title or "",
        "status":           c.status or "",
        "responsible_unit": c.responsible_unit or "",
        "pending_days":     pending_days,
        "is_overdue":       is_overdue,
        "synced_at":        c.synced_at.isoformat() if c.synced_at else None,
        "finance_note":     c.finance_note or "",
        "ragic_url":        _build_ragic_url(source, c.ragic_id or ""),
    }


# ── 核心查詢 ──────────────────────────────────────────────────────────────────

def get_unfinished_cases(
    db: Session,
    year: int,
    month: int,
    source: str = "all",        # all / hotel / mall
    status_filter: Optional[str] = None,
    overdue_only: bool = False,
    repair_type_filter: Optional[str] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
) -> dict:
    """
    聚合飯店 + 商場未完成案件，套用與各模組 dashboard 相同的判斷邏輯。
    過濾條件依 occurred_at（報修日期）的年月分組。
    """
    today = date.today()
    cases: list[dict] = []

    # 報表月底截止時間：occurred_at > 月底的案件不納入（支援查詢歷史月份）
    _last_day = calendar.monthrange(year, month)[1]
    _cutoff   = datetime(year, month, _last_day, 23, 59, 59)

    # ── 飯店（大直工務部）──────────────────────────────────────────────────
    if source in ("all", "hotel"):
        for c in db.query(DazhiRepairCase).all():
            if dazhi_is_excluded(c.status):
                continue
            if dazhi_is_completed(c.status):
                continue
            if not c.occurred_at:
                continue
            # 只納入報表月底前的案件（跨月歷史未結清一併計入）
            if c.occurred_at > _cutoff:
                continue
            # 待辦驗：與 dashboard 口徑一致，獨立類別，不算未完成
            if (c.status or "").strip() == "待辦驗":
                continue
            cases.append(_normalize(c, "hotel"))

    # ── 商場（商場工務報修）─────────────────────────────────────────────────
    if source in ("all", "mall"):
        for c in db.query(LuqunRepairCase).all():
            if luqun_is_excluded(c.status):
                continue
            if luqun_is_completed(c.status):
                continue
            if not c.occurred_at:
                continue
            # 只納入報表月底前的案件（跨月歷史未結清一併計入）
            if c.occurred_at > _cutoff:
                continue
            # 待辦驗：與 dashboard 口徑一致，獨立類別，不算未完成
            if (c.status or "").strip() == "待辦驗":
                continue
            cases.append(_normalize(c, "mall"))

    # ── 額外過濾條件 ──────────────────────────────────────────────────────────
    if status_filter:
        cases = [c for c in cases if c["status"] == status_filter]

    if overdue_only:
        cases = [c for c in cases if c["is_overdue"]]

    if repair_type_filter:
        cases = [c for c in cases if c["repair_type"] == repair_type_filter]

    if keyword:
        kw = keyword.lower()
        searchable_fields = ("case_no", "floor", "title", "status", "responsible_unit", "finance_note")
        cases = [
            c for c in cases
            if any(kw in (c.get(f) or "").lower() for f in searchable_fields)
        ]

    # ── KPI 計算 ──────────────────────────────────────────────────────────────
    kpi = _calc_kpi(cases, year, month, today)

    # ── 動態過濾選項（從未過濾的全量資料取得，以支援多選組合）─────────────────
    # 這裡直接從當前結果取，若要全量選項可分開查詢（本次選擇簡化方式）
    filter_options = {
        "statuses":     sorted({c["status"] for c in cases if c["status"]}),
        "repair_types": sorted({c["repair_type"] for c in cases if c["repair_type"]}),
    }

    # ── 分頁 ─────────────────────────────────────────────────────────────────
    total = len(cases)
    start = (page - 1) * page_size
    end   = start + page_size

    return {
        "items":          cases[start:end],
        "total":          total,
        "kpi":            kpi,
        "filter_options": filter_options,
    }


def _calc_kpi(cases: list[dict], year: int, month: int, today: date) -> dict:
    hotel_cases  = [c for c in cases if c["source"] == "hotel"]
    mall_cases   = [c for c in cases if c["source"] == "mall"]
    overdue      = [c for c in cases if c["is_overdue"]]
    pending_days = [c["pending_days"] for c in cases if c["pending_days"] is not None]
    today_str    = today.isoformat()

    month_prefix = f"{year}-{month:02d}"
    new_today = sum(
        1 for c in cases
        if c["occurred_at"] and c["occurred_at"][:10] == today_str
    )
    # new_this_month = occurred_at 落在報表年月的未完成案件數（歷史跨月案件不計入）
    new_this_month = sum(
        1 for c in cases
        if c["occurred_at"] and c["occurred_at"][:7] == month_prefix
    )

    return {
        "total_unfinished":  len(cases),
        "hotel_unfinished":  len(hotel_cases),
        "mall_unfinished":   len(mall_cases),
        "overdue_count":     len(overdue),
        "avg_pending_days":  round(sum(pending_days) / len(pending_days), 1) if pending_days else 0.0,
        "max_pending_days":  max(pending_days) if pending_days else 0,
        "new_this_month":    new_this_month,
        "new_today":         new_today,
    }


# ── 不分頁版本（供排程寄信使用）──────────────────────────────────────────────

def get_all_unfinished_cases(
    db: Session,
    year: int,
    month: int,
    include_hotel: bool = True,
    include_mall: bool = True,
) -> list[dict]:
    """取得所有未完成案件（不分頁），供排程寄信與 Excel 匯出使用。"""
    source = "all"
    if include_hotel and not include_mall:
        source = "hotel"
    elif include_mall and not include_hotel:
        source = "mall"

    result = get_unfinished_cases(
        db=db,
        year=year,
        month=month,
        source=source,
        page=1,
        page_size=99999,
    )
    return result["items"]


# ── Excel 匯出 ────────────────────────────────────────────────────────────────

_HEADER_FILL  = PatternFill("solid", fgColor="1B3A5C")
_HEADER_FONT  = Font(bold=True, color="FFFFFF")
_OVERDUE_FILL = PatternFill("solid", fgColor="FFF5F5")
_DATE_FORMAT  = "YYYY-MM-DD"

COLUMN_DEFS = [
    ("來源",           "source_label",     15),
    ("案件編號",       "case_no",          18),
    ("報修日期",       "occurred_at",      15),
    ("報修地點",       "floor",            20),
    ("工項類別",       "repair_type",      20),
    ("報修內容",       "title",            35),
    ("狀態",           "status",           12),
    ("工務處理人員",   "responsible_unit", 18),
    ("已等待天數",     "pending_days",     12),
    ("可能逾期",       "is_overdue",       10),
    ("最後更新時間",   "synced_at",        18),
    ("備註/處理說明",  "finance_note",     35),
    ("原始資料連結",   "ragic_url",        50),
]


def generate_excel(cases: list[dict], year: int, month: int) -> bytes:
    """產生 Excel 報表，含 5 個工作表。"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 移除預設工作表

    hotel_cases   = [c for c in cases if c["source"] == "hotel"]
    mall_cases    = [c for c in cases if c["source"] == "mall"]
    overdue_cases = [c for c in cases if c["is_overdue"]]

    pending_days_all = [c["pending_days"] for c in cases if c["pending_days"] is not None]
    summary = [
        ("項目", "數值"),
        ("報表年月", f"{year}年{month}月"),
        ("未完成案件總數", len(cases)),
        ("飯店未完成案件數", len(hotel_cases)),
        ("商場未完成案件數", len(mall_cases)),
        ("可能逾期案件數", len(overdue_cases)),
        ("平均等待天數", round(sum(pending_days_all) / len(pending_days_all), 1) if pending_days_all else 0),
        ("最長等待天數", max(pending_days_all) if pending_days_all else 0),
        ("⚠️ 逾期標準", f"等待天數 > {OVERDUE_THRESHOLD_DAYS} 天（暫定規則）"),
    ]

    _add_sheet(wb, f"{year}-{month:02d} 總表",     cases)
    _add_sheet(wb, "飯店未完成明細",                hotel_cases)
    _add_sheet(wb, "商場未完成明細",                mall_cases)
    _add_sheet(wb, "可能逾期案件",                   overdue_cases)
    _add_summary_sheet(wb, "統計摘要",               summary)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


_HYPERLINK_FONT = Font(color="4BA8E8", underline="single")


def _add_sheet(wb: openpyxl.Workbook, title: str, cases: list[dict]) -> None:
    ws = wb.create_sheet(title=title)

    headers = [col[0] for col in COLUMN_DEFS]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 凍結表頭（第 1 列）
    ws.freeze_panes = "A2"

    for row_idx, c in enumerate(cases, start=2):
        is_overdue = c.get("is_overdue", False)
        for col_idx, (_, field, _) in enumerate(COLUMN_DEFS, start=1):
            val = c.get(field)
            if field == "is_overdue":
                val = "是" if val else ""
            elif field in ("occurred_at", "synced_at") and isinstance(val, str) and len(val) >= 10:
                val = val[:10]
            elif field == "pending_days" and val is None:
                val = "-"

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # 原始資料連結 → 超連結
            if field == "ragic_url" and val:
                cell.value = "開啟"
                cell.hyperlink = val
                cell.font = _HYPERLINK_FONT
            elif is_overdue:
                cell.fill = _OVERDUE_FILL

    # 欄寬
    for col_idx, (_, _, width) in enumerate(COLUMN_DEFS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def _add_summary_sheet(wb: openpyxl.Workbook, title: str, rows: list[tuple]) -> None:
    ws = wb.create_sheet(title=title)
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = Font(bold=True)
    # 凍結表頭
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 25


# ── 郵件寄送 ──────────────────────────────────────────────────────────────────

def _build_html_body(
    year: int,
    month: int,
    kpi: dict,
    portal_url: str = "",
    has_attachment: bool = True,
) -> str:
    overdue_note = f"（等待天數 > {OVERDUE_THRESHOLD_DAYS} 天，暫定規則）"
    attach_note  = "📎 Excel 報表已附件寄出" if has_attachment else "（本次未附加 Excel 報表）"
    link_html    = f'<a href="{portal_url}" style="color:#4BA8E8;">點此開啟系統報表</a>' if portal_url else ""

    return f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head><meta charset="utf-8"><title>報修未完成報表</title></head>
<body style="font-family:Arial,sans-serif;color:#333;max-width:640px;margin:0 auto;padding:20px;">
  <div style="background:linear-gradient(135deg,#1B3A5C,#4BA8E8);padding:20px;border-radius:8px 8px 0 0;">
    <h2 style="color:#fff;margin:0;">📋 報修未完成報表</h2>
    <p style="color:#dce9f5;margin:4px 0 0;">{year}年{month:02d}月｜系統自動產出</p>
  </div>
  <div style="border:1px solid #ddd;border-top:none;padding:20px;border-radius:0 0 8px 8px;">
    <table style="width:100%;border-collapse:collapse;margin-bottom:16px;">
      <tr style="background:#1B3A5C;color:#fff;">
        <th style="padding:8px;text-align:left;">項目</th>
        <th style="padding:8px;text-align:right;">數值</th>
      </tr>
      <tr style="background:#f0f4f8;">
        <td style="padding:8px;">未完成案件總數</td>
        <td style="padding:8px;text-align:right;font-weight:bold;font-size:18px;">{kpi['total_unfinished']}</td>
      </tr>
      <tr>
        <td style="padding:8px;">飯店未完成案件數</td>
        <td style="padding:8px;text-align:right;">{kpi['hotel_unfinished']}</td>
      </tr>
      <tr style="background:#f0f4f8;">
        <td style="padding:8px;">商場未完成案件數</td>
        <td style="padding:8px;text-align:right;">{kpi['mall_unfinished']}</td>
      </tr>
      <tr>
        <td style="padding:8px;">可能逾期案件數 {overdue_note}</td>
        <td style="padding:8px;text-align:right;color:#e74c3c;font-weight:bold;">{kpi['overdue_count']}</td>
      </tr>
      <tr style="background:#f0f4f8;">
        <td style="padding:8px;">最長等待天數</td>
        <td style="padding:8px;text-align:right;">{kpi['max_pending_days']} 天</td>
      </tr>
    </table>
    <p style="color:#666;font-size:13px;">{attach_note}</p>
    {f'<p>{link_html}</p>' if link_html else ''}
    <hr style="border:none;border-top:1px solid #eee;margin:16px 0;">
    <p style="font-size:12px;color:#999;">此郵件由維春集團管理 Portal 自動寄出，請勿直接回覆。</p>
  </div>
</body>
</html>
"""


def _build_text_body(year: int, month: int, kpi: dict) -> str:
    overdue_note = f"（等待 > {OVERDUE_THRESHOLD_DAYS} 天，暫定）"
    return (
        f"【報修未完成報表】{year}年{month:02d}月\n"
        f"{'='*40}\n"
        f"未完成案件總數：{kpi['total_unfinished']}\n"
        f"飯店未完成案件數：{kpi['hotel_unfinished']}\n"
        f"商場未完成案件數：{kpi['mall_unfinished']}\n"
        f"可能逾期案件數 {overdue_note}：{kpi['overdue_count']}\n"
        f"最長等待天數：{kpi['max_pending_days']} 天\n"
        f"{'='*40}\n"
        f"此郵件由維春集團管理 Portal 自動寄出。\n"
    )


def send_single_email(
    to_email: str,
    to_name: str,
    subject: str,
    html_body: str,
    text_body: str,
    attachment: Optional[tuple[str, bytes]] = None,
) -> None:
    """
    寄送單封郵件。

    :param attachment: (filename, bytes) 或 None
    :raises: Exception — 任何 SMTP 錯誤直接往上拋
    """
    if not settings.MAIL_HOST:
        raise RuntimeError("MAIL_HOST 未設定，請在 .env 中設定郵件伺服器參數")

    msg = MIMEMultipart("mixed")
    from_addr = settings.MAIL_FROM or settings.MAIL_USERNAME
    msg["From"]    = formataddr((settings.MAIL_FROM_NAME, from_addr))
    msg["To"]      = to_email
    msg["Subject"] = subject

    # HTML + 純文字 fallback
    alt = MIMEMultipart("alternative")
    alt.attach(MIMEText(text_body, "plain", "utf-8"))
    alt.attach(MIMEText(html_body, "html",  "utf-8"))
    msg.attach(alt)

    # 附件
    if attachment:
        filename, data = attachment
        part = MIMEBase("application", "octet-stream")
        part.set_payload(data)
        encoders.encode_base64(part)
        # RFC 5987 編碼：正確傳遞含中文的檔名，避免 Gmail 顯示 "noname"
        part.add_header(
            "Content-Disposition",
            "attachment",
            filename=("utf-8", "", filename),
        )
        msg.attach(part)

    # SMTP 連線
    with smtplib.SMTP(settings.MAIL_HOST, settings.MAIL_SMTP_PORT, timeout=30) as server:
        server.ehlo()
        if settings.MAIL_USE_TLS:
            # 使用寬鬆 SSL Context，相容弱 DH Key 的舊郵件伺服器（SECLEVEL=1）
            ctx = ssl.create_default_context()
            ctx.set_ciphers("DEFAULT@SECLEVEL=1")
            ctx.check_hostname = False
            ctx.verify_mode = ssl.CERT_NONE
            server.starttls(context=ctx)
            server.ehlo()
        if settings.MAIL_USERNAME and settings.MAIL_PASSWORD:
            server.login(settings.MAIL_USERNAME, settings.MAIL_PASSWORD)
        server.sendmail(
            msg["From"],
            to_email,
            msg.as_string(),
        )


# ── 排程寄信主函數 ────────────────────────────────────────────────────────────

def run_daily_send(db: Session) -> None:
    """
    每日排程寄信主流程。
    由 APScheduler 呼叫，需自行建立 DB Session。

    流程：
      1. 讀取排程設定，確認 is_enabled
      2. 確認今日未重複寄送
      3. 取得 active 收件人
      4. 依 report_year_month_mode 決定報表年月
      5. 聚合案件、產生 Excel（若啟用）
      6. 逐一寄送，每人各寫一筆 log
    """
    now   = datetime.now()
    today = now.date()

    # 讀取排程設定
    sched = db.query(RepairReportScheduleSettings).first()
    if not sched:
        logger.info("[RepairReport] 尚未設定排程，略過")
        return

    if not sched.is_enabled:
        logger.info("[RepairReport] 排程未啟用（is_enabled=False），略過")
        return

    # 防重複：確認今日已無「排程」成功紀錄（manual force-send 不算，不影響排程）
    existing = (
        db.query(RepairReportMailLog)
        .filter(
            RepairReportMailLog.send_date == today,
            RepairReportMailLog.status == "success",
            RepairReportMailLog.send_source == "scheduled",
        )
        .first()
    )
    if existing:
        logger.info(f"[RepairReport] 今日（{today}）已有排程成功寄送紀錄，略過")
        return

    # 決定報表年月
    if sched.report_year_month_mode == "previous_month":
        if now.month == 1:
            report_year, report_month = now.year - 1, 12
        else:
            report_year, report_month = now.year, now.month - 1
    else:
        report_year, report_month = now.year, now.month

    # 取得 active 收件人
    recipients = (
        db.query(RepairReportRecipient)
        .filter(RepairReportRecipient.is_active == True)
        .all()
    )
    if not recipients:
        logger.warning("[RepairReport] 無 active 收件人，略過寄送")
        _write_skipped_log(db, today, now, report_year, report_month, "無 active 收件人")
        return

    # 聚合案件資料
    cases = get_all_unfinished_cases(
        db,
        report_year,
        report_month,
        include_hotel=sched.include_hotel,
        include_mall=sched.include_mall,
    )
    today_date = today
    kpi = _calc_kpi(cases, report_year, report_month, today_date)

    hotel_count = kpi["hotel_unfinished"]
    mall_count  = kpi["mall_unfinished"]
    total       = kpi["total_unfinished"]

    # 主旨
    subject = sched.email_subject_template.format(
        year=report_year,
        month=f"{report_month:02d}",
        total=total,
        hotel_count=hotel_count,
        mall_count=mall_count,
    )

    # HTML / 文字內容
    html_body = _build_html_body(report_year, report_month, kpi, has_attachment=sched.include_excel_attachment)
    text_body = _build_text_body(report_year, report_month, kpi)

    # Excel 附件
    attachment: Optional[tuple[str, bytes]] = None
    attachment_filename: Optional[str] = None
    if sched.include_excel_attachment:
        excel_bytes = generate_excel(cases, report_year, report_month)
        today_str = now.strftime("%Y%m%d")
        attachment_filename = f"報修未完成報表_{report_year}{report_month:02d}_{today_str}.xlsx"
        attachment = (attachment_filename, excel_bytes)

    send_time_str = now.strftime("%H:%M")

    # 逐一寄送
    for rcpt in recipients:
        status  = "failed"
        err_msg = None
        try:
            send_single_email(
                to_email=rcpt.email,
                to_name=rcpt.name,
                subject=subject,
                html_body=html_body,
                text_body=text_body,
                attachment=attachment,
            )
            status = "success"
            logger.info(f"[RepairReport] 寄送成功 → {rcpt.email}")
        except Exception as exc:
            err_msg = str(exc)[:500]
            logger.error(f"[RepairReport] 寄送失敗 → {rcpt.email}: {exc}")

        log_entry = RepairReportMailLog(
            send_date=today,
            send_time=send_time_str,
            report_year=report_year,
            report_month=report_month,
            recipient_email=rcpt.email,
            recipient_name=rcpt.name,
            subject=subject,
            status=status,
            error_message=err_msg,
            hotel_unfinished_count=hotel_count,
            mall_unfinished_count=mall_count,
            total_unfinished_count=total,
            attachment_filename=attachment_filename,
            send_source="scheduled",
        )
        db.add(log_entry)

    db.commit()
    logger.info(f"[RepairReport] 排程寄送完成，共 {len(recipients)} 位收件人")


def force_send_now(db: Session, year: int, month: int) -> dict:
    """
    強制立即寄送報修未完成報表，不檢查 is_enabled / 不防重複寄送。
    供 sync_tool 手動按鈕使用。
    回傳 {"sent": N, "failed": M}
    """
    now   = datetime.now()
    today = now.date()

    sched = db.query(RepairReportScheduleSettings).first()
    if not sched:
        raise RuntimeError("尚未設定排程，請先在 Portal 儲存排程設定")

    recipients = (
        db.query(RepairReportRecipient)
        .filter(RepairReportRecipient.is_active == True)
        .all()
    )
    if not recipients:
        logger.warning("[RepairReport] force_send_now: 無 active 收件人")
        return {"sent": 0, "failed": 0}

    cases = get_all_unfinished_cases(
        db, year, month,
        include_hotel=sched.include_hotel,
        include_mall=sched.include_mall,
    )
    kpi         = _calc_kpi(cases, year, month, today)
    hotel_count = kpi["hotel_unfinished"]
    mall_count  = kpi["mall_unfinished"]
    total       = kpi["total_unfinished"]

    subject = sched.email_subject_template.format(
        year=year, month=f"{month:02d}",
        total=total, hotel_count=hotel_count, mall_count=mall_count,
    )
    html_body = _build_html_body(year, month, kpi, has_attachment=sched.include_excel_attachment)
    text_body = _build_text_body(year, month, kpi)

    attachment: Optional[tuple[str, bytes]] = None
    attachment_filename: Optional[str] = None
    if sched.include_excel_attachment:
        excel_bytes = generate_excel(cases, year, month)
        today_str   = now.strftime("%Y%m%d")
        attachment_filename = f"報修未完成報表_{year}{month:02d}_{today_str}.xlsx"
        attachment  = (attachment_filename, excel_bytes)

    send_time_str = now.strftime("%H:%M")
    sent = failed = 0

    for rcpt in recipients:
        status  = "failed"
        err_msg = None
        try:
            send_single_email(
                to_email=rcpt.email,
                to_name=rcpt.name,
                subject=subject,
                html_body=html_body,
                text_body=text_body,
                attachment=attachment,
            )
            status = "success"
            sent  += 1
            logger.info("[RepairReport] force_send_now ✓ → %s", rcpt.email)
        except Exception as exc:
            err_msg = str(exc)[:500]
            failed += 1
            logger.error("[RepairReport] force_send_now ✗ → %s: %s", rcpt.email, exc)

        db.add(RepairReportMailLog(
            send_date=today, send_time=send_time_str,
            report_year=year, report_month=month,
            recipient_email=rcpt.email, recipient_name=rcpt.name,
            subject=subject, status=status, error_message=err_msg,
            hotel_unfinished_count=hotel_count,
            mall_unfinished_count=mall_count,
            total_unfinished_count=total,
            attachment_filename=attachment_filename,
            send_source="manual",
        ))

    db.commit()
    logger.info("[RepairReport] force_send_now 完成：sent=%d failed=%d", sent, failed)
    return {"sent": sent, "failed": failed}


def _write_skipped_log(
    db: Session,
    today: date,
    now: datetime,
    report_year: int,
    report_month: int,
    reason: str,
) -> None:
    db.add(RepairReportMailLog(
        send_date=today,
        send_time=now.strftime("%H:%M"),
        report_year=report_year,
        report_month=report_month,
        recipient_email="(skipped)",
        recipient_name="",
        subject="",
        status="skipped",
        error_message=reason,
        hotel_unfinished_count=0,
        mall_unfinished_count=0,
        total_unfinished_count=0,
    ))
    db.commit()


# ── 確保預設排程設定存在 ──────────────────────────────────────────────────────

def ensure_default_schedule(db: Session) -> None:
    """若 repair_report_schedule_settings 為空，插入預設一筆（is_enabled=False）。
    同時確保 repair_report_mail_logs 有 send_source 欄位（SQLite 向後相容遷移）。
    """
    # ── SQLite 向後相容：補 send_source 欄位 ──────────────────────────────────
    try:
        db.execute(
            __import__("sqlalchemy").text(
                "ALTER TABLE repair_report_mail_logs ADD COLUMN send_source TEXT DEFAULT 'scheduled'"
            )
        )
        db.commit()
        logger.info("[RepairReport] 已為 repair_report_mail_logs 加入 send_source 欄位")
    except Exception:
        # 欄位已存在時 SQLite 會拋例外，直接忽略
        db.rollback()

    existing = db.query(RepairReportScheduleSettings).first()
    if not existing:
        db.add(RepairReportScheduleSettings())
        db.commit()
        logger.info("[RepairReport] 已建立預設排程設定（is_enabled=False）")


def get_schedule_send_time(db: Session) -> tuple[int, int]:
    """讀取排程設定的 send_time，回傳 (hour, minute)，預設 08:30。"""
    sched = db.query(RepairReportScheduleSettings).first()
    if sched and sched.send_time:
        try:
            h, m = sched.send_time.split(":")
            return int(h), int(m)
        except Exception:
            pass
    return 8, 30
