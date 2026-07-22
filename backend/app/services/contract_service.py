"""
合約管理系統 - 服務層

實現所有合約相關的業務邏輯，包括新增、編輯、查詢、刪除等核心功能。
所有的資料驗證、外鍵檢查、業務規則都在此層實現。
"""

import json
from datetime import datetime, date, timedelta
from typing import Optional, List, Dict, Any
from sqlalchemy import desc, asc, and_, func
from sqlalchemy.orm import Session

from app.core.exceptions import (
    ContractNotFound,
    ContractAlreadyExists,
    ContractVendorNotFound,
    ContractBudgetCategoryNotFound,
    InvalidContractDates,
    InvalidContractAmount,
    VendorNotFound,
    BudgetCategoryNotFound,
    BudgetCategoryDisabled,
    InvalidInputData,
)
from app.models.contract import (
    Contract,
    ContractItem,
    Vendor,
    BudgetCategory,
    ContractClaim,
    ContractCostAllocation,
    ContractPaymentSchedule,
)
from app.schemas.contract import (
    ContractCreate,
    ContractUpdate,
    ContractResponse,
    ContractDetailResponse,
    VendorCreate,
    VendorUpdate,
    VendorResponse,
    BudgetCategoryCreate,
    BudgetCategoryUpdate,
    BudgetCategoryResponse,
    ContractItemCreate,
    ContractItemUpdate,
    ContractItemResponse,
    ContractClaimCreate,
    ContractClaimUpdate,
    ContractClaimResponse,
    ContractClaimReviewRequest,
)


# ─────────────────────────────────────────────────────────────────────────────
# 合約服務層
# ─────────────────────────────────────────────────────────────────────────────

class ContractService:
    """合約管理服務"""

    @staticmethod
    def list_contracts(
        db: Session,
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = None,
        status: Optional[str] = None,
        vendor_id: Optional[str] = None,
        risk_level: Optional[str] = None,
        budget_year: Optional[int] = None,
        responsible_dept: Optional[str] = None,
        manager: Optional[str] = None,
        renewal_filter: Optional[str] = None,
        sort_by: str = "updated_at",
        sort_order: str = "desc",
    ) -> tuple[List[ContractDetailResponse], int]:
        """
        查詢合約列表（含分頁、篩選、排序）

        Args:
            db: 資料庫連線
            skip: 跳過筆數
            limit: 返回筆數
            search: 搜尋編號或名稱（模糊）
            status: 狀態篩選（草稿/簽訂中/生效中/已結束/已終止）
            vendor_id: 廠商 ID 篩選
            risk_level: 風險等級篩選（低/中/高/關鍵）
            budget_year: 預算年度篩選
            responsible_dept: 負責部門篩選
            renewal_filter: 續約鏈篩選（2026-07-21；只認明確的 renewed_from_contract_id
                             關聯，2026-07-22 撤銷編號字串規律推斷，見
                             _compute_renewal_relation_flags 說明）
                is_copy    — 只看複製續約產生的合約（renewed_from_contract_id 非空）
                has_copies — 只看已被複製續約過的合約（有其他合約的
                             renewed_from_contract_id 指向自己）
            sort_by: 排序欄位（contract_id/created_at/updated_at/start_date/total_amount_tax_included）
            sort_order: 排序順序（asc/desc）

        Returns:
            (合約列表, 總筆數)
        """
        query = db.query(Contract)

        # 應用篩選條件 — G3 全文搜尋（跨 8 個欄位）
        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                (Contract.contract_id.ilike(search_pattern))
                | (Contract.contract_name.ilike(search_pattern))
                | (Contract.vendor_name.ilike(search_pattern))
                | (Contract.responsible_dept.ilike(search_pattern))
                | (Contract.manager.ilike(search_pattern))
                | (Contract.reviewer.ilike(search_pattern))
                | (Contract.remarks.ilike(search_pattern))
                | (Contract.signing_company.ilike(search_pattern))
            )

        if status:
            query = query.filter(Contract.contract_status == status)

        if vendor_id:
            query = query.filter(Contract.vendor_id == vendor_id)

        if risk_level:
            query = query.filter(Contract.risk_level == risk_level)

        if budget_year:
            query = query.filter(Contract.budget_year == budget_year)

        if responsible_dept:
            query = query.filter(Contract.responsible_dept == responsible_dept)

        # J5 — 個人化篩選：只看我的合約
        if manager:
            query = query.filter(Contract.manager == manager)

        # 續約鏈篩選（2026-07-21）：只認明確的 renewed_from_contract_id 關聯
        relation_flags = ContractService._compute_renewal_relation_flags(db)
        if renewal_filter == "is_copy":
            matching_ids = [cid for cid, f in relation_flags.items() if f["is_copy"]]
            query = query.filter(Contract.contract_id.in_(matching_ids))
        elif renewal_filter == "has_copies":
            matching_ids = [cid for cid, f in relation_flags.items() if f["has_children"]]
            query = query.filter(Contract.contract_id.in_(matching_ids))

        # 計算總筆數
        total = query.count()

        # 排序
        if sort_by == "contract_id":
            sort_col = Contract.contract_id
        elif sort_by == "created_at":
            sort_col = Contract.created_at
        elif sort_by == "start_date":
            sort_col = Contract.start_date
        elif sort_by == "total_amount_tax_included":
            sort_col = Contract.total_amount_tax_included
        else:  # updated_at
            sort_col = Contract.updated_at

        if sort_order.lower() == "asc":
            query = query.order_by(asc(sort_col))
        else:
            query = query.order_by(desc(sort_col))

        # 分頁
        contracts = query.offset(skip).limit(limit).all()

        # 轉換為 Response Schema
        result = [ContractService._make_contract_detail_response(c) for c in contracts]

        # 標記本頁合約的續約鏈相關旗標（2026-07-21，只認明確的 FK 關聯）
        for r in result:
            f = relation_flags.get(r.contract_id) or {"is_copy": False, "has_children": False, "hint": None}
            r.is_renewal_copy = f["is_copy"]
            r.has_renewal_children = f["has_children"]
            r.renewal_related_hint = f["hint"]

        return result, total

    @staticmethod
    def _compute_renewal_relation_flags(db: Session) -> Dict[str, Dict[str, Any]]:
        """
        計算每份合約的「相關合約」旗標（供合約清單 icon／篩選使用，2026-07-21）。

        2026-07-22 撤銷編號規律推斷：Samuel 確認「續約」關聯只認手動執行「複製續約」
        產生的明確 renewed_from_contract_id 記錄，不可用合約編號字串規律猜測（實測
        COP006-2026-02／COP002-2025-2／COP005-2026-1／COP018-2026 等大量互不相關的
        合約，只因編號格式相似就被誤判為相關，不正確）。本方法現在只讀取真正的 FK
        關聯，不做任何字串比對。

        Returns:
            {contract_id: {"is_copy": bool, "has_children": bool, "hint": Optional[str]}}
            hint 為提示用的關聯合約編號（複製來源／第一個被複製出的子合約）。
        """
        rows = db.query(
            Contract.contract_id, Contract.renewed_from_contract_id
        ).all()

        flags: Dict[str, Dict[str, Any]] = {
            cid: {"is_copy": False, "has_children": False, "hint": None} for cid, _ in rows
        }

        for cid, parent_id in rows:
            if parent_id:
                flags[cid]["is_copy"] = True
                flags[cid]["hint"] = parent_id
                if parent_id in flags:
                    flags[parent_id]["has_children"] = True
                    if not flags[parent_id]["hint"]:
                        flags[parent_id]["hint"] = cid

        return flags

    @staticmethod
    def get_contract(db: Session, contract_id: str) -> ContractDetailResponse:
        """
        獲取單筆合約詳情（含所有明細）

        Args:
            db: 資料庫連線
            contract_id: 合約編號

        Returns:
            合約詳情 Response
        """
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)

        return ContractService._make_contract_detail_response(contract)

    @staticmethod
    def create_contract(db: Session, contract_data: ContractCreate) -> ContractDetailResponse:
        """
        建立新合約

        Args:
            db: 資料庫連線
            contract_data: 合約建立 Schema

        Returns:
            新建的合約詳情 Response

        Raises:
            ContractAlreadyExists: 合約編號已存在
            ContractVendorNotFound: 廠商不存在
            ContractBudgetCategoryNotFound: 預算科目不存在
            InvalidContractDates: 日期無效
            InvalidContractAmount: 金額無效
        """
        # 檢查合約編號是否已存在
        existing = db.query(Contract).filter(Contract.contract_id == contract_data.contract_id).first()
        if existing:
            raise ContractAlreadyExists(contract_data.contract_id)

        # 驗證日期
        if contract_data.start_date > contract_data.end_date:
            raise InvalidContractDates(
                str(contract_data.start_date),
                str(contract_data.end_date),
            )

        # 驗證金額
        if contract_data.total_amount_tax_included <= 0:
            raise InvalidContractAmount(contract_data.total_amount_tax_included)

        # 檢查廠商是否存在
        vendor = db.query(Vendor).filter(Vendor.vendor_id == contract_data.vendor_id).first()
        if not vendor:
            raise ContractVendorNotFound(contract_data.vendor_id)

        # 檢查預算科目是否存在且啟用
        budget_category = (
            db.query(BudgetCategory)
            .filter(
                BudgetCategory.budget_year == contract_data.budget_year,
                BudgetCategory.category_l1 == contract_data.budget_category_l1,
                BudgetCategory.category_l2 == contract_data.budget_category_l2,
            )
            .first()
        )
        if not budget_category:
            raise ContractBudgetCategoryNotFound(
                contract_data.budget_year,
                contract_data.budget_category_l1,
                contract_data.budget_category_l2,
            )

        if not budget_category.is_enabled:
            raise BudgetCategoryDisabled(
                contract_data.budget_year,
                contract_data.budget_category_l1,
                contract_data.budget_category_l2,
            )

        # 構建 detail dict（Drawer 用）
        detail = ContractService._build_contract_detail_dict(
            contract_data,
            vendor,
        )

        # 創建合約記錄
        now = datetime.now()
        contract = Contract(
            contract_id=contract_data.contract_id,
            contract_name=contract_data.contract_name,
            contract_type=contract_data.contract_type,
            contract_status=contract_data.contract_status or "草稿",
            responsible_dept=contract_data.responsible_dept,
            using_depts=contract_data.using_depts,
            vendor_id=contract_data.vendor_id,
            vendor_name=vendor.vendor_name,
            start_date=contract_data.start_date,
            end_date=contract_data.end_date,
            notification_days=contract_data.notification_days or 0,
            auto_renewal=contract_data.auto_renewal or False,
            currency=contract_data.currency or "TWD",
            total_amount_tax_included=contract_data.total_amount_tax_included,
            monthly_fixed_amount=contract_data.monthly_fixed_amount,
            pricing_method=contract_data.pricing_method,
            needs_purchase_order=contract_data.needs_purchase_order or False,
            can_claim_without_po=contract_data.can_claim_without_po or False,
            needs_allocation=contract_data.needs_allocation or False,
            allocation_method=contract_data.allocation_method,
            budget_year=contract_data.budget_year,
            budget_category_l1=contract_data.budget_category_l1,
            budget_category_l2=contract_data.budget_category_l2,
            accounting_code=budget_category.accounting_code,
            budget_source=contract_data.budget_source or "年度預算",
            budget_control_method=contract_data.budget_control_method or "提醒",
            require_acceptance=contract_data.require_acceptance or False,
            risk_level=contract_data.risk_level or "中",
            manager=contract_data.manager,
            reviewer=contract_data.reviewer,
            attachment_url=contract_data.attachment_url,
            remarks=contract_data.remarks,
            detail=json.dumps(detail, ensure_ascii=False),
            created_at=now,
            updated_at=now,
        )

        db.add(contract)
        db.commit()
        db.refresh(contract)

        return ContractService._make_contract_detail_response(contract)

    @staticmethod
    def copy_renew_contract(
        db: Session,
        source_contract_id: str,
        contract_data: ContractCreate,
    ) -> ContractDetailResponse:
        """
        複製原合約建立續約新合約（「原合約複製續約」功能，2026-07-21）。

        前端表單會預先帶入原合約所有欄位值供使用者編輯（含新合約編號），
        送出時走跟「新增合約」完全相同的查重／日期／金額／廠商／預算科目驗證，
        建立成功後：
          1. 在新合約標記 renewed_from_contract_id = 原合約編號（用於上下層級查詢）
          2. 複製原合約的子資料：合約項目 / 費用分攤 / 付款計劃

        不複製的資料（屬於原合約自己的歷史交易記錄，非可續用的範本資料）：
          請款紀錄、附件、變更歷程、稽核日誌、驗收記錄、保證金、SLA 指標與記錄。

        Args:
            db: 資料庫連線
            source_contract_id: 原合約編號
            contract_data: 新合約建立 Schema（含新合約編號，可與原合約不同）

        Raises:
            ContractNotFound: 原合約不存在
            ContractAlreadyExists: 新合約編號已存在
            ContractVendorNotFound / ContractBudgetCategoryNotFound /
            InvalidContractDates / InvalidContractAmount: 同「新增合約」驗證
        """
        source = db.query(Contract).filter(Contract.contract_id == source_contract_id).first()
        if not source:
            raise ContractNotFound(source_contract_id)

        # 沿用既有新增合約邏輯（查重、日期/金額驗證、廠商/預算科目檢查、建立主檔）
        ContractService.create_contract(db, contract_data)

        new_contract = db.query(Contract).filter(Contract.contract_id == contract_data.contract_id).first()
        new_contract.renewed_from_contract_id = source_contract_id
        db.commit()

        # 複製子資料：合約項目
        for item in db.query(ContractItem).filter(ContractItem.contract_id == source_contract_id).all():
            db.add(ContractItem(
                contract_id=new_contract.contract_id,
                item_seq=item.item_seq,
                item_name=item.item_name,
                item_category=item.item_category,
                unit_price_tax_excluded=item.unit_price_tax_excluded,
                quantity=item.quantity,
                unit=item.unit,
                tax_rate=item.tax_rate,
                amount_tax_excluded=item.amount_tax_excluded,
                amount_tax_included=item.amount_tax_included,
                is_fixed=item.is_fixed,
                is_floating=item.is_floating,
            ))

        # 複製子資料：費用分攤
        for alloc in db.query(ContractCostAllocation).filter(ContractCostAllocation.contract_id == source_contract_id).all():
            db.add(ContractCostAllocation(
                contract_id=new_contract.contract_id,
                company_name=alloc.company_name,
                allocation_type=alloc.allocation_type,
                value=alloc.value,
            ))

        # 複製子資料：付款計劃（狀態重置為「待付款」，不沿用原合約的付款記錄）
        for sched in db.query(ContractPaymentSchedule).filter(ContractPaymentSchedule.contract_id == source_contract_id).all():
            db.add(ContractPaymentSchedule(
                contract_id=new_contract.contract_id,
                milestone_name=sched.milestone_name,
                due_date=sched.due_date,
                amount=sched.amount,
                status="待付款",
                notes=sched.notes,
            ))

        db.commit()
        db.refresh(new_contract)

        return ContractService._make_contract_detail_response(new_contract)

    @staticmethod
    def get_renewal_chain(db: Session, contract_id: str) -> List[Contract]:
        """
        查詢合約的完整續約鏈（上下層級）。

        先往上溯源找到最源頭的合約（renewed_from_contract_id 為 NULL 者），
        再從源頭開始往下展開整條鏈（含分支：同一份合約可能被複製續約多次）。

        Args:
            db: 資料庫連線
            contract_id: 查詢起點合約編號

        Returns:
            依起日由舊到新排序的合約清單（至少含查詢起點本身一筆）

        Raises:
            ContractNotFound: 合約不存在
        """
        current = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not current:
            raise ContractNotFound(contract_id)

        # 往上溯源找最源頭的合約（防呆：避免資料異常造成無窮迴圈）
        root = current
        visited_up = {root.contract_id}
        while root.renewed_from_contract_id:
            if root.renewed_from_contract_id in visited_up:
                break
            parent = db.query(Contract).filter(Contract.contract_id == root.renewed_from_contract_id).first()
            if not parent:
                break
            visited_up.add(parent.contract_id)
            root = parent

        # 從最源頭開始 BFS 往下展開整條鏈（含分支）
        chain: List[Contract] = []
        seen: set = set()
        queue = [root]
        while queue:
            node = queue.pop(0)
            if node.contract_id in seen:
                continue
            seen.add(node.contract_id)
            chain.append(node)
            children = db.query(Contract).filter(Contract.renewed_from_contract_id == node.contract_id).all()
            queue.extend(children)

        chain.sort(key=lambda c: (c.start_date, c.contract_id))
        return chain

    @staticmethod
    def update_contract(
        db: Session,
        contract_id: str,
        contract_data: ContractUpdate,
    ) -> ContractDetailResponse:
        """
        編輯現有合約

        Args:
            db: 資料庫連線
            contract_id: 合約編號
            contract_data: 合約更新 Schema

        Returns:
            更新後的合約詳情 Response

        Raises:
            ContractNotFound: 合約不存在
            InvalidContractDates: 日期無效
            InvalidContractAmount: 金額無效
        """
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)

        # 驗證日期（如果提供）
        start_date = contract_data.start_date or contract.start_date
        end_date = contract_data.end_date or contract.end_date
        if start_date > end_date:
            raise InvalidContractDates(str(start_date), str(end_date))

        # 驗證金額（如果提供）
        amount = contract_data.total_amount_tax_included or contract.total_amount_tax_included
        if amount <= 0:
            raise InvalidContractAmount(amount)

        # 檢查廠商（如果變更）
        if contract_data.vendor_id and contract_data.vendor_id != contract.vendor_id:
            vendor = db.query(Vendor).filter(Vendor.vendor_id == contract_data.vendor_id).first()
            if not vendor:
                raise ContractVendorNotFound(contract_data.vendor_id)
            contract.vendor_id = contract_data.vendor_id
            contract.vendor_name = vendor.vendor_name
        elif contract_data.vendor_id:
            vendor = db.query(Vendor).filter(Vendor.vendor_id == contract.vendor_id).first()
        else:
            vendor = db.query(Vendor).filter(Vendor.vendor_id == contract.vendor_id).first()

        # 更新欄位
        update_data = contract_data.dict(exclude_unset=True)
        for key, value in update_data.items():
            if value is not None:
                setattr(contract, key, value)

        # 更新 detail dict
        contract.detail = json.dumps(
            ContractService._build_contract_detail_dict(contract_data, vendor),
            ensure_ascii=False,
        )
        contract.updated_at = datetime.now()

        db.commit()
        db.refresh(contract)

        return ContractService._make_contract_detail_response(contract)

    @staticmethod
    def delete_contract(db: Session, contract_id: str) -> None:
        """
        邏輯刪除合約

        Args:
            db: 資料庫連線
            contract_id: 合約編號

        Raises:
            ContractNotFound: 合約不存在
        """
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)

        # 邏輯刪除：標記為已終止
        contract.contract_status = "已終止"
        contract.updated_at = datetime.now()

        db.commit()

    @staticmethod
    def submit_for_review(db: Session, contract_id: str, submitter: str) -> "ContractDetailResponse":
        """
        將草稿合約送審（草稿 → 審核中）。

        Args:
            db: 資料庫連線
            contract_id: 合約編號
            submitter: 送審人（顯示用，非強制入庫）

        Raises:
            ContractNotFound: 合約不存在
            InvalidInputData: 狀態不符（非草稿）
        """
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)
        if contract.contract_status != "草稿":
            raise InvalidInputData("contract_status", f"目前狀態「{contract.contract_status}」不可送審，僅限「草稿」")

        contract.contract_status = "審核中"
        contract.updated_at = datetime.now()
        db.commit()
        db.refresh(contract)
        return ContractService._to_detail(db, contract)

    @staticmethod
    def approve_contract(db: Session, contract_id: str, approver: str, comment: Optional[str] = None) -> "ContractDetailResponse":
        """
        核准合約（審核中 → 生效中）。

        Args:
            db: 資料庫連線
            contract_id: 合約編號
            approver: 審核人姓名
            comment: 審核意見（可選）
        """
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)
        if contract.contract_status != "審核中":
            raise InvalidInputData("contract_status", f"目前狀態「{contract.contract_status}」不可核准，僅限「審核中」")

        now = datetime.now()
        contract.contract_status = "生效中"
        contract.approved_by = approver
        contract.approved_at = now
        contract.approval_comment = comment
        contract.updated_at = now
        db.commit()
        db.refresh(contract)
        return ContractService._to_detail(db, contract)

    @staticmethod
    def reject_contract(db: Session, contract_id: str, rejector: str, comment: Optional[str] = None) -> "ContractDetailResponse":
        """
        拒絕合約（審核中 → 草稿，清除審核資訊）。

        Args:
            db: 資料庫連線
            contract_id: 合約編號
            rejector: 拒絕人姓名
            comment: 拒絕原因（可選）
        """
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)
        if contract.contract_status != "審核中":
            raise InvalidInputData("contract_status", f"目前狀態「{contract.contract_status}」不可拒絕，僅限「審核中」")

        contract.contract_status = "草稿"
        contract.approved_by = rejector
        contract.approved_at = None
        contract.approval_comment = comment
        contract.updated_at = datetime.now()
        db.commit()
        db.refresh(contract)
        return ContractService._to_detail(db, contract)

    @staticmethod
    def _to_detail(db: Session, contract: Contract) -> "ContractDetailResponse":
        """將 ORM Contract 轉為 ContractDetailResponse（含 detail dict）。"""
        import json as _json
        vendor = db.query(Vendor).filter(Vendor.vendor_id == contract.vendor_id).first()
        try:
            detail_dict = _json.loads(contract.detail) if contract.detail else {}
        except Exception:
            detail_dict = {}
        resp = ContractDetailResponse.from_orm(contract)
        resp.detail = detail_dict
        return resp

    # ─────────────────────────────────────────────────────────────────────────
    # 輔助函數
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def _build_contract_detail_dict(
        contract_data: Any,
        vendor: Vendor,
    ) -> Dict[str, str]:
        """構建 detail dict（用於 Drawer 顯示）"""
        detail = {
            "編號": contract_data.contract_id if hasattr(contract_data, "contract_id") else "",
            "合約名稱": contract_data.contract_name if hasattr(contract_data, "contract_name") else "",
            "合約類型": contract_data.contract_type if hasattr(contract_data, "contract_type") else "",
            "廠商": vendor.vendor_name if vendor else "",
            "廠商ID": vendor.vendor_id if vendor else "",
            "統一編號": vendor.tax_id if vendor else "",
            "聯絡人": vendor.contact_person if vendor else "—",
            "電話": vendor.phone if vendor else "—",
            "Email": vendor.email if vendor else "—",
            "銀行": vendor.bank_name if vendor else "—",
            "帳號": vendor.bank_account if vendor else "—",
            "起日": str(contract_data.start_date) if hasattr(contract_data, "start_date") else "",
            "迄日": str(contract_data.end_date) if hasattr(contract_data, "end_date") else "",
            "總額": str(contract_data.total_amount_tax_included) if hasattr(contract_data, "total_amount_tax_included") else "0",
            "計價方式": contract_data.pricing_method if hasattr(contract_data, "pricing_method") else "",
            "預算年度": str(contract_data.budget_year) if hasattr(contract_data, "budget_year") else "",
            "科目大項": contract_data.budget_category_l1 if hasattr(contract_data, "budget_category_l1") else "",
            "科目細項": contract_data.budget_category_l2 if hasattr(contract_data, "budget_category_l2") else "",
            "負責部門": contract_data.responsible_dept if hasattr(contract_data, "responsible_dept") else "",
            "風險等級": contract_data.risk_level if hasattr(contract_data, "risk_level") else "中",
            "備註": contract_data.remarks if hasattr(contract_data, "remarks") else "—",
        }
        return detail

    @staticmethod
    def _make_contract_detail_response(contract: Contract) -> ContractDetailResponse:
        """將 ORM Contract 對象轉換為 ContractDetailResponse"""
        return ContractDetailResponse(
            contract_id=contract.contract_id,
            contract_name=contract.contract_name,
            contract_type=contract.contract_type,
            contract_status=contract.contract_status,
            responsible_dept=contract.responsible_dept,
            using_depts=contract.using_depts,
            vendor_id=contract.vendor_id,
            vendor_name=contract.vendor_name,
            start_date=contract.start_date,
            end_date=contract.end_date,
            notification_days=contract.notification_days,
            auto_renewal=contract.auto_renewal,
            currency=contract.currency,
            total_amount_tax_included=contract.total_amount_tax_included,
            monthly_fixed_amount=contract.monthly_fixed_amount,
            pricing_method=contract.pricing_method,
            needs_purchase_order=contract.needs_purchase_order,
            can_claim_without_po=contract.can_claim_without_po,
            needs_allocation=contract.needs_allocation,
            allocation_method=contract.allocation_method,
            budget_year=contract.budget_year,
            budget_category_l1=contract.budget_category_l1,
            budget_category_l2=contract.budget_category_l2,
            accounting_code=contract.accounting_code,
            budget_source=contract.budget_source,
            budget_control_method=contract.budget_control_method,
            require_acceptance=contract.require_acceptance,
            risk_level=contract.risk_level,
            manager=contract.manager,
            reviewer=contract.reviewer,
            attachment_url=contract.attachment_url,
            remarks=contract.remarks,
            detail=(
                json.loads(contract.detail)
                if isinstance(contract.detail, str) else (contract.detail or {})
            ),
            created_at=contract.created_at,
            updated_at=contract.updated_at,
        )


    # ─────────────────────────────────────────────────────────────────────────
    # 統計 / Dashboard 方法
    # ─────────────────────────────────────────────────────────────────────────

    @staticmethod
    def get_stats(db: Session) -> Dict[str, Any]:
        """合約統計：總數、各狀態數、各風險等級數"""
        total = db.query(func.count(Contract.contract_id)).scalar() or 0

        by_status: Dict[str, int] = {}
        for status_val, cnt in db.query(Contract.contract_status, func.count(Contract.contract_id)).group_by(Contract.contract_status).all():
            by_status[status_val or "未知"] = cnt

        by_risk: Dict[str, int] = {}
        for risk_val, cnt in db.query(Contract.risk_level, func.count(Contract.contract_id)).group_by(Contract.risk_level).all():
            by_risk[risk_val or "未設定"] = cnt

        return {"total": total, "by_status": by_status, "by_risk_level": by_risk}

    @staticmethod
    def get_dashboard_kpi(db: Session, budget_year: Optional[int] = None) -> Dict[str, Any]:
        """Dashboard KPI 指標"""
        today = date.today()
        year = budget_year or today.year

        active_contracts = db.query(func.count(Contract.contract_id)).filter(
            Contract.contract_status == "生效中"
        ).scalar() or 0

        total_annual_amount = db.query(func.sum(Contract.total_amount_tax_included)).filter(
            Contract.budget_year == year,
            Contract.contract_status.in_(["生效中", "草稿", "簽訂中"])
        ).scalar() or 0.0

        high_risk_count = db.query(func.count(Contract.contract_id)).filter(
            Contract.risk_level.in_(["高", "關鍵"]),
            Contract.contract_status == "生效中"
        ).scalar() or 0

        expiring_in_90days = db.query(func.count(Contract.contract_id)).filter(
            Contract.contract_status == "生效中",
            Contract.end_date >= today,
            Contract.end_date <= today + timedelta(days=90)
        ).scalar() or 0

        return {
            "active_contracts": active_contracts,
            "total_annual_amount": float(total_annual_amount),
            "high_risk_count": high_risk_count,
            "expiring_in_90days": expiring_in_90days,
            "monthly_claim_amount": 0.0,  # Phase 3
            "accrual_amount": 0.0,        # Phase 3
            "budget_year": year,
        }

    @staticmethod
    def get_by_dept(
        db: Session,
        budget_year: Optional[int] = None,
        company: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Dashboard 部門金額分組（可依 signing_company 篩選）"""
        year = budget_year or date.today().year

        q = db.query(
            Contract.responsible_dept,
            func.sum(Contract.total_amount_tax_included).label("amount"),
            func.count(Contract.contract_id).label("count"),
        ).filter(Contract.budget_year == year)

        if company:
            q = q.filter(Contract.signing_company == company)

        rows = q.group_by(Contract.responsible_dept).all()

        items = [
            {"dept": row.responsible_dept or "未分類", "amount": float(row.amount or 0), "count": row.count}
            for row in rows
        ]
        return {"budget_year": year, "items": items, "company": company or ""}

    @staticmethod
    def get_expiring(db: Session, days: int = 90) -> Dict[str, Any]:
        """
        取得即將到期合約清單。
        到期日在今天 ~ 今天+days 之間且狀態不是「已終止」。
        依剩餘天數升冪排列。
        """
        today = date.today()
        deadline = today + timedelta(days=days)

        rows = db.query(Contract).filter(
            Contract.end_date >= today,
            Contract.end_date <= deadline,
            Contract.contract_status != "已終止",
        ).order_by(asc(Contract.end_date)).all()

        items = []
        for c in rows:
            end = c.end_date
            # end_date 可能是 date 或 str
            if isinstance(end, str):
                end = date.fromisoformat(end)
            remaining = (end - today).days if end else None
            items.append({
                "contract_id":               c.contract_id,
                "contract_name":             c.contract_name,
                "contract_type":             c.contract_type,
                "contract_status":           c.contract_status,
                "responsible_dept":          c.responsible_dept,
                "vendor_name":               c.vendor.vendor_name if c.vendor else "",
                "end_date":                  str(c.end_date) if c.end_date else None,
                "remaining_days":            remaining,
                "total_amount_tax_included": float(c.total_amount_tax_included or 0),
                "risk_level":                c.risk_level,
                "manager":                   c.manager,
            })

        return {"total": len(items), "days": days, "items": items}


# ─────────────────────────────────────────────────────────────────────────────
# 廠商服務層
# ─────────────────────────────────────────────────────────────────────────────

class VendorService:
    """廠商管理服務"""

    @staticmethod
    def list_vendors(
        db: Session,
        skip: int = 0,
        limit: int = 100,
        search: Optional[str] = None,
        vendor_type: Optional[str] = None,
        risk_level: Optional[str] = None,
        is_critical: Optional[bool] = None,
    ) -> tuple[List[VendorResponse], int]:
        """
        查詢廠商列表

        Args:
            db: 資料庫連線
            skip: 跳過筆數
            limit: 返回筆數
            search: 搜尋編號或名稱（模糊）
            vendor_type: 廠商類型篩選
            risk_level: 風險等級篩選
            is_critical: 是否關鍵廠商篩選

        Returns:
            (廠商列表, 總筆數)
        """
        query = db.query(Vendor)

        if search:
            search_pattern = f"%{search}%"
            query = query.filter(
                (Vendor.vendor_id.ilike(search_pattern))
                | (Vendor.vendor_name.ilike(search_pattern))
            )

        if vendor_type:
            query = query.filter(Vendor.vendor_type == vendor_type)

        if risk_level:
            query = query.filter(Vendor.risk_level == risk_level)

        if is_critical is not None:
            query = query.filter(Vendor.is_critical == is_critical)

        total = query.count()
        vendors = query.order_by(Vendor.created_at.desc()).offset(skip).limit(limit).all()

        result = [VendorResponse.from_orm(v) for v in vendors]
        return result, total

    @staticmethod
    def get_vendor(db: Session, vendor_id: str) -> VendorResponse:
        """獲取單筆廠商詳情"""
        vendor = db.query(Vendor).filter(Vendor.vendor_id == vendor_id).first()
        if not vendor:
            raise VendorNotFound(vendor_id)

        return VendorResponse.from_orm(vendor)

    @staticmethod
    def create_vendor(db: Session, vendor_data: VendorCreate) -> VendorResponse:
        """建立新廠商"""
        existing = db.query(Vendor).filter(Vendor.vendor_id == vendor_data.vendor_id).first()
        if existing:
            raise InvalidInputData("vendor_id", "廠商編號已存在")

        existing_name = db.query(Vendor).filter(Vendor.vendor_name == vendor_data.vendor_name).first()
        if existing_name:
            raise InvalidInputData("vendor_name", "廠商名稱已存在")

        now = datetime.now()
        vendor = Vendor(
            vendor_id=vendor_data.vendor_id,
            vendor_name=vendor_data.vendor_name,
            tax_id=vendor_data.tax_id,
            contact_person=vendor_data.contact_person,
            phone=vendor_data.phone,
            email=vendor_data.email,
            address=vendor_data.address,
            payment_terms=vendor_data.payment_terms,
            bank_name=vendor_data.bank_name,
            bank_account=vendor_data.bank_account,
            vendor_type=vendor_data.vendor_type,
            risk_level=vendor_data.risk_level or "中",
            is_critical=vendor_data.is_critical or False,
            created_at=now,
            updated_at=now,
        )

        db.add(vendor)
        db.commit()
        db.refresh(vendor)

        return VendorResponse.from_orm(vendor)

    @staticmethod
    def update_vendor(
        db: Session,
        vendor_id: str,
        vendor_data: VendorUpdate,
    ) -> VendorResponse:
        """編輯廠商"""
        vendor = db.query(Vendor).filter(Vendor.vendor_id == vendor_id).first()
        if not vendor:
            raise VendorNotFound(vendor_id)

        # 檢查名稱是否被使用
        if vendor_data.vendor_name and vendor_data.vendor_name != vendor.vendor_name:
            existing = db.query(Vendor).filter(Vendor.vendor_name == vendor_data.vendor_name).first()
            if existing:
                raise InvalidInputData("vendor_name", "廠商名稱已存在")

        update_data = vendor_data.dict(exclude_unset=True)
        for key, value in update_data.items():
            if value is not None:
                setattr(vendor, key, value)

        vendor.updated_at = datetime.now()
        db.commit()
        db.refresh(vendor)

        return VendorResponse.from_orm(vendor)

    @staticmethod
    def delete_vendor(db: Session, vendor_id: str) -> None:
        """刪除廠商（檢查是否有關聯合約）"""
        vendor = db.query(Vendor).filter(Vendor.vendor_id == vendor_id).first()
        if not vendor:
            raise VendorNotFound(vendor_id)

        contract_count = db.query(Contract).filter(Contract.vendor_id == vendor_id).count()
        if contract_count > 0:
            raise InvalidInputData(
                "vendor_id",
                f"廠商有 {contract_count} 份合約，無法刪除"
            )

        db.delete(vendor)
        db.commit()

    @staticmethod
    def import_vendors_from_excel(db: Session, file_bytes: bytes) -> dict:
        """
        從 Excel 檔批次匯入廠商（upsert）。

        Excel 欄位順序（第1列為標題，從第2列起為資料）：
          A 廠商編號  B 廠商名稱  C 統一編號  D 聯絡人  E 聯絡電話
          F Email    G 地址      H 付款條件  I 銀行名稱 J 銀行帳號
          K 廠商類別  L 風險等級  M 關鍵供應商（是/否）

        回傳 dict：{total_rows, created, updated, skipped, errors:[{row, vendor_id, message}]}
        """
        import io as _io
        import openpyxl as _xl
        import re as _re

        wb = _xl.load_workbook(filename=_io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active

        COLUMNS = [
            "廠商編號", "廠商名稱", "統一編號", "聯絡人", "聯絡電話",
            "Email", "地址", "付款條件", "銀行名稱", "銀行帳號",
            "廠商類別", "風險等級", "關鍵供應商",
        ]

        result = {"total_rows": 0, "created": 0, "updated": 0, "skipped": 0, "errors": []}
        now = datetime.now()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # 跳過全空列
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            result["total_rows"] += 1

            def _cell(col: int) -> str:
                v = row[col] if col < len(row) else None
                return str(v).strip() if v is not None else ""

            vendor_id   = _cell(0)
            vendor_name = _cell(1)

            # 基本驗證
            if not vendor_id:
                result["errors"].append({"row": row_idx, "vendor_id": "", "message": "廠商編號不可空白"})
                continue
            if not _re.match(r"^VND-\d{4}$", vendor_id):
                result["errors"].append({"row": row_idx, "vendor_id": vendor_id, "message": "廠商編號格式錯誤（應為 VND-NNNN）"})
                continue
            if not vendor_name:
                result["errors"].append({"row": row_idx, "vendor_id": vendor_id, "message": "廠商名稱不可空白"})
                continue

            risk_level_raw = _cell(11)
            is_critical_raw = _cell(12).upper()
            is_critical = is_critical_raw in ("是", "Y", "YES", "TRUE", "1")

            existing = db.query(Vendor).filter(Vendor.vendor_id == vendor_id).first()

            if existing:
                # 更新
                existing.vendor_name    = vendor_name
                existing.tax_id         = _cell(2) or existing.tax_id
                existing.contact_person = _cell(3) or existing.contact_person
                existing.phone          = _cell(4) or existing.phone
                existing.email          = _cell(5) or existing.email
                existing.address        = _cell(6) or existing.address
                existing.payment_terms  = _cell(7) or existing.payment_terms
                existing.bank_name      = _cell(8) or existing.bank_name
                existing.bank_account   = _cell(9) or existing.bank_account
                existing.vendor_type    = _cell(10) or existing.vendor_type
                existing.risk_level     = risk_level_raw or existing.risk_level
                existing.is_critical    = is_critical
                existing.updated_at     = now
                result["updated"] += 1
            else:
                # 檢查名稱衝突
                name_conflict = db.query(Vendor).filter(Vendor.vendor_name == vendor_name).first()
                if name_conflict:
                    result["errors"].append({"row": row_idx, "vendor_id": vendor_id, "message": f"廠商名稱「{vendor_name}」已被 {name_conflict.vendor_id} 使用"})
                    continue

                new_vendor = Vendor(
                    vendor_id       = vendor_id,
                    vendor_name     = vendor_name,
                    tax_id          = _cell(2),
                    contact_person  = _cell(3) or None,
                    phone           = _cell(4) or None,
                    email           = _cell(5) or None,
                    address         = _cell(6) or None,
                    payment_terms   = _cell(7) or None,
                    bank_name       = _cell(8) or None,
                    bank_account    = _cell(9) or None,
                    vendor_type     = _cell(10) or None,
                    risk_level      = risk_level_raw or "中",
                    is_critical     = is_critical,
                    created_at      = now,
                    updated_at      = now,
                )
                db.add(new_vendor)
                result["created"] += 1

        try:
            db.commit()
        except Exception as exc:
            db.rollback()
            raise InvalidInputData("file", f"資料庫寫入失敗：{exc}") from exc

        return result

    @staticmethod
    def get_vendor_performance(db: Session, vendor_id: str) -> dict:
        """
        計算廠商績效指標（基於 ContractClaim 紀錄）。

        指標定義：
          - 準時率：（已核准 + 已付款）/ 總請款筆數
          - 爭議率：已拒絕 / 總請款筆數
          - 平均處理天數：已完結請款（非待審核）的 (updated_at - created_at).days 平均
          - 評分等級：A / B / C / D（依準時率＋爭議率組合判斷）
        """
        vendor = db.query(Vendor).filter(Vendor.vendor_id == vendor_id).first()
        if not vendor:
            raise VendorNotFound(vendor_id)

        # 取得該廠商所有合約 ID
        contract_ids = [
            row[0]
            for row in db.query(Contract.contract_id)
            .filter(Contract.vendor_id == vendor_id)
            .all()
        ]

        if not contract_ids:
            return _empty_performance(vendor)

        # 取得所有請款紀錄
        claims = (
            db.query(ContractClaim)
            .filter(ContractClaim.contract_id.in_(contract_ids))
            .all()
        )

        total = len(claims)
        if total == 0:
            return _empty_performance(vendor)

        approved = sum(1 for c in claims if c.status in ("已核准", "已付款"))
        rejected = sum(1 for c in claims if c.status == "已拒絕")
        paid     = sum(1 for c in claims if c.status == "已付款")
        pending  = sum(1 for c in claims if c.status == "待審核")

        ontime_rate  = round(approved / total, 4)
        dispute_rate = round(rejected / total, 4)

        # 平均處理天數（已完結才算）
        resolved = [c for c in claims if c.status != "待審核"]
        if resolved:
            total_days = sum(
                max((c.updated_at - c.created_at).days, 0) for c in resolved
            )
            avg_days = round(total_days / len(resolved), 1)
        else:
            avg_days = None

        # 評分等級
        if ontime_rate >= 0.85 and dispute_rate <= 0.05:
            grade = "A"
        elif ontime_rate >= 0.70 and dispute_rate <= 0.15:
            grade = "B"
        elif ontime_rate >= 0.50 and dispute_rate <= 0.30:
            grade = "C"
        else:
            grade = "D"

        # 金額統計
        total_amount   = float(sum(c.amount or 0 for c in claims))
        approved_amount = float(sum(c.amount or 0 for c in claims if c.status in ("已核准", "已付款")))
        paid_amount    = float(sum(c.amount or 0 for c in claims if c.status == "已付款"))

        return {
            "vendor_id":        vendor.vendor_id,
            "vendor_name":      vendor.vendor_name,
            "total_claims":     total,
            "approved_count":   approved,
            "rejected_count":   rejected,
            "paid_count":       paid,
            "pending_count":    pending,
            "ontime_rate":      ontime_rate,
            "dispute_rate":     dispute_rate,
            "avg_process_days": avg_days,
            "grade":            grade,
            "total_amount":     total_amount,
            "approved_amount":  approved_amount,
            "paid_amount":      paid_amount,
            "contract_count":   len(contract_ids),
        }


def _empty_performance(vendor: Vendor) -> dict:
    return {
        "vendor_id":        vendor.vendor_id,
        "vendor_name":      vendor.vendor_name,
        "total_claims":     0,
        "approved_count":   0,
        "rejected_count":   0,
        "paid_count":       0,
        "pending_count":    0,
        "ontime_rate":      None,
        "dispute_rate":     None,
        "avg_process_days": None,
        "grade":            None,
        "total_amount":     0.0,
        "approved_amount":  0.0,
        "paid_amount":      0.0,
        "contract_count":   0,
    }


# ─────────────────────────────────────────────────────────────────────────────
# 預算科目服務層
# ─────────────────────────────────────────────────────────────────────────────

class BudgetCategoryService:
    """預算科目管理服務"""

    @staticmethod
    def list_budget_categories(
        db: Session,
        skip: int = 0,
        limit: int = 100,
        budget_year: Optional[int] = None,
        dept: Optional[str] = None,
        category_l1: Optional[str] = None,
        is_enabled: Optional[bool] = None,
    ):
        """查詢預算科目列表"""
        query = db.query(BudgetCategory)

        if budget_year:
            query = query.filter(BudgetCategory.budget_year == budget_year)

        if dept:
            query = query.filter(BudgetCategory.dept == dept)

        if category_l1:
            query = query.filter(BudgetCategory.category_l1 == category_l1)

        if is_enabled is not None:
            query = query.filter(BudgetCategory.is_enabled == is_enabled)

        total = query.count()
        categories = (
            query.order_by(BudgetCategory.budget_year.desc(), BudgetCategory.created_at.desc())
            .offset(skip)
            .limit(limit)
            .all()
        )

        result = [BudgetCategoryResponse.from_orm(c) for c in categories]
        return result, total

    @staticmethod
    def get_budget_category(
        db: Session,
        budget_year: int,
        category_l1: str,
        category_l2: str,
    ) -> BudgetCategoryResponse:
        """獲取單筆預算科目"""
        category = (
            db.query(BudgetCategory)
            .filter(
                BudgetCategory.budget_year == budget_year,
                BudgetCategory.category_l1 == category_l1,
                BudgetCategory.category_l2 == category_l2,
            )
            .first()
        )
        if not category:
            raise BudgetCategoryNotFound(budget_year, category_l1, category_l2)

        return BudgetCategoryResponse.from_orm(category)

    @staticmethod
    def create_budget_category(
        db: Session,
        category_data: BudgetCategoryCreate,
    ) -> BudgetCategoryResponse:
        """建立預算科目"""
        existing = (
            db.query(BudgetCategory)
            .filter(
                BudgetCategory.budget_year == category_data.budget_year,
                BudgetCategory.category_l1 == category_data.category_l1,
                BudgetCategory.category_l2 == category_data.category_l2,
                BudgetCategory.dept == category_data.dept,
            )
            .first()
        )
        if existing:
            raise InvalidInputData(
                "budget_category",
                f"預算科目已存在：{category_data.budget_year} / {category_data.category_l1} / {category_data.category_l2} / {category_data.dept}"
            )


        now = datetime.now()
        category = BudgetCategory(
            budget_year=category_data.budget_year,
            dept=category_data.dept,
            category_l1=category_data.category_l1,
            category_l2=category_data.category_l2,
            accounting_code=category_data.accounting_code,
            payment_code=category_data.payment_code,
            is_enabled=category_data.is_enabled or True,
            effective_date=category_data.effective_date,
            disabled_date=category_data.disabled_date,
            maintain_unit=category_data.maintain_unit,
            created_at=now,
            updated_at=now,
        )

        db.add(category)
        db.commit()
        db.refresh(category)

        return BudgetCategoryResponse.from_orm(category)

    @staticmethod
    def update_budget_category(
        db: Session,
        category_id: int,
        data: BudgetCategoryUpdate,
    ) -> BudgetCategoryResponse:
        """更新預算科目"""
        category = db.query(BudgetCategory).filter(BudgetCategory.id == category_id).first()
        if not category:
            raise BudgetCategoryNotFound(0, "ID", str(category_id))

        update_fields = data.dict(exclude_unset=True)
        for key, value in update_fields.items():
            setattr(category, key, value)
        category.updated_at = datetime.now()

        db.commit()
        db.refresh(category)
        return BudgetCategoryResponse.from_orm(category)

    @staticmethod
    def delete_budget_category(db: Session, category_id: int) -> None:
        """刪除預算科目"""
        category = db.query(BudgetCategory).filter(BudgetCategory.id == category_id).first()
        if not category:
            raise BudgetCategoryNotFound(0, "ID", str(category_id))
        db.delete(category)
        db.commit()


# ─────────────────────────────────────────────────────────────────────────────
# 請款 / 核銷服務層
# ─────────────────────────────────────────────────────────────────────────────

from app.models.contract import ContractClaim  # noqa: E402
from app.schemas.contract import (             # noqa: E402
    ContractClaimCreate,
    ContractClaimUpdate,
    ContractClaimResponse,
    ContractClaimReviewRequest,
)


class ClaimService:
    """請款 / 核銷記錄管理服務"""

    @staticmethod
    def get_stats(db: Session) -> Dict[str, Any]:
        """請款統計：各狀態筆數/金額、當月請款金額"""
        from decimal import Decimal
        from calendar import monthrange

        today = date.today()
        month_start = today.replace(day=1).strftime("%Y-%m-%d")
        _, last_day = monthrange(today.year, today.month)
        month_end = today.replace(day=last_day).strftime("%Y-%m-%d")

        # 各狀態筆數
        by_status: Dict[str, int] = {}
        by_status_amount: Dict[str, float] = {}
        for status_val, cnt, amt in db.query(
            ContractClaim.status,
            func.count(ContractClaim.id),
            func.sum(ContractClaim.amount),
        ).group_by(ContractClaim.status).all():
            key = status_val or "未知"
            by_status[key] = cnt
            by_status_amount[key] = float(amt or 0)

        # 當月請款金額
        monthly_result = db.query(func.sum(ContractClaim.amount)).filter(
            ContractClaim.claim_date >= month_start,
            ContractClaim.claim_date <= month_end,
        ).scalar()
        monthly_amount = float(monthly_result or 0)

        # 待審核筆數
        pending_count = by_status.get("待審核", 0)

        # 總計
        total_claims = sum(by_status.values())
        total_amount = sum(by_status_amount.values())

        return {
            "total_claims": total_claims,
            "total_amount": total_amount,
            "monthly_amount": monthly_amount,
            "pending_count": pending_count,
            "by_status": by_status,
            "by_status_amount": by_status_amount,
        }

    @staticmethod
    def list_claims(
        db: Session,
        contract_id: Optional[str] = None,
        status: Optional[str] = None,
        skip: int = 0,
        limit: int = 50,
    ) -> tuple[List[ContractClaimResponse], int]:
        q = db.query(ContractClaim)
        if contract_id:
            q = q.filter(ContractClaim.contract_id == contract_id)
        if status:
            q = q.filter(ContractClaim.status == status)
        total = q.count()
        rows = q.order_by(desc(ContractClaim.claim_date)).offset(skip).limit(limit).all()

        # 批次取合約名稱（避免 N+1）
        cids = list({r.contract_id for r in rows})
        name_map: Dict[str, str] = {}
        if cids:
            contracts = db.query(Contract.contract_id, Contract.contract_name).filter(
                Contract.contract_id.in_(cids)
            ).all()
            name_map = {c.contract_id: c.contract_name for c in contracts}

        result = []
        for r in rows:
            resp = ContractClaimResponse.from_orm(r)
            resp = resp.model_copy(update={"contract_name": name_map.get(r.contract_id, "")})
            result.append(resp)
        return result, total

    @staticmethod
    def create_claim(db: Session, data: ContractClaimCreate) -> ContractClaimResponse:
        # 確認合約存在
        contract = db.query(Contract).filter(Contract.contract_id == data.contract_id).first()
        if not contract:
            raise ContractNotFound(data.contract_id)
        now = datetime.now()
        claim = ContractClaim(
            contract_id=data.contract_id,
            claim_type=data.claim_type or "請款",
            claim_date=data.claim_date,
            invoice_no=data.invoice_no,
            amount=data.amount,
            status=data.status or "待審核",
            approver=data.approver,
            remarks=data.remarks,
            created_at=now,
            updated_at=now,
        )
        db.add(claim)
        db.commit()
        db.refresh(claim)
        return ContractClaimResponse.from_orm(claim)

    @staticmethod
    def get_claim(db: Session, claim_id: int) -> ContractClaimResponse:
        claim = db.query(ContractClaim).filter(ContractClaim.id == claim_id).first()
        if not claim:
            raise InvalidInputData("claim_id", f"請款記錄不存在：{claim_id}")
        return ContractClaimResponse.from_orm(claim)

    @staticmethod
    def update_claim(
        db: Session, claim_id: int, data: ContractClaimUpdate
    ) -> ContractClaimResponse:
        claim = db.query(ContractClaim).filter(ContractClaim.id == claim_id).first()
        if not claim:
            raise InvalidInputData("claim_id", f"請款記錄不存在：{claim_id}")
        for key, value in data.dict(exclude_unset=True).items():
            setattr(claim, key, value)
        claim.updated_at = datetime.now()
        db.commit()
        db.refresh(claim)
        return ContractClaimResponse.from_orm(claim)

    @staticmethod
    def delete_claim(db: Session, claim_id: int) -> None:
        claim = db.query(ContractClaim).filter(ContractClaim.id == claim_id).first()
        if not claim:
            raise InvalidInputData("claim_id", f"請款記錄不存在：{claim_id}")
        db.delete(claim)
        db.commit()

    @staticmethod
    def review_claim(
        db: Session,
        claim_id: int,
        data: ContractClaimReviewRequest,
        actor: str,
    ) -> ContractClaimResponse:
        """執行審核動作並寫入稽核軌跡"""
        from fastapi import HTTPException

        claim = db.query(ContractClaim).filter(ContractClaim.id == claim_id).first()
        if not claim:
            raise InvalidInputData("claim_id", f"請款記錄不存在：{claim_id}")

        # ── 狀態機 ─────────────────────────────────────────────────────
        TRANSITIONS = {
            "approve":    ("待審核",  "已核准"),
            "reject":     ("待審核",  "已拒絕"),
            "mark_paid":  ("已核准",  "已付款"),
            "resubmit":   ("已拒絕",  "待審核"),
        }
        if data.action not in TRANSITIONS:
            raise HTTPException(status_code=400, detail=f"不支援的審核動作：{data.action}")

        required_from, to_status = TRANSITIONS[data.action]
        if claim.status != required_from:
            raise HTTPException(
                status_code=422,
                detail=f"目前狀態「{claim.status}」不允許執行「{data.action}」操作（需為「{required_from}」）"
            )

        # ── 寫入稽核軌跡 ───────────────────────────────────────────────
        from_status = claim.status
        log_entry = {
            "action":      data.action,
            "actor":       actor,
            "from_status": from_status,
            "to_status":   to_status,
            "comment":     data.comment or "",
            "timestamp":   datetime.now().isoformat(timespec="seconds"),
        }
        existing_log: list = json.loads(claim.review_log or "[]")
        existing_log.append(log_entry)

        # ── 更新欄位 ───────────────────────────────────────────────────
        claim.status = to_status
        claim.review_log = json.dumps(existing_log, ensure_ascii=False)
        if data.action == "approve" and data.approver:
            claim.approver = data.approver
        elif data.action == "approve" and actor:
            claim.approver = actor
        claim.updated_at = datetime.now()
        db.commit()
        db.refresh(claim)

        resp = ContractClaimResponse.from_orm(claim)
        # 補 contract_name
        contract = db.query(Contract.contract_name).filter(
            Contract.contract_id == claim.contract_id
        ).first()
        resp = resp.model_copy(update={"contract_name": contract.contract_name if contract else ""})
        return resp

    @staticmethod
    def batch_review_claims(
        db: Session,
        data: "ContractClaimBatchReviewRequest",
        actor: str,
    ) -> dict:
        """批次審核（approve / reject only）"""
        from fastapi import HTTPException

        TRANSITIONS = {
            "approve": ("待審核", "已核准"),
            "reject":  ("待審核", "已拒絕"),
        }
        if data.action not in TRANSITIONS:
            raise HTTPException(status_code=400, detail=f"批次審核不支援動作：{data.action}")

        required_from, to_status = TRANSITIONS[data.action]
        success_ids, skipped_ids = [], []

        for cid in data.claim_ids:
            claim = db.query(ContractClaim).filter(ContractClaim.id == cid).first()
            if not claim or claim.status != required_from:
                skipped_ids.append(cid)
                continue
            log_entry = {
                "action":      data.action,
                "actor":       actor,
                "from_status": claim.status,
                "to_status":   to_status,
                "comment":     data.comment or "",
                "timestamp":   datetime.now().isoformat(timespec="seconds"),
            }
            existing_log: list = json.loads(claim.review_log or "[]")
            existing_log.append(log_entry)
            claim.status     = to_status
            claim.review_log = json.dumps(existing_log, ensure_ascii=False)
            if data.action == "approve":
                claim.approver = data.approver or actor
            claim.updated_at = datetime.now()
            success_ids.append(cid)

        db.commit()
        return {"success_count": len(success_ids), "skipped_count": len(skipped_ids),
                "success_ids": success_ids, "skipped_ids": skipped_ids}


# ─────────────────────────────────────────────────────────────────────────────
# 合約項目服務層
# ─────────────────────────────────────────────────────────────────────────────

class ContractItemService:
    """合約項目（Line Items）管理"""

    @staticmethod
    def list_items(db: Session, contract_id: str) -> List[ContractItemResponse]:
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)
        items = (
            db.query(ContractItem)
            .filter(ContractItem.contract_id == contract_id)
            .order_by(ContractItem.item_seq, ContractItem.id)
            .all()
        )
        return [ContractItemResponse.from_orm(i) for i in items]

    @staticmethod
    def create_item(
        db: Session, contract_id: str, data: ContractItemCreate
    ) -> ContractItemResponse:
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ContractNotFound(contract_id)
        # 自動指派 item_seq
        if data.item_seq is None:
            max_seq = (
                db.query(func.max(ContractItem.item_seq))
                .filter(ContractItem.contract_id == contract_id)
                .scalar()
            ) or 0
            item_seq = max_seq + 1
        else:
            item_seq = data.item_seq
        now = datetime.now()
        item = ContractItem(
            contract_id=contract_id,
            item_seq=item_seq,
            item_name=data.item_name,
            item_category=data.item_category or "",
            unit_price_tax_excluded=data.unit_price_tax_excluded,
            quantity=data.quantity,
            unit=data.unit,
            tax_rate=data.tax_rate,
            amount_tax_excluded=data.amount_tax_excluded,
            amount_tax_included=data.amount_tax_included,
            is_fixed=data.is_fixed,
            is_floating=data.is_floating,
            created_at=now,
            updated_at=now,
        )
        db.add(item)
        db.commit()
        db.refresh(item)
        return ContractItemResponse.from_orm(item)

    @staticmethod
    def update_item(
        db: Session, contract_id: str, item_id: int, data: ContractItemUpdate
    ) -> ContractItemResponse:
        from fastapi import HTTPException
        item = (
            db.query(ContractItem)
            .filter(ContractItem.id == item_id, ContractItem.contract_id == contract_id)
            .first()
        )
        if not item:
            raise HTTPException(status_code=404, detail="合約項目不存在")
        for key, value in data.dict(exclude_unset=True).items():
            setattr(item, key, value)
        item.updated_at = datetime.now()
        db.commit()
        db.refresh(item)
        return ContractItemResponse.from_orm(item)

    @staticmethod
    def delete_item(db: Session, contract_id: str, item_id: int) -> None:
        from fastapi import HTTPException
        item = (
            db.query(ContractItem)
            .filter(ContractItem.id == item_id, ContractItem.contract_id == contract_id)
            .first()
        )
        if not item:
            raise HTTPException(status_code=404, detail="合約項目不存在")
        elete(item)
        db.commit()



# ══════════════════════════════════════════════════════════════════════════
# 預算執行率分析
# ══════════════════════════════════════════════════════════════════════════

def get_budget_analysis(db: Session, budget_year: int) -> list:
    """
    按 (budget_year, category_l1, category_l2) 聚合請款金額。
    回傳 list of dict，每筆含：
      category_l1, category_l2, accounting_code,
      contract_count, total_claimed, paid_amount, pending_amount, approved_amount
    """
    from sqlalchemy import case as sa_case

    rows = (
        db.query(
            Contract.budget_category_l1.label("category_l1"),
            Contract.budget_category_l2.label("category_l2"),
            Contract.accounting_code.label("accounting_code"),
            func.count(Contract.id.distinct()).label("contract_count"),
            func.coalesce(func.sum(ContractClaim.amount), 0).label("total_claimed"),
            func.coalesce(
                func.sum(
                    sa_case((ContractClaim.status == "已付款", ContractClaim.amount), else_=0)
                ), 0
            ).label("paid_amount"),
            func.coalesce(
                func.sum(
                    sa_case((ContractClaim.status == "待審核", ContractClaim.amount), else_=0)
                ), 0
            ).label("pending_amount"),
            func.coalesce(
                func.sum(
                    sa_case((ContractClaim.status == "已核准", ContractClaim.amount), else_=0)
                ), 0
            ).label("approved_amount"),
        )
        .outerjoin(ContractClaim, ContractClaim.contract_id == Contract.contract_id)
        .filter(Contract.budget_year == budget_year)
        .group_by(
            Contract.budget_category_l1,
            Contract.budget_category_l2,
            Contract.accounting_code,
        )
        .order_by(
            Contract.budget_category_l1,
            Contract.budget_category_l2,
        )
        .all()
    )

    result = []
    for r in rows:
        result.append({
            "category_l1":    r.category_l1 or "",
            "category_l2":    r.category_l2 or "",
            "accounting_code": r.accounting_code or "",
            "contract_count": int(r.contract_count),
            "total_claimed":  float(r.total_claimed),
            "paid_amount":    float(r.paid_amount),
            "pending_amount": float(r.pending_amount),
            "approved_amount": float(r.approved_amount),
        })
    return result


# ══════════════════════════════════════════════════════════════════════════
# Excel 匯出工具函數
# ══════════════════════════════════════════════════════════════════════════

import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1B3A5C")
_LINK_FONT   = Font(color="4BA8E8", underline="single")


def generate_contract_excel(contracts: list, db=None) -> bytes:
    """產生合約列表 Excel，單一工作表。contracts 為 ContractDetailResponse list。
    db: 若傳入，則一併查詢費用分攤明細（多欄展開）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合約列表"

    # ── 固定欄位 ────────────────────────────────────────────────────────────
    FIXED_COLS = [
        ("合約編號",       "contract_id",               20),
        ("合約名稱",       "contract_name",             35),
        ("合約類型",       "contract_type",             15),
        ("狀態",           "contract_status",           12),
        ("廠商名稱",       "vendor_name",               25),
        ("負責部門",       "responsible_dept",          15),
        ("簽約公司",       "signing_company",           15),  # F3
        ("簽約權責部門",   "signing_dept",              15),  # F3
        ("預算使用部門",   "budget_dept",               15),  # F3
        ("計價規格",       "pricing_spec",              15),  # F3
        ("開始日期",       "start_date",                14),
        ("結束日期",       "end_date",                  14),
        ("合約金額（含稅）", "total_amount_tax_included", 20),
        ("幣別",           "currency",                   8),
        ("風險等級",       "risk_level",                10),
        ("預算年度",       "budget_year",               10),
        ("預算大項",       "budget_category_l1",        20),
        ("預算細項",       "budget_category_l2",        25),
        ("會計科目",       "accounting_code",           15),
        ("主要負責人",     "manager",                   12),
        ("備註",           "remarks",                   30),
        ("建立時間",       "created_at",                18),
    ]

    # ── 費用分攤：查詢並建立 dict {contract_id: [rows]} ─────────────────────
    alloc_map: dict = {}
    max_alloc = 0
    if db is not None:
        try:
            from app.models.contract import ContractCostAllocation
            cids = [getattr(c, "contract_id", None) for c in contracts]
            cids = [c for c in cids if c]
            rows = db.query(ContractCostAllocation).filter(
                ContractCostAllocation.contract_id.in_(cids)
            ).order_by(ContractCostAllocation.contract_id, ContractCostAllocation.id).all()
            for r in rows:
                alloc_map.setdefault(r.contract_id, []).append(r)
            max_alloc = max((len(v) for v in alloc_map.values()), default=0)
        except Exception:
            pass

    # ── 動態費用分攤欄 ───────────────────────────────────────────────────────
    ALLOC_COLS = []  # (header, width) tuples for allocation expansion
    for i in range(1, max_alloc + 1):
        ALLOC_COLS.append((f"分攤公司{i}", 14))
        ALLOC_COLS.append((f"分攤類型{i}", 12))
        ALLOC_COLS.append((f"分攤數值{i}", 12))

    # ── 表頭 ─────────────────────────────────────────────────────────────────
    all_headers = [(h, w) for h, _, w in FIXED_COLS] + ALLOC_COLS
    for col_idx, (header, width) in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    # ── 資料列 ───────────────────────────────────────────────────────────────
    for row_idx, c in enumerate(contracts, start=2):
        # 固定欄
        for col_idx, (_, field, _) in enumerate(FIXED_COLS, start=1):
            val = getattr(c, field, None)
            if val is None:
                val = ""
            elif hasattr(val, "isoformat"):
                val = str(val)[:10] if field in ("start_date", "end_date", "created_at") else str(val)[:19]
            ws.cell(row=row_idx, column=col_idx, value=val).alignment = Alignment(vertical="top")

        # 費用分攤欄（多欄展開）
        if max_alloc > 0:
            cid = getattr(c, "contract_id", None)
            allocations = alloc_map.get(cid, [])
            base_col = len(FIXED_COLS) + 1
            for i, alloc in enumerate(allocations):
                offset = i * 3
                ws.cell(row=row_idx, column=base_col + offset,     value=alloc.company_name).alignment = Alignment(vertical="top")
                type_label = "比例%" if alloc.allocation_type == "percentage" else "固定金額"
                ws.cell(row=row_idx, column=base_col + offset + 1, value=type_label).alignment = Alignment(vertical="top")
                ws.cell(row=row_idx, column=base_col + offset + 2, value=float(alloc.value)).alignment = Alignment(vertical="top")

    # 金額欄格式（第 13 欄，固定欄含新增後的位置）
    amt_col_idx = next(i for i, (_, f, _) in enumerate(FIXED_COLS, start=1) if f == "total_amount_tax_included")
    amount_col_letter = openpyxl.utils.get_column_letter(amt_col_idx)
    for r in range(2, len(contracts) + 2):
        ws[f"{amount_col_letter}{r}"].number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_claims_excel(claims: list) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "請款清單"

    COLS = [
        ("請款ID",     "id",             10),
        ("合約編號",   "contract_id",    20),
        ("合約名稱",   "contract_name",  35),
        ("請款類型",   "claim_type",     12),
        ("請款日期",   "claim_date",     14),
        ("費用歸屬公司", "cost_company", 15),  # F6
        ("發票號碼",   "invoice_no",     18),
        ("請款金額",   "amount",         16),
        ("狀態",       "status",         12),
        ("審核人",     "approver",       12),
        ("備註",       "remarks",        30),
        ("建立時間",   "created_at",     18),
    ]

    for col_idx, (header, _, width) in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = _HEADER_FONT
        cell.fill  = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22

    for row_idx, c in enumerate(claims, start=2):
        for col_idx, (_, field, _) in enumerate(COLS, start=1):
            val = getattr(c, field, None)
            if val is None:
                val = ""
            elif hasattr(val, "isoformat"):
                val = str(val)[:10] if field in ("claim_date", "created_at") else str(val)[:19]
            ws.cell(row=row_idx, column=col_idx, value=val).alignment = Alignment(vertical="top")

    # amount 欄動態定位（避免加欄後 hardcode 出錯）
    amt_col_idx = next(i for i, (_, f, _) in enumerate(COLS, start=1) if f == "amount")
    amount_col = openpyxl.utils.get_column_letter(amt_col_idx)
    for r in range(2, len(claims) + 2):
        ws[f"{amount_col}{r}"].number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class RenewalService:
    """合約續約申請 CRUD + 審核流程"""

    @staticmethod
    def _log_entry(actor: str, action: str, comment: str | None = None) -> dict:
        from app.core.time import twnow
        entry = {"actor": actor, "action": action, "ts": twnow().isoformat()}
        if comment:
            entry["comment"] = comment
        return entry

    @staticmethod
    def create(db: Session, contract_id: str, data, applicant: str) -> "ContractRenewal":
        from app.models.contract import ContractRenewal, Contract
        import json
        contract = db.query(Contract).filter(Contract.contract_id == contract_id).first()
        if not contract:
            raise ValueError(f"合約 {contract_id} 不存在")
        log = [RenewalService._log_entry(applicant, "申請續約")]
        renewal = ContractRenewal(
            contract_id=contract_id,
            renewal_start_date=data.renewal_start_date,
            renewal_end_date=data.renewal_end_date,
            new_amount=data.new_amount,
            renewal_reason=data.renewal_reason,
            remarks=data.remarks,
            applicant=applicant,
            applicant_dept=data.applicant_dept,
            status="待審核",
            review_log=json.dumps(log, ensure_ascii=False),
        )
        db.add(renewal)
        db.commit()
        db.refresh(renewal)
        return renewal

    @staticmethod
    def list_by_contract(db: Session, contract_id: str) -> list:
        from app.models.contract import ContractRenewal
        return (
            db.query(ContractRenewal)
            .filter(ContractRenewal.contract_id == contract_id)
            .order_by(ContractRenewal.created_at.desc())
            .all()
        )

    @staticmethod
    def list_all(db: Session, status: str | None = None, skip: int = 0, limit: int = 50) -> tuple:
        from app.models.contract import ContractRenewal
        q = db.query(ContractRenewal)
        if status:
            q = q.filter(ContractRenewal.status == status)
        total = q.count()
        items = q.order_by(ContractRenewal.created_at.desc()).offset(skip).limit(limit).all()
        return items, total

    @staticmethod
    def review(db: Session, renewal_id: int, action: str, reviewer: str, comment: str | None) -> "ContractRenewal":
        from app.models.contract import ContractRenewal
        from app.core.time import twnow
        import json
        renewal = db.query(ContractRenewal).filter(ContractRenewal.id == renewal_id).first()
        if not renewal:
            raise ValueError(f"續約申請 #{renewal_id} 不存在")
        allowed = {"approve": ("待審核",), "reject": ("待審核",), "withdraw": ("待審核",)}
        if action not in allowed:
            raise ValueError(f"不支援的操作：{action}")
        if renewal.status not in allowed[action]:
            raise ValueError(f"目前狀態「{renewal.status}」不允許執行「{action}」")
        status_map = {"approve": "已核准", "reject": "已拒絕", "withdraw": "已撤回"}
        renewal.status = status_map[action]
        renewal.reviewer = reviewer
        renewal.reviewed_at = twnow()
        renewal.review_comment = comment
        log = json.loads(renewal.review_log or "[]")
        log.append(RenewalService._log_entry(reviewer, action, comment))
        renewal.review_log = json.dumps(log, ensure_ascii=False)
        db.commit()
        db.refresh(renewal)
        return renewal

    @staticmethod
    def get(db: Session, renewal_id: int) -> "ContractRenewal | None":
        from app.models.contract import ContractRenewal
        return db.query(ContractRenewal).filter(ContractRenewal.id == renewal_id).first()
