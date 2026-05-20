"""
Ragic 與 Portal 欄位比對 Service

功能：
  1. 從 ragic_app_portal_annotations 讀取模組清單（結合前端靜態資料的 Python 鏡像）
  2. 讀取本地 DB schema（SQLAlchemy Inspector）
  3. 讀取已知 Portal API 欄位（config 設定）
  4. 比對 Ragic 欄位 vs Portal DB / API / 前端欄位
  5. 依據比對規則判斷異常
  6. 將結果儲存至 ragic_portal_field_mappings
  7. 產生 Excel 稽核報告

第一版採用：
  - LOCAL_TABLE_MAP：Python 鏡像前端靜態資料
  - PORTAL_FIELD_CONFIG：手工登記各模組的 API / 前端欄位（可逐步完善）
  - DB schema inspection：自動掃描本地 SQLite 欄位
"""

import json
import io
import httpx
from datetime import datetime
from typing import Optional
from sqlalchemy.orm import Session
from sqlalchemy import inspect, text

from app.core.time import twnow
from app.core.config import settings
from app.models.ragic_app_directory import RagicAppPortalAnnotation
from app.models.ragic_field_audit import (
    RagicPortalAuditRun,
    RagicPortalFieldMapping,
    RagicPortalKpiMapping,
)


# ── 常數：Ragic App itemNo → 本地 DB table（對應前端 LOCAL_TABLE_MAP）──────────
LOCAL_TABLE_MAP: dict[int, list[str]] = {
    # 例行抄表/設備檢查
    1: ["hotel_mr_batch", "hotel_mr_reading"],
    2: ["hotel_mr_batch", "hotel_mr_reading"],
    3: ["hotel_mr_batch", "hotel_mr_reading"],
    4: ["hotel_mr_batch", "hotel_mr_reading"],
    # 保全巡檢
    5:  ["security_patrol_batch", "security_patrol_item"],
    6:  ["security_patrol_batch", "security_patrol_item"],
    7:  ["security_patrol_batch", "security_patrol_item"],
    8:  ["security_patrol_batch", "security_patrol_item"],
    9:  ["security_patrol_batch", "security_patrol_item"],
    10: ["security_patrol_batch", "security_patrol_item"],
    13: ["security_patrol_batch", "security_patrol_item"],
    # 工務/保養報表
    68: ["room_maintenance_detail_records"],
    # 整棟工務每日巡檢
    71: ["b1f_inspection_batch", "b1f_inspection_item"],
    72: ["b2f_inspection_batch", "b2f_inspection_item"],
    73: ["b4f_inspection_batch", "b4f_inspection_item"],
    74: ["rf_inspection_batch", "rf_inspection_item"],
    # 春大直報修
    77: ["luqun_repair_case"],
    # 大直工務部
    87: ["dazhi_repair_case"],
    88: ["dazhi_repair_case"],
    # 商場工務每日巡檢
    51: ["mall_fi_inspection_batch", "mall_fi_inspection_item"],
    52: ["mall_fi_inspection_batch", "mall_fi_inspection_item"],
    53: ["mall_fi_inspection_batch", "mall_fi_inspection_item"],
    54: ["mall_fi_inspection_batch", "mall_fi_inspection_item"],
    55: ["mall_fi_inspection_batch", "mall_fi_inspection_item"],
    # 核准請購單
    84:  ["approved_purchase_requests", "approved_purchase_request_items"],
    99:  ["approved_purchase_requests", "approved_purchase_request_items"],
    107: ["approved_purchase_requests", "approved_purchase_request_items"],
    114: ["approved_purchase_requests", "approved_purchase_request_items"],
    122: ["approved_purchase_requests", "approved_purchase_request_items"],
    127: ["approved_purchase_requests", "approved_purchase_request_items"],
    132: ["approved_purchase_requests", "approved_purchase_request_items"],
    199: ["approved_purchase_requests", "approved_purchase_request_items"],
}

# ── 常數：Ragic App itemNo → 模組名稱 + Portal 路由 ───────────────────────────
# 對應前端 PORTAL_DEFAULTS（僅保留有 localTable 的主要模組）
PORTAL_MODULE_MAP: dict[int, dict] = {
    1:   {"module": "例行抄表/設備檢查", "portal_name": "每日數值登錄表",      "portal_url": "/hotel/daily-meter-readings", "company": "飯店"},
    5:   {"module": "保全巡檢",           "portal_name": "保全巡檢",            "portal_url": "/security/dashboard",         "company": "全棟"},
    51:  {"module": "商場工務每日巡檢",   "portal_name": "商場工務巡檢",         "portal_url": "/mall-facility-inspection/dashboard", "company": "商場"},
    68:  {"module": "工務保養",           "portal_name": "客房保養明細",         "portal_url": "/hotel/room-maintenance-detail", "company": "飯店"},
    71:  {"module": "整棟工務每日巡檢",   "portal_name": "整棟巡檢 - B1F",       "portal_url": "/full-building-inspection/dashboard", "company": "全棟"},
    74:  {"module": "整棟工務每日巡檢",   "portal_name": "整棟巡檢 - RF",        "portal_url": "/full-building-inspection/dashboard", "company": "全棟"},
    77:  {"module": "春大直報修",         "portal_name": "商場工務報修",         "portal_url": "/luqun-repair/dashboard",     "company": "商場"},
    87:  {"module": "大直工務部",         "portal_name": "大直工務報修",         "portal_url": "/dazhi-repair/dashboard",     "company": "飯店"},
    84:  {"module": "核准請購單",         "portal_name": "核准請購單月報表",     "portal_url": "/purchase-report/monthly",    "company": "商場"},
    199: {"module": "核准請購單",         "portal_name": "核准請購單月報表",     "portal_url": "/purchase-report/monthly",    "company": "資訊部"},
}

# ── 常數：Ragic 表單 URL 對照表（itemNo → 直接可點擊的 Ragic 表單連結）──────────
# 供 Tab 1「在 Ragic 查看」連結 + 「同步 Ragic 欄位」按鈕使用
# URL 格式：https://{server}/{account}/{tab_folder}/{sheet_index}
RAGIC_URL_MAP: dict[int, str] = {
    # 客房保養明細（report2 sheet 2）
    68: "https://ap12.ragic.com/soutlet001/report2/2",
    # 保全巡檢（security-patrol 各班別 sheet）
    5:  "https://ap12.ragic.com/soutlet001/security-patrol/2",
    6:  "https://ap12.ragic.com/soutlet001/security-patrol/3",
    7:  "https://ap12.ragic.com/soutlet001/security-patrol/4",
    8:  "https://ap12.ragic.com/soutlet001/security-patrol/5",
    9:  "https://ap12.ragic.com/soutlet001/security-patrol/6",
    10: "https://ap12.ragic.com/soutlet001/security-patrol/9",
    13: "https://ap12.ragic.com/soutlet001/security-patrol/1",
    # 春大直工務報修（luqun repair sheet 6）
    77: "https://ap12.ragic.com/soutlet001/luqun-public-works-repair-reporting-system/6",
    # 大直工務部（lequn-public-works sheet 8）
    87: "https://ap12.ragic.com/soutlet001/lequn-public-works/8",
    88: "https://ap12.ragic.com/soutlet001/lequn-public-works/8",
    # 核准請購單（可視實際 Ragic 表單補充）
    # 84: "https://ap16.ragic.com/intraragicapp/...",
}

# ── 常數：已知 Portal API 欄位（手工登記，逐步完善）────────────────────────────
# 格式：{ table_name: { "api_fields": [...], "frontend_fields": [...] } }
PORTAL_FIELD_CONFIG: dict[str, dict] = {
    "security_patrol_batch": {
        "api_fields":      ["id", "batch_date", "sheet_key", "start_time", "end_time", "work_hours", "synced_at"],
        "frontend_fields": ["日期", "班次", "開始時間", "結束時間", "工時"],
        "display_names":   {"batch_date": "日期", "sheet_key": "班次", "start_time": "開始", "end_time": "結束", "work_hours": "工時"},
    },
    "security_patrol_item": {
        "api_fields":      ["id", "batch_id", "check_point", "check_result", "remark"],
        "frontend_fields": ["巡檢點", "結果", "備註"],
        "display_names":   {"check_point": "巡檢點", "check_result": "結果", "remark": "備註"},
    },
    "luqun_repair_case": {
        # api_fields = 後端 API response 中出現的欄位（即 to_dict() 輸出欄位）
        "api_fields": [
            "ragic_id", "case_no", "title", "reporter_name", "repair_type",
            "floor", "floor_normalized", "occurred_at", "responsible_unit",
            "work_hours", "status", "outsource_fee", "maintenance_fee", "total_fee",
            "deduction_item", "deduction_fee", "deduction_counter", "deduction_counter_name",
            "acceptor", "accept_status", "closer", "finance_note", "mgmt_response",
            "is_completed", "completed_at", "close_days", "year", "month",
            "occ_year", "occ_month", "is_room_case", "room_no", "room_category",
            "images_json", "ragic_url",
        ],
        # frontend_fields = 前端 Table/Drawer/Filter 顯示的中文欄位名稱
        "frontend_fields": [
            "報修編號", "標題", "報修人姓名", "報修類型", "發生樓層",
            "發生時間", "負責單位", "花費工時", "處理狀況", "委外費用",
            "維修費用", "費用合計", "驗收者", "驗收", "結案人",
            "扣款事項", "扣款費用", "扣款專櫃", "財務備註", "管理單位回應",
            "結案時間", "結案天數",
        ],
        # display_names = DB欄位 → 中文名稱（供 Ragic 欄位比對使用）
        "display_names": {
            "case_no":               "報修編號",
            "title":                 "標題",
            "reporter_name":         "報修人姓名",
            "repair_type":           "報修類型",
            "floor":                 "發生樓層",
            "occurred_at":           "發生時間",
            "responsible_unit":      "負責單位",
            "work_hours":            "花費工時",
            "status":                "處理狀況",
            "outsource_fee":         "委外費用",
            "maintenance_fee":       "維修費用",
            "total_fee":             "費用合計",
            "acceptor":              "驗收者",
            "accept_status":         "驗收",
            "closer":                "結案人",
            "deduction_item":        "扣款事項",
            "deduction_fee":         "扣款費用",
            "deduction_counter":     "扣款專櫃原始值",
            "deduction_counter_name":"扣款專櫃",
            "finance_note":          "財務備註",
            "mgmt_response":         "管理單位回應",
            "completed_at":          "結案時間",
            "close_days":            "結案天數",
            "is_completed":          "是否結案",
            "year":                  "年度",
            "month":                 "月份",
            "occ_year":              "發生年度",
            "occ_month":             "發生月份",
            "is_room_case":          "是否房務報修",
            "room_no":               "房號",
            "room_category":         "房型",
            "images_json":           "附圖",
            "floor_normalized":      "樓層(標準化)",
        },
        # filter_fields = 前端篩選面板使用的欄位
        "filter_fields": ["repair_type", "floor", "status"],
        # calculated_fields = Portal 後端計算/衍生欄位（Ragic 無直接對應）
        "calculated_fields": [
            "total_fee", "floor_normalized", "close_days", "is_completed",
            "year", "month", "occ_year", "occ_month",
            "is_room_case", "room_no", "room_category",
        ],
        # export_fields = 匯出 Excel 包含的欄位
        "export_fields": [
            "case_no", "title", "repair_type", "floor", "occurred_at", "status",
            "closer", "acceptor", "accept_status", "is_completed",
            "deduction_item", "deduction_fee", "deduction_counter_name",
            "work_hours", "close_days", "outsource_fee", "maintenance_fee",
            "total_fee", "finance_note",
        ],
    },
    "dazhi_repair_case": {
        "api_fields": [
            "ragic_id", "case_no", "title", "reporter_name", "repair_type",
            "floor", "floor_normalized", "occurred_at", "responsible_unit",
            "work_hours", "status", "outsource_fee", "maintenance_fee", "total_fee",
            "deduction_item", "deduction_fee",
            "acceptor", "accept_status", "closer", "finance_note",
            "is_completed", "completed_at", "close_days", "year", "month",
            "is_room_case", "room_no", "room_category",
            "images_json", "ragic_url",
        ],
        "frontend_fields": [
            "報修編號", "標題", "報修人姓名", "報修類型", "發生樓層",
            "發生時間", "負責單位", "花費工時", "處理狀況", "委外費用",
            "維修費用", "費用合計", "驗收者", "驗收", "結案人",
            "扣款事項", "扣款費用", "財務備註", "結案時間", "結案天數",
        ],
        "display_names": {
            "case_no":           "報修編號",
            "title":             "標題",
            "reporter_name":     "報修人姓名",
            "repair_type":       "報修類型",
            "floor":             "發生樓層",
            "occurred_at":       "發生時間",
            "responsible_unit":  "負責單位",
            "work_hours":        "花費工時",
            "status":            "處理狀況",
            "outsource_fee":     "委外費用",
            "maintenance_fee":   "維修費用",
            "total_fee":         "費用合計",
            "acceptor":          "驗收者",
            "accept_status":     "驗收",
            "closer":            "結案人",
            "deduction_item":    "扣款事項",
            "deduction_fee":     "扣款費用",
            "finance_note":      "財務備註",
            "completed_at":      "結案時間",
            "close_days":        "結案天數",
            "is_completed":      "是否結案",
            "year":              "年度",
            "month":             "月份",
            "is_room_case":      "是否房務報修",
            "room_no":           "房號",
            "room_category":     "房型",
            "images_json":       "附圖",
            "floor_normalized":  "樓層(標準化)",
        },
        "filter_fields": ["repair_type", "floor", "status"],
        "calculated_fields": [
            "total_fee", "floor_normalized", "close_days", "is_completed",
            "year", "month", "is_room_case", "room_no", "room_category",
        ],
        "export_fields": [
            "case_no", "title", "repair_type", "floor", "occurred_at", "status",
            "closer", "acceptor", "accept_status", "is_completed",
            "deduction_item", "deduction_fee",
            "work_hours", "close_days", "outsource_fee", "maintenance_fee",
            "total_fee", "finance_note",
        ],
    },
    "approved_purchase_requests": {
        "api_fields":      ["id", "ragic_id", "order_no", "dept", "purchase_date", "approved_date",
                            "total_amount", "status", "applicant", "ragic_url"],
        "frontend_fields": ["請購單號", "部門", "日期", "核准日", "金額", "狀態", "申請人"],
        "display_names":   {"order_no": "請購單號", "total_amount": "總金額", "status": "狀態"},
    },
    "approved_purchase_request_items": {
        "api_fields":      ["id", "request_id", "item_name", "quantity", "unit", "unit_price", "amount"],
        "frontend_fields": ["品項", "數量", "單位", "單價", "金額"],
        "display_names":   {"item_name": "品項名稱", "unit_price": "單價", "amount": "金額"},
    },
    "hotel_mr_batch": {
        "api_fields":      ["id", "ragic_id", "batch_date", "sheet_key", "synced_at"],
        "frontend_fields": ["日期", "表單類別"],
        "display_names":   {"batch_date": "日期", "sheet_key": "表單類別"},
    },
    "hotel_mr_reading": {
        "api_fields":      ["id", "batch_id", "meter_name", "reading_value", "unit"],
        "frontend_fields": ["錶名稱", "讀數", "單位"],
        "display_names":   {"meter_name": "錶名稱", "reading_value": "讀數"},
    },
    "room_maintenance_detail_records": {
        "api_fields":      ["id", "ragic_id", "room_no", "maintenance_date", "item", "status", "cost", "ragic_url"],
        "frontend_fields": ["房號", "保養日期", "保養項目", "狀態", "費用"],
        "display_names":   {"room_no": "房號", "maintenance_date": "保養日期", "status": "狀態"},
    },
    "b1f_inspection_batch": {
        "api_fields":      ["id", "ragic_id", "batch_date", "start_time", "end_time", "work_hours", "synced_at"],
        "frontend_fields": ["日期", "開始", "結束", "工時"],
        "display_names":   {"batch_date": "日期", "work_hours": "工時"},
    },
    "b1f_inspection_item": {
        "api_fields":      ["id", "batch_id", "item_name", "result", "remark"],
        "frontend_fields": ["巡檢項目", "結果", "備註"],
        "display_names":   {"item_name": "巡檢項目", "result": "結果"},
    },
    "mall_fi_inspection_batch": {
        "api_fields":      ["id", "ragic_id", "batch_date", "sheet_key", "start_time", "end_time", "work_hours", "synced_at"],
        "frontend_fields": ["日期", "樓層", "開始", "結束", "工時"],
        "display_names":   {"batch_date": "日期", "sheet_key": "樓層", "work_hours": "工時"},
    },
    "mall_fi_inspection_item": {
        "api_fields":      ["id", "batch_id", "item_name", "result", "remark"],
        "frontend_fields": ["巡檢項目", "結果", "備註"],
        "display_names":   {"item_name": "巡檢項目", "result": "結果"},
    },
}

# ── 常數：KPI 計算追溯（手工登記 Dashboard KPI 與 DB 欄位的關係）──────────────
KPI_DEFINITIONS: list[dict] = [
    {
        "module_name": "保全巡檢",
        "portal_route": "/security/dashboard",
        "kpi_name": "本月巡檢總場次",
        "page_section": "KPI Card",
        "api_endpoint": "/api/v1/security-patrol/stats",
        "db_table": "security_patrol_batch",
        "source_fields": ["id", "batch_date"],
        "date_field": "batch_date",
        "filters": {"本月": "batch_date >= month_start"},
        "formula": "COUNT(id) WHERE batch_date IN 本月",
        "ragic_source_fields": ["巡檢日期", "班次"],
        "trace_status": "traceable",
    },
    {
        "module_name": "保全巡檢",
        "portal_route": "/security/dashboard",
        "kpi_name": "本月巡檢總工時",
        "page_section": "KPI Card",
        "api_endpoint": "/api/v1/security-patrol/stats",
        "db_table": "security_patrol_batch",
        "source_fields": ["work_hours", "batch_date"],
        "date_field": "batch_date",
        "filters": {"本月": "batch_date >= month_start"},
        "formula": "SUM(work_hours) WHERE batch_date IN 本月",
        "ragic_source_fields": ["工時", "開始時間", "結束時間"],
        "trace_status": "traceable",
    },
    {
        "module_name": "商場工務報修",
        "portal_route": "/luqun-repair/dashboard",
        "kpi_name": "本月報修件數",
        "page_section": "KPI Card",
        "api_endpoint": "/api/v1/luqun-repair/stats",
        "db_table": "luqun_repair_case",
        "source_fields": ["ragic_id", "occurred_at"],
        "date_field": "occurred_at",
        "filters": {"本月": "occurred_at >= month_start"},
        "formula": "COUNT(ragic_id) WHERE occurred_at IN 本月",
        "ragic_source_fields": ["發生時間", "報修編號"],
        "trace_status": "traceable",
    },
    {
        "module_name": "飯店工務報修",
        "portal_route": "/dazhi-repair/dashboard",
        "kpi_name": "本月報修件數",
        "page_section": "KPI Card",
        "api_endpoint": "/api/v1/dazhi-repair/stats",
        "db_table": "dazhi_repair_case",
        "source_fields": ["ragic_id", "occurred_at"],
        "date_field": "occurred_at",
        "filters": {"本月": "occurred_at >= month_start"},
        "formula": "COUNT(ragic_id) WHERE occurred_at IN 本月",
        "ragic_source_fields": ["發生時間", "報修編號"],
        "trace_status": "traceable",
    },
    {
        "module_name": "核准請購單月報表",
        "portal_route": "/purchase-report/monthly",
        "kpi_name": "月報表總金額",
        "page_section": "統計表頭",
        "api_endpoint": "/api/v1/purchase-report/monthly",
        "db_table": "approved_purchase_requests",
        "source_fields": ["total_amount", "approved_date"],
        "date_field": "approved_date",
        "filters": {"月份": "approved_date LIKE YYYY-MM%"},
        "formula": "SUM(total_amount) GROUP BY dept",
        "ragic_source_fields": ["核准日期", "核准金額"],
        "trace_status": "traceable",
    },
]


def _get_db_inspector(db: Session):
    """取得 SQLAlchemy Inspector，用於讀取 DB schema。"""
    from sqlalchemy import inspect as sa_inspect
    return sa_inspect(db.bind)


def get_all_db_tables(db: Session) -> dict[str, list[dict]]:
    """
    掃描本地 DB 所有 table 的欄位清單。
    回傳格式：{ table_name: [{ name, type, nullable, primary_key }, ...] }
    """
    try:
        inspector = _get_db_inspector(db)
        result = {}
        for table_name in inspector.get_table_names():
            cols = inspector.get_columns(table_name)
            result[table_name] = [
                {
                    "name": c["name"],
                    "type": str(c["type"]),
                    "nullable": c.get("nullable", True),
                    "primary_key": c.get("primary_key", False),
                }
                for c in cols
            ]
        return result
    except Exception as e:
        return {}


def _get_annotated_modules(db: Session) -> list[dict]:
    """
    從 ragic_app_portal_annotations 讀取有 Portal 對應的模組，
    結合 PORTAL_MODULE_MAP 取得完整資訊。
    只回傳有 LOCAL_TABLE 的模組（實際有同步資料者）。
    """
    rows = db.query(RagicAppPortalAnnotation).all()
    db_map = {r.item_no: r for r in rows}

    modules = {}  # key = portal_route + table，去重
    for item_no, tables in LOCAL_TABLE_MAP.items():
        info = PORTAL_MODULE_MAP.get(item_no, {})
        annotation = db_map.get(item_no)

        portal_name = info.get("portal_name", "") or (annotation.portal_name if annotation else "")
        portal_url = info.get("portal_url", "") or (annotation.portal_url if annotation else "")
        module_label = info.get("module", "")
        company = info.get("company", "")

        # Ragic URL 優先順序：
        #  1. 使用者在 DB 手動設定（annotation.ragic_url）
        #  2. 程式碼常數 RAGIC_URL_MAP
        db_ragic_url = (annotation.ragic_url or "") if annotation else ""
        ragic_url = db_ragic_url or RAGIC_URL_MAP.get(item_no, "")

        key = (portal_url, tuple(tables))
        if key not in modules:
            modules[key] = {
                "item_no": item_no,
                "company": company,
                "module_name": module_label,
                "portal_name": portal_name,
                "portal_url": portal_url,
                "local_tables": tables,
                "ragic_url": ragic_url,
                "is_active": True,
            }

    return list(modules.values())


def get_module_overview(db: Session) -> list[dict]:
    """
    取得模組比對總覽（Tab 1 資料）。
    每個模組彙整欄位數量與比對狀態。
    """
    modules = _get_annotated_modules(db)
    db_tables = get_all_db_tables(db)

    result = []
    seen_routes: set[str] = set()   # 防止同路由重複出現（Tab 1 / Tab 2 Select 去重）

    for mod in modules:
        route = mod["portal_url"]
        # 跳過沒有 Portal 路由的模組（LOCAL_TABLE_MAP 有、PORTAL_MODULE_MAP 無的項目）
        if not route:
            continue
        # 同一路由只回傳第一筆（多個 itemNo 指向同路由時合併顯示）
        if route in seen_routes:
            continue
        seen_routes.add(route)

        tables = mod["local_tables"]
        total_db_cols = 0
        for t in tables:
            total_db_cols += len(db_tables.get(t, []))

        # 查詢已存儲的 mapping 紀錄
        existing = db.query(RagicPortalFieldMapping).filter(
            RagicPortalFieldMapping.portal_route == route
        ).all()

        normal = sum(1 for m in existing if m.mapping_status == "normal")
        issues = sum(1 for m in existing if m.mapping_status not in ("normal", "unmapped"))
        unmapped = sum(1 for m in existing if m.mapping_status == "unmapped")

        # 取最近比對時間
        last_checked = max((m.last_checked_at for m in existing if m.last_checked_at), default=None)

        if not existing:
            status = "not_audited"
        elif issues > 0:
            status = "error" if any(m.severity == "high" for m in existing if m.mapping_status not in ("normal",)) else "warning"
        else:
            status = "normal"

        # API 欄位數（從設定取）
        portal_api_cols = 0
        for t in tables:
            cfg = PORTAL_FIELD_CONFIG.get(t, {})
            portal_api_cols += len(cfg.get("api_fields", []))

        # 前端欄位數
        portal_fe_cols = 0
        for t in tables:
            cfg = PORTAL_FIELD_CONFIG.get(t, {})
            portal_fe_cols += len(cfg.get("frontend_fields", []))

        result.append({
            "item_no": mod["item_no"],
            "company": mod["company"],
            "module_name": mod["module_name"],
            "portal_name": mod["portal_name"],
            "portal_route": route,
            "local_tables": tables,
            "ragic_url": mod.get("ragic_url", ""),
            "ragic_field_count": 0,        # 需要 Ragic API 才能取得，先填 0
            "portal_db_field_count": total_db_cols,
            "portal_api_field_count": portal_api_cols,
            "portal_fe_field_count": portal_fe_cols,
            "normal_count": normal,
            "issue_count": issues,
            "unmapped_count": unmapped,
            "total_mapping_count": len(existing),
            "last_checked_at": last_checked.isoformat() if last_checked else None,
            "status": status,
            "is_active": mod["is_active"],
        })

    return result


def get_module_field_detail(db: Session, portal_route: str) -> list[dict]:
    """
    取得單一模組的欄位 Mapping 明細（Tab 2 資料）。
    若 DB 沒有 mapping 紀錄，自動從 DB schema 生成草稿。
    """
    existing = db.query(RagicPortalFieldMapping).filter(
        RagicPortalFieldMapping.portal_route == portal_route
    ).order_by(RagicPortalFieldMapping.portal_db_table, RagicPortalFieldMapping.id).all()

    if existing:
        return [_mapping_to_dict(m) for m in existing]

    # 自動從 DB schema 生成（未執行比對時的草稿）
    db_tables = get_all_db_tables(db)
    result = []

    # 找到此路由對應的 tables
    target_tables: list[str] = []
    for item_no, tables in LOCAL_TABLE_MAP.items():
        info = PORTAL_MODULE_MAP.get(item_no, {})
        if info.get("portal_url") == portal_route:
            for t in tables:
                if t not in target_tables:
                    target_tables.append(t)

    for table in target_tables:
        cols = db_tables.get(table, [])
        cfg = PORTAL_FIELD_CONFIG.get(table, {})
        api_fields = cfg.get("api_fields", [])
        fe_fields = cfg.get("frontend_fields", [])
        display_names = cfg.get("display_names", {})

        for col in cols:
            fname = col["name"]
            result.append({
                "id": None,
                "category": "DB 欄位",
                "portal_db_table": table,
                "portal_db_field": fname,
                "portal_api_field": fname if fname in api_fields else "",
                "portal_frontend_field": display_names.get(fname, ""),
                "display_name": display_names.get(fname, ""),
                "ragic_field_id": "",
                "ragic_field_name": "",
                "ragic_field_type": "",
                "is_ragic_required": False,
                "is_ragic_formula": False,
                "is_ragic_subtable": False,
                "is_displayed": bool(display_names.get(fname)),
                "is_filter": False,
                "is_export": fname in api_fields,
                "is_calculated": False,
                "mapping_status": "unmapped",
                "severity": None,
                "issue_type": None,
                "issue_message": "尚未執行欄位比對，請點選「執行比對」",
                "suggestion": "執行比對後系統會自動判斷對應狀態",
                "is_resolved": False,
                "last_checked_at": None,
            })

    return result


def get_all_issues(db: Session, severity: Optional[str] = None, is_resolved: Optional[bool] = None) -> list[dict]:
    """
    取得所有異常欄位清單（Tab 3 資料）。
    """
    q = db.query(RagicPortalFieldMapping).filter(
        RagicPortalFieldMapping.mapping_status.notin_(["normal", "unmapped"])
    )
    if severity:
        q = q.filter(RagicPortalFieldMapping.severity == severity)
    if is_resolved is not None:
        q = q.filter(RagicPortalFieldMapping.is_resolved == is_resolved)

    rows = q.order_by(
        RagicPortalFieldMapping.severity,
        RagicPortalFieldMapping.module_name
    ).all()

    return [_mapping_to_dict(m) for m in rows]


def get_kpi_mappings(db: Session, module_name: Optional[str] = None) -> list[dict]:
    """
    取得 KPI / Dashboard 計算追溯清單（Tab 4 資料）。
    優先從 DB 讀取，若無則回傳 KPI_DEFINITIONS 靜態設定。
    """
    q = db.query(RagicPortalKpiMapping)
    if module_name:
        q = q.filter(RagicPortalKpiMapping.module_name == module_name)
    rows = q.all()

    if rows:
        return [_kpi_to_dict(k) for k in rows]

    # 回傳靜態定義（尚未執行 seed 時）
    result = []
    for kpi in KPI_DEFINITIONS:
        if module_name and kpi.get("module_name") != module_name:
            continue
        result.append({
            "id": None,
            "module_name": kpi.get("module_name"),
            "portal_route": kpi.get("portal_route"),
            "kpi_name": kpi.get("kpi_name"),
            "page_section": kpi.get("page_section"),
            "api_endpoint": kpi.get("api_endpoint"),
            "db_table": kpi.get("db_table"),
            "source_fields": json.dumps(kpi.get("source_fields", []), ensure_ascii=False),
            "date_field": kpi.get("date_field"),
            "filters": json.dumps(kpi.get("filters", {}), ensure_ascii=False),
            "formula": kpi.get("formula"),
            "ragic_source_fields": json.dumps(kpi.get("ragic_source_fields", []), ensure_ascii=False),
            "trace_status": kpi.get("trace_status", "unknown"),
            "issue_message": None,
            "suggestion": None,
            "last_checked_at": None,
        })
    return result


def run_audit(db: Session, triggered_by: str = "system", scope: str = "all") -> dict:
    """
    執行欄位比對稽核任務。
    第一版邏輯：
      1. 掃描本地 DB schema
      2. 對比 PORTAL_FIELD_CONFIG 已知 API / 前端欄位
      3. 生成 mapping 紀錄（只寫入尚未存在者，避免覆蓋手工標記）
      4. 記錄 audit_run
      5. Seed KPI mappings
    """
    now = twnow()
    run = RagicPortalAuditRun(
        run_time=now,
        triggered_by=triggered_by,
        scope=scope,
        status="running",
    )
    db.add(run)
    db.flush()

    db_tables = get_all_db_tables(db)
    modules = _get_annotated_modules(db)

    total_modules = 0
    normal_count = 0
    warning_count = 0
    error_count = 0
    created_count = 0

    processed_routes: set[str] = set()

    for mod in modules:
        route = mod["portal_url"]
        if not route or route in processed_routes:
            continue
        processed_routes.add(route)
        total_modules += 1

        tables = mod["local_tables"]
        module_has_issue = False
        module_has_error = False

        for table in tables:
            cols = db_tables.get(table, [])
            cfg = PORTAL_FIELD_CONFIG.get(table, {})
            api_fields      = set(cfg.get("api_fields", []))
            fe_fields        = set(cfg.get("frontend_fields", []))
            display_names    = cfg.get("display_names", {})
            filter_fields    = set(cfg.get("filter_fields", []))
            calculated_fields = set(cfg.get("calculated_fields", []))
            # export_fields 若未定義，退回用 api_fields 判斷
            export_fields    = set(cfg.get("export_fields", api_fields))

            for col in cols:
                fname = col["name"]
                col_type = col["type"].lower()

                # 計算四個 flag（新建與更新都需要）
                _is_displayed   = bool(display_names.get(fname))
                _is_filter      = fname in filter_fields
                _is_export      = fname in export_fields
                _is_calculated  = fname in calculated_fields

                # 若已有紀錄：更新 last_checked_at 與 flags（config 修正後可立即生效）
                existing = db.query(RagicPortalFieldMapping).filter(
                    RagicPortalFieldMapping.portal_route == route,
                    RagicPortalFieldMapping.portal_db_table == table,
                    RagicPortalFieldMapping.portal_db_field == fname,
                ).first()
                if existing:
                    existing.last_checked_at = now
                    existing.is_displayed   = _is_displayed
                    existing.is_filter      = _is_filter
                    existing.is_export      = _is_export
                    existing.is_calculated  = _is_calculated
                    # 同步更新 display_name / portal_api_field（config 改正後也跟著更新）
                    existing.display_name          = display_names.get(fname, existing.display_name)
                    existing.portal_frontend_field = display_names.get(fname, existing.portal_frontend_field)
                    existing.portal_api_field      = fname if fname in api_fields else existing.portal_api_field
                    continue

                # 判斷 mapping status 與異常
                status, severity, issue_type, issue_msg, suggestion = _evaluate_field(
                    fname, col_type, api_fields, fe_fields, display_names
                )

                if severity == "high":
                    module_has_error = True
                elif severity in ("medium", "low"):
                    module_has_issue = True

                mapping = RagicPortalFieldMapping(
                    app_directory_id=mod["item_no"],
                    company=mod["company"],
                    module_name=mod["module_name"],
                    portal_route=route,
                    portal_db_table=table,
                    portal_db_field=fname,
                    portal_api_field=fname if fname in api_fields else "",
                    portal_frontend_field=display_names.get(fname, ""),
                    display_name=display_names.get(fname, ""),
                    ragic_field_id="",
                    ragic_field_name="",
                    ragic_field_type="",
                    is_displayed=_is_displayed,
                    is_filter=_is_filter,
                    is_export=_is_export,
                    is_calculated=_is_calculated,
                    mapping_status=status,
                    severity=severity,
                    issue_type=issue_type,
                    issue_message=issue_msg,
                    suggestion=suggestion,
                    is_resolved=False,
                    last_checked_at=now,
                )
                db.add(mapping)
                created_count += 1

        if module_has_error:
            error_count += 1
        elif module_has_issue:
            warning_count += 1
        else:
            normal_count += 1

    # Seed KPI mappings（若不存在）
    for kpi in KPI_DEFINITIONS:
        exists = db.query(RagicPortalKpiMapping).filter(
            RagicPortalKpiMapping.module_name == kpi["module_name"],
            RagicPortalKpiMapping.kpi_name == kpi["kpi_name"],
        ).first()
        if not exists:
            k = RagicPortalKpiMapping(
                module_name=kpi["module_name"],
                portal_route=kpi["portal_route"],
                kpi_name=kpi["kpi_name"],
                page_section=kpi.get("page_section"),
                api_endpoint=kpi.get("api_endpoint"),
                db_table=kpi.get("db_table"),
                source_fields=json.dumps(kpi.get("source_fields", []), ensure_ascii=False),
                date_field=kpi.get("date_field"),
                filters=json.dumps(kpi.get("filters", {}), ensure_ascii=False),
                formula=kpi.get("formula"),
                ragic_source_fields=json.dumps(kpi.get("ragic_source_fields", []), ensure_ascii=False),
                trace_status=kpi.get("trace_status", "unknown"),
                last_checked_at=now,
            )
            db.add(k)

    # 更新 audit run 結果
    run.total_modules = total_modules
    run.normal_count = normal_count
    run.warning_count = warning_count
    run.error_count = error_count
    run.status = "completed"
    run.notes = f"新建 {created_count} 筆 mapping 紀錄"

    db.commit()

    return {
        "run_id": run.id,
        "run_time": run.run_time.isoformat(),
        "total_modules": total_modules,
        "normal_count": normal_count,
        "warning_count": warning_count,
        "error_count": error_count,
        "created_mappings": created_count,
        "status": "completed",
    }


def _evaluate_field(
    fname: str,
    col_type: str,
    api_fields: set[str],
    fe_fields: set[str],
    display_names: dict,
) -> tuple[str, Optional[str], Optional[str], Optional[str], Optional[str]]:
    """
    判斷單一欄位的比對狀態與異常。
    回傳：(mapping_status, severity, issue_type, issue_message, suggestion)
    """
    # 系統欄位直接視為正常
    system_fields = {"id", "created_at", "updated_at", "synced_at", "ragic_id", "ragic_url"}
    if fname in system_fields:
        return "normal", None, None, None, None

    in_api = fname in api_fields
    in_fe = bool(display_names.get(fname))

    # 金額欄位型態檢查
    if any(k in fname for k in ("amount", "cost", "price", "fee", "金額", "費用", "單價")):
        if "char" in col_type or "text" in col_type:
            return (
                "type_mismatch",
                "high",
                "金額欄位型態疑似非數值",
                f"欄位「{fname}」看起來是金額，但資料庫型態為文字（{col_type}），可能導致 Dashboard 計算錯誤。",
                "請確認 Ragic 原始欄位型態，若為數值請修改 DB 欄位型態或在 sync 時做型態轉換。",
            )

    # 日期欄位型態檢查
    if any(k in fname for k in ("date", "time", "日期", "時間")):
        if "int" in col_type or ("char" in col_type and "date" not in fname):
            return (
                "type_mismatch",
                "medium",
                "時間欄位型態疑似非日期",
                f"欄位「{fname}」看起來是時間，但資料庫型態為 {col_type}，可能影響日期篩選功能。",
                "請確認此欄位是否需要轉為 DATE / DATETIME 型態。",
            )

    # 在 DB 有，但 API / 前端都沒有
    if not in_api and not in_fe:
        return (
            "portal_only",
            "low",
            "Portal DB 有，API 與前端未使用",
            f"欄位「{fname}」存在於資料庫 ({col_type})，但 API 回傳和前端顯示都沒有使用，可能是暫存或廢棄欄位。",
            "請確認此欄位是否仍需保留，若已廢棄可考慮移除（需確認無其他依賴）。",
        )

    # 在 DB 和 API 有，但前端沒有顯示
    if in_api and not in_fe:
        return (
            "normal",  # API 有就算正常（前端可能刻意不顯示）
            None, None, None, None,
        )

    return "normal", None, None, None, None


def generate_excel_report(db: Session) -> bytes:
    """
    產生 Excel 稽核報告（Tab 5 匯出功能）。
    包含 5 個工作表：模組總覽、欄位 Mapping、異常清單、KPI追溯、建議清單。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl 未安裝，請執行 pip install openpyxl")

    wb = openpyxl.Workbook()

    # ── 樣式定義 ──────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B3A5C")
    warning_fill = PatternFill("solid", fgColor="FFF5F5")
    error_fill = PatternFill("solid", fgColor="FFE4E4")
    normal_fill = PatternFill("solid", fgColor="F0FFF0")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_header(ws, headers: list[str], fill=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = fill or header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def auto_width(ws):
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1：模組總覽 ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "1_模組總覽"
    headers1 = ["公司/據點", "Portal 模組名稱", "Portal 路由", "本地 DB Table",
                "DB 欄位數", "API 欄位數", "前端欄位數",
                "正常對應數", "異常數", "未對應數", "最近比對時間", "狀態"]
    write_header(ws1, headers1)

    overview = get_module_overview(db)
    status_map = {"normal": "正常", "warning": "注意", "error": "異常", "not_audited": "尚未比對"}
    for r, mod in enumerate(overview, 2):
        row_data = [
            mod["company"], mod["portal_name"], mod["portal_route"],
            ", ".join(mod["local_tables"]),
            mod["portal_db_field_count"], mod["portal_api_field_count"], mod["portal_fe_field_count"],
            mod["normal_count"], mod["issue_count"], mod["unmapped_count"],
            mod["last_checked_at"] or "—",
            status_map.get(mod["status"], mod["status"]),
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if mod["status"] == "error":
                cell.fill = error_fill
            elif mod["status"] == "warning":
                cell.fill = warning_fill
            elif mod["status"] == "normal":
                cell.fill = normal_fill
    auto_width(ws1)

    # ── Sheet 2：欄位 Mapping 明細 ────────────────────────────────────────────
    ws2 = wb.create_sheet("2_欄位Mapping明細")
    headers2 = ["模組名稱", "DB Table", "DB 欄位", "DB 型態",
                "API 欄位", "前端欄位", "中文顯示名稱",
                "Ragic 欄位名稱", "Ragic 欄位 ID", "Ragic 型態",
                "是否顯示", "是否計算", "是否篩選", "是否匯出",
                "對應狀態", "異常說明", "建議處理"]
    write_header(ws2, headers2)

    all_mappings = db.query(RagicPortalFieldMapping).order_by(
        RagicPortalFieldMapping.module_name,
        RagicPortalFieldMapping.portal_db_table,
        RagicPortalFieldMapping.id,
    ).all()

    status_label = {
        "normal": "正常", "ragic_only": "Ragic 有，Portal 無",
        "portal_only": "Portal 有，Ragic 無", "name_mismatch": "名稱疑似不同",
        "type_mismatch": "型態不一致", "null_rate_high": "空值率異常",
        "formula_unmarked": "公式欄位未標示", "subtable_unmarked": "子表格未處理",
        "unmapped": "未建立 Mapping",
    }

    for r, m in enumerate(all_mappings, 2):
        row_data = [
            m.module_name, m.portal_db_table, m.portal_db_field, m.ragic_field_type or "—",
            m.portal_api_field or "—", m.portal_frontend_field or "—", m.display_name or "—",
            m.ragic_field_name or "—", m.ragic_field_id or "—", m.ragic_field_type or "—",
            "是" if m.is_displayed else "否",
            "是" if m.is_calculated else "否",
            "是" if m.is_filter else "否",
            "是" if m.is_export else "否",
            status_label.get(m.mapping_status, m.mapping_status),
            m.issue_message or "—",
            m.suggestion or "—",
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if m.mapping_status not in ("normal", "unmapped"):
                if m.severity == "high":
                    cell.fill = error_fill
                elif m.severity == "medium":
                    cell.fill = warning_fill
    auto_width(ws2)

    # ── Sheet 3：異常清單 ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("3_異常清單")
    headers3 = ["公司/據點", "模組名稱", "DB Table", "欄位名稱",
                "異常類型", "異常說明", "建議處理", "嚴重程度", "是否已處理"]
    write_header(ws3, headers3)

    issues = db.query(RagicPortalFieldMapping).filter(
        RagicPortalFieldMapping.mapping_status.notin_(["normal", "unmapped"])
    ).order_by(RagicPortalFieldMapping.severity, RagicPortalFieldMapping.module_name).all()

    sev_map = {"high": "高", "medium": "中", "low": "低"}
    for r, m in enumerate(issues, 2):
        row_data = [
            m.company, m.module_name, m.portal_db_table, m.portal_db_field,
            m.issue_type or "—", m.issue_message or "—", m.suggestion or "—",
            sev_map.get(m.severity or "", m.severity or "—"),
            "是" if m.is_resolved else "否",
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if m.severity == "high":
                cell.fill = error_fill
            elif m.severity == "medium":
                cell.fill = warning_fill
    auto_width(ws3)

    # ── Sheet 4：KPI 計算追溯 ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("4_KPI計算追溯")
    headers4 = ["模組名稱", "Portal 路由", "KPI 名稱", "顯示位置",
                "API Endpoint", "DB Table", "使用欄位", "日期依據",
                "篩選條件", "計算公式", "Ragic 原始欄位", "追溯狀態", "備註"]
    write_header(ws4, headers4)

    kpi_rows = db.query(RagicPortalKpiMapping).order_by(RagicPortalKpiMapping.module_name).all()
    trace_map = {"traceable": "可追溯", "partial": "部分可追溯", "untraceable": "無法追溯", "unknown": "未確認"}
    for r, k in enumerate(kpi_rows, 2):
        row_data = [
            k.module_name, k.portal_route, k.kpi_name, k.page_section or "—",
            k.api_endpoint or "—", k.db_table or "—",
            k.source_fields or "—", k.date_field or "—",
            k.filters or "—", k.formula or "—",
            k.ragic_source_fields or "—",
            trace_map.get(k.trace_status, k.trace_status),
            k.issue_message or "—",
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=v)
            cell.border = thin_border
    auto_width(ws4)

    # ── Sheet 5：建議修正清單 ─────────────────────────────────────────────────
    ws5 = wb.create_sheet("5_建議修正清單")
    headers5 = ["優先順序", "嚴重程度", "模組名稱", "欄位名稱", "問題說明", "建議處理", "處理狀態"]
    write_header(ws5, headers5)

    # 依嚴重程度排序：high > medium > low
    sev_order = {"high": 1, "medium": 2, "low": 3}
    sorted_issues = sorted(
        [m for m in all_mappings if m.mapping_status not in ("normal",) and m.issue_message],
        key=lambda m: sev_order.get(m.severity or "low", 9),
    )
    for r, m in enumerate(sorted_issues, 2):
        row_data = [
            r - 1,
            sev_map.get(m.severity or "", "—"),
            m.module_name,
            f"{m.portal_db_table}.{m.portal_db_field}",
            m.issue_message or "—",
            m.suggestion or "—",
            "已處理" if m.is_resolved else "待處理",
        ]
        for c, v in enumerate(row_data, 1):
            cell = ws5.cell(row=r, column=c, value=v)
            cell.border = thin_border
            if m.severity == "high" and not m.is_resolved:
                cell.fill = error_fill
            elif m.severity == "medium" and not m.is_resolved:
                cell.fill = warning_fill
    auto_width(ws5)

    # 輸出 bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def get_audit_summary(db: Session) -> dict:
    """取得整體稽核摘要（首頁 KPI Card）。"""
    # 已設定模組數
    total_modules = len(set(
        info.get("portal_url", "") for info in PORTAL_MODULE_MAP.values()
        if info.get("portal_url")
    ))

    # 已比對模組數
    audited_routes = db.execute(
        text("SELECT COUNT(DISTINCT portal_route) FROM ragic_portal_field_mappings")
    ).scalar() or 0

    # 正常 / 異常模組數
    normal_count = 0
    error_count = 0
    for mod in get_module_overview(db):
        if mod["status"] == "normal":
            normal_count += 1
        elif mod["status"] in ("error", "warning"):
            error_count += 1

    # 未對應欄位數
    unmapped_count = db.query(RagicPortalFieldMapping).filter(
        RagicPortalFieldMapping.mapping_status == "unmapped"
    ).count()

    # 高風險異常數
    high_risk = db.query(RagicPortalFieldMapping).filter(
        RagicPortalFieldMapping.severity == "high",
        RagicPortalFieldMapping.is_resolved == False,
    ).count()

    # 最近一次稽核
    last_run = db.query(RagicPortalAuditRun).order_by(
        RagicPortalAuditRun.run_time.desc()
    ).first()

    return {
        "total_modules": total_modules,
        "audited_modules": audited_routes,
        "normal_modules": normal_count,
        "error_modules": error_count,
        "unmapped_fields": unmapped_count,
        "high_risk_issues": high_risk,
        "last_run_time": last_run.run_time.isoformat() if last_run else None,
        "last_run_status": last_run.status if last_run else None,
    }


# ── Ragic API 欄位同步 ────────────────────────────────────────────────────────

def sync_ragic_fields_from_url(
    db: Session,
    item_no: int,
    ragic_url: str,
    triggered_by: str = "system",
) -> dict:
    """
    從 Ragic API 抓取指定表單的所有欄位定義，儲存至 ragic_portal_field_mappings。

    步驟：
      1. GET {ragic_url}?api=&limit=1  ── 取得一筆資料，從 key 推斷欄位名稱與初步型態
      2. GET {ragic_url}?info=1        ── 嘗試取得欄位 metadata（型態、必填、公式等）
      3. 將 Ragic 欄位與現有 Portal mapping 做比對
         - 有對應的 Portal DB 欄位 → 更新 ragic_* 資訊，mapping_status 改為 normal
         - Ragic 有、Portal 無 → 新建 mapping_status="ragic_only" 的紀錄
      4. 更新同路由所有 mapping 的 ragic_url 欄位（讓 Drawer 可以顯示連結）

    回傳：
      { item_no, ragic_url, portal_route, ragic_field_count,
        synced_count, updated_count, ragic_fields }
    """
    now = twnow()
    headers = {"Authorization": f"Basic {settings.RAGIC_API_KEY}"}

    # ── Step 1：抓一筆資料，推斷欄位名稱與初步型態 ────────────────────────────
    ragic_fields: dict[str, dict] = {}  # { field_name: { field_id, field_type, is_required, is_formula, is_subtable } }
    fetch_error: str = ""

    try:
        resp = httpx.get(
            ragic_url,
            params={"api": "", "limit": "1"},
            headers=headers,
            timeout=30,
            verify=settings.RAGIC_VERIFY_SSL,
        )
        resp.raise_for_status()
        data = resp.json()

        # Ragic 回傳：{ "record_id": { "欄位名": "值", "_ragicId": 123, ... }, ... }
        for record_id, record in data.items():
            if not isinstance(record, dict):
                continue
            for field_key, field_val in record.items():
                if field_key.startswith("_") or field_key.isdigit():
                    continue  # 跳過系統欄位（_ragicId 等）與數字 key
                if field_key not in ragic_fields:
                    ragic_fields[field_key] = {
                        "field_id": "",
                        "field_type": _infer_ragic_type(field_val),
                        "is_required": False,
                        "is_formula": False,
                        "is_subtable": isinstance(field_val, (list, dict)),
                    }
            break  # 只處理第一筆即可
    except Exception as e:
        fetch_error = str(e)

    # ── Step 2：嘗試 ?info=1 取得欄位 metadata（更精確）─────────────────────
    try:
        info_resp = httpx.get(
            ragic_url,
            params={"info": "1"},
            headers=headers,
            timeout=30,
            verify=settings.RAGIC_VERIFY_SSL,
        )
        if info_resp.status_code == 200:
            info_data = info_resp.json()
            if isinstance(info_data, dict):
                # 嘗試兩種常見的 Ragic info 回傳格式
                fields_info = (
                    info_data.get("fields")
                    or info_data.get("column")
                    or {}
                )
                if isinstance(fields_info, dict):
                    for fid, finfo in fields_info.items():
                        if not isinstance(finfo, dict):
                            continue
                        fname = (
                            finfo.get("name")
                            or finfo.get("label")
                            or finfo.get("fieldName")
                        )
                        if not fname or str(fname).isdigit():
                            continue
                        ragic_fields[fname] = {
                            "field_id": str(fid),
                            "field_type": (
                                finfo.get("type")
                                or finfo.get("fieldType")
                                or ragic_fields.get(fname, {}).get("field_type", "text")
                            ),
                            "is_required": bool(
                                finfo.get("required") or finfo.get("isRequired")
                            ),
                            "is_formula": bool(
                                finfo.get("formula") or finfo.get("isFormula")
                            ),
                            "is_subtable": bool(
                                finfo.get("subtable") or finfo.get("isSubtable")
                                or (ragic_fields.get(fname, {}).get("is_subtable", False))
                            ),
                        }
    except Exception:
        pass  # info 端點可選，失敗不影響主流程

    # ── Step 3：取得模組資訊 ──────────────────────────────────────────────────
    mod_info = PORTAL_MODULE_MAP.get(item_no, {})
    portal_route = mod_info.get("portal_url", "")
    module_name = mod_info.get("module", "")
    company = mod_info.get("company", "")
    local_tables = LOCAL_TABLE_MAP.get(item_no, [])

    synced_count = 0   # 新建 ragic_only 紀錄數
    updated_count = 0  # 更新現有 mapping 的 ragic 欄位資訊數

    for fname, fmeta in ragic_fields.items():
        # 尋找此 Ragic 欄位對應的 Portal mapping（依 display_name 或 portal_frontend_field 比對）
        portal_match = _find_portal_match(fname, local_tables, db, portal_route)

        if portal_match:
            # 更新現有 mapping 的 Ragic 欄位資訊
            portal_match.ragic_url = ragic_url
            portal_match.ragic_field_id = fmeta.get("field_id", "")
            portal_match.ragic_field_name = fname
            portal_match.ragic_field_type = fmeta.get("field_type", "")
            portal_match.is_ragic_required = fmeta.get("is_required", False)
            portal_match.is_ragic_formula = fmeta.get("is_formula", False)
            portal_match.is_ragic_subtable = fmeta.get("is_subtable", False)
            portal_match.last_checked_at = now
            # 若之前是 unmapped，有 Ragic 欄位後改為 normal
            if portal_match.mapping_status == "unmapped":
                portal_match.mapping_status = "normal"
                portal_match.severity = None
                portal_match.issue_type = None
                portal_match.issue_message = None
                portal_match.suggestion = None
            updated_count += 1
        else:
            # Ragic 有此欄位，Portal 無對應 → 建立 ragic_only 紀錄（避免重複）
            already = db.query(RagicPortalFieldMapping).filter(
                RagicPortalFieldMapping.portal_route == portal_route,
                RagicPortalFieldMapping.ragic_field_name == fname,
                RagicPortalFieldMapping.mapping_status == "ragic_only",
            ).first()

            if not already:
                new_m = RagicPortalFieldMapping(
                    app_directory_id=item_no,
                    company=company,
                    module_name=module_name,
                    portal_route=portal_route,
                    ragic_url=ragic_url,
                    ragic_field_id=fmeta.get("field_id", ""),
                    ragic_field_name=fname,
                    ragic_field_type=fmeta.get("field_type", "text"),
                    is_ragic_required=fmeta.get("is_required", False),
                    is_ragic_formula=fmeta.get("is_formula", False),
                    is_ragic_subtable=fmeta.get("is_subtable", False),
                    portal_db_table="",
                    portal_db_field="",
                    portal_api_field="",
                    portal_frontend_field="",
                    display_name=fname,   # 以 Ragic 欄位名作為顯示名稱
                    is_displayed=False,
                    is_filter=False,
                    is_export=False,
                    is_calculated=False,
                    mapping_status="ragic_only",
                    severity="medium",
                    issue_type="ragic_only",
                    issue_message=f"Ragic 欄位「{fname}」在 Portal 尚無對應 DB 欄位",
                    suggestion="請確認此欄位是否需要同步到 Portal；若需要請建立對應的 DB 欄位",
                    is_resolved=False,
                    last_checked_at=now,
                )
                db.add(new_m)
                synced_count += 1

    # ── Step 4：為此路由所有 mapping 補上 ragic_url（確保 Drawer 能顯示連結）───
    if portal_route and ragic_url:
        db.query(RagicPortalFieldMapping).filter(
            RagicPortalFieldMapping.portal_route == portal_route,
            RagicPortalFieldMapping.ragic_url.is_(None),
        ).update({"ragic_url": ragic_url})

    db.commit()

    return {
        "item_no": item_no,
        "ragic_url": ragic_url,
        "portal_route": portal_route,
        "ragic_field_count": len(ragic_fields),
        "synced_count": synced_count,
        "updated_count": updated_count,
        "ragic_fields": list(ragic_fields.keys()),
        "fetch_error": fetch_error or None,
    }


def _infer_ragic_type(value) -> str:
    """從 Ragic API 回傳值推斷欄位型態（用於 ?limit=1 回傳）。"""
    if isinstance(value, bool):
        return "checkbox"
    if isinstance(value, (int, float)):
        return "number"
    if isinstance(value, list):
        return "subtable"
    if isinstance(value, dict):
        return "subtable"
    if isinstance(value, str):
        v = value.strip()
        # YYYY-MM-DD
        if len(v) == 10 and v[4:5] == "-" and v[7:8] == "-" and v[:4].isdigit():
            return "date"
        # YYYY-MM-DD HH:MM:SS
        if len(v) >= 16 and v[4:5] == "-" and v[10:11] == " ":
            return "datetime"
    return "text"


def _find_portal_match(
    ragic_field_name: str,
    local_tables: list[str],
    db: Session,
    portal_route: str,
) -> "Optional[RagicPortalFieldMapping]":
    """
    根據 Ragic 欄位名稱（中文），尋找對應的 Portal DB mapping 紀錄。
    比對策略（依優先順序）：
      1. DB 現有 mapping 的 display_name == ragic_field_name
      2. DB 現有 mapping 的 portal_frontend_field == ragic_field_name
      3. PORTAL_FIELD_CONFIG display_names 中有對應的 DB 欄位，且 DB 有此 mapping
      4. PORTAL_FIELD_CONFIG field_aliases 別名對應（處理 Ragic 表單用不同名稱的情況）
    """
    # 1 & 2：從已存在的 mapping 比對
    existing = (
        db.query(RagicPortalFieldMapping)
        .filter(
            RagicPortalFieldMapping.portal_route == portal_route,
            RagicPortalFieldMapping.mapping_status != "ragic_only",
        )
        .all()
    )
    for m in existing:
        if (m.display_name and m.display_name == ragic_field_name) or (
            m.portal_frontend_field and m.portal_frontend_field == ragic_field_name
        ):
            return m

    # 3：查 PORTAL_FIELD_CONFIG display_names
    for table in local_tables:
        cfg = PORTAL_FIELD_CONFIG.get(table, {})
        display_names = cfg.get("display_names", {})
        for db_field, display_name in display_names.items():
            if display_name == ragic_field_name:
                m = (
                    db.query(RagicPortalFieldMapping)
                    .filter(
                        RagicPortalFieldMapping.portal_route == portal_route,
                        RagicPortalFieldMapping.portal_db_table == table,
                        RagicPortalFieldMapping.portal_db_field == db_field,
                    )
                    .first()
                )
                if m:
                    return m

        # 4：查 field_aliases（Ragic 欄位別名）
        aliases = cfg.get("field_aliases", {})
        if ragic_field_name in aliases:
            db_field = aliases[ragic_field_name]
            m = (
                db.query(RagicPortalFieldMapping)
                .filter(
                    RagicPortalFieldMapping.portal_route == portal_route,
                    RagicPortalFieldMapping.portal_db_table == table,
                    RagicPortalFieldMapping.portal_db_field == db_field,
                )
                .first()
            )
            if m:
                return m

    return None


# ── 內部工具函數 ──────────────────────────────────────────────────────────────


def _mapping_to_dict(m: RagicPortalFieldMapping) -> dict:
    return {
        "id": m.id,
        "app_directory_id": m.app_directory_id,
        "company": m.company,
        "module_name": m.module_name,
        "portal_route": m.portal_route,
        "ragic_url": m.ragic_url,
        "ragic_form_name": m.ragic_form_name,
        "ragic_field_id": m.ragic_field_id,
        "ragic_field_name": m.ragic_field_name,
        "ragic_field_type": m.ragic_field_type,
        "is_ragic_required": m.is_ragic_required,
        "is_ragic_formula": m.is_ragic_formula,
        "is_ragic_subtable": m.is_ragic_subtable,
        "portal_db_table": m.portal_db_table,
        "portal_db_field": m.portal_db_field,
        "portal_api_field": m.portal_api_field,
        "portal_frontend_field": m.portal_frontend_field,
        "display_name": m.display_name,
        "is_displayed": m.is_displayed,
        "is_filter": m.is_filter,
        "is_export": m.is_export,
        "is_calculated": m.is_calculated,
        "mapping_status": m.mapping_status,
        "severity": m.severity,
        "issue_type": m.issue_type,
        "issue_message": m.issue_message,
        "suggestion": m.suggestion,
        "notes": m.notes,
        "is_resolved": m.is_resolved,
        "last_checked_at": m.last_checked_at.isoformat() if m.last_checked_at else None,
    }


def _kpi_to_dict(k: RagicPortalKpiMapping) -> dict:
    import json as _json
    def _safe_loads(s):
        if not s:
            return []
        try:
            return _json.loads(s)
        except Exception:
            return []

def _kpi_to_dict(k: RagicPortalKpiMapping) -> dict:
    import json as _json
    def _safe_loads(s):
        if not s:
            return []
        try:
            return _json.loads(s)
        except Exception:
            return []

    return {
        "id": k.id,
        "module_name": k.module_name,
        "portal_route": k.portal_route,
        "kpi_name": k.kpi_name,
        "page_section": k.page_section,
        "api_endpoint": k.api_endpoint,
        "db_table": k.db_table,
        "source_fields": _safe_loads(k.source_fields),
        "date_field": k.date_field,
        "filters": _safe_loads(k.filters),
        "formula": k.formula,
        "ragic_source_fields": _safe_loads(k.ragic_source_fields),
        "trace_status": k.trace_status,
        "last_checked_at": k.last_checked_at.isoformat() if k.last_checked_at else None,
    }
