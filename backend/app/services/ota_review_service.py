"""
OTA 口碑分析 — 評論查詢 service

規格書：`docs/SPEC_ota_reviews.md` §8.1、§9.3
"""
from __future__ import annotations

import io
import json

from sqlalchemy import String, func, or_, select
from sqlalchemy.orm import Session

from app.models.ota_review import OtaReview, OtaSource
from app.schemas.ota_review import (OtaReviewDetailOut, OtaReviewListOut,
                                    OtaReviewRow)
from app.services.ota_normalize import PLATFORM_LABEL

SUMMARY_LIMIT = 80


def _fmt_dt(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def _topics(review: OtaReview) -> list[str]:
    if not review.topics_json:
        return []
    try:
        data = json.loads(review.topics_json)
        return [str(t) for t in data] if isinstance(data, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _summary(review: OtaReview) -> str:
    text = (
        review.title
        or review.negative_text     # 負評優先顯示 —— 列表上最需要一眼看到的是問題
        or review.positive_text
        or review.comment
        or "（僅評分，無留言）"
    )
    text = " ".join(text.split())
    return text[:SUMMARY_LIMIT] + "…" if len(text) > SUMMARY_LIMIT else text


def _build_query(
    *,
    hotel_code: str = "",
    platform: str = "",
    start: str = "",
    end: str = "",
    min_score: float | None = None,
    max_score: float | None = None,
    score_below: float | None = None,
    sentiment: str = "",
    topic: str = "",
    keyword: str = "",
    alert_only: bool = False,
    alert_status: str = "",
    include_duplicate: bool = False,
):
    """
    組出 join 後的基礎查詢。

    ⚠️ 預設 `is_duplicate = False`：跨 OTA 重複的評論不該被算兩次（§5.3）。
       評論清單頁提供「顯示重複」開關，統計端一律不放行。
    """
    stmt = select(OtaReview, OtaSource).join(OtaSource, OtaSource.id == OtaReview.source_id)

    if not include_duplicate:
        stmt = stmt.where(OtaReview.is_duplicate.is_(False))
    if hotel_code:
        stmt = stmt.where(OtaReview.hotel_code == hotel_code)
    if platform:
        stmt = stmt.where(OtaReview.platform == platform)

    # ⚠️ 空日期的評論不落在任何區間內。這是刻意的 ——
    #    解析不出日期的評論不該混進「某個月」的統計（§5.4）。
    if start:
        stmt = stmt.where(OtaReview.review_date >= start, OtaReview.review_date != "")
    if end:
        stmt = stmt.where(OtaReview.review_date <= end, OtaReview.review_date != "")

    # 分數篩選一律用 score_10（統一 10 分制），不可用 score_raw
    if min_score is not None:
        stmt = stmt.where(OtaReview.score_10 >= min_score)
    if max_score is not None:
        stmt = stmt.where(OtaReview.score_10 <= max_score)
    if score_below is not None:
        # ⭐ 「低於某分」的篩選（2026-08-23）。門檻由呼叫端指定，
        #    Dashboard 的「負面評論」KPI 下鑽時帶的是 `NEGATIVE_SCORE_MAX`（6.0）。
        #
        # ⚠️ 是 **`<` 不是 `<=`** —— 與 `get_overview()` 的負評判定一致。
        #    用 `max_score`（`<=`）代替的話會多含一批剛好等於門檻的評論，
        #    於是點了 Dashboard 的「24 則」跳過來卻看到 31 則。
        #    數字對不起來比不能點還糟：使用者會同時不信任兩個數字。
        stmt = stmt.where(OtaReview.score_10 < score_below)

    if sentiment:
        stmt = stmt.where(OtaReview.sentiment_label == sentiment)
    if topic:
        # topics_json 是 JSON array 字串，用 LIKE 比對即可（資料量小，不需另建關聯表）
        stmt = stmt.where(OtaReview.topics_json.cast(String).like(f'%"{topic}%'))
    if alert_only:
        stmt = stmt.where(OtaReview.is_alert.is_(True))
    if alert_status:
        stmt = stmt.where(OtaReview.alert_status == alert_status)

    if keyword:
        term = f"%{keyword.strip()}%"
        stmt = stmt.where(or_(
            OtaReview.author.like(term),
            OtaReview.title.like(term),
            OtaReview.positive_text.like(term),
            OtaReview.negative_text.like(term),
            OtaReview.comment.like(term),
        ))
    return stmt


def list_reviews(db: Session, *, page: int = 1, page_size: int = 50, **filters) -> OtaReviewListOut:
    stmt = _build_query(**filters)

    total = db.execute(
        select(func.count()).select_from(stmt.subquery())
    ).scalar_one()

    # 空日期排最後（比照原本的 Tkinter 工具作法），其餘依日期新到舊
    stmt = stmt.order_by(
        (OtaReview.review_date == "").asc(),
        OtaReview.review_date.desc(),
        OtaReview.id.desc(),
    ).limit(page_size).offset((max(page, 1) - 1) * page_size)

    rows = []
    for review, source in db.execute(stmt).all():
        rows.append(OtaReviewRow(
            id=review.id,
            hotel_code=review.hotel_code,
            hotel_name=source.hotel_name,
            platform=review.platform,
            platform_label=PLATFORM_LABEL.get(review.platform, review.platform),
            author=review.author,
            score_raw=float(review.score_raw) if review.score_raw is not None else None,
            score_scale=review.score_scale,
            score_10=float(review.score_10) if review.score_10 is not None else None,
            summary=_summary(review),
            review_date=review.review_date,
            sentiment_label=review.sentiment_label,
            topics=_topics(review),
            is_alert=review.is_alert,
            alert_status=review.alert_status,
            is_duplicate=review.is_duplicate,
        ))

    return OtaReviewListOut(rows=rows, total=total, page=page, page_size=page_size)


def get_review_detail(db: Session, review_id: int) -> OtaReviewDetailOut | None:
    """
    明細 Drawer 資料（規格書 §9.3）。

    `detail` 用中文欄位名稱作 key，前端逐項渲染 —— 這是 CLAUDE.md §7 的規範，
    只是把 `ragic_url` 換成 `review_url`（本模組沒有 Ragic）。
    """
    row = db.execute(
        select(OtaReview, OtaSource)
        .join(OtaSource, OtaSource.id == OtaReview.source_id)
        .where(OtaReview.id == review_id)
    ).first()
    if row is None:
        return None

    review, source = row
    platform_label = PLATFORM_LABEL.get(review.platform, review.platform)

    # 分數顯示：同時給原始值與換算值，避免使用者看到 Tripadvisor 9.0 分而困惑
    if review.score_10 is None:
        score_text = "—"
    elif review.score_scale == 5:
        score_text = f"{float(review.score_10):.1f} / 10（原始 {float(review.score_raw):.1f}，5 分制）"
    else:
        score_text = f"{float(review.score_10):.1f} / 10"

    detail: dict[str, str] = {
        "評分": score_text,
        "旅客暱稱": review.author or "—",
        "旅客類型": review.traveler_type or "—",
        "國籍": review.nationality or "—",
        "房型": review.room_type or "—",
        "入住晚數": f"{review.nights} 晚" if review.nights else "—",
        "評論日期": review.review_date or "—（原始頁未提供或無法解析）",
        "入住年月": review.stay_month or "—",
        "正面評語": review.positive_text or "—",
        "負面評語": review.negative_text or "—",
        "完整留言": review.comment or "—",
        "情緒判定": review.sentiment_label or "—（尚未分析）",
        "判定來源": {"rule": "規則字典", "ai": "AI 補判", "manual": "人工"}.get(
            review.sentiment_engine, "—"
        ),
        "主題標籤": "、".join(_topics(review)) or "—",
        "跨站重複": "是（已排除於統計外）" if review.is_duplicate else "否",
        "擷取時間": _fmt_dt(review.fetched_at) or "—",
    }

    return OtaReviewDetailOut(
        id=review.id,
        hotel_code=review.hotel_code,
        hotel_name=source.hotel_name,
        platform=review.platform,
        platform_label=platform_label,
        review_url=review.review_url or source.url,
        author=review.author,
        score_raw=float(review.score_raw) if review.score_raw is not None else None,
        score_scale=review.score_scale,
        score_10=float(review.score_10) if review.score_10 is not None else None,
        title=review.title,
        positive_text=review.positive_text,
        negative_text=review.negative_text,
        comment=review.comment,
        review_date=review.review_date,
        stay_month=review.stay_month,
        sentiment_label=review.sentiment_label,
        sentiment_score=float(review.sentiment_score) if review.sentiment_score is not None else None,
        sentiment_engine=review.sentiment_engine,
        topics=_topics(review),
        is_alert=review.is_alert,
        alert_status=review.alert_status,
        alert_note=review.alert_note,
        is_duplicate=review.is_duplicate,
        fetched_at=_fmt_dt(review.fetched_at),
        detail=detail,
    )


def update_alert(
    db: Session,
    review_id: int,
    *,
    alert_status: str,
    alert_note: str,
    user_id: str | None,
) -> OtaReviewDetailOut | None:
    """
    更新負評警示的處理狀態。

    這四欄是**人工營運欄位**，`ota_ingest_service.upsert_reviews` 明確不碰它們。
    """
    from app.core.time import twnow

    review = db.get(OtaReview, review_id)
    if review is None:
        return None

    review.alert_status = alert_status
    review.alert_note = alert_note or ""
    review.alert_handler_id = user_id
    review.alert_handled_at = twnow()
    db.commit()
    return get_review_detail(db, review_id)


# ══════════════════════════════════════════════════════════════════════════
# 匯出
# ══════════════════════════════════════════════════════════════════════════
EXPORT_HEADERS = [
    "飯店代碼", "飯店名稱", "OTA平台", "原始分數", "分制", "統一分數(10分制)",
    "旅客暱稱", "旅客類型", "國籍", "房型", "入住晚數",
    "標題", "正評", "負評", "完整留言",
    "評論日期", "入住年月", "情緒", "主題", "跨站重複", "評論網址", "擷取時間",
]


def export_rows(db: Session, **filters) -> list[list]:
    """
    匯出資料列（欄位與 CSV 匯入格式對齊，可匯出後修改再匯回）。

    刻意不在 service 產生檔案 —— router 決定要輸出 Excel 還是 CSV。
    """
    stmt = _build_query(**filters).order_by(
        (OtaReview.review_date == "").asc(),
        OtaReview.review_date.desc(),
        OtaReview.id.desc(),
    )
    rows = []
    for review, source in db.execute(stmt).all():
        rows.append([
            review.hotel_code,
            source.hotel_name,
            PLATFORM_LABEL.get(review.platform, review.platform),
            float(review.score_raw) if review.score_raw is not None else "",
            review.score_scale,
            float(review.score_10) if review.score_10 is not None else "",
            review.author,
            review.traveler_type,
            review.nationality,
            review.room_type,
            review.nights if review.nights is not None else "",
            review.title,
            review.positive_text,
            review.negative_text,
            review.comment,
            review.review_date,
            review.stay_month,
            review.sentiment_label,
            "、".join(_topics(review)),
            "是" if review.is_duplicate else "否",
            review.review_url or source.url,
            _fmt_dt(review.fetched_at),
        ])
    return rows


def export_xlsx(db: Session, **filters) -> bytes:
    """用 openpyxl 產生 Excel（套件已在 requirements，不新增依賴）。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "OTA評論"

    ws.append(EXPORT_HEADERS)
    header_fill = PatternFill("solid", fgColor="1B3A5C")   # 品牌主色（PROTECTED）
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in export_rows(db, **filters):
        ws.append(row)

    widths = [12, 14, 14, 10, 8, 16, 16, 12, 10, 18, 10,
              24, 40, 40, 40, 12, 12, 10, 20, 10, 40, 20]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
