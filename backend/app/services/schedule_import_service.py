"""
班表 Excel 匯入 Service
支援民國年 Sheet 格式（如：115年5月）
自動辨識日期列、姓名欄，處理未知班別與新增人員。
"""
import re
import uuid
import logging
from datetime import date, datetime
from typing import IO, Any

import openpyxl
from sqlalchemy.orm import Session

from app.models.schedule import (
    Schedule, ScheduleDetail, ScheduleImportLog,
    StaffMember, ShiftType,
)

logger = logging.getLogger(__name__)

# 統計欄位關鍵字（遇到時停止讀取班別）
STAT_KEYWORDS = {"應出勤", "實出勤", "超時加班", "加班", "請假", "見習", "備註", "合計"}

# 星期對應（辨識星期列用）
WEEKDAY_CHARS = {"一", "二", "三", "四", "五", "六", "日", "Mon", "Tue", "Wed",
                 "Thu", "Fri", "Sat", "Sun"}


# ─────────────────────────────────────────────────────────────
# 輔助函數
# ─────────────────────────────────────────────────────────────

def _parse_roc_year_month(sheet_name: str) -> tuple[int, int] | None:
    """從 Sheet 名稱解析民國年月，轉換為西元年。
    支援格式：115年5月、115年05月、115/5、1145
    回傳 (year, month) 或 None
    """
    m = re.search(r"(\d{2,3})[年/](\d{1,2})月?", sheet_name)
    if m:
        roc = int(m.group(1))
        month = int(m.group(2))
        western = roc + 1911
        if 1 <= month <= 12:
            return western, month
    return None


def _parse_name(raw: str) -> dict[str, str]:
    """解析姓名字串，提取 name / employment_type / remark
    例：李宗銘(PT) → name=李宗銘, employment_type=PT
        吳友仁(福群) → name=吳友仁, remark=福群
    """
    raw = str(raw).strip()
    m = re.match(r"^(.+?)[（(](.+?)[）)]$", raw)
    if not m:
        return {"name": raw, "employment_type": "正職", "remark": ""}
    name = m.group(1).strip()
    tag = m.group(2).strip()
    # 判斷是 employment_type 還是 remark
    if tag.upper() in ("PT", "兼職"):
        return {"name": name, "employment_type": "PT", "remark": ""}
    elif tag.upper() in ("支援", "支援人員"):
        return {"name": name, "employment_type": "支援人員", "remark": ""}
    else:
        return {"name": name, "employment_type": "正職", "remark": tag}


def _is_date_row(row_values: list) -> bool:
    """判斷是否為日期列：至少 10 個值落在 1~31"""
    count = 0
    for v in row_values:
        try:
            n = int(str(v).strip())
            if 1 <= n <= 31:
                count += 1
        except (ValueError, TypeError):
            pass
    return count >= 10


def _is_weekday_row(row_values: list) -> bool:
    """判斷是否為星期列：至少 5 個值是星期字元"""
    count = sum(1 for v in row_values if str(v).strip() in WEEKDAY_CHARS)
    return count >= 5


def _is_stat_col(header: str) -> bool:
    """判斷欄位是否為統計欄（應出勤、加班等）"""
    return any(kw in str(header) for kw in STAT_KEYWORDS)


def _is_name_like(val: Any) -> bool:
    """判斷是否為人名（非數字、非空、非星期）"""
    s = str(val).strip()
    if not s or s in WEEKDAY_CHARS:
        return False
    try:
        int(s)
        return False
    except ValueError:
        pass
    # 排除純英文單字班別代碼（2~3 個字母+數字）
    if re.match(r"^[A-Za-z][A-Za-z0-9]{0,3}$", s):
        return False
    return True


# ─────────────────────────────────────────────────────────────
# 主要匯入函數
# ─────────────────────────────────────────────────────────────

def import_excel(
    db: Session,
    file_content: bytes,
    file_name: str,
    override_year: int | None = None,
    override_month: int | None = None,
) -> dict:
    """
    解析 Excel 班表並寫入資料庫。

    回傳結果摘要：
    {
        "schedule_id": str,
        "schedule_year": int,
        "schedule_month": int,
        "total_rows": int,
        "total_details": int,
        "success_count": int,
        "warning_count": int,
        "error_count": int,
        "unknown_shift_codes": list[str],
        "new_staff_names": list[str],
        "message": str,
        "import_batch_id": str,
        "year_month_detected": bool,   # 是否成功自動偵測年月
    }
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    batch_id = str(uuid.uuid4())
    warnings: list[str] = []
    errors: list[str] = []
    unknown_codes: set[str] = set()
    new_staff_names: list[str] = []

    # ── 1. 解析年月 ──────────────────────────────────────────
    ym = _parse_roc_year_month(sheet_name)
    year_month_detected = ym is not None
    if override_year and override_month:
        year, month = override_year, override_month
        year_month_detected = True
    elif ym:
        year, month = ym
    else:
        return {
            "schedule_id": None,
            "year_month_detected": False,
            "message": f"無法從 Sheet 名稱「{sheet_name}」解析年月，請手動指定年月後重試。",
            "import_batch_id": batch_id,
        }

    # ── 2. 確認同月班表是否已存在 ─────────────────────────────
    existing = (
        db.query(Schedule)
        .filter(
            Schedule.schedule_year == year,
            Schedule.schedule_month == month,
            Schedule.is_deleted == False,
        )
        .first()
    )
    if existing:
        return {
            "schedule_id": None,
            "year_month_detected": True,
            "schedule_year": year,
            "schedule_month": month,
            "message": (
                f"{year} 年 {month} 月班表已存在（ID: {existing.id}），"
                "請先刪除舊班表後再重新匯入。"
            ),
            "already_exists": True,
            "import_batch_id": batch_id,
        }

    # ── 3. 讀取工作表資料 ────────────────────────────────────
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return {
            "schedule_id": None,
            "year_month_detected": True,
            "message": "工作表為空，無資料可匯入。",
            "import_batch_id": batch_id,
        }

    # ── 4. 找日期列（掃描前 8 列）────────────────────────────
    date_row_idx: int | None = None
    date_col_map: dict[int, int] = {}   # col_index → day
    weekday_row_idx: int | None = None
    title_text = ""

    for i, row in enumerate(all_rows[:8]):
        vals = [v for v in row]
        if i == 0:
            # 第一列可能是標題
            title_text = " ".join(str(v) for v in vals if v)
        if _is_date_row(vals):
            date_row_idx = i
            # 建立 col → day 對應（跳過第一欄姓名欄）
            for col_i, v in enumerate(vals):
                try:
                    day = int(str(v).strip())
                    if 1 <= day <= 31:
                        date_col_map[col_i] = day
                except (ValueError, TypeError):
                    pass
            break

    if date_row_idx is None:
        return {
            "schedule_id": None,
            "year_month_detected": True,
            "schedule_year": year,
            "schedule_month": month,
            "message": "無法自動辨識日期列（找不到連續 10 個以上的 1~31 數字），請確認 Excel 格式。",
            "import_batch_id": batch_id,
        }

    # 找星期列（日期列相鄰下一列）
    if date_row_idx + 1 < len(all_rows):
        next_row = list(all_rows[date_row_idx + 1])
        if _is_weekday_row(next_row):
            weekday_row_idx = date_row_idx + 1

    # 確定資料起始列（日期列 / 星期列 之後）
    data_start_row = (weekday_row_idx if weekday_row_idx is not None else date_row_idx) + 1

    # ── 5. 統計欄位識別（哪些 col 是統計欄）────────────────────
    stat_col_indices: set[int] = set()
    date_row_vals = list(all_rows[date_row_idx])
    for col_i, v in enumerate(date_row_vals):
        if _is_stat_col(str(v)):
            stat_col_indices.add(col_i)

    # ── 6. 讀取班別 lookup ──────────────────────────────────
    shift_lookup: dict[str, ShiftType] = {
        s.code.upper(): s
        for s in db.query(ShiftType).filter(ShiftType.is_deleted == False).all()
    }

    # ── 7. 讀取人員 lookup ──────────────────────────────────
    staff_lookup: dict[str, StaffMember] = {
        s.source_name: s
        for s in db.query(StaffMember).filter(StaffMember.is_deleted == False).all()
    }

    # ── 8. 建立 Schedule 主檔 ────────────────────────────────
    schedule = Schedule(
        schedule_year=year,
        schedule_month=month,
        title=title_text[:200] if title_text else f"{year}年{month}月班表",
        source_file_name=file_name,
        import_batch_id=batch_id,
        status="imported",
        raw_summary={},
    )
    db.add(schedule)
    db.flush()  # 取得 schedule.id

    # ── 9. 逐列讀取班表資料 ──────────────────────────────────
    details: list[ScheduleDetail] = []
    raw_summary: dict[str, dict] = {}
    total_rows = 0

    for row_vals in all_rows[data_start_row:]:
        if not row_vals:
            continue
        # 第一欄為姓名
        raw_name = row_vals[0]
        if not raw_name or not _is_name_like(raw_name):
            continue

        raw_name_str = str(raw_name).strip()
        total_rows += 1

        # 解析姓名
        parsed = _parse_name(raw_name_str)
        name = parsed["name"]
        employment_type = parsed["employment_type"]
        remark = parsed["remark"]

        # 取得或建立人員
        staff = staff_lookup.get(raw_name_str)
        if not staff:
            staff = StaffMember(
                name=name,
                source_name=raw_name_str,
                employment_type=employment_type,
                remark=remark,
            )
            db.add(staff)
            db.flush()
            staff_lookup[raw_name_str] = staff
            new_staff_names.append(raw_name_str)
            logger.info("schedule_import: 新增人員 %s", raw_name_str)

        # 統計欄位暫存
        person_stats: dict[str, Any] = {}

        # 逐欄讀取班別
        for col_i, val in enumerate(row_vals):
            if col_i == 0:
                continue  # 姓名欄跳過

            # 統計欄位
            if col_i in stat_col_indices:
                header = str(date_row_vals[col_i]).strip()
                if val is not None:
                    try:
                        person_stats[header] = float(val)
                    except (ValueError, TypeError):
                        person_stats[header] = str(val)
                continue

            day = date_col_map.get(col_i)
            if day is None:
                continue

            # 空白 = 未排班，跳過
            if val is None or str(val).strip() == "":
                continue

            shift_code = str(val).strip()
            raw_value = shift_code

            # 嘗試建立工作日期
            try:
                work_date = date(year, month, day)
            except ValueError:
                warnings.append(f"{year}/{month}/{day} 日期無效，已略過")
                continue

            # 比對班別
            shift = shift_lookup.get(shift_code.upper())
            shift_type_id = shift.id if shift else None
            start_time = shift.start_time if shift else ""
            end_time = shift.end_time if shift else ""
            work_minutes = shift.work_minutes if shift else 0

            if not shift:
                unknown_codes.add(shift_code)
                warnings.append(f"未知班別代碼「{shift_code}」（{name} {work_date}）")

            detail = ScheduleDetail(
                schedule_id=schedule.id,
                work_date=work_date,
                staff_id=staff.id,
                staff_name=name,   # 純姓名（已去除括號備註），source_name 保存在 schedule_staff_members
                shift_code=shift_code,
                shift_type_id=shift_type_id,
                start_time=start_time,
                end_time=end_time,
                work_minutes=work_minutes,
                raw_value=raw_value,
            )
            details.append(detail)

        # 儲存統計欄位
        if person_stats:
            raw_summary[raw_name_str] = person_stats

    # ── 10. 批次寫入明細 ─────────────────────────────────────
    db.bulk_save_objects(details)
    schedule.raw_summary = raw_summary
    db.flush()

    # ── 11. 寫入匯入紀錄 ─────────────────────────────────────
    import_log = ScheduleImportLog(
        import_batch_id=batch_id,
        file_name=file_name,
        sheet_name=sheet_name,
        schedule_year=year,
        schedule_month=month,
        total_rows=total_rows,
        total_details=len(details),
        success_count=len(details) - len(warnings),
        warning_count=len(warnings),
        error_count=len(errors),
        unknown_shift_codes=sorted(unknown_codes),
        new_staff_names=new_staff_names,
        message="\n".join(warnings[:50]),  # 最多記 50 筆警告
    )
    db.add(import_log)
    db.commit()

    logger.info(
        "schedule_import: %d/%d 完成，明細 %d 筆，警告 %d，未知班別 %s",
        year, month, len(details), len(warnings), sorted(unknown_codes),
    )

    return {
        "schedule_id": schedule.id,
        "schedule_year": year,
        "schedule_month": month,
        "year_month_detected": year_month_detected,
        "already_exists": False,
        "total_rows": total_rows,
        "total_details": len(details),
        "success_count": len(details),
        "warning_count": len(warnings),
        "error_count": len(errors),
        "unknown_shift_codes": sorted(unknown_codes),
        "new_staff_names": new_staff_names,
        "message": f"匯入完成：{year}年{month}月，共 {len(details)} 筆明細",
        "import_batch_id": batch_id,
        "warnings": warnings[:50],
    }
