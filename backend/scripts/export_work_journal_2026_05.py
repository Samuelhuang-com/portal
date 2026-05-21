"""
export_work_journal_2026_05.py
═══════════════════════════════════════════════════════════════════════
匯出「工作日誌」TAB 2026/05 所有明細 → Excel 6 個 Sheet

執行方式（在 backend/ 目錄下）：
    python scripts/export_work_journal_2026_05.py

輸出：portal/exports/工作日誌明細_2026_05.xlsx
      （若 exports/ 不存在會自動建立）

資料來源：
  - luqun_repair_case       → 商場工務（依關鍵字分類）
  - dazhi_repair_case       → 大直工務（依關鍵字分類）
  - ihg_rm_master           → IHG 客房保養（固定「例行維護」）
  - hotel_di_inspection_batch → 飯店每日巡檢（固定「每日巡檢」）

統計月份口徑（與 work_category_analysis.py 完全一致）：
  luqun / dazhi：completed_at（已完成）/ occurred_at（未完成）
  ihg_room     ：maint_date
  hotel_di     ：inspection_date（double fallback → start_time）
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path

# ── 確保 import app.* 可用 ──────────────────────────────────────────────────
BACKEND_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.models.luqun_repair import LuqunRepairCase
from app.models.dazhi_repair import DazhiRepairCase
from app.models.ihg_room_maintenance import IHGRoomMaintenanceMaster
from app.models.hotel_daily_inspection import HotelDIBatch
from app.services.time_utils import parse_minutes as _parse_di_minutes

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════════
# 常數
# ══════════════════════════════════════════════════════════════════════════════

TARGET_YEAR  = 2026
TARGET_MONTH = 5

CATEGORIES = ["現場報修", "上級交辦", "緊急事件", "例行維護", "每日巡檢"]

SOURCE_LABELS = {
    "luqun":    "商場工務",
    "dazhi":    "大直工務",
    "ihg_room": "IHG客房保養",
    "hotel_di": "飯店每日巡檢",
}

_CATEGORY_RULES: list[tuple[str, list[str]]] = [
    ("緊急事件",  ["緊急", "急修", "突發", "漏電緊急", "火警", "停電"]),
    ("每日巡檢",  ["巡檢", "巡視", "例巡", "日巡"]),
    ("例行維護",  ["例行", "定期", "保養", "維護", "定保", "年保", "季保", "月保"]),
    ("上級交辦",  ["交辦", "上級", "主管指示", "主管交辦", "院長", "指示", "指派"]),
]

# 欄位定義
REPAIR_COLS   = ["#", "案件編號", "標題", "報修類型", "來源", "人員", "樓層/區域", "狀態",
                 "報修日期", "完工日期", "工時(HR)", "工項類別"]
IHG_COLS      = ["#", "房號", "樓層", "保養月份", "實際保養日", "人員", "狀態", "工時(HR)", "工項類別"]
HOTEL_DI_COLS = ["#", "Sheet區域", "巡檢人員", "巡檢日期", "開始時間", "結束時間", "工時(HR)", "工項類別"]

# ══════════════════════════════════════════════════════════════════════════════
# 工具函式（與 work_category_analysis.py 相同邏輯）
# ══════════════════════════════════════════════════════════════════════════════

def _classify(title: str, repair_type: str) -> str:
    text = (title or "") + (repair_type or "")
    for cat, keywords in _CATEGORY_RULES:
        if any(kw in text for kw in keywords):
            return cat
    return "現場報修"


def _parse_minutes_to_hours(val: str) -> float:
    if not val:
        return 0.0
    m = re.search(r"[\d.]+", str(val))
    return float(m.group()) / 60.0 if m else 0.0


def _parse_hotel_date(date_str: str):
    if not date_str:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d"):
        try:
            d = datetime.strptime(date_str.strip()[:10].replace("-", "/"), "%Y/%m/%d")
            return (d.year, d.month, d.day)
        except ValueError:
            pass
    return None


def _time_only(s: str) -> str:
    s = (s or "").strip()
    return s.rsplit(" ", 1)[-1] if " " in s else s


def _stat_dt(c) -> datetime | None:
    return c.completed_at if c.completed_at else c.occurred_at


def _fmt_dt(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y/%m/%d")
    return str(dt)


# ══════════════════════════════════════════════════════════════════════════════
# 資料載入
# ══════════════════════════════════════════════════════════════════════════════

def load_repair_rows(db) -> list[dict]:
    """luqun + dazhi → 依類別分組的明細列"""
    rows = []

    for c in db.query(LuqunRepairCase).all():
        if not c.occurred_at or not c.work_hours or c.work_hours <= 0:
            continue
        dt = _stat_dt(c)
        if not dt:
            continue
        if dt.year != TARGET_YEAR or dt.month != TARGET_MONTH:
            continue
        cat = _classify(c.title or "", c.repair_type or "")
        rows.append({
            "案件編號":  c.case_no or c.ragic_id,
            "標題":      c.title or "",
            "報修類型":  c.repair_type or "",
            "來源":      SOURCE_LABELS["luqun"],
            "人員":      (c.responsible_unit or "").strip() or "未指定",
            "樓層/區域": c.floor or "",
            "狀態":      c.status or "",
            "報修日期":  _fmt_dt(c.occurred_at),
            "完工日期":  _fmt_dt(c.completed_at),
            "工時(HR)":  round(c.work_hours, 2),
            "工項類別":  cat,
        })

    for c in db.query(DazhiRepairCase).all():
        if not c.occurred_at or not c.work_hours or c.work_hours <= 0:
            continue
        dt = _stat_dt(c)
        if not dt:
            continue
        if dt.year != TARGET_YEAR or dt.month != TARGET_MONTH:
            continue
        cat = _classify(c.title or "", c.repair_type or "")
        rows.append({
            "案件編號":  c.case_no or c.ragic_id,
            "標題":      c.title or "",
            "報修類型":  c.repair_type or "",
            "來源":      SOURCE_LABELS["dazhi"],
            "人員":      (c.closer or "").strip() or "未指定",
            "樓層/區域": c.floor or "",
            "狀態":      c.status or "",
            "報修日期":  _fmt_dt(c.occurred_at),
            "完工日期":  _fmt_dt(c.completed_at),
            "工時(HR)":  round(c.work_hours, 2),
            "工項類別":  cat,
        })

    return rows


def load_ihg_rows(db) -> list[dict]:
    rows = []
    for rec in db.query(IHGRoomMaintenanceMaster).all():
        # 工時
        if rec.work_minutes is not None and rec.work_minutes > 0:
            hours = rec.work_minutes / 60.0
        else:
            try:
                raw = rec.get_raw()
                mins_val = raw.get("工時計算", "")
                if mins_val in (None, "", "None"):
                    continue
                hours = _parse_minutes_to_hours(str(mins_val))
            except Exception:
                continue
        if hours <= 0:
            continue
        # 日期
        yd = _parse_hotel_date(rec.maint_date) if rec.maint_date else None
        if yd is None:
            try:
                yd = (int(rec.maint_year), int(rec.maint_month), 1)
            except (ValueError, TypeError):
                continue
        if yd[0] != TARGET_YEAR or yd[1] != TARGET_MONTH:
            continue
        rows.append({
            "房號":      rec.room_no or "",
            "樓層":      rec.floor or "",
            "保養月份":  f"{yd[0]}/{yd[1]:02d}",
            "實際保養日": rec.maint_date or f"{yd[0]}/{yd[1]:02d}/01",
            "人員":      (rec.assignee_name or "").strip() or "未指定",
            "狀態":      rec.status or "",
            "工時(HR)":  round(hours, 2),
            "工項類別":  "例行維護",
        })
    return rows


def load_hotel_di_rows(db) -> list[dict]:
    rows = []
    for b in db.query(HotelDIBatch).all():
        yd = _parse_hotel_date(b.inspection_date) if b.inspection_date else None
        if yd is None and b.start_time:
            yd = _parse_hotel_date(b.start_time)
        if not yd:
            continue
        if yd[0] != TARGET_YEAR or yd[1] != TARGET_MONTH:
            continue

        mins = _parse_di_minutes(
            _time_only(b.start_time or ""),
            _time_only(b.end_time or ""),
        )
        if mins > 0:
            hours = mins / 60.0
        else:
            m = re.search(r"[\d.]+", b.work_hours or "")
            hours = float(m.group()) if m else 0.0

        rows.append({
            "Sheet區域":  b.sheet_name or b.sheet_key or "",
            "巡檢人員":   (b.inspector_name or "").strip() or "未指定",
            "巡檢日期":   b.inspection_date or "",
            "開始時間":   _time_only(b.start_time or ""),
            "結束時間":   _time_only(b.end_time or ""),
            "工時(HR)":   round(hours, 2),
            "工項類別":   "每日巡檢",
        })
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# Excel 格式工具
# ══════════════════════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill("solid", start_color="1B3A5C")
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
DATA_FONT    = Font(name="Arial", size=10)
ALT_FILL     = PatternFill("solid", start_color="EEF4FB")
TOTAL_FILL   = PatternFill("solid", start_color="4BA8E8")
TOTAL_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
THIN_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                      top=BORDER_SIDE, bottom=BORDER_SIDE)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN   = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGN  = Alignment(horizontal="right", vertical="center")

CAT_COLORS = {
    "現場報修": "1B3A5C",
    "上級交辦": "D46B08",
    "緊急事件": "CF1322",
    "例行維護": "389E0D",
    "每日巡檢": "096DD9",
}


def _write_header(ws, cols: list[str]) -> None:
    for ci, col in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22


def _write_data_row(ws, row_no: int, values: list, alt: bool = False) -> None:
    fill = ALT_FILL if alt else None
    for ci, val in enumerate(values, start=1):
        cell = ws.cell(row=row_no, column=ci, value=val)
        cell.font = DATA_FONT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        # 數字靠右
        if isinstance(val, (int, float)):
            cell.alignment = RIGHT_ALIGN
        elif ci == 1:  # 序號置中
            cell.alignment = CENTER_ALIGN
        else:
            cell.alignment = LEFT_ALIGN
    ws.row_dimensions[row_no].height = 18


def _write_total_row(ws, row_no: int, label: str, count: int,
                     hours: float, cols: list[str]) -> None:
    ws.cell(row=row_no, column=1, value="合計")
    # 找件數欄（無則不填）
    n_col  = None
    hr_col = None
    for i, c in enumerate(cols, start=1):
        if c == "#":
            n_col = i
        if "工時" in c:
            hr_col = i

    for ci in range(1, len(cols) + 1):
        cell = ws.cell(row=row_no, column=ci)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = RIGHT_ALIGN

    ws.cell(row=row_no, column=1).value = "合計"
    ws.cell(row=row_no, column=1).alignment = CENTER_ALIGN
    if n_col:
        ws.cell(row=row_no, column=n_col).value = f"{count} 件"
    if hr_col:
        ws.cell(row=row_no, column=hr_col).value = round(hours, 2)
    ws.row_dimensions[row_no].height = 22


def _auto_col_width(ws, cols: list[str]) -> None:
    MIN_W, MAX_W = 8, 40
    for ci, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(ci)
        max_len = len(col_name)
        for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        # CJK 字元約 2 倍寬
        cjk_count = sum(1 for c in col_name if '一' <= c <= '鿿')
        adjusted = max_len + cjk_count * 0.8
        ws.column_dimensions[col_letter].width = min(MAX_W, max(MIN_W, adjusted + 2))


# ══════════════════════════════════════════════════════════════════════════════
# Sheet 寫入
# ══════════════════════════════════════════════════════════════════════════════

def write_repair_sheet(ws, rows: list[dict], category: str) -> tuple[int, float]:
    """寫入修繕類 sheet（現場報修 / 上級交辦 / 緊急事件 / 例行維護-報修部分）"""
    filtered = [r for r in rows if r["工項類別"] == category]
    cols = REPAIR_COLS
    _write_header(ws, cols)

    total_hrs = 0.0
    for i, r in enumerate(filtered, start=1):
        vals = [
            i,
            r["案件編號"],
            r["標題"],
            r["報修類型"],
            r["來源"],
            r["人員"],
            r["樓層/區域"],
            r["狀態"],
            r["報修日期"],
            r["完工日期"],
            r["工時(HR)"],
            r["工項類別"],
        ]
        _write_data_row(ws, i + 1, vals, alt=(i % 2 == 0))
        total_hrs += r["工時(HR)"]

    total_row = len(filtered) + 2
    _write_total_row(ws, total_row, "合計", len(filtered), total_hrs, cols)
    _auto_col_width(ws, cols)
    ws.freeze_panes = "A2"
    return len(filtered), total_hrs


def write_ihg_sheet(ws, rows: list[dict]) -> tuple[int, float]:
    cols = IHG_COLS
    _write_header(ws, cols)
    total_hrs = 0.0
    for i, r in enumerate(rows, start=1):
        vals = [i, r["房號"], r["樓層"], r["保養月份"], r["實際保養日"],
                r["人員"], r["狀態"], r["工時(HR)"], r["工項類別"]]
        _write_data_row(ws, i + 1, vals, alt=(i % 2 == 0))
        total_hrs += r["工時(HR)"]
    total_row = len(rows) + 2
    _write_total_row(ws, total_row, "合計", len(rows), total_hrs, cols)
    _auto_col_width(ws, cols)
    ws.freeze_panes = "A2"
    return len(rows), total_hrs


def write_hotel_di_sheet(ws, rows: list[dict]) -> tuple[int, float]:
    cols = HOTEL_DI_COLS
    _write_header(ws, cols)
    total_hrs = 0.0
    for i, r in enumerate(rows, start=1):
        vals = [i, r["Sheet區域"], r["巡檢人員"], r["巡檢日期"],
                r["開始時間"], r["結束時間"], r["工時(HR)"], r["工項類別"]]
        _write_data_row(ws, i + 1, vals, alt=(i % 2 == 0))
        total_hrs += r["工時(HR)"]
    total_row = len(rows) + 2
    _write_total_row(ws, total_row, "合計", len(rows), total_hrs, cols)
    _auto_col_width(ws, cols)
    ws.freeze_panes = "A2"
    return len(rows), total_hrs


def write_summary_sheet(ws, summary: dict) -> None:
    """加總 sheet：各類別件數 + 工時"""
    ws.title = "加總"

    # 標題列
    header_cols = ["工項類別", "件數", "工時(HR)", "平均工時/件"]
    _write_header(ws, header_cols)

    grand_count = 0
    grand_hrs   = 0.0

    for i, cat in enumerate(CATEGORIES, start=1):
        count, hrs = summary.get(cat, (0, 0.0))
        avg = round(hrs / count, 2) if count > 0 else 0.0
        color = CAT_COLORS.get(cat, "000000")
        vals = [cat, count, round(hrs, 2), avg]
        for ci, val in enumerate(vals, start=1):
            cell = ws.cell(row=i + 1, column=ci, value=val)
            cell.font = Font(name="Arial", size=10,
                             color=color if ci == 1 else "000000",
                             bold=(ci == 1))
            cell.fill = ALT_FILL if i % 2 == 0 else PatternFill()
            cell.border = THIN_BORDER
            cell.alignment = RIGHT_ALIGN if isinstance(val, (int, float)) else LEFT_ALIGN
        ws.row_dimensions[i + 1].height = 18
        grand_count += count
        grand_hrs   += hrs

    # 總計列
    total_row = len(CATEGORIES) + 2
    total_avg = round(grand_hrs / grand_count, 2) if grand_count > 0 else 0.0
    for ci, val in enumerate(["總計", grand_count, round(grand_hrs, 2), total_avg], start=1):
        cell = ws.cell(row=total_row, column=ci, value=val)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER
        cell.alignment = RIGHT_ALIGN if isinstance(val, (int, float)) else CENTER_ALIGN
    ws.row_dimensions[total_row].height = 22

    # 欄寬
    for ci, col_name in enumerate(header_cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = [18, 10, 12, 14][ci - 1]


# ══════════════════════════════════════════════════════════════════════════════
# 主程式
# ══════════════════════════════════════════════════════════════════════════════

def main():
    # ── DB 連線 ───────────────────────────────────────────────────────────────
    db_url = settings.DATABASE_URL
    print(f"[INFO] 連接資料庫：{db_url}")
    engine = create_engine(db_url, connect_args={"check_same_thread": False})
    Session = sessionmaker(bind=engine)
    db = Session()

    try:
        print(f"[INFO] 載入資料（{TARGET_YEAR}/{TARGET_MONTH:02d}）…")
        repair_rows   = load_repair_rows(db)
        ihg_rows      = load_ihg_rows(db)
        hotel_di_rows = load_hotel_di_rows(db)
    finally:
        db.close()

    print(f"[INFO]   報修類（luqun+dazhi）: {len(repair_rows)} 筆")
    print(f"[INFO]   IHG 客房保養         : {len(ihg_rows)} 筆")
    print(f"[INFO]   飯店每日巡檢         : {len(hotel_di_rows)} 筆")

    # ── 建立 Excel ────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 刪除預設空 sheet

    summary: dict[str, tuple[int, float]] = {}

    # 1. 現場報修
    ws1 = wb.create_sheet("現場報修")
    c, h = write_repair_sheet(ws1, repair_rows, "現場報修")
    summary["現場報修"] = (c, h)
    print(f"[INFO] 現場報修 → {c} 件 / {h:.2f} HR")

    # 2. 上級交辦
    ws2 = wb.create_sheet("上級交辦")
    c, h = write_repair_sheet(ws2, repair_rows, "上級交辦")
    summary["上級交辦"] = (c, h)
    print(f"[INFO] 上級交辦 → {c} 件 / {h:.2f} HR")

    # 3. 緊急事件
    ws3 = wb.create_sheet("緊急事件")
    c, h = write_repair_sheet(ws3, repair_rows, "緊急事件")
    summary["緊急事件"] = (c, h)
    print(f"[INFO] 緊急事件 → {c} 件 / {h:.2f} HR")

    # 4. 例行維護（報修來源 keyword + IHG）
    ws4 = wb.create_sheet("例行維護")
    # 先寫 repair 部分（例行維護 keyword 命中）
    repair_routine = [r for r in repair_rows if r["工項類別"] == "例行維護"]

    # 合併 IHG 轉成統一格式
    combined_routine: list[dict] = []
    for r in repair_routine:
        combined_routine.append({
            "type": "repair",
            "案件編號": r["案件編號"],
            "標題":     r["標題"],
            "報修類型": r["報修類型"],
            "來源":     r["來源"],
            "人員":     r["人員"],
            "樓層/區域": r["樓層/區域"],
            "狀態":     r["狀態"],
            "報修日期": r["報修日期"],
            "完工日期": r["完工日期"],
            "工時(HR)": r["工時(HR)"],
        })
    for r in ihg_rows:
        combined_routine.append({
            "type": "ihg",
            "案件編號": r["房號"],
            "標題":     f"IHG客房保養 {r['房號']}",
            "報修類型": "客房保養",
            "來源":     "IHG客房保養",
            "人員":     r["人員"],
            "樓層/區域": r["樓層"],
            "狀態":     r["狀態"],
            "報修日期": r["實際保養日"],
            "完工日期": r["實際保養日"],
            "工時(HR)": r["工時(HR)"],
        })

    # 用統一欄位寫入
    cols_r = REPAIR_COLS
    _write_header(ws4, cols_r)
    total_hrs_r = 0.0
    for i, r in enumerate(combined_routine, start=1):
        vals = [
            i, r["案件編號"], r["標題"], r["報修類型"],
            r["來源"], r["人員"], r["樓層/區域"], r["狀態"],
            r["報修日期"], r["完工日期"], r["工時(HR)"], "例行維護",
        ]
        _write_data_row(ws4, i + 1, vals, alt=(i % 2 == 0))
        total_hrs_r += r["工時(HR)"]
    _write_total_row(ws4, len(combined_routine) + 2, "合計",
                     len(combined_routine), total_hrs_r, cols_r)
    _auto_col_width(ws4, cols_r)
    ws4.freeze_panes = "A2"
    summary["例行維護"] = (len(combined_routine), total_hrs_r)
    print(f"[INFO] 例行維護 → {len(combined_routine)} 件 / {total_hrs_r:.2f} HR "
          f"（報修 {len(repair_routine)} + IHG {len(ihg_rows)}）")

    # 5. 每日巡檢
    ws5 = wb.create_sheet("每日巡檢")
    # 也包含 repair 裡的「每日巡檢」keyword 命中（通常 0 筆，但保留一致性）
    repair_di = [r for r in repair_rows if r["工項類別"] == "每日巡檢"]

    combined_di: list[dict] = []
    for r in repair_di:
        combined_di.append({
            "Sheet區域": r["來源"],
            "巡檢人員":  r["人員"],
            "巡檢日期":  r["報修日期"],
            "開始時間":  "",
            "結束時間":  "",
            "工時(HR)":  r["工時(HR)"],
        })
    for r in hotel_di_rows:
        combined_di.append(r)

    cols_di = HOTEL_DI_COLS
    _write_header(ws5, cols_di)
    total_hrs_di = 0.0
    for i, r in enumerate(combined_di, start=1):
        vals = [
            i, r["Sheet區域"], r["巡檢人員"], r["巡檢日期"],
            r["開始時間"], r["結束時間"], r["工時(HR)"], "每日巡檢",
        ]
        _write_data_row(ws5, i + 1, vals, alt=(i % 2 == 0))
        total_hrs_di += r["工時(HR)"]
    _write_total_row(ws5, len(combined_di) + 2, "合計",
                     len(combined_di), total_hrs_di, cols_di)
    _auto_col_width(ws5, cols_di)
    ws5.freeze_panes = "A2"
    summary["每日巡檢"] = (len(combined_di), total_hrs_di)
    print(f"[INFO] 每日巡檢 → {len(combined_di)} 件 / {total_hrs_di:.2f} HR "
          f"（報修 {len(repair_di)} + 每日巡檢 {len(hotel_di_rows)}）")

    # 6. 加總
    ws6 = wb.create_sheet("加總")
    write_summary_sheet(ws6, summary)

    # ── 輸出 ────────────────────────────────────────────────────────────────
    out_dir = BACKEND_DIR.parent / "exports"
    out_dir.mkdir(exist_ok=True)
    out_path = out_dir / f"工作日誌明細_{TARGET_YEAR}_{TARGET_MONTH:02d}.xlsx"
    wb.save(str(out_path))
    print(f"\n✅ 已輸出：{out_path}")
    print("   請用 Excel 開啟確認。")


if __name__ == "__main__":
    main()
