"""
합約管理 Excel 샘플 데이터 → Portal DB 匯入 스크립트
사용법: cd portal/backend && python scripts/import_contract_sample.py
"""
import sys
import os
import json
from datetime import datetime

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.dirname(SCRIPT_DIR)
PORTAL_DIR  = os.path.dirname(BACKEND_DIR)

# Excel 파일 위치
EXCEL_PATH = os.path.join(PORTAL_DIR, '合約管理Excel模板_各部門審閱版_含範本資料_已修改.xlsx')

# 실제 서버 DB를 사용하도록 app 경로 설정
sys.path.insert(0, BACKEND_DIR)

# dotenv 로드 (DATABASE_URL 읽기)
try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(BACKEND_DIR, '.env'))
except ImportError:
    pass  # python-dotenv 없어도 환경변수 직접 사용

# ── DB 연결 ──────────────────────────────────────────────────────────────────
from app.core.config import settings
from app.core.database import engine
from sqlalchemy import text

# 테이블 자동 생성
import app.models.contract  # noqa — ContractClaim, Contract, Vendor 등 등록
from app.core.database import Base
Base.metadata.create_all(bind=engine)
print(f"DB: {settings.DATABASE_URL}")

from sqlalchemy.orm import Session

# ── 유틸 함수 ─────────────────────────────────────────────────────────────────
def yn_to_bool(v):
    if v is None:
        return None
    return str(v).strip() in ('是', 'Y', 'y', 'True', '1', 'true')

def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    if len(s) == 10:
        return s
    return s[:10]

# ── Excel 읽기 ────────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("openpyxl 없음. 설치: pip install openpyxl")
    sys.exit(1)

if not os.path.exists(EXCEL_PATH):
    print(f"Excel 파일을 찾을 수 없습니다: {EXCEL_PATH}")
    print("포털 루트 폴더에 '합約管理Excel模板_各部門審閱版_含範本資料_已修改.xlsx' 파일이 있어야 합니다.")
    sys.exit(1)

print(f"Excel: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)

# ── 1. 廠商 匯入 ──────────────────────────────────────────────────────────────
ws_v = wb['10_供應商主檔']
rows_v = list(ws_v.iter_rows(values_only=True))
header_idx = next(i for i, r in enumerate(rows_v) if r[0] == '廠商編號')
data_v = [r for r in rows_v[header_idx + 1:] if r[0] is not None]

vendor_inserted = vendor_skipped = 0
now = datetime.now().isoformat()

with Session(engine) as db:
    for row in data_v:
        vid = str(row[0]).strip()
        exists = db.execute(text("SELECT 1 FROM vendors WHERE vendor_id=:id"), {"id": vid}).fetchone()
        if exists:
            vendor_skipped += 1
            continue
        db.execute(text("""
            INSERT INTO vendors
            (vendor_id, vendor_name, tax_id, contact_person, phone, email, address,
             payment_terms, bank_name, bank_account, vendor_type, risk_level, is_critical,
             created_at, updated_at)
            VALUES (:vid,:vname,:tax,:contact,:phone,:email,:addr,
                    :pt,:bank,:bacct,:vtype,:risk,:critical,:now,:now2)
        """), {
            "vid":      vid,
            "vname":    row[1],
            "tax":      row[2],
            "contact":  row[3],
            "phone":    row[4],
            "email":    row[5],
            "addr":     row[6],
            "pt":       row[7],
            "bank":     row[8],
            "bacct":    row[9],
            "vtype":    row[10],
            "risk":     row[11],
            "critical": yn_to_bool(row[12]),
            "now":      now,
            "now2":     now,
        })
        vendor_inserted += 1

    db.commit()

print(f"廠商：新增 {vendor_inserted} 筆，跳過（已存在）{vendor_skipped} 筆")

# ── 2. 合約 匯入 ──────────────────────────────────────────────────────────────
ws_c = wb['01_合約主檔']
rows_c = list(ws_c.iter_rows(values_only=True))
header_idx_c = next(i for i, r in enumerate(rows_c) if r[0] == '合約編號')
data_c = [r for r in rows_c[header_idx_c + 1:] if r[0] is not None]

contract_inserted = contract_skipped = 0

with Session(engine) as db:
    for row in data_c:
        cid = str(row[0]).strip()
        exists = db.execute(text("SELECT 1 FROM contracts WHERE contract_id=:id"), {"id": cid}).fetchone()
        if exists:
            contract_skipped += 1
            continue

        detail = {}
        for label, val in [
            ('個資或資安條款', row[30]),
            ('保密條款',       row[31]),
            ('違約金條款',     row[32]),
            ('SLA服務水準',    row[33]),
            ('保固條款',       row[34]),
            ('是否需請購單',   row[17]),
            ('是否可無請購請款', row[18]),
            ('是否需分攤',     row[19]),
            ('分攤方式',       row[20]),
            ('是否需驗收',     row[37]),
        ]:
            if val is not None:
                detail[label] = str(val)

        db.execute(text("""
            INSERT INTO contracts
            (contract_id, contract_name, contract_type, contract_status,
             responsible_dept, using_depts, vendor_id, vendor_name,
             start_date, end_date, notification_days, auto_renewal,
             currency, total_amount_tax_included, monthly_fixed_amount, pricing_method,
             needs_purchase_order, can_claim_without_po, needs_allocation, allocation_method,
             budget_year, budget_category_l1, budget_category_l2, accounting_code,
             budget_source, budget_control_method, require_acceptance,
             risk_level, manager, reviewer, attachment_url, remarks,
             detail, created_at, updated_at)
            VALUES
            (:cid,:cname,:ctype,:cstatus,
             :rdept,:udepts,:vid,:vname,
             :sdate,:edate,:ndays,:autorenewal,
             :currency,:total,:monthly,:pricing,
             :needs_po,:can_claim,:needs_alloc,:alloc_method,
             :byear,:bl1,:bl2,:accounting,
             :bsource,:bctrl,:require_acc,
             :risk,:manager,:reviewer,:attach,:remarks,
             :detail,:now,:now2)
        """), {
            "cid":         cid,
            "cname":       row[1],
            "ctype":       row[2],
            "cstatus":     row[3],
            "rdept":       row[4],
            "udepts":      row[5],
            "vid":         str(row[6]) if row[6] else None,
            "vname":       row[7],
            "sdate":       fmt_date(row[8]),
            "edate":       fmt_date(row[9]),
            "ndays":       row[10],
            "autorenewal": yn_to_bool(row[12]),
            "currency":    row[13],
            "total":       float(row[14]) if row[14] else None,
            "monthly":     float(row[15]) if row[15] else None,
            "pricing":     row[16],
            "needs_po":    yn_to_bool(row[17]),
            "can_claim":   yn_to_bool(row[18]),
            "needs_alloc": yn_to_bool(row[19]),
            "alloc_method":row[20],
            "byear":       row[21],
            "bl1":         row[22],
            "bl2":         row[23],
            "accounting":  row[24],
            "bsource":     row[35],
            "bctrl":       row[36],
            "require_acc": yn_to_bool(row[37]),
            "risk":        row[25],
            "manager":     row[26],
            "reviewer":    row[27],
            "attach":      row[28],
            "remarks":     row[29],
            "detail":      json.dumps(detail, ensure_ascii=False),
            "now":         now,
            "now2":        now,
        })
        contract_inserted += 1

    db.commit()

print(f"合約：新增 {contract_inserted} 筆，跳過（已存在）{contract_skipped} 筆")

# ── 검증 ──────────────────────────────────────────────────────────────────────
print("\n=== 匯入結果 ===")
with Session(engine) as db:
    contracts = db.execute(text(
        "SELECT contract_id, contract_name, contract_status, total_amount_tax_included FROM contracts ORDER BY contract_id"
    )).fetchall()
    print(f"合約共 {len(contracts)} 筆：")
    for c in contracts:
        amt = f"{float(c[3]):,.0f}" if c[3] else "—"
        print(f"  {c[0]}  {c[1]}  [{c[2]}]  ${amt}")

    vendors = db.execute(text(
        "SELECT vendor_id, vendor_name, risk_level FROM vendors ORDER BY vendor_id"
    )).fetchall()
    print(f"\n廠商共 {len(vendors)} 筆：")
    for v in vendors:
        print(f"  {v[0]}  {v[1]}  [{v[2]}]")

print("\n✅ 完成！請重新整理 Portal 合約管理頁面。")
