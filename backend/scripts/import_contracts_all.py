"""
import_contracts_all.py
────────────────────────────────────────────────────────────────────────────────
三步驟完整匯入：
  Step 1 — Vendors      從 01_合約主檔 抽取廠商，插入 vendors 表
  Step 2 — Contracts    從 01_合約主檔 插入 contracts 表（同一合約編號多行取第一筆）
  Step 3 — ContractItems 從 02_合約明細項目 插入 contract_items 表

使用方式（在 backend/ 目錄下執行）：
    python scripts/import_contracts_all.py [EXCEL_PATH] [--clear] [--step STEP]

    EXCEL_PATH        預設為 ../20260601合約管理_with02.xlsx
    --clear           各表清空後再匯入（幂等重跑用）
    --step 1|2|3      只跑指定步驟（不指定則全部跑）

範例：
    cd backend
    python scripts/import_contracts_all.py
    python scripts/import_contracts_all.py --clear
    python scripts/import_contracts_all.py --step 2
"""

import sys, os, argparse, json, math
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime, date, timedelta

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.models.contract import Vendor, Contract, ContractItem

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("excel_path", nargs="?",
    default=os.path.join(os.path.dirname(BACKEND_DIR), "20260601合約管理_with02.xlsx"))
parser.add_argument("--clear", action="store_true")
parser.add_argument("--step", type=int, choices=[1, 2, 3], default=None)
args = parser.parse_args()

EXCEL   = args.excel_path
SHEET01 = "01_合約主檔"
SHEET02 = "02_合約明細項目"

# ── DB ────────────────────────────────────────────────────────────────────────
_url = settings.DATABASE_URL.replace("sqlite+aiosqlite", "sqlite")
engine = create_engine(_url, connect_args={"check_same_thread": False} if "sqlite" in _url else {})
# SQLite FK
if "sqlite" in _url:
    from sqlalchemy import event as _event
    @_event.listens_for(engine, "connect")
    def _fk_on(conn, _):
        conn.execute("PRAGMA foreign_keys=ON")

Session = sessionmaker(bind=engine)

# ── 工具函式 ──────────────────────────────────────────────────────────────────
def to_dec(v, default=Decimal("0")) -> Decimal:
    if v is None or v == "": return default
    try: return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except: return default

def to_bool(v) -> bool:
    return str(v).strip() in ("是", "True", "1")

def safe_str(v) -> str:
    return str(v).strip() if v is not None else ""

def safe_date(v) -> date | None:
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try:
        s = str(v).strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try: return datetime.strptime(s, fmt).date()
            except: pass
    except: pass
    return None  # 無效日期（如 2028/16/14）


# ════════════════════════════════════════════════════════════════════════════
# 讀 Excel
# ════════════════════════════════════════════════════════════════════════════
print(f"\n📂  Excel：{EXCEL}")
if not os.path.exists(EXCEL):
    print("❌  找不到 Excel 檔案"); sys.exit(1)

wb_data = openpyxl.load_workbook(EXCEL, data_only=True)

# ── Sheet 01 ──────────────────────────────────────────────────────────────────
ws01       = wb_data[SHEET01]
raw01      = list(ws01.iter_rows(values_only=True))
headers01  = raw01[3]
h01        = {str(h).strip(): i for i, h in enumerate(headers01) if h}

def g01(row, field):
    idx = h01.get(field)
    if idx is None: return None
    v = row[idx]
    if v is None: return None
    if isinstance(v, float) and math.isnan(v): return None
    return v

rows01 = [(i+5, r) for i, r in enumerate(raw01[4:]) if any(v for v in r if v is not None)]

# ── Sheet 02 ──────────────────────────────────────────────────────────────────
ws02  = wb_data[SHEET02]
raw02 = list(ws02.iter_rows(min_row=5, values_only=True))
rows02 = [r for r in raw02 if r[0] is not None]

print(f"✓  01 共 {len(rows01)} 筆 | 02 共 {len(rows02)} 筆")


# ════════════════════════════════════════════════════════════════════════════
# Step 1 — Vendors
# ════════════════════════════════════════════════════════════════════════════
def step1_vendors():
    print("\n── Step 1：Vendors ───────────────────────────────────────────────")

    # 抽取唯一廠商名稱（保持出現順序）
    seen = {}
    for _, row in rows01:
        name = safe_str(g01(row, "廠商名稱"))
        raw_id = safe_str(g01(row, "廠商編號"))
        if name and name not in seen:
            seen[name] = raw_id  # 可能是 VND-XXXX 或空

    # 為空 ID 者自動分配，從目前 DB 最大號+1 開始
    with Session() as db:
        existing = {r[0]: r[1] for r in db.execute(text("SELECT vendor_id, vendor_name FROM vendors"))}
        existing_names = set(existing.values())

        # 找最大號
        max_seq = 0
        for vid in existing:
            try: max_seq = max(max_seq, int(vid.replace("VND-", "")))
            except: pass

        inserted = skipped = 0
        for name, raw_id in seen.items():
            if name in existing_names:
                skipped += 1
                continue
            max_seq += 1
            vid = raw_id if raw_id.startswith("VND-") else f"VND-{max_seq:04d}"
            # 確保 VND-ID 不重複
            while vid in existing:
                max_seq += 1
                vid = f"VND-{max_seq:04d}"

            db.add(Vendor(
                vendor_id   = vid,
                vendor_name = name,
                created_at  = datetime.now(),
                updated_at  = datetime.now(),
            ))
            existing[vid] = name
            existing_names.add(name)
            inserted += 1

        db.commit()
    print(f"✅  插入：{inserted} | 跳過（已存在）：{skipped}")


# ════════════════════════════════════════════════════════════════════════════
# Step 2 — Contracts
# ════════════════════════════════════════════════════════════════════════════
def step2_contracts():
    print("\n── Step 2：Contracts ─────────────────────────────────────────────")

    # 依 contract_id 分組，取第一筆為代表，收集 using_depts/budget_dept
    contract_map: dict[str, dict] = {}   # contract_id → row data
    for _, row in rows01:
        raw_id   = safe_str(g01(row, "合約編號"))
        cid      = raw_id if raw_id else safe_str(g01(row, "合約名稱"))
        dept_use = safe_str(g01(row, "預算使用部門"))
        bco      = safe_str(g01(row, "費用分公司"))

        if cid not in contract_map:
            contract_map[cid] = {"row": row, "using_depts": set(), "budget_companies": set()}
        if dept_use: contract_map[cid]["using_depts"].add(dept_use)
        if bco:      contract_map[cid]["budget_companies"].add(bco)

    # 讀廠商 ID 對應表
    with Session() as db:
        vnd_map = {r[1]: r[0] for r in db.execute(text("SELECT vendor_id, vendor_name FROM vendors"))}

    with Session() as db:
        existing_cids = {r[0] for r in db.execute(text("SELECT contract_id FROM contracts"))}
        now = datetime.now()
        inserted = skipped = error_count = 0

        for cid, meta in contract_map.items():
            if cid in existing_cids:
                skipped += 1
                continue

            row = meta["row"]
            using_depts = ";".join(sorted(meta["using_depts"])) or safe_str(g01(row, "預算使用部門"))
            budget_cos  = ";".join(sorted(meta["budget_companies"]))

            vendor_name = safe_str(g01(row, "廠商名稱"))
            vendor_id   = vnd_map.get(vendor_name, "")

            # 日期（Row 89 有無效 end_date "2028/16/14"，safe_date 回 None）
            start = safe_date(g01(row, "合約起日"))
            end   = safe_date(g01(row, "合約迄日"))
            safe_start = start or date(2026, 1, 1)
            safe_end   = max(end, safe_start) if end else max(date(2026, 12, 31), safe_start)
            notif = int(float(g01(row, "到期前通知天數") or 0)) if g01(row, "到期前通知天數") is not None else 0
            latest_term = (safe_end - timedelta(days=notif)) if notif else None

            # 金額
            total_raw    = g01(row, "合約總金額未稅")
            monthly_raw  = g01(row, "每月固定金額\n未稅")
            total_inc    = to_dec(total_raw) * Decimal("1.05") if total_raw else Decimal("0")
            monthly      = to_dec(monthly_raw) if monthly_raw else None

            ct = safe_str(g01(row, "合約類型"))

            try:
                db.add(Contract(
                    contract_id             = cid,
                    contract_name           = safe_str(g01(row, "合約名稱")),
                    contract_type           = ct,
                    contract_status         = safe_str(g01(row, "合約狀態")) or "生效中",
                    responsible_dept        = safe_str(g01(row, "簽約權責部門")),
                    using_depts             = using_depts,
                    vendor_id               = vendor_id,
                    vendor_name             = vendor_name,
                    start_date              = safe_start,
                    end_date                = safe_end,
                    notification_days       = notif,
                    latest_termination_date = latest_term,
                    auto_renewal            = to_bool(g01(row, "是否自動續約")),
                    currency                = safe_str(g01(row, "幣別")) or "TWD",
                    total_amount_tax_included = total_inc,
                    monthly_fixed_amount    = monthly,
                    pricing_method          = safe_str(g01(row, "計價方式")),
                    needs_purchase_order    = to_bool(g01(row, "是否需請購單")),
                    can_claim_without_po    = to_bool(g01(row, "是否可無請購請款")),
                    needs_allocation        = to_bool(g01(row, "是否需分攤")),
                    allocation_method       = safe_str(g01(row, "分攤方式")) or None,
                    budget_year             = int(float(g01(row, "預算年度") or 2026)),
                    budget_category_l1      = safe_str(g01(row, "預算大項")),
                    budget_category_l2      = safe_str(g01(row, "預算細項")),
                    accounting_code         = safe_str(g01(row, "會計科目")),
                    risk_level              = safe_str(g01(row, "風險等級")) or "低",
                    manager                 = safe_str(g01(row, "管理人")),
                    reviewer                = safe_str(g01(row, "覆核人")),
                    signing_company         = safe_str(g01(row, "簽約公司")) or None,
                    signing_dept            = safe_str(g01(row, "簽約權責部門")) or None,
                    budget_company          = budget_cos or None,
                    budget_dept             = using_depts or None,
                    pricing_spec            = safe_str(g01(row, "計價規格")) or None,
                    attachment_url          = safe_str(g01(row, "附件連結")) or None,
                    remarks                 = safe_str(g01(row, "備註")),
                    detail                  = "{}",
                    created_at              = now,
                    updated_at              = now,
                ))
                existing_cids.add(cid)
                inserted += 1
            except Exception as e:
                print(f"  ⚠️  合約 {cid} 失敗：{e}")
                db.rollback()
                error_count += 1
                continue

        db.commit()

    print(f"✅  插入：{inserted} | 跳過（已存在）：{skipped} | 錯誤：{error_count}")


# ════════════════════════════════════════════════════════════════════════════
# Step 3 — ContractItems（讀 02 Sheet）
# ════════════════════════════════════════════════════════════════════════════
def step3_items():
    print("\n── Step 3：ContractItems ─────────────────────────────────────────")

    # 02 欄位索引（A~T，0-based）
    IDX = {"cid":0,"seq":1,"name":2,"cat":3,"unit_price":7,"qty":8,"unit":9,
           "tax_rate":10,"amount_ex":11,"amount_inc":12,"is_fixed":13,"is_float":14}

    with Session() as db:
        existing_cids = {r[0] for r in db.execute(text("SELECT contract_id FROM contracts"))}
        existing_keys = {(r[0], r[1]) for r in db.execute(text("SELECT contract_id, item_seq FROM contract_items"))}
        now = datetime.now()
        inserted = skipped_fk = skipped_dup = errors = 0

        for row in rows02:
            def col(k): return row[IDX[k]]
            cid = safe_str(col("cid"))
            seq = col("seq")

            if cid not in existing_cids:
                skipped_fk += 1
                continue
            key = (cid, seq)
            if key in existing_keys:
                skipped_dup += 1
                continue

            unit_price = to_dec(col("unit_price"))
            quantity   = to_dec(col("qty"))
            tax_pct    = Decimal(str(round(float(col("tax_rate") or 0.05) * 100, 2)))
            raw_ex     = col("amount_ex")
            raw_inc    = col("amount_inc")
            if raw_ex not in (None, ""):
                amt_ex  = to_dec(raw_ex)
                amt_inc = to_dec(raw_inc)
            else:
                amt_ex  = (unit_price * quantity).quantize(Decimal("0.01"), ROUND_HALF_UP) if quantity else Decimal("0")
                amt_inc = (amt_ex * (1 + tax_pct / 100)).quantize(Decimal("0.01"), ROUND_HALF_UP)

            try:
                db.add(ContractItem(
                    contract_id             = cid,
                    item_seq                = int(seq) if seq else 1,
                    item_name               = safe_str(col("name")),
                    item_category           = safe_str(col("cat")),
                    unit_price_tax_excluded = unit_price if unit_price else None,
                    quantity                = quantity if quantity else None,
                    unit                    = safe_str(col("unit")) or None,
                    tax_rate                = tax_pct,
                    amount_tax_excluded     = amt_ex,
                    amount_tax_included     = amt_inc,
                    is_fixed                = to_bool(col("is_fixed")),
                    is_floating             = to_bool(col("is_float")),
                    created_at              = now,
                    updated_at              = now,
                ))
                existing_keys.add(key)
                inserted += 1
            except Exception as e:
                print(f"  ⚠️  {cid}#{seq} 失敗：{e}")
                errors += 1

        db.commit()

    print(f"✅  插入：{inserted} | 跳過（合約不存在）：{skipped_fk} | 跳過（重複）：{skipped_dup} | 錯誤：{errors}")


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════
# --clear 時依 FK 反向順序一次清空，再逐步插入
if args.clear:
    with Session() as db:
        db.execute(text("DELETE FROM contract_items"))
        db.execute(text("DELETE FROM contracts"))
        db.execute(text("DELETE FROM vendors"))
        db.commit()
        print("🗑️   已清空 contract_items → contracts → vendors（FK 順序）")

steps = [args.step] if args.step else [1, 2, 3]
for s in steps:
    if s == 1: step1_vendors()
    elif s == 2: step2_contracts()
    elif s == 3: step3_items()

print("\n🎉  完成\n")
