"""
import_contract_items.py
────────────────────────
從 Excel 「02_合約明細項目」Sheet 匯入 contract_items 資料表。

使用方式（在 backend/ 目錄下執行）：
    python scripts/import_contract_items.py [EXCEL_PATH] [--clear]

    EXCEL_PATH  預設為 ../20260601合約管理_with02.xlsx
    --clear     清空 contract_items 後再 INSERT（預設：跳過已有 contract_id+item_seq 的重複）

範例：
    cd backend
    python scripts/import_contract_items.py
    python scripts/import_contract_items.py --clear
"""

import sys
import os
import argparse
from decimal import Decimal, ROUND_HALF_UP
from datetime import datetime

# ── 路徑設定 ─────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BACKEND_DIR = os.path.dirname(SCRIPT_DIR)
sys.path.insert(0, BACKEND_DIR)

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.models.contract import ContractItem

# ── 參數解析 ──────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description="匯入合約明細項目至資料庫")
parser.add_argument(
    "excel_path",
    nargs="?",
    default=os.path.join(os.path.dirname(BACKEND_DIR), "20260601合約管理_with02.xlsx"),
    help="Excel 檔案路徑（預設：../20260601合約管理_with02.xlsx）",
)
parser.add_argument("--clear", action="store_true", help="匯入前清空 contract_items 資料表")
args = parser.parse_args()

EXCEL_PATH = args.excel_path
SHEET_NAME = "02_合約明細項目"

# ── 資料庫連線 ────────────────────────────────────────────────────────────────
_db_url = settings.DATABASE_URL.replace("sqlite+aiosqlite", "sqlite")
engine = create_engine(
    _db_url,
    connect_args={"check_same_thread": False} if "sqlite" in _db_url else {},
)
Session = sessionmaker(bind=engine)

# ── 讀取 Excel ────────────────────────────────────────────────────────────────
print(f"📂 讀取 Excel：{EXCEL_PATH}")
if not os.path.exists(EXCEL_PATH):
    print(f"❌ 找不到檔案：{EXCEL_PATH}")
    sys.exit(1)

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
if SHEET_NAME not in wb.sheetnames:
    print(f"❌ 找不到 Sheet：{SHEET_NAME}")
    sys.exit(1)

ws = wb[SHEET_NAME]
rows = list(ws.iter_rows(min_row=5, values_only=True))
data_rows = [r for r in rows if r[0] is not None]  # A 欄 = 合約編號，非空才算資料
print(f"✓ 讀到 {len(data_rows)} 筆資料")

# 欄位索引（依 02 sheet 欄位順序）
# A  B    C     D    E   F   G    H        I    J    K     L            M            N         O         P        Q        R        S        T
# 合約編號 項次 項目名稱 類別 大項 細項 會計 單價未稅 數量 單位 稅率 金額未稅(公式) 金額含稅(公式) 固定費用 浮動費用 浮動基礎 納入預算 納入應計 預算月份 備註
IDX = {
    "contract_id":   0,   # A
    "item_seq":      1,   # B
    "item_name":     2,   # C
    "item_category": 3,   # D
    "unit_price":    7,   # H
    "quantity":      8,   # I
    "unit":          9,   # J
    "tax_rate":      10,  # K（0.05 → DB 存 5）
    "amount_ex":     11,  # L（公式計算值）
    "amount_inc":    12,  # M（公式計算值）
    "is_fixed":      13,  # N
    "is_floating":   14,  # O
}


def to_dec(v, default=Decimal("0")) -> Decimal:
    if v is None or v == "":
        return default
    try:
        return Decimal(str(v)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except Exception:
        return default


def to_bool(v) -> bool:
    return str(v).strip() in ("是", "True", "1", "YES", "true")


# ── 匯入 ──────────────────────────────────────────────────────────────────────
with Session() as db:
    # ── 取得現有合約 ID 集合 ──────────────────────────────────────────────────
    existing_contract_ids: set[str] = set()
    try:
        result = db.execute(text("SELECT contract_id FROM contracts"))
        existing_contract_ids = {row[0] for row in result}
    except Exception as e:
        print(f"⚠️  讀取 contracts 表失敗（{e}），將跳過 FK 檢查，直接嘗試插入")

    # ── 可選：清空舊資料 ─────────────────────────────────────────────────────
    if args.clear:
        deleted = db.execute(text("DELETE FROM contract_items"))
        db.commit()
        print(f"🗑️  已清空 contract_items（{deleted.rowcount} 筆）")

    # ── 取得已存在的 (contract_id, item_seq) 組合 ──────────────────────────
    existing_keys: set[tuple] = set()
    if not args.clear:
        result = db.execute(text("SELECT contract_id, item_seq FROM contract_items"))
        existing_keys = {(row[0], row[1]) for row in result}

    # ── 逐筆插入 ────────────────────────────────────────────────────────────
    inserted = skipped_fk = skipped_dup = error_count = 0
    now = datetime.now()

    for row in data_rows:
        def col(k):
            return row[IDX[k]]

        cid = str(col("contract_id")).strip()
        seq = col("item_seq")

        # FK 檢查（有取到合約清單時才擋）
        if existing_contract_ids and cid not in existing_contract_ids:
            skipped_fk += 1
            continue

        # 重複檢查
        key = (cid, seq)
        if key in existing_keys:
            skipped_dup += 1
            continue

        # 計算金額
        unit_price = to_dec(col("unit_price"))
        quantity   = to_dec(col("quantity"))
        tax_rate_pct = Decimal(str(round(float(col("tax_rate") or 0.05) * 100, 2)))

        # 優先用 Excel 計算值，若空才自算
        raw_ex  = col("amount_ex")
        raw_inc = col("amount_inc")
        if raw_ex not in (None, "", ""):
            amount_ex  = to_dec(raw_ex)
            amount_inc = to_dec(raw_inc)
        else:
            amount_ex  = (unit_price * quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if quantity else Decimal("0")
            amount_inc = (amount_ex * (1 + tax_rate_pct / 100)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        try:
            item = ContractItem(
                contract_id             = cid,
                item_seq                = int(seq) if seq else 1,
                item_name               = str(col("item_name") or "").strip(),
                item_category           = str(col("item_category") or "").strip(),
                unit_price_tax_excluded = unit_price if unit_price else None,
                quantity                = quantity if quantity else None,
                unit                    = str(col("unit") or "").strip() or None,
                tax_rate                = tax_rate_pct,
                amount_tax_excluded     = amount_ex,
                amount_tax_included     = amount_inc,
                is_fixed                = to_bool(col("is_fixed")),
                is_floating             = to_bool(col("is_floating")),
                created_at              = now,
                updated_at              = now,
            )
            db.add(item)
            existing_keys.add(key)
            inserted += 1

        except Exception as e:
            print(f"  ⚠️  Row {cid}#{seq} 插入失敗：{e}")
            error_count += 1

    db.commit()

# ── 結果摘要 ─────────────────────────────────────────────────────────────────
print()
print("═" * 50)
print(f"✅  插入成功：{inserted} 筆")
if skipped_fk:
    print(f"⚠️   跳過（合約不存在）：{skipped_fk} 筆")
if skipped_dup:
    print(f"ℹ️   跳過（已存在）：{skipped_dup} 筆")
if error_count:
    print(f"❌  錯誤：{error_count} 筆")
print("═" * 50)
