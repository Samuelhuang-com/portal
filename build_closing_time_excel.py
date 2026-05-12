"""
3.2 結案時間明細比對 + 3.3 報修類型統計 Excel 產生器
執行方式：python build_closing_time_excel.py
輸出：3.2結案時間_明細比對_202601-202605.xlsx（與本腳本同目錄）
"""
import sys, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── 路徑設定 ──────────────────────────────────────────────────────────────────
HERE   = Path(__file__).parent
SOURCE = HERE.parent / "uploads" / "工程維修單 房務部查閱用_202605112243-83a7f7a2.xlsx"
if not SOURCE.exists():
    # fallback: 同資料夾
    SOURCE = HERE / "工程維修單 房務部查閱用_202605112243-83a7f7a2.xlsx"
OUTPUT = HERE / "3.2結案時間_明細比對_202601-202605.xlsx"

COMPLETED_STATUSES = {"結案","已辦驗","已驗收","已結案","完修","已完成","完成"}
EXCLUDED_STATUSES  = {"取消"}

# ── 報修類型標準化 mapping（與 dazhi_repair_service.py 一致）──────────────────
REPAIR_TYPE_MAPPING = {
    "建築":"建築","結構":"建築","外牆":"建築","玻璃":"建築","門窗":"建築",
    "電梯":"建築","手扶梯":"建築","招牌":"建築",
    "衛廁":"衛廁","廁所":"衛廁","馬桶":"衛廁","洗手":"衛廁","洗手槽":"衛廁",
    "烘手機":"衛廁","哺乳":"衛廁","蓮蓬頭":"衛廁","浴缸":"衛廁","花灑":"衛廁",
    "淋浴":"衛廁","水龍頭":"衛廁","洗臉":"衛廁","面盆":"衛廁",
    "消防":"消防","瓦斯":"消防","偵煙":"消防","灑水":"消防","鐵捲門":"消防",
    "安全門":"消防","緊急":"消防",
    "空調":"空調","冷氣":"空調","冷卻":"空調","送風":"空調","補風":"空調",
    "導流":"空調","分離式":"空調",
    "機電":"機電","機房":"機電","發電機":"機電","配電":"機電",
    "給排水":"給排水","漏水":"給排水","水塔":"給排水","污水":"給排水",
    "排水":"給排水","給水":"給排水",
    "排煙":"排煙","靜電機":"排煙","水洗機":"排煙","截油槽":"排煙","排煙管":"排煙",
    "監控":"監控","cctv":"監控","攝影":"監控","監視":"監控",
    "弱電":"弱電","交換機":"弱電","電話":"弱電","網路":"弱電",
    "照明":"照明","燈":"照明","燈具":"照明","路燈":"照明",
    "停車":"停車","車牌":"停車","繳費機":"停車","柵欄":"停車",
    "人流":"其他","入金機":"其他",
    "專櫃":"專櫃","租戶":"專櫃","承租":"專櫃",
    "公區":"公區","公共":"公區","lobby":"公區","大廳":"公區",
    "廣場":"公區","梯廳":"公區","接待":"公區","露台":"公區",
    "後勤":"後勤空間","辦公室":"後勤空間","儲藏":"後勤空間","員工餐":"後勤空間",
}

REPAIR_TYPE_ORDER = [
    "建築","衛廁","消防","空調","機電","給排水",
    "排煙","監控","弱電","照明","停車","其他",
    "專櫃","凍&藏類設備","內裝","廚房&吧台設備",
    "會議設備","瓦斯類設備","公區","後勤空間",
]

def normalize_repair_type(raw_type: str, title: str = "") -> str:
    """Portal dazhi_repair_service.normalize_repair_type() 的 Python 複製。"""
    combined_src = raw_type.lower()
    for kw, std in REPAIR_TYPE_MAPPING.items():
        if kw.lower() in combined_src:
            return std
    combined_fb = (title + " " + raw_type).lower()
    for kw, std in REPAIR_TYPE_MAPPING.items():
        if kw.lower() in combined_fb:
            return std
    return raw_type.strip() if raw_type.strip() else "其他"

# 類型顏色對應（統計 Sheet 用）
TYPE_COLORS = {
    "建築":"4472C4","衛廁":"ED7D31","消防":"FF0000","空調":"70AD47",
    "機電":"FFC000","給排水":"00B0F0","排煙":"7030A0","監控":"00B050",
    "弱電":"FF7C80","照明":"FFFF00","停車":"A9D18E","其他":"BFBFBF",
    "專櫃":"D6BCF5","公區":"9DC3E6","後勤空間":"F4B183",
}

PORTAL_VALUES = {
    1: {"count": 240, "total_days": 366.0, "avg": 1.52},
    2: {"count": 230, "total_days": 710.9, "avg": 3.09},
    3: {"count": 209, "total_days": 483.1, "avg": 2.31},
    4: {"count": 121, "total_days": 233.2, "avg": 1.93},
    5: {"count": 52,  "total_days": 194.4, "avg": 3.74},
}

C_HEADER_BG  = "1B3A5C"
C_HEADER_FG  = "FFFFFF"
C_ALT_ROW    = "EBF5FB"
C_WHITE      = "FFFFFF"
C_GREEN_FILL = "D5F5E3"
C_RED_FILL   = "FADBD8"
C_SUBTOTAL   = "D6EAF8"

# ── 讀取來源 ──────────────────────────────────────────────────────────────────
print(f"Reading: {SOURCE}")
df_raw = pd.read_excel(SOURCE, sheet_name=0, header=0)
print(f"Shape: {df_raw.shape}  Columns: {list(df_raw.columns)}")

def find_col(df, candidates, exact_first=False):
    """先嘗試完全相符，找不到再做 substring。"""
    cols = list(df.columns)
    # 完全相符優先
    for name in candidates:
        if name in cols:
            return name
    if exact_first:
        return None
    # substring fallback
    for name in candidates:
        for c in cols:
            if name in str(c):
                return c
    return None

col_order_no   = find_col(df_raw, ["報修單編號","單號","維修單號"])
col_occurred   = find_col(df_raw, ["報修日期","報修時間","發生日期"])
col_completed  = find_col(df_raw, ["完工時間","完工日期","完成時間"])
col_close_time = find_col(df_raw, ["結案時間","結案日期"])
# 處理狀態：必須完全相符，避免誤抓「檢查處理狀態」
col_status     = find_col(df_raw, ["處理狀態","狀態"], exact_first=True)
col_repair_fee = find_col(df_raw, ["維修費用","修繕費用"])
col_outsource  = find_col(df_raw, ["委外費用","委外金額"])
col_type       = find_col(df_raw, ["類型","報修類型","類別"], exact_first=True)
col_title      = find_col(df_raw, ["(備註/詳細說明)","備註","標題","報修名稱"])

print(f"cols → order:{col_order_no} occ:{col_occurred} comp:{col_completed} "
      f"close:{col_close_time} status:{col_status} "
      f"repair:{col_repair_fee} outsrc:{col_outsource}")

def to_ts(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        ts = pd.Timestamp(val)
        return ts if not pd.isnull(ts) else None
    except Exception:
        return None

def to_float(val):
    try:
        f = float(val)
        return 0.0 if pd.isna(f) else f
    except Exception:
        return 0.0

# ── 過濾邏輯（與 Portal compute_closing_time 小型一致）────────────────────────
rows = []
for _, r in df_raw.iterrows():
    status = str(r[col_status]).strip() if col_status else ""
    if status in EXCLUDED_STATUSES or status not in COMPLETED_STATUSES:
        continue
    completed_at = to_ts(r[col_completed]) if col_completed else None
    if completed_at is None and col_close_time:
        completed_at = to_ts(r[col_close_time])
    occurred_at = to_ts(r[col_occurred]) if col_occurred else None
    if completed_at is None or occurred_at is None:
        continue
    repair_fee    = to_float(r[col_repair_fee])   if col_repair_fee else 0.0
    outsource_fee = to_float(r[col_outsource])    if col_outsource  else 0.0
    total_fee     = repair_fee + outsource_fee
    if total_fee != 0:
        continue  # 只留小型（無費用）
    close_days = round((completed_at - occurred_at).total_seconds() / 86400, 2)
    raw_type = str(r[col_type]).strip()  if col_type  else ""
    raw_title= str(r[col_title]).strip() if col_title else ""
    repair_type = normalize_repair_type(raw_type, raw_title)

    rows.append({
        "order_no":      str(r[col_order_no]).strip() if col_order_no else "",
        "occurred_at":   occurred_at,
        "completed_at":  completed_at,
        "status":        status,
        "repair_fee":    repair_fee,
        "outsource_fee": outsource_fee,
        "total_fee":     total_fee,
        "close_days":    close_days,
        "repair_type":   repair_type,
        "year":          completed_at.year,
        "month":         completed_at.month,
    })

df = pd.DataFrame(rows)
print(f"Filtered rows (small, non-cancelled, both dates): {len(df)}")

# ── 每月統計 ──────────────────────────────────────────────────────────────────
month_stats = {}
month_dfs   = {}
for m in range(1, 6):
    mdf = df[(df["year"]==2026) & (df["month"]==m)].sort_values("occurred_at").reset_index(drop=True)
    month_dfs[m] = mdf
    n = len(mdf)
    month_stats[m] = {
        "count":      n,
        "total_days": round(float(mdf["close_days"].sum()), 2) if n else 0.0,
        "avg":        round(float(mdf["close_days"].mean()), 2) if n else 0.0,
    }
    p = PORTAL_VALUES[m]
    s = month_stats[m]
    cnt_ok = "✓" if s["count"]==p["count"]                         else f"✗ Excel={s['count']} Portal={p['count']}"
    tot_ok = "✓" if abs(s["total_days"]-p["total_days"]) <= 0.1    else f"✗ Excel={s['total_days']} Portal={p['total_days']}"
    avg_ok = "✓" if abs(s["avg"]-p["avg"]) <= 0.02                 else f"✗ Excel={s['avg']} Portal={p['avg']}"
    print(f"  {m}月: count {cnt_ok}  totalDays {tot_ok}  avg {avg_ok}")

# ── 樣式輔助 ──────────────────────────────────────────────────────────────────
def mfont(bold=False, color="000000"):
    return Font(name="Arial", size=10, bold=bold, color=color)

def mfill(hex_c):
    return PatternFill("solid", start_color=hex_c, fgColor=hex_c)

def malign(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=False)

def set_hdr(cell, text, align="center"):
    cell.value     = text
    cell.font      = Font(name="Arial", size=10, bold=True, color=C_HEADER_FG)
    cell.fill      = mfill(C_HEADER_BG)
    cell.alignment = malign(align)

# ── 建立活頁簿 ────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

MONTH_NAMES = {1:"1月",2:"2月",3:"3月",4:"4月",5:"5月"}
month_sum_rows = {}  # m -> (data_n, sum_r)

# ── 各月明細 Sheet ────────────────────────────────────────────────────────────
for m in range(1, 6):
    ws = wb.create_sheet(title=f"{m}月明細")
    mdf = month_dfs[m]
    n   = len(mdf)

    hdrs   = ["序號","報修單編號","報修日期","完工時間","處理狀態",
              "維修費用","委外費用","合計費用","結案天數","類型（標準化）","備註"]
    widths = [6, 20, 13, 13, 10, 11, 11, 11, 11, 14, 16]

    ws.row_dimensions[1].height = 20
    for ci, (h, w) in enumerate(zip(hdrs, widths), 1):
        set_hdr(ws.cell(row=1, column=ci), h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    for ri, (_, row) in enumerate(mdf.iterrows(), 1):
        er  = ri + 1
        bg  = C_ALT_ROW if ri % 2 == 0 else C_WHITE

        def sc(c, val, align="left", fmt=None, bold=False):
            cell = ws.cell(row=er, column=c)
            cell.value     = val
            cell.font      = mfont(bold)
            cell.fill      = mfill(bg)
            cell.alignment = malign(align)
            if fmt: cell.number_format = fmt

        sc(1, ri, "center")
        sc(2, row["order_no"])

        for ci, dt_val, fmt_str in [(3, row["occurred_at"], "YYYY/MM/DD"),
                                    (4, row["completed_at"], "YYYY/MM/DD")]:
            c2 = ws.cell(row=er, column=ci)
            c2.value          = dt_val.date()
            c2.font           = mfont()
            c2.fill           = mfill(bg)
            c2.alignment      = malign("left")
            c2.number_format  = fmt_str

        sc(5, row["status"],        "center")
        sc(6, row["repair_fee"],    "right", "#,##0.00")
        sc(7, row["outsource_fee"], "right", "#,##0.00")

        hc = ws.cell(row=er, column=8)
        hc.value = f"=F{er}+G{er}"; hc.font = mfont(); hc.fill = mfill(bg)
        hc.alignment = malign("right"); hc.number_format = "#,##0.00"

        ic = ws.cell(row=er, column=9)
        ic.value = row["close_days"]; ic.font = mfont(); ic.fill = mfill(bg)
        ic.alignment = malign("right"); ic.number_format = "0.00"

        sc(10, row["repair_type"], "center")
        sc(11, "", "left")

    if n > 0:
        sub_r = n + 2
        sum_r = n + 3
        ds, de = 2, n + 1

        for c in range(1, 12):
            ws.cell(row=sub_r, column=c).fill = mfill(C_SUBTOTAL)
            ws.cell(row=sum_r, column=c).fill = mfill(C_SUBTOTAL)

        ws.cell(row=sub_r, column=1).value = "小計"
        ws.cell(row=sub_r, column=1).font  = mfont(bold=True)

        for c, L in [(6,"F"),(7,"G"),(8,"H"),(9,"I")]:
            cell = ws.cell(row=sub_r, column=c)
            cell.value = f"=SUM({L}{ds}:{L}{de})"; cell.font = mfont(bold=True)
            cell.alignment = malign("right")
            cell.number_format = "#,##0.00" if c < 9 else "0.00"

        ws.cell(row=sum_r, column=1).value = "結案數";   ws.cell(row=sum_r, column=1).font = mfont(bold=True)
        ws.cell(row=sum_r, column=2).value = f"=COUNTA(B{ds}:B{de})"; ws.cell(row=sum_r, column=2).font = mfont(bold=True); ws.cell(row=sum_r, column=2).alignment = malign("right")
        ws.cell(row=sum_r, column=3).value = "總天數";   ws.cell(row=sum_r, column=3).font = mfont(bold=True)
        ws.cell(row=sum_r, column=4).value = f"=SUM(I{ds}:I{de})";    ws.cell(row=sum_r, column=4).font = mfont(bold=True); ws.cell(row=sum_r, column=4).alignment = malign("right"); ws.cell(row=sum_r, column=4).number_format = "0.0"
        ws.cell(row=sum_r, column=5).value = "平均天數"; ws.cell(row=sum_r, column=5).font = mfont(bold=True)
        ws.cell(row=sum_r, column=6).value = f"=AVERAGE(I{ds}:I{de})"; ws.cell(row=sum_r, column=6).font = mfont(bold=True); ws.cell(row=sum_r, column=6).alignment = malign("right"); ws.cell(row=sum_r, column=6).number_format = "0.00"

        month_sum_rows[m] = (n, sum_r)
    else:
        ws.cell(row=2, column=1).value = "(本月無資料)"
        month_sum_rows[m] = (0, None)

    ws.freeze_panes = "A2"

# ── 比對摘要 Sheet ─────────────────────────────────────────────────────────────
ws_s = wb.create_sheet(title="比對摘要", index=0)

ws_s.merge_cells("A1:J1")
tc = ws_s.cell(row=1, column=1)
tc.value     = "3.2 結案時間明細比對摘要 (2026年1月–5月，小型報修 total_fee=0)"
tc.font      = Font(name="Arial", size=12, bold=True, color=C_HEADER_FG)
tc.fill      = mfill(C_HEADER_BG)
tc.alignment = malign("center")
ws_s.row_dimensions[1].height = 26

hdrs2  = ["月份","結案數(Excel)","結案數(Portal)","一致",
          "總天數(Excel)","總天數(Portal)","一致",
          "均天數(Excel)","均天數(Portal)","一致"]
widths2= [8, 15, 15, 7, 15, 15, 7, 15, 15, 7]
ws_s.row_dimensions[2].height = 20
for ci, (h, w) in enumerate(zip(hdrs2, widths2), 1):
    set_hdr(ws_s.cell(row=2, column=ci), h)
    ws_s.column_dimensions[get_column_letter(ci)].width = w

for i, m in enumerate(range(1, 6), 1):
    dr     = i + 2
    portal = PORTAL_VALUES[m]
    stats  = month_stats[m]
    n, sum_r = month_sum_rows[m]
    sname  = f"{m}月明細"
    bg     = C_ALT_ROW if i % 2 == 0 else C_WHITE

    def sc2(c, val, align="right", fmt=None, bold=False, fill_c=None):
        cell = ws_s.cell(row=dr, column=c)
        cell.value     = val
        cell.font      = mfont(bold)
        cell.fill      = mfill(fill_c if fill_c else bg)
        cell.alignment = malign(align)
        if fmt: cell.number_format = fmt

    sc2(1, MONTH_NAMES[m], "center", bold=True)

    # Excel 計算值（參照明細 Sheet 的彙總列）
    ef_count = f"='{sname}'!B{sum_r}" if sum_r else "=0"
    ef_total = f"='{sname}'!D{sum_r}" if sum_r else "=0"
    ef_avg   = f"='{sname}'!F{sum_r}" if sum_r else "=0"

    sc2(2, ef_count, "right", "#,##0")
    sc2(3, portal["count"], "right", "#,##0")
    count_ok = abs(stats["count"] - portal["count"]) == 0
    sc2(4, f'=IF(ABS(B{dr}-C{dr})<=0.5,"✓","✗")', "center", bold=True,
        fill_c=C_GREEN_FILL if count_ok else C_RED_FILL)

    sc2(5, ef_total, "right", "#,##0.0")
    sc2(6, portal["total_days"], "right", "#,##0.0")
    tot_ok = abs(stats["total_days"] - portal["total_days"]) <= 0.1
    sc2(7, f'=IF(ABS(E{dr}-F{dr})<=0.1,"✓","✗")', "center", bold=True,
        fill_c=C_GREEN_FILL if tot_ok else C_RED_FILL)

    sc2(8, ef_avg, "right", "0.00")
    sc2(9, portal["avg"], "right", "0.00")
    avg_ok = abs(stats["avg"] - portal["avg"]) <= 0.02
    sc2(10, f'=IF(ABS(H{dr}-I{dr})<=0.02,"✓","✗")', "center", bold=True,
        fill_c=C_GREEN_FILL if avg_ok else C_RED_FILL)

ws_s.freeze_panes = "A3"

# ── 3.3 類型統計 Sheet ─────────────────────────────────────────────────────────
# 口徑：completed_at 落在各月（小型，非取消），以 repair_type 為維度計數
ws_t = wb.create_sheet(title="3.3類型統計")

# 建立類型 × 月份交叉表
from collections import defaultdict
type_month_cnt = defaultdict(lambda: defaultdict(int))  # type_month_cnt[type][month]
all_types_seen = set()

for m in range(1, 6):
    mdf = month_dfs[m]
    for _, row in mdf.iterrows():
        rt = row["repair_type"]
        type_month_cnt[rt][m] += 1
        all_types_seen.add(rt)

# 依 REPAIR_TYPE_ORDER 排序，其餘追加
ordered_types = [t for t in REPAIR_TYPE_ORDER if t in all_types_seen]
ordered_types += sorted([t for t in all_types_seen if t not in REPAIR_TYPE_ORDER])

months = [1, 2, 3, 4, 5]
month_labels = ["1月", "2月", "3月", "4月", "5月"]

# 標題列
ws_t.merge_cells(f"A1:{get_column_letter(2 + len(months))}1")
t2 = ws_t.cell(row=1, column=1)
t2.value     = "3.3 報修類型統計（2026年1–5月，小型報修，完工時間口徑）"
t2.font      = Font(name="Arial", size=12, bold=True, color=C_HEADER_FG)
t2.fill      = mfill(C_HEADER_BG)
t2.alignment = malign("center")
ws_t.row_dimensions[1].height = 26

# 欄標題
hdr_row = ["類型（標準化）"] + month_labels + ["合計"]
ws_t.row_dimensions[2].height = 20
col_widths_t = [16] + [10]*len(months) + [10]
for ci, (h, w) in enumerate(zip(hdr_row, col_widths_t), 1):
    set_hdr(ws_t.cell(row=2, column=ci), h)
    ws_t.column_dimensions[get_column_letter(ci)].width = w

# 資料列
total_row_by_month = defaultdict(int)
for ri, rt in enumerate(ordered_types, 1):
    dr  = ri + 2
    bg  = C_ALT_ROW if ri % 2 == 0 else C_WHITE
    # 類型名稱 + 色塊
    tc = ws_t.cell(row=dr, column=1)
    tc.value     = rt
    tc.font      = Font(name="Arial", size=10, bold=True,
                        color=TYPE_COLORS.get(rt, "000000"))
    tc.fill      = mfill(bg)
    tc.alignment = malign("left")

    for ci, m in enumerate(months, 2):
        cnt = type_month_cnt[rt][m]
        total_row_by_month[m] += cnt
        cell = ws_t.cell(row=dr, column=ci)
        cell.value     = cnt if cnt else "-"
        cell.font      = mfont()
        cell.fill      = mfill(bg)
        cell.alignment = malign("center")

    # 合計欄（Excel SUM formula）
    data_start = get_column_letter(2)
    data_end   = get_column_letter(1 + len(months))
    sc_sum = ws_t.cell(row=dr, column=2 + len(months))
    sc_sum.value          = f"=SUM({data_start}{dr}:{data_end}{dr})"
    sc_sum.font           = mfont(bold=True)
    sc_sum.fill           = mfill(bg)
    sc_sum.alignment      = malign("center")
    sc_sum.number_format  = "#,##0"

# 合計列
tot_r = len(ordered_types) + 3
for c in range(1, 2 + len(months) + 1):
    ws_t.cell(row=tot_r, column=c).fill = mfill(C_SUBTOTAL)

ws_t.cell(row=tot_r, column=1).value = "合計"
ws_t.cell(row=tot_r, column=1).font  = mfont(bold=True)

data_r_start = 3
data_r_end   = len(ordered_types) + 2
for ci in range(2, 2 + len(months) + 1):
    col_letter = get_column_letter(ci)
    c2 = ws_t.cell(row=tot_r, column=ci)
    c2.value          = f"=SUM({col_letter}{data_r_start}:{col_letter}{data_r_end})"
    c2.font           = mfont(bold=True)
    c2.alignment      = malign("center")
    c2.number_format  = "#,##0"

ws_t.freeze_panes = "B3"

# Console 輸出預覽
print("\n3.3 類型統計預覽（小型，completed_at 口徑）:")
print(f"{'類型':<12}", end="")
for ml in month_labels: print(f"{ml:>6}", end="")
print(f"{'合計':>6}")
for rt in ordered_types:
    cnts = [type_month_cnt[rt][m] for m in months]
    if sum(cnts) == 0: continue
    print(f"{rt:<12}", end="")
    for c in cnts: print(f"{c:>6}", end="")
    print(f"{sum(cnts):>6}")

# ── 儲存 ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"\n✓ 已產生：{OUTPUT}")
print("  Sheets:", [s.title for s in wb.worksheets])
