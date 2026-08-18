#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
週期採購 — 春大直料號／部門／類別主檔對齊腳本（2026-08-18）
==============================================================

依 2026-08-18 與 Samuel 確認的四點：

  1. 非附件料號一律刪除（附件＝《20260716春大直設料號明細表-加代碼_20260818.xlsx》）
  2. Excel 分頁 → 類別／公司／部門對應：
        維修備品-工程部需求   → 工程部
        清潔用品-管理部需求   → 管理部
        文具用品-所有部門需求 → 所有部門（＝對四個部門各建一筆對照）
        營業備品-營業部需求   → 營業部
  3. 缺的類別補上（料號的空白 category＋類別主檔補齊編碼原則全部類別）
  4. 新增「類別主檔」（cycle_purchase_categories），先按上述建立，未來料號直接連結

⚠️ 本腳本放在 docs/ 而不是 Temp/：正式區要靠 git pull 拿到它才跑得起來，
而 Temp/ 完全不進版控（CLAUDE.md §10 規則 4：需要跨機器的 migration 腳本放 docs/）。
同目錄的《20260716春大直設料號明細表-加代碼_20260818.xlsx》是它的輸入來源，
一併納入版控，正式區才不用另外傳檔。先例見 docs/sync_hotel_schedule_to_mall.py。

────────────────────────────────────────────────────────────────
先讀這段：這次跟 2026-08-13 的 reimport_chunda_items.py 差在哪
────────────────────────────────────────────────────────────────
附件的 178 筆料號內容與上次匯入用的 `20260417春大直設料號明細表-加代碼.xlsx`
**逐欄比對完全相同**（無新增、無刪除、無任何欄位變動，已於 2026-08-18 驗證）。
所以這支腳本不是「重新匯入料號」，而是三件事：

  (a) 部門歸屬改名（工務部→工程部、清潔部→管理部），文具改掛全部門
  (b) 把 DB 裡多出來、不在附件裡的春大直料號清掉
  (c) 建立類別主檔

因此**預設不覆寫料號的品名／規格**——上次匯入時有 9 筆「用第一個空白切品名/
規格」切壞的資料，很可能已經被人工修過，這裡再洗一次就等於把人家的修正
蓋掉。要連同單價／庫存量／MOQ 一起從 Excel 刷新，另外加 `--refresh-fields`。

────────────────────────────────────────────────────────────────
使用方式
────────────────────────────────────────────────────────────────

  1) 盤點模式（預設，不寫入任何資料）：

       python align_chunda_items_20260818.py ^
           --excel-path "20260716春大直設料號明細表-加代碼_20260818.xlsx"

  2) 確認盤點報告沒問題後：

       python align_chunda_items_20260818.py ^
           --excel-path "20260716春大直設料號明細表-加代碼_20260818.xlsx" --execute

  預設 DB 路徑 C:\\portal_data\\cycle-purchase.db（同 backend/.env），
  要換用 --db-path。execute 模式會先備份（含 -wal/-shm）再動手，
  全程單一 transaction，任何一步失敗都 rollback。

  ⚠️ 執行前請先停掉後端服務。跑完後**必須重啟後端**，讓 main.py 的
  `_migrate_cycle_purchase_item_mapping_unique` 把料號對照表的唯一鍵
  從 (item_id, company) 放寬成 (item_id, company, department_id)——
  沒放寬之前，文具料號的第二筆部門對照會插不進去。本腳本會在執行前自己
  檢查唯一鍵版本，還是舊版就直接中止並告訴你先重啟後端。

────────────────────────────────────────────────────────────────
安全邊界（刻意不做的事）
────────────────────────────────────────────────────────────────
- **不刪任何單據**。上次的 reimport 腳本會把公司別＝春大直的請購／彙整／
  採購／驗收／請款單整批清掉（當時全是測試資料）。現在這些可能是真的，
  所以：要刪的料號若已被單據引用，本腳本**不刪、改成停用**（is_active=0）
  並在報告裡列出來，由人工決定。
- **不改 portal.db**。部門主檔已於 2026-08-17 改成鏡像自「系統設定 →
  公司/部門管理」（portal.db RefDepartment）。若要改名的部門
  `source_department_id` 非 NULL，改了下次同步就被蓋回去，所以這種情況
  直接中止並請你去系統設定那邊改，不在這裡硬改。
- **不改 items.category 既有的非空值**。只補空的。類別字串是週期設定
  `applicable_categories` 的比對鍵，改字串會讓既有週期設定篩不到料號
  （與 2026-07-16 彙整單期別字串對不上是同一種病灶）。補上來的細分類
  名稱只寫進類別主檔的 sub_name。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，請先執行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)


COMPANY = "春大直"
CODE_PREFIX = "CH-"          # items.item_code 前綴（見 models/cycle_purchase_item.py 2026-08-13）
DEFAULT_DB = r"C:\portal_data\cycle-purchase.db"

# ⚠️ 2026-08-18 第二版（跑過正式區盤點後修正）
# ────────────────────────────────────────────────────────────────
# 第一版假設要「把工務部改名成工程部、清潔部改名成管理部」。實際盤點發現
# **工程部與管理部本來就已經存在**——2026-08-17 部門主檔改成鏡像同步之後，
# 「系統設定 → 公司/部門管理」已經把春大直的 13 個部門灌進來（管理／資訊／
# 工程／財務／採購／房務／商場／客房／餐飲／客服／安全／工務／營業），
# 週採早期本地自建的「清潔部」「文具印刷部」則留在旁邊。
#
# 所以這不是改名，是**把料號對照從舊部門搬到既有部門**（與 Samuel 確認）：
#   工務部(同步) → 工程部(同步)   清潔部(本地) → 管理部(同步)
#   文具印刷部(本地) → 四個部門    營業部(同步) 不變
# 搬完把三個空掉的舊部門停用（不刪除，歷史單據還指著它們）。
#
# 改名的做法在這裡是錯的：目標名稱已被別筆佔用，改下去會出現兩個同名部門；
# 而且工務部是同步來的，dept_name 由同步覆蓋，改了也會被蓋回去。
# 現在這版完全不動 dept_name，只動 mapping 的 department_id 與 is_active
# （is_active 是週採自己維護的欄位，同步不碰，見 cycle_purchase_department_sync.py）。

# ⚠️ 2026-08-18 第三版（正式區盤點後再修一次）
# ────────────────────────────────────────────────────────────────
# 第二版假設「工程部／管理部已經存在，所以只要搬 mapping、不要改名」。
# 那是**開發機**的樣子。正式區（D:\portal）跑下來才發現：部門同步從沒在正式區
# 跑過，春大直底下只有週採早期自建的 4 個本地部門
# （清潔部 / 工務部 / 營業部 / 文具印刷部），13 個同步部門一個都沒有。
#
# 兩台機器狀態不同，所以腳本改成**依現況自動選做法**（與 Samuel 確認）：
#
#   目標部門已存在                → 直接用它，只搬 mapping（開發機走這條）
#   目標部門不存在，舊部門是本地自建 → 把舊部門改名成目標名稱（正式區走這條）
#   目標部門不存在，舊部門是同步來的 → 中止。改名會被下次同步蓋回去，
#                                    這種情況要去「系統設定 → 公司/部門管理」處理
#   目標部門不存在，也沒有舊部門可改名 → 本地自建一個（正式區的「客服部」走這條）
#
# 改名只發生在「本地自建」的部門上，所以不會跟同步打架，也不會出現兩個同名部門。
ALL_DEPTS = ["工程部", "管理部", "營業部", "客服部"]

# 舊部門 → 這次要對應到的部門。
# 目標存在時＝把 mapping 搬過去；目標不存在且舊部門是本地自建時＝直接改名。
DEPT_REMAP = {
    "工務部": "工程部",
    "清潔部": "管理部",
}
# 搬完之後不會再有料號掛在底下的舊部門，停用不刪除：硬刪會被既有單據的
# FK（RESTRICT）擋住，也會弄丟歷史單據的部門名稱。
# 停用前會先確認它底下真的沒有剩餘的春大直對照，還有殘留就跳過並回報。
DEPT_TO_DEACTIVATE = ["工務部", "清潔部", "文具印刷部"]

# Excel 分頁 → (大分類名稱, 該分頁料號要掛的部門清單)
SHEET_SPEC = {
    "維修備品-工程部需求":   ("工程",     ["工程部"]),
    "清潔用品-管理部需求":   ("清潔",     ["管理部"]),
    "文具用品-所有部門需求": ("文具",     ALL_DEPTS),
    "營業備品-營業部需求":   ("營業用品", ["營業部"]),
}

# 大分類代碼（料號首字）→ 大分類名稱
MAJOR_NAME = {"E": "工程", "C": "清潔", "G": "文具", "S": "營業用品"}

# ── 類別主檔：中分類名稱 ─────────────────────────────────────────────────────
# 來源 Excel 的「類別」欄是「中分類名稱[-細分類名稱]」黏成的一個字串，
# 用「-」機械式切會切錯（例：「文具-筆-螢光」的中分類是「文具-筆」不是「文具」）。
# 因此中分類名稱在這裡逐一寫死，對照《設料號明細表》的「編碼原則」分頁。
MID_NAME = {
    ("E", "01"): "空調備品",   ("E", "02"): "公區照明", ("E", "03"): "停車場照明",
    ("E", "04"): "濾芯",
    ("C", "01"): "客廁備品",   ("C", "02"): "公區備品", ("C", "03"): "公區",
    ("G", "01"): "印刷",       ("G", "02"): "紙",       ("G", "03"): "電池",
    ("G", "04"): "膠帶",       ("G", "05"): "束線帶",   ("G", "06"): "便利貼",
    ("G", "07"): "標籤貼",     ("G", "08"): "護貝膠膜", ("G", "09"): "碳粉匣",
    ("G", "10"): "文具-筆",    ("G", "11"): "文具-修正", ("G", "12"): "文具-整理",
    ("G", "13"): "文具-刀",    ("G", "14"): "文具-尺",  ("G", "15"): "文具",
    ("G", "16"): "收納",       ("G", "17"): "資料檔案夾",
    ("S", "01"): "包耗材",     ("S", "02"): "餐具",     ("S", "03"): "公區用品",
}

# ── 類別主檔：細分類名稱「補上」的部分 ────────────────────────────────────────
# 來源 Excel 有 30 個細分類碼沒有自己的名稱（類別字串只寫到中分類，例如
# G12 底下八個細分類全都叫「文具-整理」）。這裡依該細分類底下的實際品名
# 命名補齊，供人工複核；**只寫進類別主檔的 sub_name，不會改動任何料號的
# category 字串**（理由見檔頭「安全邊界」）。
# 註記格式：(大分類, 中分類, 細分類): (細分類名稱, 命名依據的品名摘要)
SUB_NAME_FILLED = {
    ("G", "01", "01"): ("信封",       "西式信封／中式信封"),
    ("G", "01", "02"): ("名片",       "名片"),
    ("G", "01", "03"): ("感熱紙卷",   "統一發票感熱紙卷"),
    ("G", "02", "01"): ("影印紙",     "A3／A4 影印紙"),
    ("G", "02", "02"): ("多用途紙",   "阿波羅影印雷射噴墨3用紙"),
    ("G", "06", "01"): ("N次貼",      "N次貼系列"),
    ("G", "06", "02"): ("便條紙",     "3M E56N-2PK便條紙"),
    ("G", "07", "01"): ("標籤",       "標籤貼／印刷標籤"),
    ("G", "07", "02"): ("索引片",     "雙面五彩／五彩索引片"),
    ("G", "07", "03"): ("透明標籤",   "LD透明標籤／保護膜標籤"),
    ("G", "11", "02"): ("替換帶",     "PLUS修正帶(內帶)"),
    ("G", "11", "03"): ("修正帶",     "SDI輕鬆按修正帶"),
    ("G", "11", "04"): ("橡皮擦",     "PLUS環保橡皮擦"),
    ("G", "12", "01"): ("膠水",       "膠水 白金50cc"),
    ("G", "12", "02"): ("口紅膠",     "UHU 口紅膠"),
    ("G", "12", "03"): ("釘書機",     "釘書機／訂書機"),
    ("G", "12", "04"): ("釘書針",     "SDI 釘書針"),
    ("G", "12", "05"): ("迴紋針",     "力大牌迴紋針"),
    ("G", "12", "06"): ("長尾夾",     "黑色長尾夾"),
    ("G", "12", "07"): ("原子夾",     "塑膠原子夾"),
    ("G", "12", "08"): ("事務帶",     "開明事務帶"),
    ("G", "13", "01"): ("剪刀",       "事務用剪刀"),
    ("G", "13", "02"): ("美工刀",     "美工刀／美工刀片"),
    ("G", "14", "01"): ("直尺",       "塑膠尺(30Cm)"),
    ("G", "14", "02"): ("捲尺",       "捲尺 5m"),
    ("G", "16", "03"): ("雜誌箱",     "雜誌箱 MF2408"),
    ("G", "17", "01"): ("內頁袋",     "11孔內頁袋"),
    ("G", "17", "02"): ("L型夾",      "透明夾L型"),
    ("G", "17", "03"): ("文件夾",     "2孔PP透明文件夾"),
    ("G", "17", "04"): ("3孔檔案夾",  "3孔檔案夾 立強牌R.870"),
    ("G", "17", "05"): ("4孔資料夾",  "4孔資料夾"),
    ("G", "17", "06"): ("卷宗",       "中式卷宗(紙質)"),
    ("S", "03", "02"): ("瓦斯桶",     "10kg瓦斯桶"),
}

# 來源資料本身的疑義，寫進類別主檔的 notes 供人工判讀，腳本不自行更正
# （CLAUDE.md §4：察覺不相關的問題只提出、不自行修改）。
CATEGORY_NOTES = {
    ("G", "04", "01"): "來源 Excel 同一個細分類有兩種寫法：G0401001「膠帶-雙面泡棉」、"
                       "G0401002「膠帶雙面泡棉」（少一個連字號）。此處以有連字號者為準，"
                       "料號端的 category 字串未更動，需人工確認要不要統一。",
    ("G", "03", "02"): "來源 Excel 類別字串為「電池-離」，品名為「離電池3V-CR1220」，"
                       "CR1220 為鋰電池，疑為「鋰」誤植成「離」。未自行更正。",
    ("C", "03", "01"): "中分類名稱在來源是「公區」（C01「客廁備品」、C02「公區備品」皆為"
                       "「◯◯備品」），命名不一致，照原樣建檔。",
}


# ════════════════════════════════════════════════════════════════════════════
# Excel 讀取
# ════════════════════════════════════════════════════════════════════════════

def parse_moq(raw) -> tuple[int, str | None]:
    """mini order 欄位在 model 是 Integer NOT NULL，但來源有一筆寫「1箱」。"""
    if raw is None or raw == "":
        return 0, None
    if isinstance(raw, (int, float)):
        return int(raw), None
    m = re.search(r"\d+", str(raw))
    if m:
        return int(m.group()), f"原始 mini order 欄位文字為「{raw}」，已取數字部分，請人工複查單位"
    return 0, f"原始 mini order 欄位文字為「{raw}」，無法解析為數字，已存 0，請人工複查"


def split_name_spec(full_name: str) -> tuple[str, str | None]:
    """用第一個空白切品名/規格；切不出來就整句當品名。"""
    full_name = (full_name or "").strip()
    parts = re.split(r"\s+", full_name, maxsplit=1)
    if len(parts) == 1 or not parts[1].strip():
        return parts[0], None
    return parts[0], parts[1].strip()


def read_excel_rows(excel_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    rows: list[dict] = []
    for sheet_name, (major_name, dept_names) in SHEET_SPEC.items():
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Excel 缺少分頁「{sheet_name}」，請確認檔案版本正確")
        ws = wb[sheet_name]
        header = [c.value for c in ws[1]]

        def col(name: str, required: bool = True):
            if name in header:
                return header.index(name)
            if required:
                raise RuntimeError(f"分頁「{sheet_name}」缺少欄位「{name}」")
            return None

        idx = {
            "code": col("料號"), "name": col("品名"),
            "cat": col("類別", False), "unit": col("單位", False),
            "vendor": col("廠商", False), "price": col("單價", False),
            "max": col("最大庫存量", False), "min": col("最小庫存量", False),
        }
        moq_idx = col("mini order ", False)
        if moq_idx is None:
            moq_idx = col("mini order", False)

        for r in ws.iter_rows(min_row=3, values_only=True):
            raw_code = r[idx["code"]]
            if not raw_code:
                continue
            code = str(raw_code).strip()
            if not re.fullmatch(r"[A-Z]\d{7}", code):
                raise RuntimeError(
                    f"料號「{code}」不符編碼原則（1 碼英文 + 中分類 2 碼 + 細分類 2 碼 + 流水 3 碼），"
                    f"分頁「{sheet_name}」。請先修正 Excel，不要讓腳本猜。"
                )
            full_name = str(r[idx["name"]] or "").strip()
            name, spec = split_name_spec(full_name)
            moq_value, moq_note = parse_moq(r[moq_idx] if moq_idx is not None else None)

            def cell(key):
                i = idx[key]
                return r[i] if i is not None else None

            rows.append({
                "sheet": sheet_name,
                "dept_names": list(dept_names),
                "major_code": code[0],
                "major_name": major_name,
                "mid_code": code[1:3],
                "sub_code": code[3:5],
                "original_code": code,
                "item_code": CODE_PREFIX + code,
                "full_name": full_name,
                "item_name": name,
                "spec": spec,
                "category": str(cell("cat") or "").strip() or None,
                "unit": (str(cell("unit")).strip() if cell("unit") else None),
                "vendor_name": (str(cell("vendor")).strip() if cell("vendor") else None),
                "unit_price": cell("price"),
                "max_stock": cell("max"),
                "min_stock": cell("min"),
                "moq": moq_value,
                "moq_note": moq_note,
            })

    codes = [r["original_code"] for r in rows]
    dup = sorted({c for c in codes if codes.count(c) > 1})
    if dup:
        raise RuntimeError(f"Excel 內部料號重複，需先處理：{dup}")
    return rows


def build_categories(rows: list[dict]) -> list[dict]:
    """由料號碼位反推三層類別，一個 (大,中,細) 組合一列。"""
    seen: dict[tuple[str, str, str], dict] = {}
    for r in rows:
        key = (r["major_code"], r["mid_code"], r["sub_code"])
        if key in seen:
            continue
        mid_name = MID_NAME.get((r["major_code"], r["mid_code"]))
        if not mid_name:
            raise RuntimeError(
                f"料號 {r['original_code']} 的中分類 {r['major_code']}{r['mid_code']} "
                f"不在 MID_NAME 對照表裡。請先在腳本裡補上正確的中分類名稱，不要用猜的。"
            )
        filled = SUB_NAME_FILLED.get(key)
        category_name = r["category"] or mid_name
        # 細分類名稱優先序：①人工補齊表 ②類別字串扣掉中分類前綴 ③留空
        if filled:
            sub_name, basis = filled
            note = f"細分類名稱為 2026-08-18 依品名補齊（依據：{basis}），來源 Excel 未命名。"
        else:
            sub_name = None
            note = None
            if category_name.startswith(mid_name + "-"):
                sub_name = category_name[len(mid_name) + 1:]
        extra = CATEGORY_NOTES.get(key)
        notes = "；".join(x for x in (note, extra) if x) or None

        seen[key] = {
            "company": COMPANY,
            "dept_names": r["dept_names"],
            "major_code": r["major_code"], "major_name": MAJOR_NAME[r["major_code"]],
            "mid_code": r["mid_code"], "mid_name": mid_name,
            "sub_code": r["sub_code"], "sub_name": sub_name,
            "category_name": category_name,
            "serial_width": 3,
            "notes": notes,
        }
    return [seen[k] for k in sorted(seen)]


# ════════════════════════════════════════════════════════════════════════════
# DB 查詢輔助
# ════════════════════════════════════════════════════════════════════════════

def mapping_unique_is_new(conn: sqlite3.Connection) -> bool:
    """料號對照表的唯一鍵是否已放寬成三欄（見 main.py 的啟動遷移）。"""
    for row in conn.execute("PRAGMA index_list(cycle_purchase_item_mappings)").fetchall():
        name, unique, origin = row[1], row[2], row[3]
        if not unique or origin != "u":
            continue
        cols = [r[2] for r in conn.execute(f"PRAGMA index_info({name})").fetchall()]
        if cols == ["item_id", "company", "department_id"]:
            return True
    return False


def get_depts(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        "SELECT id, dept_code, dept_name, is_active, source_department_id "
        "FROM cycle_purchase_departments WHERE company = ? ORDER BY dept_code",
        (COMPANY,),
    ).fetchall()


def plan_departments(conn: sqlite3.Connection) -> tuple[list[dict], list[str]]:
    """
    依部門主檔現況決定每個目標部門要怎麼生出來（見檔案上方第三版說明）。

    回傳 (actions, blockers)。actions 每筆是
      {"target": 目標部門名, "action": "use"|"rename"|"create",
       "from": 來源部門名或 None, "dept_id": 現有 id 或 None,
       "note": 給報告用的說明}
    盤點與 execute 共用同一份決策，避免兩邊邏輯漂移。
    """
    by_name = {d[2]: d for d in get_depts(conn)}
    rename_source = {new: old for old, new in DEPT_REMAP.items()}
    actions: list[dict] = []
    blockers: list[str] = []

    for target in ALL_DEPTS:
        if target in by_name:
            row = by_name[target]
            actions.append({
                "target": target, "action": "use", "from": None, "dept_id": row[0],
                "note": f"已存在 id={row[0]}（{'同步' if row[4] else '本地自建'}）",
            })
            continue

        old = rename_source.get(target)
        old_row = by_name.get(old) if old else None
        if old_row is None:
            actions.append({
                "target": target, "action": "create", "from": None, "dept_id": None,
                "note": "部門主檔沒有、也沒有可改名的舊部門 → 本地自建一個",
            })
            continue

        if old_row[4]:
            blockers.append(
                f"部門「{old}」是同步來源（source_department_id={old_row[4]}），"
                f"不能在這裡改名成「{target}」——改了下次同步會被蓋回去。"
                f"請到「系統設定 → 公司/部門管理」處理。"
            )
            actions.append({
                "target": target, "action": "blocked", "from": old, "dept_id": old_row[0],
                "note": "❌ 舊部門是同步來源，不可改名",
            })
            continue

        actions.append({
            "target": target, "action": "rename", "from": old, "dept_id": old_row[0],
            "note": f"把本地自建的「{old}」(id={old_row[0]}) 改名為「{target}」",
        })
    return actions, blockers


def chunda_only_item_ids(conn: sqlite3.Connection) -> set[int]:
    """只屬於春大直（沒跟其他公司共用）的料號 id。"""
    return {
        row[0] for row in conn.execute(
            """
            SELECT id FROM cycle_purchase_items
            WHERE id IN (SELECT item_id FROM cycle_purchase_item_mappings WHERE company = ?)
              AND id NOT IN (SELECT item_id FROM cycle_purchase_item_mappings WHERE company != ?)
            """,
            (COMPANY, COMPANY),
        ).fetchall()
    }


REFERENCING_TABLES = [
    ("cycle_purchase_request_items", "item_id", "請購單明細"),
    ("cycle_purchase_po_items", "item_id", "採購單明細"),
    ("cycle_purchase_receiving_items", "item_id", "驗收單明細"),
    ("cycle_purchase_summary", "item_id", "彙整單"),
]


def item_reference_counts(conn: sqlite3.Connection, item_ids: set[int]) -> dict[int, list[str]]:
    """回傳 {item_id: ["請購單明細×3", ...]}，沒被引用的不會出現在 dict 裡。"""
    result: dict[int, list[str]] = {}
    if not item_ids:
        return result
    ids = list(item_ids)
    chunk = 500
    for table, col, label in REFERENCING_TABLES:
        try:
            for i in range(0, len(ids), chunk):
                part = ids[i:i + chunk]
                ph = ",".join("?" * len(part))
                for row in conn.execute(
                    f"SELECT {col}, COUNT(*) FROM {table} WHERE {col} IN ({ph}) GROUP BY {col}",
                    part,
                ).fetchall():
                    result.setdefault(row[0], []).append(f"{label}×{row[1]}")
        except sqlite3.OperationalError:
            continue  # 該表尚未建立（舊環境），跳過
    return result


# ════════════════════════════════════════════════════════════════════════════
# 盤點報告
# ════════════════════════════════════════════════════════════════════════════

def report(conn: sqlite3.Connection, rows: list[dict], categories: list[dict]) -> None:
    sep = "=" * 74
    print(sep)
    print("盤點模式（不會寫入任何資料）")
    print(sep)

    print("\n--- 0. 前置條件 ---")
    ok = mapping_unique_is_new(conn)
    print(f"  料號對照表唯一鍵已放寬成 (item_id, company, department_id)："
          f"{'是' if ok else '❌ 否 —— execute 會中止，請先重啟後端讓自動遷移跑完'}")

    print("\n--- 1. 春大直部門主檔現況 ---")
    depts = get_depts(conn)
    by_name = {d[2]: d for d in depts}
    for d in depts:
        src = "同步" if d[4] else "本地自建"
        print(f"  id={d[0]:<4} code={d[1]:<12} name={d[2]:<10} is_active={d[3]} 來源={src}")

    print("\n  這次要做的部門調整（做法依現況自動選，見腳本上方第三版說明）：")
    actions, blockers = plan_departments(conn)
    label = {"use": "沿用", "rename": "改名", "create": "新建", "blocked": "中止"}
    for a in actions:
        print(f"    目標部門「{a['target']}」→ {label[a['action']]}：{a['note']}")

    # 改名之後，原本掛在舊部門底下的對照會自動變成掛在新名稱底下（同一個 id），
    # 不需要也不會逐筆搬移，所以搬移筆數只算「目標部門已存在」那種情況。
    scope_ids = chunda_only_item_ids(conn)
    renamed_from = {a["from"] for a in actions if a["action"] == "rename"}
    for old, new in DEPT_REMAP.items():
        row = by_name.get(old)
        if row is None:
            continue
        if old in renamed_from:
            cnt = conn.execute(
                "SELECT COUNT(*) FROM cycle_purchase_item_mappings "
                "WHERE company = ? AND department_id = ?",
                (COMPANY, row[0]),
            ).fetchone()[0]
            print(f"    {old}(id={row[0]}) 改名為 {new}　"
                  f"[底下 {cnt} 筆對照跟著改歸屬，不需逐筆搬移]")
            continue
        moving = out_of_scope = 0
        for (item_id_,) in conn.execute(
            "SELECT item_id FROM cycle_purchase_item_mappings "
            "WHERE company = ? AND department_id = ?",
            (COMPANY, row[0]),
        ).fetchall():
            if item_id_ in scope_ids:
                moving += 1
            else:
                out_of_scope += 1
        extra = f"，另有 {out_of_scope} 筆範圍外對照留在原地" if out_of_scope else ""
        print(f"    {old}(id={row[0]}) → {new}　[搬移 {moving} 筆料號對照{extra}]")

    for name in DEPT_TO_DEACTIVATE:
        row = by_name.get(name)
        if not row or not row[3]:
            continue
        if name in renamed_from:
            continue  # 這個部門會被改名成目標部門，當然不會停用
        left = sum(
            1 for (item_id_,) in conn.execute(
                "SELECT item_id FROM cycle_purchase_item_mappings "
                "WHERE company = ? AND department_id = ?",
                (COMPANY, row[0]),
            ).fetchall()
            if item_id_ not in scope_ids
        )
        if left:
            print(f"    ⚠ 部門「{name}」搬完後仍有 {left} 筆範圍外對照（兩公司共用的統購料號），"
                  f"**不會停用**——停用會讓那些料號在請購時整批消失")
        else:
            print(f"    停用部門「{name}」（搬完底下就沒有料號了；不刪除，保留歷史單據）")

    print("\n--- 2. 料號比對 ---")
    excel_codes = {r["original_code"] for r in rows}
    target_ids = chunda_only_item_ids(conn)
    db_rows = {}
    if target_ids:
        ph = ",".join("?" * len(target_ids))
        for row in conn.execute(
            f"""
            SELECT i.id, i.item_code, i.item_name, i.category, m.original_code
            FROM cycle_purchase_items i
            JOIN cycle_purchase_item_mappings m ON m.item_id = i.id AND m.company = ?
            WHERE i.id IN ({ph})
            """,
            [COMPANY, *target_ids],
        ).fetchall():
            db_rows.setdefault(row[0], row)

    db_codes = {r[4] for r in db_rows.values() if r[4]}
    to_delete = {iid: r for iid, r in db_rows.items() if (r[4] or "") not in excel_codes}
    to_create = sorted(excel_codes - db_codes)

    print(f"  Excel 料號：{len(excel_codes)} 筆")
    print(f"  DB「只屬於春大直」料號：{len(db_rows)} 筆")
    print(f"  → 需新增：{len(to_create)} 筆" + (f"　{to_create[:10]}" if to_create else ""))
    print(f"  → 不在附件、需刪除：{len(to_delete)} 筆")
    if to_delete:
        refs = item_reference_counts(conn, set(to_delete))
        for iid, r in sorted(to_delete.items()):
            used = "、".join(refs.get(iid, []))
            action = f"⚠ 已被引用（{used}）→ 改為停用" if used else "可直接刪除"
            print(f"      id={iid} {r[1]} {r[2]}　[{action}]")

    # 只算「留下來的」料號；要刪/要停用的那些補 category 沒有意義
    empty_cat = [
        r for iid, r in db_rows.items()
        if iid not in to_delete and not (r[3] or "").strip()
    ]
    print(f"  → category 為空、可依 Excel 補上：{len(empty_cat)} 筆")
    for r in empty_cat[:15]:
        print(f"      {r[1]} {r[2]}")

    print("\n--- 3. 料號對照表（部門展開） ---")
    per_dept: dict[str, int] = {}
    for r in rows:
        for d in r["dept_names"]:
            per_dept[d] = per_dept.get(d, 0) + 1
    total = sum(per_dept.values())
    for d, n in sorted(per_dept.items(), key=lambda x: -x[1]):
        print(f"    {d}：{n} 筆")
    print(f"    合計 {total} 筆對照（{len(rows)} 個料號；文具用品 "
          f"{sum(1 for r in rows if r['major_code'] == 'G')} 筆 × {len(ALL_DEPTS)} 個部門）")
    now = conn.execute(
        "SELECT COUNT(*) FROM cycle_purchase_item_mappings WHERE company = ?", (COMPANY,)
    ).fetchone()[0]
    keep_ids = {iid for code, iid in
                {r[4]: iid for iid, r in db_rows.items()}.items() if code in excel_codes}
    in_scope = 0
    if keep_ids:
        ph2 = ",".join("?" * len(keep_ids))
        in_scope = conn.execute(
            f"SELECT COUNT(*) FROM cycle_purchase_item_mappings "
            f"WHERE company = ? AND item_id IN ({ph2})",
            [COMPANY, *keep_ids],
        ).fetchone()[0]
    print(f"    目前 DB 春大直對照共 {now} 筆，其中屬於本次範圍的 {in_scope} 筆 → 會變成 {total} 筆")
    print(f"    範圍外的 {now - in_scope} 筆（兩公司共用的統購料號、已停用但仍被單據引用的料號）不會被動到")

    print("\n--- 4. 類別主檔 ---")
    try:
        exists = conn.execute(
            "SELECT COUNT(*) FROM cycle_purchase_categories WHERE company = ?", (COMPANY,)
        ).fetchone()[0]
    except sqlite3.OperationalError:
        print("  ❌ 資料表 cycle_purchase_categories 不存在 —— 請先重啟後端讓 create_all 建表")
        blockers.append("cycle_purchase_categories 資料表不存在")
        exists = 0
    print(f"  將建立/更新 {len(categories)} 個類別（目前 DB 有 {exists} 筆）")
    filled = [c for c in categories if c["sub_name"] and SUB_NAME_FILLED.get(
        (c["major_code"], c["mid_code"], c["sub_code"]))]
    unnamed = [c for c in categories if not c["sub_name"]]
    print(f"  其中細分類名稱由本次補齊的有 {len(filled)} 個（來源 Excel 未命名，"
          f"依品名命名，請人工複核）：")
    for c in filled:
        print(f"    {c['major_code']}{c['mid_code']}{c['sub_code']}　"
              f"{c['mid_name']} → {c['sub_name']}　（類別字串仍為「{c['category_name']}」）")
    if unnamed:
        print(f"  仍無細分類名稱（該中分類只有一個細分類，類別字串＝中分類）：{len(unnamed)} 個")
        print("    " + "、".join(f"{c['major_code']}{c['mid_code']}{c['sub_code']}"
                                 f" {c['category_name']}" for c in unnamed))
    noted = [c for c in categories if c["notes"] and CATEGORY_NOTES.get(
        (c["major_code"], c["mid_code"], c["sub_code"]))]
    if noted:
        print(f"\n  來源資料疑義（已寫進 notes，腳本不自行更正）：")
        for c in noted:
            print(f"    {c['major_code']}{c['mid_code']}{c['sub_code']}："
                  f"{CATEGORY_NOTES[(c['major_code'], c['mid_code'], c['sub_code'])]}")

    moq_issues = [r for r in rows if r["moq_note"]]
    if moq_issues:
        print(f"\n--- 5. Excel 欄位格式問題（{len(moq_issues)} 筆） ---")
        for r in moq_issues:
            print(f"    {r['original_code']}: {r['moq_note']}")

    print("\n" + sep)
    if blockers:
        print("❌ 有以下阻擋項，execute 會中止：")
        for b in blockers:
            print(f"   - {b}")
    else:
        print("✅ 沒有阻擋項。確認以上都沒問題後，加 --execute 才會真的動資料。")
    print(sep)


# ════════════════════════════════════════════════════════════════════════════
# 執行
# ════════════════════════════════════════════════════════════════════════════

def backup_db(db_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = db_path.with_name(f"{db_path.name}.bak-{ts}")
    shutil.copy2(db_path, backup)
    for suffix in ("-wal", "-shm"):
        side = db_path.with_name(db_path.name + suffix)
        if side.exists():
            shutil.copy2(side, backup.with_name(backup.name + suffix))
    print(f"已備份到：{backup}")
    return backup


def execute(conn: sqlite3.Connection, rows: list[dict], categories: list[dict],
            refresh_fields: bool) -> None:
    if not mapping_unique_is_new(conn):
        raise RuntimeError(
            "料號對照表的唯一鍵還是舊版 (item_id, company)。請先重啟後端服務，"
            "讓 main.py 的 _migrate_cycle_purchase_item_mapping_unique 跑完再執行本腳本。"
        )
    try:
        conn.execute("SELECT 1 FROM cycle_purchase_categories LIMIT 1")
    except sqlite3.OperationalError:
        raise RuntimeError(
            "資料表 cycle_purchase_categories 不存在。請先重啟後端服務，"
            "讓 create_all 建好類別主檔資料表再執行本腳本。"
        )

    dept_actions, dept_blockers = plan_departments(conn)
    if dept_blockers:
        raise RuntimeError("中止執行：\n  - " + "\n  - ".join(dept_blockers))

    conn.execute("BEGIN")
    try:
        stats = {k: 0 for k in (
            "dept_renamed", "dept_created", "dept_deactivated",
            "item_created", "item_updated", "item_deleted", "item_deactivated",
            "category_filled", "mapping_created", "mapping_updated", "mapping_moved",
            "mapping_deleted",
            "cat_created", "cat_updated",
        )}

        # ── 1. 部門 ──────────────────────────────────────────────────────────
        # 依 plan_departments() 的決策執行（沿用／改名／新建）。改名只會發生在
        # 本地自建的部門上，同步來源的在上面就已經擋掉了。
        dept_id: dict[str, int] = {}
        renamed_from: set[str] = set()
        for a in dept_actions:
            if a["action"] == "use":
                dept_id[a["target"]] = a["dept_id"]
            elif a["action"] == "rename":
                conn.execute(
                    "UPDATE cycle_purchase_departments SET dept_name = ? WHERE id = ?",
                    (a["target"], a["dept_id"]),
                )
                dept_id[a["target"]] = a["dept_id"]
                renamed_from.add(a["from"])
                stats["dept_renamed"] += 1
            elif a["action"] == "create":
                cur = conn.execute(
                    "INSERT INTO cycle_purchase_departments "
                    "(company, dept_code, dept_name, is_active) VALUES (?, ?, ?, 1)",
                    (COMPANY, f"CH-{a['target']}", a["target"]),
                )
                dept_id[a["target"]] = cur.lastrowid
                stats["dept_created"] += 1

        # 四個目標部門都必須啟用，否則請購單產不出來
        for did in dept_id.values():
            conn.execute(
                "UPDATE cycle_purchase_departments SET is_active = 1 WHERE id = ?", (did,)
            )
        depts = {d[2]: d for d in get_depts(conn)}
        # 舊部門的停用留到最後（步驟 7）：要先把料號搬完，才知道它底下是不是
        # 真的空了。

        # ── 2. 現有料號盤點 ─────────────────────────────────────────────────
        target_ids = chunda_only_item_ids(conn)
        db_by_code: dict[str, int] = {}
        db_meta: dict[int, tuple] = {}
        if target_ids:
            ph = ",".join("?" * len(target_ids))
            for row in conn.execute(
                f"""
                SELECT i.id, i.category, m.original_code
                FROM cycle_purchase_items i
                JOIN cycle_purchase_item_mappings m ON m.item_id = i.id AND m.company = ?
                WHERE i.id IN ({ph})
                """,
                [COMPANY, *target_ids],
            ).fetchall():
                db_meta[row[0]] = row
                if row[2]:
                    db_by_code[row[2]] = row[0]

        excel_by_code = {r["original_code"]: r for r in rows}

        # ── 3. 刪除／停用不在附件裡的料號 ───────────────────────────────────
        obsolete = {iid for code, iid in db_by_code.items() if code not in excel_by_code}
        obsolete |= {iid for iid in target_ids if iid not in db_meta}  # 沒有 mapping 的孤兒
        refs = item_reference_counts(conn, obsolete)
        for iid in sorted(obsolete):
            if iid in refs:
                # 只在還是啟用中時才動，否則重跑一次就會再貼一段同樣的備註
                cur = conn.execute(
                    "UPDATE cycle_purchase_items SET is_active = 0, "
                    "notes = COALESCE(notes || char(10), '') || ? "
                    "WHERE id = ? AND is_active = 1",
                    (f"[2026-08-18] 不在春大直料號明細表內，但已被{('、'.join(refs[iid]))}引用，"
                     f"故停用而非刪除。", iid),
                )
                stats["item_deactivated"] += cur.rowcount
            else:
                conn.execute("DELETE FROM cycle_purchase_items WHERE id = ?", (iid,))
                stats["item_deleted"] += 1

        # ── 4. 新增缺少的料號 ／ 補空白 category ────────────────────────────
        for code, r in excel_by_code.items():
            iid = db_by_code.get(code)
            if iid is None:
                cur = conn.execute(
                    """
                    INSERT INTO cycle_purchase_items
                        (item_code, item_name, spec, category, unit, default_qty, moq,
                         max_stock, min_stock, unit_price, is_active, is_cycle_item, notes)
                    VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?, ?, 1, 1, ?)
                    """,
                    (r["item_code"], r["item_name"], r["spec"], r["category"], r["unit"],
                     r["moq"], r["max_stock"], r["min_stock"], r["unit_price"],
                     f"[2026-08-18] 由春大直料號明細表新增。{r['moq_note'] or ''}".strip()),
                )
                db_by_code[code] = cur.lastrowid
                stats["item_created"] += 1
                continue

            existing_cat = (db_meta.get(iid, (None, None, None))[1] or "").strip()
            if not existing_cat and r["category"]:
                conn.execute(
                    "UPDATE cycle_purchase_items SET category = ? WHERE id = ?",
                    (r["category"], iid),
                )
                stats["category_filled"] += 1
            if refresh_fields:
                conn.execute(
                    "UPDATE cycle_purchase_items SET unit = ?, moq = ?, max_stock = ?, "
                    "min_stock = ?, unit_price = ?, is_active = 1 WHERE id = ?",
                    (r["unit"], r["moq"], r["max_stock"], r["min_stock"],
                     r["unit_price"], iid),
                )
                stats["item_updated"] += 1

        # ── 5. 料號對照表：依分頁展開到部門 ─────────────────────────────────
        vendor_by_name = {
            (row[1] or "").strip(): row[0]
            for row in conn.execute(
                "SELECT id, vendor_name FROM cycle_purchase_vendors"
            ).fetchall()
        }
        # ⚠️ 只處理「本次範圍內」的料號（＝附件裡的 178 筆對應到的 item_id）。
        # 不能拿 company='春大直' 全撈：兩家公司共用的 7 筆統購料號（如永豐餘
        # 衛生紙）也有春大直的對照，那 7 筆依 2026-08-13 決議完全不碰，
        # 全撈會把它們的春大直對照當成「多餘的」刪掉，日耀那邊的對照還在、
        # 春大直卻突然請購不到那些品項，而且不會有任何錯誤訊息。
        # 範圍再收一次：只有「附件裡有的料號」才算數。被刪掉的料號其對照已隨
        # ON DELETE CASCADE 一起消失；被停用（因為還被單據引用）的料號則刻意
        # 保留它的對照——那是歷史單據追溯的依據，砍掉只會讓舊單顯示不出料號。
        scope_item_ids = {
            db_by_code[code] for code in excel_by_code if code in db_by_code
        }
        existing_maps = {
            (row[1], row[2]): row[0]
            for row in conn.execute(
                "SELECT id, item_id, department_id FROM cycle_purchase_item_mappings "
                "WHERE company = ?",
                (COMPANY,),
            ).fetchall()
            if row[1] in scope_item_ids
        }

        # ⚠️ 先「原地搬移」再新增/刪除，而不是一律砍掉重建。
        # mapping 的 id 被 cycle_purchase_request_items.item_mapping_id（RESTRICT）
        # 指著，砍掉重建會被既有請購單擋住；就算沒被擋住，重建也會讓舊單失去
        # 對照追溯。所以對每個料號：先把「還掛在舊部門、而該舊部門正是這次要
        # 搬走的」那幾筆，直接 UPDATE 成尚未被佔用的目標部門，id 原封不動。
        by_item: dict[int, list[tuple[int, int]]] = {}   # item_id -> [(dept_id, mapping_id)]
        for (iid_, did_), mid_ in existing_maps.items():
            by_item.setdefault(iid_, []).append((did_, mid_))

        target_ids_set = set(dept_id.values())
        for code, r in excel_by_code.items():
            iid = db_by_code.get(code)
            if iid is None:
                continue
            want_dids = [dept_id[d] for d in r["dept_names"]]
            current = by_item.get(iid, [])
            covered = {d for d, _ in current if d in want_dids}
            # 可以拿來改派的：目前掛在「不在目標清單裡」的部門那幾筆
            movable = [(d, m) for d, m in current if d not in target_ids_set]
            for did in want_dids:
                if did in covered or not movable:
                    continue
                old_did, mid = movable.pop(0)
                conn.execute(
                    "UPDATE cycle_purchase_item_mappings SET department_id = ? WHERE id = ?",
                    (did, mid),
                )
                existing_maps.pop((iid, old_did), None)
                existing_maps[(iid, did)] = mid
                covered.add(did)
                stats["mapping_moved"] += 1

        wanted: set[tuple[int, int]] = set()
        for code, r in excel_by_code.items():
            iid = db_by_code.get(code)
            if iid is None:
                continue
            vid = vendor_by_name.get(r["vendor_name"] or "")
            for dname in r["dept_names"]:
                did = dept_id[dname]
                wanted.add((iid, did))
                key = (iid, did)
                if key in existing_maps:
                    conn.execute(
                        "UPDATE cycle_purchase_item_mappings SET original_code = ?, "
                        "original_name = ?, original_vendor_name = ?, "
                        "original_unit_price = ?, vendor_id = COALESCE(?, vendor_id) "
                        "WHERE id = ?",
                        (code, r["full_name"], r["vendor_name"], r["unit_price"],
                         vid, existing_maps[key]),
                    )
                    stats["mapping_updated"] += 1
                else:
                    conn.execute(
                        """
                        INSERT INTO cycle_purchase_item_mappings
                            (item_id, company, department_id, original_code, original_name,
                             original_vendor_name, vendor_id, original_unit_price,
                             is_confirmed, notes)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                        """,
                        (iid, COMPANY, did, code, r["full_name"], r["vendor_name"], vid,
                         r["unit_price"], "[2026-08-18] 依春大直料號明細表分頁建立部門對照。"),
                    )
                    stats["mapping_created"] += 1

        # 多餘的春大直對照（部門已改名/文具舊的單一部門對照）刪掉。
        # 被單據 RESTRICT 綁住的（request_items.item_mapping_id）刪不掉，
        # 這裡不硬刪，留著並在結尾回報。
        blocked_mappings = []
        for (iid, did), mid in existing_maps.items():
            if (iid, did) in wanted:
                continue
            try:
                conn.execute("DELETE FROM cycle_purchase_item_mappings WHERE id = ?", (mid,))
                stats["mapping_deleted"] += 1
            except sqlite3.IntegrityError:
                blocked_mappings.append(mid)

        # ── 6. 類別主檔 ─────────────────────────────────────────────────────
        for c in categories:
            # dept_names 為四個部門時代表「不限部門」→ department_id 留 NULL
            did = None
            if len(c["dept_names"]) == 1:
                did = dept_id[c["dept_names"][0]]
            existing = conn.execute(
                "SELECT id FROM cycle_purchase_categories "
                "WHERE company = ? AND major_code = ? AND mid_code = ? AND sub_code = ?",
                (COMPANY, c["major_code"], c["mid_code"], c["sub_code"]),
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE cycle_purchase_categories SET department_id = ?, major_name = ?, "
                    "mid_name = ?, sub_name = ?, category_name = ?, serial_width = ?, "
                    "notes = ?, is_active = 1 WHERE id = ?",
                    (did, c["major_name"], c["mid_name"], c["sub_name"],
                     c["category_name"], c["serial_width"], c["notes"], existing[0]),
                )
                stats["cat_updated"] += 1
            else:
                conn.execute(
                    """
                    INSERT INTO cycle_purchase_categories
                        (company, department_id, major_code, major_name, mid_code, mid_name,
                         sub_code, sub_name, category_name, serial_width, is_active, notes)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                    """,
                    (COMPANY, did, c["major_code"], c["major_name"], c["mid_code"],
                     c["mid_name"], c["sub_code"], c["sub_name"], c["category_name"],
                     c["serial_width"], c["notes"]),
                )
                stats["cat_created"] += 1

        # ── 7. 停用搬空的舊部門 ─────────────────────────────────────────────
        # 放在最後才做：要先搬完料號，才知道它底下是不是真的空了。
        # 還有殘留（多半是兩公司共用的統購料號，或被單據卡住沒搬成的）就跳過，
        # 停用一個底下還有料號的部門，會讓那些料號在請購時整批消失。
        skipped_depts = []
        for name in DEPT_TO_DEACTIVATE:
            if name in renamed_from:
                continue  # 這個部門已經被改名成目標部門，它現在是啟用中的目標之一
            row = depts.get(name)
            if row is None or not row[3]:
                continue
            left = conn.execute(
                "SELECT COUNT(*) FROM cycle_purchase_item_mappings "
                "WHERE company = ? AND department_id = ?",
                (COMPANY, row[0]),
            ).fetchone()[0]
            if left:
                skipped_depts.append((name, left))
                continue
            conn.execute(
                "UPDATE cycle_purchase_departments SET is_active = 0 WHERE id = ?",
                (row[0],),
            )
            stats["dept_deactivated"] += 1

        conn.execute("COMMIT")
    except Exception:
        conn.execute("ROLLBACK")
        raise

    print("\n執行完成：")
    labels = {
        "dept_renamed": "本地自建部門改名", "dept_created": "部門新建（本地自建）",
        "dept_deactivated": "舊部門停用（已搬空）",
        "item_created": "料號新增", "item_updated": "料號欄位刷新", "item_deleted": "料號刪除",
        "item_deactivated": "料號停用（被單據引用，不能刪）",
        "category_filled": "料號 category 補空白",
        "mapping_moved": "料號對照原地改派部門（保留 id 與單據追溯）",
        "mapping_created": "料號對照新增", "mapping_updated": "料號對照更新",
        "mapping_deleted": "料號對照刪除",
        "cat_created": "類別主檔新增", "cat_updated": "類別主檔更新",
    }
    for k, label in labels.items():
        print(f"  {label}：{stats[k]}")
    if skipped_depts:
        print("\n⚠ 以下舊部門底下還有料號對照，沒有停用（多半是兩公司共用的統購料號）：")
        for name, left in skipped_depts:
            print(f"    {name}：仍有 {left} 筆對照")
    if blocked_mappings:
        print(f"\n⚠ 有 {len(blocked_mappings)} 筆多餘的料號對照被既有單據引用而無法刪除"
              f"（mapping id：{blocked_mappings}），請人工確認。")
    print("\n請重啟後端服務，並到「週期採購 → 類別主檔」複核細分類名稱。")


def main() -> int:
    ap = argparse.ArgumentParser(description="春大直料號／部門／類別主檔對齊（2026-08-18）")
    ap.add_argument("--excel-path", required=True, help="春大直設料號明細表 .xlsx")
    ap.add_argument("--db-path", default=DEFAULT_DB, help=f"cycle-purchase.db 路徑（預設 {DEFAULT_DB}）")
    ap.add_argument("--execute", action="store_true", help="真的寫入（預設只盤點）")
    ap.add_argument("--refresh-fields", action="store_true",
                    help="連同單價／庫存量／MOQ 一起從 Excel 刷新（預設不動，避免蓋掉人工修正）")
    args = ap.parse_args()

    excel_path = Path(args.excel_path)
    db_path = Path(args.db_path)
    if not excel_path.exists():
        print(f"找不到 Excel：{excel_path}", file=sys.stderr)
        return 1
    if not db_path.exists():
        print(f"找不到資料庫：{db_path}", file=sys.stderr)
        return 1

    rows = read_excel_rows(excel_path)
    categories = build_categories(rows)

    if args.execute:
        backup_db(db_path)

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        if args.execute:
            execute(conn, rows, categories, args.refresh_fields)
        else:
            report(conn, rows, categories)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
