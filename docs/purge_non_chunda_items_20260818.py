#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
週期採購 — 清掉「不在春大直料號明細表裡」的所有料號（2026-08-18）
==================================================================

依 Samuel 指示：料號主檔只保留《20260716春大直設料號明細表-加代碼_20260818.xlsx》
裡的 178 筆，其餘（含日曜天地的料號、兩公司共用的統購料號）全部刪除。

⚠️ 這支是**破壞性**腳本，而且刪的不只是春大直的範圍。請務必先看完盤點報告。
⚠️ 放在 docs/ 而不是 Temp/：正式區要靠 git pull 拿到它（CLAUDE.md §10 規則 4）。

────────────────────────────────────────────────────────────────
執行順序（很重要）
────────────────────────────────────────────────────────────────
必須先跑完 `docs/align_chunda_items_20260818.py --execute`，再跑這一支。
理由：本腳本的「要保留哪些」是靠「該料號有一筆春大直對照，且原始碼在 Excel 裡」
判定的，align 腳本負責把對照建好。順序反過來會保留錯的東西。
腳本會自己檢查，判定結果湊不出 178 筆就直接中止。

    1) 盤點（預設，不寫入）：
       python purge_non_chunda_items_20260818.py ^
           --excel-path "..\\Temp\\20260716春大直設料號明細表-加代碼_20260818.xlsx"

    2) 確認報告後：
       python purge_non_chunda_items_20260818.py ^
           --excel-path "..." --execute

────────────────────────────────────────────────────────────────
安全邊界
────────────────────────────────────────────────────────────────
- **只刪沒有被任何單據引用的料號**。被請購／彙整／採購／驗收單引用的料號，
  硬刪會被 FK RESTRICT 擋住（或者更糟：把別人的歷史單據拆爛）。這種料號本腳本
  **不刪、不停用、原封不動**，只在報告裡逐筆列出是哪張單在用。
  要不要連單據一起清，是另一個決定，請看完清單再說，不要讓腳本代勞。
- 刪除前會驗證「保留清單確實涵蓋 Excel 的全部 178 個料號」。少一個就中止——
  寧可什麼都不做，也不要刪到一半才發現保留集合是錯的。
- execute 前自動備份（含 -wal/-shm），全程單一 transaction。
"""
from __future__ import annotations

import argparse
import re
import shutil
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("缺少 openpyxl，請先執行：pip install openpyxl", file=sys.stderr)
    sys.exit(1)

COMPANY = "春大直"
DEFAULT_DB = r"C:\portal_data\cycle-purchase.db"

# 料號被哪些表引用（表名, 欄位, 顯示名稱, 單據編號來源）
REFERENCING = [
    ("cycle_purchase_request_items", "item_id", "請購單明細",
     "SELECT r.request_no FROM cycle_purchase_request_items ri "
     "JOIN cycle_purchase_requests r ON r.id = ri.request_id WHERE ri.item_id = ?"),
    ("cycle_purchase_summary", "item_id", "彙整單",
     "SELECT DISTINCT period_label FROM cycle_purchase_summary WHERE item_id = ?"),
    ("cycle_purchase_po_items", "item_id", "採購單明細",
     "SELECT p.po_no FROM cycle_purchase_po_items pi "
     "JOIN cycle_purchase_pos p ON p.id = pi.po_id WHERE pi.item_id = ?"),
    ("cycle_purchase_receiving_items", "item_id", "驗收單明細",
     "SELECT rc.receiving_no FROM cycle_purchase_receiving_items rci "
     "JOIN cycle_purchase_receiving rc ON rc.id = rci.receiving_id WHERE rci.item_id = ?"),
]


def read_excel_codes(excel_path: Path) -> set[str]:
    """讀出 Excel 四個分頁的全部料號（原始碼，如 E0101001）。"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    codes: set[str] = set()
    for ws in wb.worksheets:
        header = [c.value for c in ws[1]]
        if "料號" not in header:
            continue
        idx = header.index("料號")
        for r in ws.iter_rows(min_row=3, values_only=True):
            v = r[idx]
            if not v:
                continue
            code = str(v).strip()
            if re.fullmatch(r"[A-Z]\d{7}", code):
                codes.add(code)
    if not codes:
        raise RuntimeError("Excel 裡讀不到任何料號，請確認檔案版本正確")
    return codes


def chunda_only_item_ids(conn: sqlite3.Connection) -> set[int]:
    """只屬於春大直（沒有跟其他公司共用）的料號 id。"""
    return {
        row[0] for row in conn.execute(
            """
            SELECT id FROM cycle_purchase_items
            WHERE id IN (SELECT item_id FROM cycle_purchase_item_mappings WHERE company = ?)
              AND id NOT IN (SELECT item_id FROM cycle_purchase_item_mappings WHERE company != ?)
            """,
            (COMPANY, COMPANY),
        ).fetchall()
    }


def build_keep_set(conn: sqlite3.Connection, excel_codes: set[str]) -> tuple[set[int], dict]:
    """
    保留清單＝「**只屬於春大直**、且春大直對照的原始碼在 Excel 裡」的料號。

    ⚠️ 為什麼要加「只屬於春大直」這個條件：
    資料庫裡另外有 7 筆兩公司共用的統購料號（2026-08-13 決議保留的集團碼，
    如永豐餘衛生紙），它們同時掛著春大直與日曜天地的對照，而春大直那筆對照的
    原始碼**跟 178 筆專用料號重複**（同一個 C0101001 之類）。
    不加這個條件的話，同一個原始碼會對到兩筆料號，判不出該留哪一個。
    既然這次的目標是「只留附件裡的 178 筆」，共用料號本來就該一起清掉
    （附件是春大直專用的明細表，共用的集團碼不在裡面），所以直接以
    「春大直專用」為準，共用料號歸到刪除清單。

    回傳 (keep_ids, code_to_ids)。code_to_ids 供檢查有沒有哪個 Excel 料號
    在資料庫裡完全找不到、或對到多筆（後者代表 178 筆自己有重複，要先處理）。
    """
    only_chunda = chunda_only_item_ids(conn)
    keep_ids: set[int] = set()
    code_to_ids: dict[str, list[int]] = {}
    for item_id, code in conn.execute(
        "SELECT item_id, original_code FROM cycle_purchase_item_mappings WHERE company = ?",
        (COMPANY,),
    ).fetchall():
        if item_id not in only_chunda:
            continue
        if code and code.strip() in excel_codes:
            keep_ids.add(item_id)
            code_to_ids.setdefault(code.strip(), [])
            if item_id not in code_to_ids[code.strip()]:
                code_to_ids[code.strip()].append(item_id)
    return keep_ids, code_to_ids


def item_refs(conn: sqlite3.Connection, item_ids: list[int]) -> dict[int, list[str]]:
    """{item_id: ["請購單明細：PR-2026-08-001×2", ...]}，沒被引用的不出現。"""
    result: dict[int, list[str]] = {}
    if not item_ids:
        return result
    chunk = 400
    for table, col, label, doc_sql in REFERENCING:
        try:
            for i in range(0, len(item_ids), chunk):
                part = item_ids[i:i + chunk]
                ph = ",".join("?" * len(part))
                rows = conn.execute(
                    f"SELECT {col}, COUNT(*) FROM {table} WHERE {col} IN ({ph}) GROUP BY {col}",
                    part,
                ).fetchall()
                for item_id, cnt in rows:
                    try:
                        docs = [str(r[0]) for r in conn.execute(doc_sql, (item_id,)).fetchall() if r[0]]
                    except sqlite3.OperationalError:
                        docs = []
                    doc_txt = "、".join(sorted(set(docs))[:5]) or "—"
                    result.setdefault(item_id, []).append(f"{label}×{cnt}（{doc_txt}）")
        except sqlite3.OperationalError:
            continue  # 該表不存在（舊環境）
    return result


def collect(conn: sqlite3.Connection, excel_codes: set[str]):
    keep_ids, code_to_ids = build_keep_set(conn, excel_codes)

    all_items = {
        row[0]: row for row in conn.execute(
            "SELECT id, item_code, item_name, is_active FROM cycle_purchase_items"
        ).fetchall()
    }
    companies: dict[int, list[str]] = {}
    for item_id, company in conn.execute(
        "SELECT item_id, company FROM cycle_purchase_item_mappings"
    ).fetchall():
        companies.setdefault(item_id, [])
        if company not in companies[item_id]:
            companies[item_id].append(company)

    drop_ids = [i for i in all_items if i not in keep_ids]
    refs = item_refs(conn, drop_ids)
    return all_items, companies, keep_ids, code_to_ids, drop_ids, refs


def report(conn: sqlite3.Connection, excel_codes: set[str]) -> bool:
    all_items, companies, keep_ids, code_to_ids, drop_ids, refs = collect(conn, excel_codes)
    sep = "=" * 74
    print(sep)
    print("盤點模式（不會寫入任何資料）")
    print(sep)

    print(f"\n--- 1. 保留清單 ---")
    print(f"  Excel 料號數：{len(excel_codes)}")
    print(f"  資料庫比對得到、要保留的料號：{len(keep_ids)} 筆")
    missing = sorted(excel_codes - set(code_to_ids))
    dup = {c: ids for c, ids in code_to_ids.items() if len(ids) > 1}
    blockers = []
    if missing:
        blockers.append(f"有 {len(missing)} 個 Excel 料號在資料庫裡找不到春大直對照")
        print(f"  ❌ 找不到春大直對照的 Excel 料號（{len(missing)} 個）：{missing[:20]}")
        print(f"     → 請先執行 align_chunda_items_20260818.py --execute")
    if dup:
        blockers.append(f"有 {len(dup)} 個原始碼對到多筆料號")
        print(f"  ❌ 同一個原始碼對到多筆料號（資料重複，要先處理）：")
        for code, ids in list(dup.items())[:20]:
            names = "、".join(f"id={i}({all_items[i][1]})" for i in ids)
            print(f"       {code} → {names}")

    print(f"\n--- 2. 要刪除的料號：{len(drop_ids)} 筆 ---")
    by_company: dict[str, list[int]] = {}
    for i in drop_ids:
        key = "／".join(companies.get(i, [])) or "（無任何公司對照）"
        by_company.setdefault(key, []).append(i)
    for key, ids in sorted(by_company.items(), key=lambda x: -len(x[1])):
        blocked = [i for i in ids if i in refs]
        print(f"    {key}：{len(ids)} 筆"
              f"（其中 {len(blocked)} 筆被單據引用，本腳本不會刪）")

    print(f"\n--- 3. 被單據引用、**不會刪**的料號：{len(refs)} 筆 ---")
    if refs:
        print("    這些料號保持原狀（不刪也不停用）。要不要連單據一起清是另一個決定，")
        print("    請看完下面清單再說：")
        for item_id in sorted(refs):
            row = all_items[item_id]
            comp = "／".join(companies.get(item_id, [])) or "—"
            print(f"      id={item_id} {row[1]} {row[2]}　[{comp}]")
            for r in refs[item_id]:
                print(f"          - {r}")
    else:
        print("    沒有任何要刪的料號被單據引用，可以乾淨刪除。")

    deletable = [i for i in drop_ids if i not in refs]
    print(f"\n--- 4. 這次實際會刪除：{len(deletable)} 筆 ---")
    print(f"    刪除後料號主檔會剩下 {len(all_items) - len(deletable)} 筆"
          f"（保留 {len(keep_ids)} 筆 + 被單據卡住的 {len(refs)} 筆）")

    print("\n" + sep)
    if blockers:
        print("❌ 有以下阻擋項，execute 會中止：")
        for b in blockers:
            print(f"   - {b}")
    else:
        print("✅ 沒有阻擋項。確認以上都沒問題後，加 --execute 才會真的刪除。")
    print(sep)
    return not blockers


def backup_db(db_path: Path) -> Path:
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = db_path.with_name(f"{db_path.name}.bak-{ts}")
    shutil.copy2(db_path, backup)
    for suffix in ("-wal", "-shm"):
        side = db_path.with_name(db_path.name + suffix)
        if side.exists():
            shutil.copy2(side, backup.with_name(backup.name + suffix))
    print(f"已備份到：{backup}")
    return backup


def execute(conn: sqlite3.Connection, excel_codes: set[str]) -> None:
    all_items, companies, keep_ids, code_to_ids, drop_ids, refs = collect(conn, excel_codes)

    missing = sorted(excel_codes - set(code_to_ids))
    if missing:
        raise RuntimeError(
            f"有 {len(missing)} 個 Excel 料號在資料庫裡找不到春大直對照"
            f"（如 {missing[:5]}），保留集合不完整，中止執行。"
            f"請先執行 align_chunda_items_20260818.py --execute。"
        )
    dup = {c: ids for c, ids in code_to_ids.items() if len(ids) > 1}
    if dup:
        raise RuntimeError(
            f"有 {len(dup)} 個原始碼對到多筆料號（如 {list(dup)[:5]}），"
            f"無法判斷該留哪一筆，中止執行。請先人工處理重複料號。"
        )
    if len(keep_ids) != len(excel_codes):
        raise RuntimeError(
            f"保留清單 {len(keep_ids)} 筆 ≠ Excel 料號 {len(excel_codes)} 筆，"
            f"判定有問題，中止執行（寧可不動，也不要刪錯）。"
        )

    deletable = [i for i in drop_ids if i not in refs]
    print(f"將刪除 {len(deletable)} 筆料號；"
          f"另有 {len(refs)} 筆被單據引用，維持原狀不動。")

    conn.execute("BEGIN")
    try:
        deleted = 0
        for item_id in deletable:
            # 對照表是 ON DELETE CASCADE，會跟著消失
            conn.execute("DELETE FROM cycle_purchase_items WHERE id = ?", (item_id,))
            deleted += 1
        conn.execute("COMMIT")
    except Exception:
        conn.execute("ROLLBACK")
        raise

    left = conn.execute("SELECT COUNT(*) FROM cycle_purchase_items").fetchone()[0]
    print(f"\n執行完成：刪除 {deleted} 筆，料號主檔現在剩 {left} 筆。")
    if refs:
        print(f"⚠ 有 {len(refs)} 筆被單據引用而保留，清單見盤點報告。")
    print("請重啟後端服務。")


def main() -> int:
    ap = argparse.ArgumentParser(description="只保留春大直料號明細表裡的料號（2026-08-18）")
    ap.add_argument("--excel-path", required=True)
    ap.add_argument("--db-path", default=DEFAULT_DB)
    ap.add_argument("--execute", action="store_true", help="真的刪除（預設只盤點）")
    args = ap.parse_args()

    excel_path, db_path = Path(args.excel_path), Path(args.db_path)
    if not excel_path.exists():
        print(f"找不到 Excel：{excel_path}", file=sys.stderr)
        return 1
    if not db_path.exists():
        print(f"找不到資料庫：{db_path}", file=sys.stderr)
        return 1

    excel_codes = read_excel_codes(excel_path)
    if args.execute:
        backup_db(db_path)

    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA foreign_keys=ON")
    try:
        if args.execute:
            execute(conn, excel_codes)
        else:
            report(conn, excel_codes)
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
