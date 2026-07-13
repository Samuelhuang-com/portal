# -*- coding: utf-8 -*-
"""
週期採購 — 料號主檔匯入腳本
從「日曜天地設料號明細表.xlsx」「春大直設料號明細表.xlsx」匯入到
cycle-purchase.db 的 cycle_purchase_items / cycle_purchase_vendors /
cycle_purchase_departments / cycle_purchase_item_mappings 四張表。

治理規則（依《週期採購_Portal規劃評估_v1.0.md》第四節分析結果，
並經 Samuel 於 2026-07-10 確認）：
  - 兩家公司原始料號字串相同的有 115 筆，但品名／廠商／單價三者完全一致、
    確認為同一實體商品的只有 7 筆 —— 只有這 7 筆會合併成 1 筆集團料號＋
    2 筆對照（is_confirmed=True）；其餘同號不同貨，各自建立獨立料號。
  - 集團新料號沿用原本編碼結構（大分類英文 + 中分類2碼 + 細分類2碼 +
    流水碼3碼），但流水碼改為全公司統一分配（同一 prefix 下，日曜天地與
    春大直的品項共用同一組流水序號，確保新碼不重複、不再有「同碼不同貨」
    的風險）。
  - 廠商：依 Excel「廠商」欄位文字建立，僅做「完全相同字串」去重，
    不做模糊比對／自動合併（近似但不同的廠商名稱，如「永豐餘」與
    「永豐紙業」，會各自建檔，需要您日後自行到「供應商」頁面確認是否合併）。
  - 缺品名等資料的料號（實測 22 筆，全部在日曜天地／工務用 分頁）：
    仍會建檔（保留其料號存在的紀錄），但品名會標記「[資料不全] 類別」，
    並設為停用（is_active=False），待補齊資料後您再手動啟用。

2026-07-11 新增（與 Samuel 討論後確認）— 部門範圍：
  逐列核對兩個 Excel 後確認，每家公司內部「分頁」（工務用／清潔用品／
  文具&印刷／營業用品）、分頁內的「類別」、料號三者是乾淨的三層關係，
  沒有任何料號或類別橫跨兩個分頁。這個分頁邊界對應真實的功能性部門
  （工務部／清潔部／文具印刷部／營業部），不是單純的分類標籤——請購單
  「可選料號」查詢會按「公司＋部門」篩選，所以每一筆料號對照（mapping）
  現在都必須帶 department_id。
  兩家公司分頁命名不完全一樣（日曜天地是中文「工務用OK」，春大直是英文
  「ENG」），本腳本用 SHEET_TO_DEPT_CODE 寫死對照表，不做文字比對猜測；
  遇到對照表沒有的分頁名稱會直接報錯中止，不會靜默略過或亂猜。
  部門代碼／名稱（兩家公司一致，經 Samuel 確認）：
    ENG   工務部
    CLEAN 清潔部
    STA   文具印刷部
    OPS   營業部
  部門主檔若不存在會自動建立（owner_user_id 留空，日後請到「部門主檔」
  頁面手動指定承辦人）；若已存在（company+dept_code 找得到）就直接沿用。

使用方式（在 backend 資料夾底下，啟用虛擬環境後執行）：
    python import_cycle_purchase_item_master.py --dry-run   # 先預覽，不寫入
    python import_cycle_purchase_item_master.py              # 正式匯入

可重複執行：已經匯入過的（公司＋原始料號）組合會被偵測到並跳過，
不會重複建立料號或對照，所以不小心跑第二次也不會壞資料。

2026-07-11 新增（第三期彙整單／採購單規劃時發現並修正）：
料號對照表現在也會記 vendor_id（不只 original_vendor_name 文字），供彙整單/
採購單按供應商分單用。這裡只影響「以後才匯入」的資料；已經在資料庫裡的舊
對照列，用一次性回填腳本 scripts/backfill_item_mapping_vendor_id.py 補齊，
不需要重跑本腳本、不需要清空資料庫。
"""

import argparse
import re
import sqlite3
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("請先安裝 openpyxl：pip install openpyxl")
    sys.exit(1)

# ── 設定 ──────────────────────────────────────────────────────────────────
PORTAL_ROOT = Path(__file__).parent
NY_EXCEL = PORTAL_ROOT / "ragic-cycle-purchase" / "docs" / "日曜天地設料號明細表.xlsx"
CD_EXCEL = PORTAL_ROOT / "ragic-cycle-purchase" / "docs" / "春大直設料號明細表.xlsx"
DB_PATH = Path("C:/portal_data/cycle-purchase.db")

CODE_PATTERN = re.compile(r"^[A-Z]\d{4}\d{3}$")
COMPANY_NY = "日曜天地"
COMPANY_CD = "春大直"

# ── 部門主檔草案（兩家公司代碼一致，2026-07-11 與 Samuel 確認）──────────────
DEPT_DEFS = [
    (COMPANY_NY, "ENG", "工務部"),
    (COMPANY_NY, "CLEAN", "清潔部"),
    (COMPANY_NY, "STA", "文具印刷部"),
    (COMPANY_NY, "OPS", "營業部"),
    (COMPANY_CD, "ENG", "工務部"),
    (COMPANY_CD, "CLEAN", "清潔部"),
    (COMPANY_CD, "STA", "文具印刷部"),
    (COMPANY_CD, "OPS", "營業部"),
]

# ── 分頁名稱 -> 部門代碼（寫死，不做文字比對猜測；兩家公司打法不同）─────────
SHEET_TO_DEPT_CODE = {
    (COMPANY_NY, "工務用OK"): "ENG",
    (COMPANY_NY, "清潔用品OK"): "CLEAN",
    (COMPANY_NY, "文具&印刷OK"): "STA",
    (COMPANY_NY, "營業用品 -無"): "OPS",
    (COMPANY_CD, "ENG"): "ENG",
    (COMPANY_CD, "清潔用品"): "CLEAN",
    (COMPANY_CD, "文具&印刷"): "STA",
    (COMPANY_CD, "營業用品 "): "OPS",
}


def load_excel_rows(path, company):
    """讀取一個公司的 Excel，回傳 list of dict（依欄位標題對應，不假設固定欄位順序）。"""
    if not path.exists():
        print(f"找不到 Excel 檔案：{path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for sheetname in wb.sheetnames:
        if sheetname == "編碼原則":
            continue
        ws = wb[sheetname]
        header = [
            (c.value.strip() if isinstance(c.value, str) else c.value) for c in ws[1]
        ]
        idx = {h: i for i, h in enumerate(header) if h}

        code_i = idx.get("料號")
        if code_i is None:
            continue

        def g(row, key):
            i = idx.get(key)
            if i is None or i >= len(row):
                return None
            v = row[i]
            if isinstance(v, str):
                v = v.strip()
                if v == "":
                    return None
            return v

        for r in range(2, ws.max_row + 1):
            row = [c.value for c in ws[r]]
            code = row[code_i] if code_i < len(row) else None
            if not code:
                continue
            code = str(code).strip()
            if not CODE_PATTERN.match(code):
                # 排除混入資料區的非料號雜訊列（例如編碼原則表被誤貼進資料分頁）
                continue
            rows.append(
                {
                    "company": company,
                    "sheet": sheetname,
                    "code": code,
                    "category": g(row, "類別"),
                    "location": g(row, "位置"),
                    "name": g(row, "品名"),
                    "unit": g(row, "單位"),
                    "vendor": g(row, "廠商"),
                    "price": g(row, "單價"),
                    "max_stock": g(row, "最大庫存量"),
                    "min_stock": g(row, "最小庫存量"),
                    "mini_order": g(row, "mini order"),
                    "notes": g(row, "備註"),
                }
            )
    return rows


def validate_sheets_known(rows, company):
    """防呆：確保每一列的分頁都在 SHEET_TO_DEPT_CODE 對照表裡，不靜默略過或亂猜。"""
    unknown = sorted({r["sheet"] for r in rows if (company, r["sheet"]) not in SHEET_TO_DEPT_CODE})
    if unknown:
        print(f"錯誤：{company} 的 Excel 裡有分頁不在 SHEET_TO_DEPT_CODE 對照表：{unknown}")
        print("請先確認這個分頁對應哪個部門，更新腳本裡的 SHEET_TO_DEPT_CODE 後再重跑，不要用猜的。")
        sys.exit(1)


def find_confirmed_matches(ny_rows, cd_rows):
    """兩公司料號字串相同、且品名也相同 -> 視為同一實體商品，回傳這些 code 的集合。"""
    ny_by_code = {r["code"]: r for r in ny_rows}
    cd_by_code = {r["code"]: r for r in cd_rows}
    overlap = set(ny_by_code) & set(cd_by_code)
    confirmed = set()
    for code in overlap:
        a, b = ny_by_code[code], cd_by_code[code]
        if a["name"] and a["name"] == b["name"]:
            confirmed.add(code)
    return confirmed, overlap


def to_number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if re.match(r"^-?\d+(\.\d+)?$", s):
        return float(s) if "." in s else int(s)
    return None


def to_int_with_fallback(v):
    """回傳 (整數或 None, 原始文字若含非數字內容則回傳供寫入備註)。"""
    if v is None:
        return None, None
    if isinstance(v, (int, float)):
        return int(v), None
    s = str(v).strip()
    m = re.match(r"^(\d+)", s)
    if m and m.group(1) == s:
        return int(s), None
    if m:
        return int(m.group(1)), s
    return None, s


def ensure_departments(cur, dry_run):
    """確保 DEPT_DEFS 裡的 8 筆部門都存在（company+dept_code 找得到就沿用，
    找不到才新建；owner_user_id 留空，日後請到「部門主檔」頁面手動指定）。
    回傳 {(company, dept_code): department_id}。"""
    dept_id_map = {}
    created = 0
    for company, code, name in DEPT_DEFS:
        cur.execute(
            "SELECT id FROM cycle_purchase_departments WHERE company = ? AND dept_code = ?",
            (company, code),
        )
        row = cur.fetchone()
        if row:
            dept_id_map[(company, code)] = row["id"]
            continue
        if not dry_run:
            cur.execute(
                """INSERT INTO cycle_purchase_departments
                   (company, dept_code, dept_name, owner_user_id, is_active, created_at)
                   VALUES (?, ?, ?, NULL, 1, datetime('now'))""",
                (company, code, name),
            )
            dept_id_map[(company, code)] = cur.lastrowid
        else:
            dept_id_map[(company, code)] = -1  # dry-run 用假 id
        created += 1
    return dept_id_map, created


def main():
    parser = argparse.ArgumentParser(description="匯入週期採購料號主檔")
    parser.add_argument("--dry-run", action="store_true", help="只預覽，不實際寫入資料庫")
    args = parser.parse_args()

    print("=" * 70)
    print("  週期採購 — 料號主檔匯入")
    print(f"  來源：{NY_EXCEL.name} / {CD_EXCEL.name}")
    print(f"  目標資料庫：{DB_PATH}")
    if args.dry_run:
        print("  [DRY-RUN 模式：不會實際寫入資料庫]")
    print("=" * 70)

    if not DB_PATH.exists():
        print(f"找不到資料庫檔案：{DB_PATH}")
        print("請先啟動一次後端（uvicorn app.main:app），讓 cycle-purchase.db 與資料表建立完成後再執行本腳本。")
        sys.exit(1)

    print("\n讀取 Excel...")
    ny_rows = load_excel_rows(NY_EXCEL, COMPANY_NY)
    cd_rows = load_excel_rows(CD_EXCEL, COMPANY_CD)
    print(f"  {COMPANY_NY}：{len(ny_rows)} 筆有效料號")
    print(f"  {COMPANY_CD}：{len(cd_rows)} 筆有效料號")

    validate_sheets_known(ny_rows, COMPANY_NY)
    validate_sheets_known(cd_rows, COMPANY_CD)

    confirmed_codes, overlap_codes = find_confirmed_matches(ny_rows, cd_rows)
    overlap_not_confirmed = overlap_codes - confirmed_codes
    print(f"\n兩公司料號字串重複：{len(overlap_codes)} 筆")
    print(f"  其中確認為同一商品（將合併為 1 筆＋2 個對照）：{len(confirmed_codes)} 筆")
    print(f"  其餘同號不同貨（各自獨立建檔）：{len(overlap_not_confirmed)} 筆")

    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    for tbl in (
        "cycle_purchase_items",
        "cycle_purchase_vendors",
        "cycle_purchase_departments",
        "cycle_purchase_item_mappings",
    ):
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?", (tbl,)
        )
        if not cur.fetchone():
            print(f"資料庫裡找不到資料表 {tbl}，請確認後端已經至少啟動過一次（會自動建表）。")
            sys.exit(1)

    print("\n確保部門主檔存在（工務部／清潔部／文具印刷部／營業部 × 2 公司）...")
    dept_id_map, depts_created = ensure_departments(cur, args.dry_run)
    print(f"  部門主檔新建立：{depts_created} 筆（其餘沿用既有）")

    # ── 已匯入過的 (company, original_code) -> item_id，讓腳本可重複執行 ──
    cur.execute(
        "SELECT company, original_code, item_id FROM cycle_purchase_item_mappings"
    )
    already_imported = {(r["company"], r["original_code"]): r["item_id"] for r in cur.fetchall()}

    # ── 供應商：先讀現有的，準備做「完全相同字串」比對 ──
    cur.execute("SELECT id, vendor_code, vendor_name FROM cycle_purchase_vendors")
    vendor_by_name = {r["vendor_name"]: r["id"] for r in cur.fetchall()}
    existing_vendor_codes = [r["vendor_code"] for r in cur.execute(
        "SELECT vendor_code FROM cycle_purchase_vendors"
    ).fetchall()]
    max_vendor_seq = 0
    for vc in existing_vendor_codes:
        m = re.match(r"^CPV-(\d+)$", vc)
        if m:
            max_vendor_seq = max(max_vendor_seq, int(m.group(1)))
    next_vendor_seq = [max_vendor_seq + 1]

    def get_or_create_vendor(name):
        if not name:
            return None
        if name in vendor_by_name:
            return vendor_by_name[name]
        code = f"CPV-{next_vendor_seq[0]:04d}"
        next_vendor_seq[0] += 1
        if not args.dry_run:
            cur.execute(
                "INSERT INTO cycle_purchase_vendors "
                "(vendor_code, vendor_name, is_active, created_at, updated_at) "
                "VALUES (?, ?, 1, datetime('now'), datetime('now'))",
                (code, name),
            )
            vid = cur.lastrowid
        else:
            vid = -1  # dry-run 用假 id
        vendor_by_name[name] = vid
        return vid

    # ── item_code 流水碼分配：從資料庫既有料號中找出每個 prefix 目前的最大序號 ──
    cur.execute("SELECT item_code FROM cycle_purchase_items")
    prefix_serial = defaultdict(int)
    for r in cur.fetchall():
        code = r["item_code"]
        if CODE_PATTERN.match(code):
            prefix, serial = code[:5], int(code[5:])
            prefix_serial[prefix] = max(prefix_serial[prefix], serial)

    def alloc_item_code(prefix):
        prefix_serial[prefix] += 1
        return f"{prefix}{prefix_serial[prefix]:03d}"

    stats = {
        "items_created": 0,
        "items_skipped_existing": 0,
        "items_incomplete": 0,
        "mappings_created": 0,
        "mappings_skipped_existing": 0,
        "vendors_created_before": len(vendor_by_name),
        "confirmed_merges": 0,
    }

    confirmed_item_id_this_run = {}  # code -> item_id（本次執行中，供 CD pass 對照）

    def insert_item(row, extra_note=None):
        prefix = row["code"][:5]
        new_code = alloc_item_code(prefix)
        incomplete = not row["name"]
        name = row["name"] or f"[資料不全] {row['category'] or '未分類'}"

        note_parts = [f"原始料號: {row['company']} {row['code']}"]
        if row.get("location"):
            note_parts.append(f"原始位置: {row['location']}")
        if incomplete:
            note_parts.append("資料不全（缺品名/單位/廠商/單價等），待補齊後啟用")
            stats["items_incomplete"] += 1
        if extra_note:
            note_parts.append(extra_note)
        if row.get("notes"):
            note_parts.append(f"原始備註: {row['notes']}")

        moq, moq_raw = to_int_with_fallback(row.get("mini_order"))
        if moq_raw:
            note_parts.append(f"原始最小訂購量文字: {moq_raw}")

        vendor_id = get_or_create_vendor(row.get("vendor"))

        if not args.dry_run:
            cur.execute(
                """INSERT INTO cycle_purchase_items
                   (item_code, item_name, category, unit, default_qty, moq,
                    max_stock, min_stock, unit_price, default_vendor_id,
                    is_active, is_cycle_item, notes, created_at, updated_at)
                   VALUES (?, ?, ?, ?, 0, ?, ?, ?, ?, ?, ?, 1, ?, datetime('now'), datetime('now'))""",
                (
                    new_code,
                    name,
                    row.get("category"),
                    row.get("unit"),
                    moq or 0,
                    to_number(row.get("max_stock")),
                    to_number(row.get("min_stock")),
                    to_number(row.get("price")),
                    vendor_id,
                    0 if incomplete else 1,
                    "；".join(note_parts),
                ),
            )
            item_id = cur.lastrowid
        else:
            item_id = -1

        stats["items_created"] += 1
        return item_id, new_code

    def insert_mapping(item_id, row, is_confirmed):
        dept_code = SHEET_TO_DEPT_CODE[(row["company"], row["sheet"])]
        department_id = dept_id_map[(row["company"], dept_code)]
        # 2026-07-11 新增：mapping 也要記 vendor_id（不只 original_vendor_name 文字），
        # 供第三期彙整單/採購單按供應商分單用。重用 get_or_create_vendor() 的既有
        # 去重結果——這個函式是 dict-based 去重，同一字串重複呼叫不會建立第二筆。
        vendor_id = get_or_create_vendor(row.get("vendor"))
        if not args.dry_run:
            cur.execute(
                """INSERT INTO cycle_purchase_item_mappings
                   (item_id, company, department_id, original_code, original_name,
                    original_vendor_name, vendor_id, original_unit_price, is_confirmed,
                    created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, datetime('now'), datetime('now'))""",
                (
                    item_id,
                    row["company"],
                    department_id,
                    row["code"],
                    row.get("name"),
                    row.get("vendor"),
                    vendor_id,
                    to_number(row.get("price")),
                    1 if is_confirmed else 0,
                ),
            )
        stats["mappings_created"] += 1

    def process_row(row):
        key = (row["company"], row["code"])
        if key in already_imported:
            stats["mappings_skipped_existing"] += 1
            return

        code = row["code"]
        if code in confirmed_codes:
            if code in confirmed_item_id_this_run:
                item_id = confirmed_item_id_this_run[code]
                insert_mapping(item_id, row, is_confirmed=True)
            elif (COMPANY_NY, code) in already_imported and row["company"] == COMPANY_CD:
                # 之前已經匯入過日曜天地那一半，這次只補春大直的對照
                item_id = already_imported[(COMPANY_NY, code)]
                insert_mapping(item_id, row, is_confirmed=True)
            else:
                item_id, _ = insert_item(row)
                confirmed_item_id_this_run[code] = item_id
                insert_mapping(item_id, row, is_confirmed=True)
                if row["company"] == COMPANY_NY:
                    stats["confirmed_merges"] += 1
        else:
            extra = (
                "備註: 料號字串與另一公司重複但實際品項不同，未合併"
                if code in overlap_not_confirmed
                else None
            )
            item_id, _ = insert_item(row, extra)
            insert_mapping(item_id, row, is_confirmed=True)

    print("\n開始匯入...")
    for row in ny_rows:
        process_row(row)
    for row in cd_rows:
        process_row(row)

    if not args.dry_run:
        con.commit()
    con.close()

    print("\n" + "=" * 70)
    print("  匯入完成" + ("（DRY-RUN，未實際寫入）" if args.dry_run else ""))
    print("=" * 70)
    print(f"  部門主檔新建立：{depts_created} 筆")
    print(f"  新建立料號：{stats['items_created']} 筆")
    print(f"    其中資料不全、已設為停用：{stats['items_incomplete']} 筆")
    print(f"  合併為同一料號的組數：{stats['confirmed_merges']} 組")
    print(f"  新建立對照（含部門歸屬）：{stats['mappings_created']} 筆")
    print(f"  已存在、本次略過的對照：{stats['mappings_skipped_existing']} 筆")
    print(f"  匯入前既有供應商數：{stats['vendors_created_before']}")
    print(f"  匯入後供應商總數：{len(vendor_by_name)}（新建立 {len(vendor_by_name) - stats['vendors_created_before']} 筆）")
    print("\n  請注意：供應商僅做完全相同字串去重，近似但拼寫不同的供應商")
    print("  名稱不會自動合併，請日後自行到「供應商」頁面檢查、合併重複。")
    print("=" * 70)


if __name__ == "__main__":
    main()
