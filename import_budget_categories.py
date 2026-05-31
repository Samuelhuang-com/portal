"""
預算科目批次匯入腳本
從 Excel 匯入預算科目到 portal.db 的 budget_categories 資料表

使用方式：
    python import_budget_categories.py
    python import_budget_categories.py --overwrite   # 遇到重複時覆寫（更新 accounting_code）
"""

import argparse
import sqlite3
import sys
from datetime import date, datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("❌ 請先安裝 openpyxl：pip install openpyxl")
    sys.exit(1)

# ── 設定 ──────────────────────────────────────────────────────────────────────
EXCEL_PATH = Path(__file__).parent / "台灣_預算科目_匯入模板.xlsx"
SHEET_NAME = "預算科目_匯入用"
DB_PATH    = Path("C:/portal_data/portal.db")

BUDGET_YEAR    = 2026
DEPT           = "管理部"
EFFECTIVE_DATE = date(2026, 1, 1)   # "2026-01-01"

# Excel 欄位對應（A=大項, B=細項, C=會計科目）
COL_L1   = 1   # 大項
COL_L2   = 2   # 細項
COL_CODE = 3   # 會計科目

HEADER_ROW = 1  # 第 1 列為標題，從第 2 列開始讀資料


def parse_args():
    p = argparse.ArgumentParser(description="匯入預算科目")
    p.add_argument("--overwrite", action="store_true",
                   help="遇到重複 (budget_year, category_l1, category_l2, dept) 時，更新 accounting_code")
    p.add_argument("--dry-run", action="store_true",
                   help="只預覽，不實際寫入資料庫")
    return p.parse_args()


def load_excel(path: Path) -> list[dict]:
    """讀取 Excel，回傳 list of dict，已去除首尾空白，跳過全空列。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"❌ 找不到工作表 '{SHEET_NAME}'，現有工作表：{wb.sheetnames}")
        sys.exit(1)
    ws = wb[SHEET_NAME]

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1):
        l1   = str(row[COL_L1 - 1] or "").strip()
        l2   = str(row[COL_L2 - 1] or "").strip()
        code = str(row[COL_CODE - 1] or "").strip()

        if not l1 and not l2:   # 全空列跳過
            continue
        if not l2:
            print(f"  ⚠️  第 {i} 列：大項='{l1}' 但細項為空，略過")
            continue

        rows.append({"l1": l1, "l2": l2, "code": code, "excel_row": i})

    wb.close()
    return rows


def deduplicate(rows: list[dict]) -> tuple[list[dict], int]:
    """同一個 Excel 檔案內 (l1, l2) 重複時只取第一筆，回傳 (去重後列表, 跳過數)。"""
    seen: set[tuple] = set()
    unique, skipped = [], 0
    for r in rows:
        key = (r["l1"], r["l2"])
        if key in seen:
            print(f"  ℹ️  Excel 內重複 (第 {r['excel_row']} 列)：{r['l1']} / {r['l2']}，已略過")
            skipped += 1
        else:
            seen.add(key)
            unique.append(r)
    return unique, skipped


def import_to_db(rows: list[dict], overwrite: bool, dry_run: bool) -> dict:
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    cur = con.cursor()

    now_str = datetime.now().isoformat(timespec="seconds")
    eff_str = EFFECTIVE_DATE.isoformat()   # "2026-01-01"

    inserted = 0
    updated  = 0
    skipped  = 0

    for r in rows:
        # 確認是否已存在
        cur.execute(
            "SELECT id, accounting_code FROM budget_categories "
            "WHERE budget_year=? AND category_l1=? AND category_l2=? AND dept=?",
            (BUDGET_YEAR, r["l1"], r["l2"], DEPT),
        )
        existing = cur.fetchone()

        if existing:
            if overwrite:
                if not dry_run:
                    cur.execute(
                        "UPDATE budget_categories SET accounting_code=?, updated_at=? "
                        "WHERE id=?",
                        (r["code"], now_str, existing["id"]),
                    )
                updated += 1
            else:
                skipped += 1
        else:
            if not dry_run:
                cur.execute(
                    """INSERT INTO budget_categories
                       (budget_year, dept, category_l1, category_l2,
                        accounting_code, payment_code, is_enabled,
                        effective_date, disabled_date, maintain_unit,
                        created_at, updated_at)
                       VALUES (?,?,?,?,?,NULL,1,?,NULL,'',?,?)""",
                    (BUDGET_YEAR, DEPT, r["l1"], r["l2"],
                     r["code"], eff_str, now_str, now_str),
                )
            inserted += 1

    if not dry_run:
        con.commit()
    con.close()

    return {"inserted": inserted, "updated": updated, "skipped": skipped}


def main():
    args = parse_args()

    print("=" * 60)
    print("  預算科目匯入腳本")
    print(f"  來源：{EXCEL_PATH.name}  →  {DB_PATH}")
    print(f"  預算年度：{BUDGET_YEAR}  部門：{DEPT}  生效日：{EFFECTIVE_DATE}")
    if args.dry_run:
        print("  ⚠️  DRY-RUN 模式：不會實際寫入資料庫")
    if args.overwrite:
        print("  ⚠️  OVERWRITE 模式：重複時會更新 accounting_code")
    print("=" * 60)

    if not EXCEL_PATH.exists():
        print(f"❌ 找不到 Excel 檔案：{EXCEL_PATH}")
        sys.exit(1)

    if not DB_PATH.exists():
        print(f"❌ 找不到資料庫：{DB_PATH}")
        sys.exit(1)

    # 讀取 Excel
    print("\n📂 讀取 Excel...")
    rows = load_excel(EXCEL_PATH)
    print(f"   讀取 {len(rows)} 筆有效資料列")

    # Excel 內部去重
    print("\n🔍 檢查 Excel 內部重複...")
    rows, excel_dupes = deduplicate(rows)
    print(f"   Excel 內部重複略過：{excel_dupes} 筆，剩餘：{len(rows)} 筆")

    # 匯入
    print("\n💾 開始匯入資料庫...")
    result = import_to_db(rows, overwrite=args.overwrite, dry_run=args.dry_run)

    # 結果摘要
    print("\n" + "=" * 60)
    print("  匯入完成！")
    print(f"  ✅ 新增：{result['inserted']} 筆")
    if args.overwrite:
        print(f"  🔄 更新：{result['updated']} 筆")
    else:
        print(f"  ⏭️  略過（已存在）：{result['skipped']} 筆")
    if args.dry_run:
        print("  （DRY-RUN，以上均未實際寫入）")
    print("=" * 60)


if __name__ == "__main__":
    main()
